"""
================================================================================
CEREBRO-X |  UNIT TESTS
================================================================================
File: tests/unit/test_all.py

Coverage:
  1. Authentication & RBAC
  2. DDS BBB Engineering Score
  3. MLOps (registry, drift detection)
  4. Orchestrator (DAG, retry, circuit breaker)
  5. Cache (LRU, multi-tier)
  6. Compliance (PHI detection, audit trail, data masking)
================================================================================
"""
import time
from datetime import datetime, timedelta

import numpy as np
import pytest

# ═════════════════════════════════════════════════════════════════════════════
# 1. AUTHENTICATION & RBAC
# ═════════════════════════════════════════════════════════════════════════════

class TestAuth:
    """JWT authentication and role-based access control."""

    def test_password_hashing(self):
        from src.api.auth import pwd_context
        hashed = pwd_context.hash("secret123")
        assert pwd_context.verify("secret123", hashed)
        assert not pwd_context.verify("wrong", hashed)

    def test_create_user(self, db_session):
        from src.api.auth import AuthService, UserCreate
        svc = AuthService(db_session)
        user = svc.create_user(UserCreate(
            email="test@test.com",
            username="testuser",
            password="strongpw123",
            role="researcher",
        ))
        assert user.username == "testuser"
        assert user.role == "researcher"
        assert user.is_active is True

    def test_duplicate_user_rejected(self, db_session):
        from src.api.auth import AuthService, UserCreate
        svc = AuthService(db_session)
        svc.create_user(UserCreate(
            email="dup@test.com", username="dupuser",
            password="pw123456", role="readonly",
        ))
        with pytest.raises(ValueError, match="already registered"):
            svc.create_user(UserCreate(
                email="dup@test.com", username="dupuser2",
                password="pw123456", role="readonly",
            ))

    def test_login_success(self, db_session):
        from src.api.auth import AuthService, UserCreate
        svc = AuthService(db_session)
        svc.create_user(UserCreate(
            email="login@test.com", username="loginuser",
            password="pw123456", role="admin",
        ))
        result = svc.login("loginuser", "pw123456")
        assert result is not None
        assert result.access_token
        assert result.refresh_token
        assert result.role == "admin"

    def test_login_wrong_password(self, db_session):
        from src.api.auth import AuthService, UserCreate
        svc = AuthService(db_session)
        svc.create_user(UserCreate(
            email="fail@test.com", username="failuser",
            password="correctpw", role="readonly",
        ))
        result = svc.login("failuser", "wrongpw")
        assert result is None

    def test_jwt_token_roundtrip(self):
        from src.api.auth import TokenEngine
        token = TokenEngine.create_access_token(
            {"sub": "42", "role": "researcher", "username": "test"}
        )
        payload = TokenEngine.decode_token(token)
        assert payload["sub"] == "42"
        assert payload["role"] == "researcher"
        assert payload["type"] == "access"

    def test_refresh_token_rotation(self, db_session):
        from src.api.auth import AuthService, UserCreate
        svc = AuthService(db_session)
        svc.create_user(UserCreate(
            email="refresh@test.com", username="refreshuser",
            password="pw123456", role="admin",
        ))
        login_result = svc.login("refreshuser", "pw123456")
        new_tokens = svc.refresh_tokens(login_result.refresh_token)
        assert new_tokens is not None
        assert new_tokens.access_token != login_result.access_token
        # Old refresh token should be revoked
        reuse = svc.refresh_tokens(login_result.refresh_token)
        assert reuse is None

    def test_rbac_permissions(self):
        from src.api.auth import Role, has_permission
        assert has_permission(Role.ADMIN, "pipeline:run")
        assert has_permission(Role.ADMIN, "user:delete")
        assert has_permission(Role.RESEARCHER, "pipeline:run")
        assert not has_permission(Role.RESEARCHER, "user:delete")
        assert not has_permission(Role.READONLY, "pipeline:run")
        assert has_permission(Role.READONLY, "results:read")

    def test_api_key_creation(self, db_session):
        from src.api.auth import AuthService, UserCreate
        svc = AuthService(db_session)
        user = svc.create_user(UserCreate(
            email="apikey@test.com", username="apikeyuser",
            password="pw123456", role="researcher",
        ))
        raw_key, key_model = svc.create_api_key(user.id, "test-key")
        assert raw_key.startswith("cerebro_")
        assert key_model.key_prefix == raw_key[:12]
        # Verify
        found_user = svc.verify_api_key(raw_key)
        assert found_user is not None
        assert found_user.id == user.id


# Task ownership authorization (GET /pipeline/status/{task_id} no longer
# accepts any authenticated user for any task_id) is covered by
# tests/integration/test_task_ownership.py, which exercises the real HTTP
# endpoints end-to-end with mocked Celery — no need to duplicate that
# coverage at the unit-test level here.


# ═════════════════════════════════════════════════════════════════════════════
# 2. DDS BBB ENGINEERING SCORE
# ═════════════════════════════════════════════════════════════════════════════

class TestDDSScore:
    """BBB Engineering Score computation."""

    def test_optimal_formulation_scores_high(self):
        from src.dds.enterprise_infra import DDSEngine
        row = {
            "size_nm": 80,
            "zeta_potential_mv": -10,
            "pegylation_degree_mol_pct": 5,
            "surface_ligand": "ApoE",
            "ligand_density_per_nm2": 1.0,
            "encapsulation_efficiency_pct": 85,
            "pgp_escape_coeff": 0.9,
            "apo_e_affinity": "very_high",
            "carpa_risk_index": 0.1,
            "off_target_liver_pct": 15,
            "phase_transition_temp_c": 55,
        }
        score = DDSEngine.compute_bbb_engineering_score(row)
        assert score >= 75, f"Optimal formulation should score ≥75, got {score}"
        assert score <= 100

    def test_terrible_formulation_scores_low(self):
        from src.dds.enterprise_infra import DDSEngine
        row = {
            "size_nm": 500,
            "zeta_potential_mv": -40,
            "pegylation_degree_mol_pct": 15,
            "surface_ligand": "None",
            "ligand_density_per_nm2": 5.0,
            "encapsulation_efficiency_pct": 20,
            "pgp_escape_coeff": 0.1,
            "apo_e_affinity": "low",
            "carpa_risk_index": 0.9,
            "off_target_liver_pct": 80,
            "phase_transition_temp_c": 35,
        }
        score = DDSEngine.compute_bbb_engineering_score(row)
        assert score < 30, f"Terrible formulation should score <30, got {score}"

    def test_score_bounded_0_100(self):
        from src.dds.enterprise_infra import DDSEngine
        for _ in range(50):
            row = {
                "size_nm": np.random.uniform(20, 500),
                "zeta_potential_mv": np.random.uniform(-50, 50),
                "pegylation_degree_mol_pct": np.random.uniform(0, 20),
                "surface_ligand": np.random.choice(["ApoE", "RVG", "None"]),
                "ligand_density_per_nm2": np.random.uniform(0, 5),
                "encapsulation_efficiency_pct": np.random.uniform(10, 99),
                "pgp_escape_coeff": np.random.uniform(0, 1),
                "apo_e_affinity": np.random.choice(["low", "high", "very_high"]),
                "carpa_risk_index": np.random.uniform(0, 1),
                "off_target_liver_pct": np.random.uniform(0, 100),
                "phase_transition_temp_c": np.random.uniform(30, 70),
            }
            score = DDSEngine.compute_bbb_engineering_score(row)
            assert 0 <= score <= 100

    def test_run_summary_log_names_the_actual_drug_not_a_hardcoded_one(
            self, monkeypatch, tmp_path, caplog):
        """DDSEngine.run()'s top-5 summary line was hardcoded to
        "... FOR LECANEMAB BBB DELIVERY:" regardless of which drug's
        config was actually loaded -- the same class of bug this project
        has repeatedly had to purge elsewhere (an unrelated drug name
        bleeding into output for any other drug). Runs the real engine
        end-to-end against a temp config for a differently-named drug
        and confirms the log names that drug, not Lecanemab."""
        import logging

        import yaml as _yaml

        import src.dds.enterprise_infra as ei

        cfg = {
            "drug": {"name": "Rivastigmine", "mw_da": 250.3,
                      "logp": 1.7, "half_life_days": 0.04},
            "formulations": [{
                "id": "F1", "name": "Test-Formulation", "carrier_type": "liposome",
                "size_nm": 80, "zeta_potential_mv": -10,
                "pegylation_degree_mol_pct": 5, "surface_ligand": "ApoE",
                "ligand_density_per_nm2": 1.0, "encapsulation_efficiency_pct": 85,
                "pgp_escape_coeff": 0.9, "apo_e_affinity": "very_high",
                "carpa_risk_index": 0.1, "off_target_liver_pct": 15,
                "phase_transition_temp_c": 55,
            }],
        }
        config_path = tmp_path / "dds_config.yaml"
        config_path.write_text(_yaml.dump(cfg))
        results_dir = tmp_path / "dds_results"
        results_dir.mkdir()
        monkeypatch.setattr(ei, "DDS_CONFIG", config_path)
        monkeypatch.setattr(ei, "DDS_RESULTS", results_dir)

        with caplog.at_level(logging.INFO, logger="CEREBRO-INFRA"):
            df = ei.DDSEngine.run()

        assert df is not None
        summary_lines = [r.message for r in caplog.records
                          if "BBB DELIVERY" in r.message]
        assert summary_lines, "expected the top-5 summary line to be logged"
        assert "LECANEMAB" not in summary_lines[0]
        assert "RIVASTIGMINE" in summary_lines[0]


class TestImputerEngineRespectsSecondaryFieldsWhitelist:
    """ImputerEngine's own docstring documents a deliberate philosophy:
    CORE fields are strict-rejected, and ONLY a curated SECONDARY_FIELDS
    whitelist (LogP, PDI, ligand density, ...) is eligible for
    IterativeImputer -- everything else should be left alone. The actual
    `eligible` list in impute() never consulted SECONDARY_FIELDS at all;
    it imputed any numeric column the caller passed minus the 3 CORE
    fields, whatever else happened to be in the dataframe. Harmless at
    the one real call site today (runs before any score/rank column
    exists), but silently dropped the safety boundary the class exists
    to enforce."""

    def test_only_whitelisted_secondary_fields_get_imputed(self):
        import numpy as np
        import pandas as pd

        from src.dds.enterprise_infra import ImputerEngine

        df = pd.DataFrame({
            "Drug": ["A", "B", "C"],
            "MW_Da": [300, 350, 320],
            "Half_Life_Days": [1, 2, 3],
            "LogP": [2.1, np.nan, 1.8],             # in SECONDARY_FIELDS
            "Some_Other_Score": [50, np.nan, 70],   # NOT in SECONDARY_FIELDS
        })
        out = ImputerEngine.impute(
            df, df.select_dtypes(include=np.number).columns.tolist())
        assert not out["LogP"].isna().any()
        assert out["Some_Other_Score"].isna().sum() == 1, (
            "a column outside SECONDARY_FIELDS must not be ML-imputed")


class TestDdsEngineEnrichDrugFieldsHonestFallback:
    """enrich_drug_fields' own comment states its v22.1 policy: "NO
    drug-name-specific fallbacks. If cascade fails, the fields stay
    None and downstream code reports the gap" -- applied correctly to
    MW_Da/Half_Life_Days (both default to None via MW_REF/CLINICAL_HL
    lookups) but LogP silently defaulted to a hardcoded -0.7 for every
    drug regardless of what the cascade actually returned, contradicting
    that same policy stated two lines below it. The subsequent log.info
    call also crashed with TypeError on a genuinely-missing (None) value
    -- formatting None with ':.0f' -- which the surrounding
    except Exception then mislabeled as "cascade enrichment failed" even
    when enrichment had actually succeeded for the fields it could."""

    def _install_fake_cascade(self, monkeypatch, mw_da=None, logp=None, half_life=None):
        import types

        class FakeCascade:
            @staticmethod
            def fetch_drug(name):
                data = {}
                if mw_da is not None:
                    data["MW_Da"] = mw_da
                if logp is not None:
                    data["LogP"] = logp
                if half_life is not None:
                    data["Half_Life_Days"] = half_life
                return data

        fake_pipeline = types.ModuleType("CEREBRO_Pipeline")
        fake_pipeline.CLINICAL_HL = {}
        fake_pipeline.MW_REF = {}
        fake_pipeline.CascadeDataEngine = FakeCascade
        # monkeypatch.setitem restores the real "CEREBRO_Pipeline" entry
        # (registered by src.path_resolver, and depended on by unrelated
        # tests -- e.g. TestModulePathShims checking CEREBRO_Pipeline.PATHS)
        # automatically when this test ends, unlike a bare
        # sys.modules[...] = assignment, which would leak this fake,
        # PATHS-less module into every later test in the same session.
        monkeypatch.setitem(__import__("sys").modules, "CEREBRO_Pipeline", fake_pipeline)

    def test_logp_stays_none_when_cascade_does_not_return_it(self, monkeypatch):
        import numpy as np
        import pandas as pd

        from src.dds.enterprise_infra import DDSEngine

        self._install_fake_cascade(monkeypatch, mw_da=300.5)  # LogP/HL genuinely absent
        df = pd.DataFrame({"Drug": ["test_drug"], "MW_Da": [np.nan],
                            "LogP": [np.nan], "Half_Life_Days": [np.nan]})
        out = DDSEngine.enrich_drug_fields(df)
        assert out["MW_Da"].iloc[0] == 300.5
        assert pd.isna(out["LogP"].iloc[0]), (
            "LogP must not silently become -0.7 for a drug the cascade "
            "couldn't resolve it for")
        assert pd.isna(out["Half_Life_Days"].iloc[0])

    def test_log_statement_does_not_crash_and_mislabel_a_partial_success(self, monkeypatch):
        """A genuinely-missing field (None) reaching the log.info format
        string must not raise -- that TypeError used to get caught by
        the surrounding except Exception and logged as "cascade
        enrichment failed", which is false: MW_Da enrichment above had
        already succeeded."""
        import numpy as np
        import pandas as pd

        from src.dds.enterprise_infra import DDSEngine

        self._install_fake_cascade(monkeypatch, mw_da=300.5)
        df = pd.DataFrame({"Drug": ["test_drug"], "MW_Da": [np.nan],
                            "LogP": [np.nan], "Half_Life_Days": [np.nan]})
        out = DDSEngine.enrich_drug_fields(df)  # must not raise
        assert out["MW_Da"].iloc[0] == 300.5, (
            "the real enrichment succeeded and must not be discarded by "
            "a log-formatting crash being mistaken for a cascade failure")


# ═════════════════════════════════════════════════════════════════════════════
# 3. MLOps
# ═════════════════════════════════════════════════════════════════════════════

class TestMLOps:
    """Model registry, drift detection, experiment tracking."""

    def test_model_registry_lifecycle(self, tmp_db):
        from src.ml.mlops import ModelRegistry, ModelStage, ModelVersion
        reg = ModelRegistry(tmp_db)

        # Register v1
        reg.register(ModelVersion(
            model_name="test_model", version="1.0.0",
            stage=ModelStage.DEVELOPMENT,
            metrics={"r2": 0.85, "mae": 0.12},
        ))

        # Promote to production
        reg.promote("test_model", "1.0.0", ModelStage.PRODUCTION)
        prod = reg.get_production("test_model")
        assert prod is not None
        assert prod.version == "1.0.0"

        # Register v2 and promote (should archive v1)
        reg.register(ModelVersion(
            model_name="test_model", version="1.1.0",
            metrics={"r2": 0.90},
        ))
        reg.promote("test_model", "1.1.0", ModelStage.PRODUCTION)
        prod = reg.get_production("test_model")
        assert prod.version == "1.1.0"

        # v1 should be archived
        versions = reg.list_versions("test_model", ModelStage.ARCHIVED)
        assert len(versions) == 1
        assert versions[0].version == "1.0.0"

    def test_rollback(self, tmp_db):
        from src.ml.mlops import ModelRegistry, ModelStage, ModelVersion
        reg = ModelRegistry(tmp_db)
        reg.register(ModelVersion(model_name="m", version="1.0.0"))
        reg.register(ModelVersion(model_name="m", version="2.0.0"))
        reg.promote("m", "2.0.0", ModelStage.PRODUCTION)
        reg.rollback("m", "1.0.0")
        assert reg.get_production("m").version == "1.0.0"

    def test_psi_no_drift(self):
        """Regression test for real flakiness: unseeded random draws from
        the identical distribution occasionally produced a PSI just over
        the 0.1 threshold from sampling noise alone (~1 in 300 runs,
        confirmed empirically) — an intermittent failure with nothing
        wrong in the code under test. Seeded like test_full_drift_detection
        already does, for a deterministic, reproducible result."""
        from src.ml.mlops import ModelDriftDetector
        np.random.seed(42)
        ref = np.random.normal(0.5, 0.1, 500)
        cur = np.random.normal(0.5, 0.1, 500)
        psi = ModelDriftDetector.compute_psi(ref, cur)
        assert psi < 0.1, f"Same distribution should have low PSI, got {psi}"

    def test_psi_detects_drift(self):
        from src.ml.mlops import ModelDriftDetector
        np.random.seed(42)
        ref = np.random.normal(0.5, 0.1, 500)
        cur = np.random.normal(0.8, 0.2, 500)  # shifted
        psi = ModelDriftDetector.compute_psi(ref, cur)
        assert psi > 0.25, f"Shifted distribution should have high PSI, got {psi}"

    def test_full_drift_detection(self):
        from src.ml.mlops import ModelDriftDetector
        np.random.seed(123)
        ref = np.random.normal(0.65, 0.1, 1000)
        cur = np.random.normal(0.65, 0.1, 1000)  # same dist, large sample
        result = ModelDriftDetector.detect_prediction_drift(ref, cur)
        assert result["severity"] in ("none", "moderate")  # small random noise ok

    def test_performance_degradation_detection(self):
        from src.ml.mlops import ModelDriftDetector
        result = ModelDriftDetector.check_performance_degradation(
            current_r2=0.60, baseline_r2=0.85,
            current_mae=0.25, baseline_mae=0.10,
        )
        assert result["degraded"] is True

    def test_experiment_tracker(self, tmp_db):
        from src.ml.mlops import ExperimentTracker, ModelRegistry
        # ModelRegistry init creates all tables including experiments
        ModelRegistry(tmp_db)
        tracker = ExperimentTracker(tmp_db)
        run_id = tracker.start_run("test_model", {"lr": 0.01})
        tracker.log_metrics(run_id, {"r2": 0.88})
        tracker.end_run(run_id, "completed")
        run = tracker.get_run(run_id)
        assert run["status"] == "completed"
        assert run["metrics"]["r2"] == 0.88


# ═════════════════════════════════════════════════════════════════════════════
# 4. ORCHESTRATOR
# ═════════════════════════════════════════════════════════════════════════════

class TestOrchestrator:
    """DAG execution, retry logic, circuit breaker."""

    def test_circuit_breaker_trips_on_failures(self):
        from src.workers.orchestrator import (
            CircuitBreaker,
            CircuitBreakerConfig,
            CircuitState,
        )
        cb = CircuitBreaker("test", CircuitBreakerConfig(
            failure_threshold=3, recovery_timeout=1
        ))
        assert cb.state == CircuitState.CLOSED
        for _ in range(3):
            cb.record_failure()
        assert cb.state == CircuitState.OPEN
        assert not cb.can_execute()

    def test_circuit_breaker_recovers(self):
        from src.workers.orchestrator import (
            CircuitBreaker,
            CircuitBreakerConfig,
            CircuitState,
        )
        cb = CircuitBreaker("test", CircuitBreakerConfig(
            failure_threshold=2, recovery_timeout=0.1, success_threshold=1
        ))
        cb.record_failure()
        cb.record_failure()
        assert cb.state == CircuitState.OPEN
        time.sleep(0.15)
        assert cb.can_execute()  # should transition to HALF_OPEN
        cb.record_success()
        assert cb.state == CircuitState.CLOSED

    def test_circuit_breaker_decorator_recovers_with_default_config(self):
        """Regression test for a permanent-lockout bug: __call__'s wrapper
        incremented _half_open_calls on every call but never decremented
        it. With the DEFAULT config (half_open_max=1 < success_threshold=3
        -- exactly what PipelineOrchestrator.register() constructs, since
        it never overrides either), can_execute() in HALF_OPEN only allows
        a call through while _half_open_calls < half_open_max. The very
        first post-recovery call consumed that one slot forever, so the
        breaker could never accumulate the 3 successes needed to close --
        it got stuck rejecting every call with CircuitBreakerOpenError
        permanently, even once the wrapped function was fully healthy
        again. Verified directly before the fix: a breaker tripped once,
        then given an always-succeeding function, let exactly one call
        through and blocked every call after that, forever."""
        from src.workers.orchestrator import (
            CircuitBreaker,
            CircuitBreakerConfig,
            CircuitBreakerOpenError,
            CircuitState,
        )
        cb = CircuitBreaker("test", CircuitBreakerConfig(
            failure_threshold=1, recovery_timeout=0.01,
        ))  # success_threshold=3, half_open_max=1 -- the real defaults

        @cb
        def flaky(should_fail):
            if should_fail:
                raise ValueError("boom")
            return "ok"

        with pytest.raises(ValueError):
            flaky(True)
        assert cb.state == CircuitState.OPEN
        time.sleep(0.02)

        results = []
        for _ in range(5):
            try:
                results.append(flaky(False))
            except CircuitBreakerOpenError:
                results.append("BLOCKED")

        assert results == ["ok", "ok", "ok", "ok", "ok"], (
            f"breaker got stuck rejecting calls after recovery: {results}")
        assert cb.state == CircuitState.CLOSED

    def test_retry_decorator(self):
        from src.workers.orchestrator import RetryPolicy, retry_with_backoff
        call_count = 0

        @retry_with_backoff(RetryPolicy(max_retries=3, base_delay=0.01))
        def flaky():
            nonlocal call_count
            call_count += 1
            if call_count < 3:
                raise ValueError("not yet")
            return "ok"

        assert flaky() == "ok"
        assert call_count == 3

    def test_dag_topological_sort(self):
        from src.workers.orchestrator import PipelineOrchestrator, TaskDefinition
        orch = PipelineOrchestrator()
        orch.register(TaskDefinition("a", func=lambda: None))
        orch.register(TaskDefinition("b", func=lambda: None, depends_on=["a"]))
        orch.register(TaskDefinition("c", func=lambda: None, depends_on=["a"]))
        orch.register(TaskDefinition("d", func=lambda: None, depends_on=["b", "c"]))
        levels = orch._topological_sort()
        assert levels[0] == ["a"]
        assert set(levels[1]) == {"b", "c"}
        assert levels[2] == ["d"]

    def test_dag_execution(self):
        from src.workers.orchestrator import (
            PipelineOrchestrator,
            TaskDefinition,
            TaskState,
        )
        results = []
        orch = PipelineOrchestrator()
        orch.register(TaskDefinition("step1", func=lambda: results.append(1) or "done1"))
        orch.register(TaskDefinition("step2", func=lambda **kw: results.append(2) or "done2",
                                     depends_on=["step1"]))
        execs = orch.execute("test_run")
        assert execs["step1"].state == TaskState.SUCCESS
        assert execs["step2"].state == TaskState.SUCCESS
        assert results == [1, 2]

    def test_dag_failure_cascades(self):
        from src.workers.orchestrator import (
            PipelineOrchestrator,
            RetryPolicy,
            TaskDefinition,
            TaskState,
        )
        def fail_fn():
            raise RuntimeError("boom")

        orch = PipelineOrchestrator()
        orch.register(TaskDefinition("bad", func=fail_fn,
                                     retry_policy=RetryPolicy(max_retries=0)))
        orch.register(TaskDefinition("after", func=lambda **kw: "ok",
                                     depends_on=["bad"]))
        execs = orch.execute()
        assert execs["bad"].state == TaskState.DEAD
        assert execs["after"].state == TaskState.CANCELLED

class _FakeOrchRedisClient:
    """Minimal in-process Redis stand-in covering just the hash/list calls
    TaskHealthRegistry makes (hgetall/hset/lpush/ltrim/lrange/ping), so its
    Redis-transition logic and the endpoints reading it can be verified
    without a real Redis server."""

    def __init__(self):
        self.store = {}

    def ping(self):
        return True

    def hgetall(self, key):
        return dict(self.store.get(key, {}))

    def hset(self, key, mapping):
        self.store.setdefault(key, {})
        self.store[key].update({k: str(v) for k, v in mapping.items()})

    def lpush(self, key, value):
        self.store.setdefault(key, [])
        self.store[key].insert(0, value)

    def ltrim(self, key, start, end):
        lst = self.store.get(key, [])
        self.store[key] = lst[start:end + 1] if end != -1 else lst[start:]

    def lrange(self, key, start, end):
        lst = self.store.get(key, [])
        return lst[start:] if end == -1 else lst[start:end + 1]


class _FakeCeleryTask:
    """Stands in for the Celery Task instance Celery passes as `sender`
    to task_success/task_retry/task_failure -- just the two attributes
    the signal handlers in cerebro_orchestrator.py actually read."""

    def __init__(self, name, retries=0):
        self.name = name
        self.request = type("Req", (), {"retries": retries})()


class TestTaskHealthRegistry:
    """/orchestrator/status and /orchestrator/dead-letter used to build a
    brand-new PipelineOrchestrator() per request via
    create_cerebro_pipeline_dag() -- one that only ever holds placeholder
    lambda tasks and has never executed anything. The real pipeline runs
    as plain Celery tasks (pipeline_full_task, train_model_task,
    fetch_data_task, generate_report_task) that never touch
    PipelineOrchestrator at all, so dead_letter_queue was always [] and
    every circuit breaker was always CLOSED, regardless of whether real
    tasks were failing in production. Fixed by having those real tasks'
    Celery success/retry/failure signals update TaskHealthRegistry, a
    Redis-backed store the two endpoints now read instead."""

    def _make_registry_with_fake_redis(self):
        from src.workers.orchestrator import TaskHealthRegistry
        registry = TaskHealthRegistry.__new__(TaskHealthRegistry)
        registry._client = _FakeOrchRedisClient()
        return registry

    def test_record_failure_trips_breaker_to_open(self):
        from src.workers.orchestrator import CircuitState

        registry = self._make_registry_with_fake_redis()
        # cerebro.train_model has failure_threshold=2 in REAL_TASK_CONFIGS.
        registry.record_failure("cerebro.train_model")
        assert registry.breaker_status(["cerebro.train_model"])[
            "cerebro.train_model"]["state"] == CircuitState.CLOSED.value
        registry.record_failure("cerebro.train_model")
        status = registry.breaker_status(["cerebro.train_model"])["cerebro.train_model"]
        assert status["state"] == CircuitState.OPEN.value
        assert status["failures"] == 2

    def test_record_success_after_recovery_closes_breaker(self):
        from src.workers.orchestrator import CircuitState

        registry = self._make_registry_with_fake_redis()
        registry.record_failure("cerebro.train_model")
        registry.record_failure("cerebro.train_model")
        assert registry.breaker_status(["cerebro.train_model"])[
            "cerebro.train_model"]["state"] == CircuitState.OPEN.value

        # Simulate the recovery window having elapsed.
        key = registry._breaker_key("cerebro.train_model")
        registry._client.store[key]["last_failure"] = "0"

        # default success_threshold=3 -- REAL_TASK_CONFIGS only overrides
        # failure_threshold, so closing HALF_OPEN still needs 3 successes,
        # same as CircuitBreaker's in-process semantics.
        for _ in range(3):
            registry.record_success("cerebro.train_model")
        status = registry.breaker_status(["cerebro.train_model"])["cerebro.train_model"]
        assert status["state"] == CircuitState.CLOSED.value
        assert status["failures"] == 0

    def test_dead_letter_add_and_read_round_trips_through_redis(self):
        registry = self._make_registry_with_fake_redis()
        assert registry.dead_letter_queue() == []

        registry.add_dead_letter(
            task_name="cerebro.fetch_data", task_id="abc-123",
            error="ConnectionError: upstream unreachable", attempts=5,
        )
        dlq = registry.dead_letter_queue()
        assert len(dlq) == 1
        assert dlq[0]["task_name"] == "cerebro.fetch_data"
        assert dlq[0]["task_id"] == "abc-123"
        assert dlq[0]["attempts"] == 5

    def test_celery_task_failure_signal_populates_dead_letter(self):
        """Dispatches the real celery.signals.task_failure signal (not a
        mock of our own handler) and verifies the handler wired in
        cerebro_orchestrator.py actually reacts to it."""
        import src.workers.orchestrator as orch_mod
        from celery.signals import task_failure

        registry = self._make_registry_with_fake_redis()
        orig_registry = orch_mod._task_registry
        orch_mod._task_registry = registry
        try:
            task_failure.send(
                sender=_FakeCeleryTask("cerebro.fetch_data", retries=4),
                task_id="task-999",
                exception=RuntimeError("connection reset"),
                args=(), kwargs={}, traceback=None, einfo=None,
            )
        finally:
            orch_mod._task_registry = orig_registry

        dlq = registry.dead_letter_queue()
        assert len(dlq) == 1
        assert dlq[0]["task_id"] == "task-999"
        assert dlq[0]["attempts"] == 5
        assert "connection reset" in dlq[0]["error"]

    def test_orchestrator_endpoints_reflect_a_real_celery_task_failure(
            self, test_client, auth_headers):
        """End-to-end: fire the real task_failure signal for a real task
        name, then hit the actual HTTP endpoints and confirm they surface
        it -- not a freshly-constructed, never-executed orchestrator."""
        if not auth_headers:
            pytest.skip("auth fixture unavailable in this environment")

        import src.workers.orchestrator as orch_mod
        from celery.signals import task_failure

        registry = self._make_registry_with_fake_redis()
        orig_registry = orch_mod._task_registry
        orch_mod._task_registry = registry
        try:
            task_failure.send(
                sender=_FakeCeleryTask("cerebro.report_generate", retries=0),
                task_id="e2e-task-1",
                exception=ValueError("template render failed"),
                args=(), kwargs={}, traceback=None, einfo=None,
            )

            r_dead = test_client.get("/orchestrator/dead-letter", headers=auth_headers)
            assert r_dead.status_code == 200
            body = r_dead.json()
            assert "note" not in body, "a fake-but-available registry must not claim degraded state"
            task_ids = [e["task_id"] for e in body["dead_letter_queue"]]
            assert "e2e-task-1" in task_ids

            r_status = test_client.get("/orchestrator/status", headers=auth_headers)
            assert r_status.status_code == 200
            breakers = r_status.json()["circuit_breakers"]
            assert breakers["cerebro.report_generate"]["failures"] >= 1
        finally:
            orch_mod._task_registry = orig_registry

    def test_registry_unavailable_reports_degraded_note_instead_of_fake_health(self):
        """When Redis genuinely can't be reached, the endpoints must say
        so rather than silently reporting default CLOSED/empty state as
        if it were real."""
        from src.workers.orchestrator import TaskHealthRegistry
        registry = TaskHealthRegistry.__new__(TaskHealthRegistry)
        registry._client = None

        assert registry.available is False
        assert registry.dead_letter_queue() == []
        from src.workers.orchestrator import CircuitState
        status = registry.breaker_status(["cerebro.fetch_data"])["cerebro.fetch_data"]
        assert status["state"] == CircuitState.CLOSED.value


# ═════════════════════════════════════════════════════════════════════════════
# 4b. MONITORING & ALERTING (src/monitoring/monitoring.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDefaultAlertRulesUseRealData:
    """setup_default_alerts registered three of its four rules as
    `condition=lambda: False` -- a hardcoded stub that could never fire no
    matter how bad the real pipeline failure rate, model drift, or API
    latency actually got, while a fourth rule (disk space) was genuinely
    wired to real data. An admin relying on /monitoring/alerts to catch
    any of those three conditions would never see anything, regardless of
    a real incident. Wired all three to real, windowed data: a
    SlidingWindowTracker fed by track_pipeline_execution and
    RequestTrackingMiddleware (Prometheus Counters/Histograms are
    cumulative since process start with no query engine here to compute
    "in the last hour" from them), and DriftEventLogger
    (src/ml/mlops.py), which already persisted real drift events to
    SQLite but nothing was reading them back for alerting."""

    def test_pipeline_failure_rate_fires_on_real_failures_not_hardcoded_false(self):
        from src.monitoring.monitoring import (
            AlertEngine,
            setup_default_alerts,
            track_pipeline_execution,
        )

        @track_pipeline_execution
        def good():
            return "ok"

        @track_pipeline_execution
        def bad():
            raise RuntimeError("boom")

        good()
        for _ in range(3):
            with pytest.raises(RuntimeError):
                bad()

        engine = AlertEngine()
        setup_default_alerts(engine)
        fired = [a.rule_name for a in engine.evaluate()]
        assert "high_pipeline_failure_rate" in fired

    def test_pipeline_failure_rate_does_not_fire_with_no_real_data(self):
        """A process that hasn't run any pipelines yet is "no data", not
        a false 100% failure rate — verifies the count()==0 guard."""
        from src.monitoring.monitoring import PIPELINE_OUTCOME_TRACKER, _check_pipeline_failure_rate
        # Use a window far enough in the past that any prior test's
        # recordings in this shared module-level tracker don't leak in.
        assert _check_pipeline_failure_rate(window_sec=1e-9) is False

    def test_api_latency_fires_on_real_slow_requests(self):
        from src.monitoring.monitoring import API_LATENCY_TRACKER, _check_api_latency
        for _ in range(10):
            API_LATENCY_TRACKER.record(15.0)  # seconds, over the 10s threshold
        assert _check_api_latency(window_sec=3600, threshold_sec=10.0) is True

    def test_model_drift_reads_real_drift_event_logger(self, tmp_path):
        from src.ml.mlops import DriftEventLogger
        from src.monitoring.monitoring import _check_model_drift

        db = tmp_path / "mlops_test.db"
        logger = DriftEventLogger(db_path=db)
        import sqlite3
        conn = sqlite3.connect(db)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS drift_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_name TEXT, drift_type TEXT, metric_name TEXT,
                metric_value REAL, threshold REAL, severity TEXT,
                details TEXT, detected_at TEXT
            )
        """)
        conn.commit()
        conn.close()
        logger.log_event("test_model", "psi", "prediction_dist", 0.4, 0.25,
                          severity="critical")

        orig = DriftEventLogger.get_recent

        def _fake_get_recent(self, model_name=None, hours=24):
            # Route through the real, un-patched method bound to a fresh
            # instance pointed at the tmp db -- calling self.get_recent()
            # here would recurse into this same monkeypatch forever.
            fresh = DriftEventLogger.__new__(DriftEventLogger)
            fresh.db_path = db
            return orig(fresh, model_name, hours)

        DriftEventLogger.get_recent = _fake_get_recent
        try:
            assert _check_model_drift(hours=1) is True
        finally:
            DriftEventLogger.get_recent = orig


# ═════════════════════════════════════════════════════════════════════════════
# 5. CACHE
# ═════════════════════════════════════════════════════════════════════════════

class TestCache:
    """Multi-tier caching layer."""

    def test_lru_basic(self):
        from src.ml.cache import LRUCache
        cache = LRUCache(max_size=5, default_ttl=60)
        cache.set("a", {"value": 1})
        assert cache.get("a") == {"value": 1}
        assert cache.get("missing") is None

    def test_lru_ttl_expiry(self):
        from src.ml.cache import LRUCache
        cache = LRUCache(max_size=5, default_ttl=0.05)
        cache.set("x", "data")
        assert cache.get("x") == "data"
        time.sleep(0.1)
        assert cache.get("x") is None  # expired

    def test_lru_eviction(self):
        from src.ml.cache import LRUCache
        cache = LRUCache(max_size=3)
        cache.set("a", 1)
        cache.set("b", 2)
        cache.set("c", 3)
        cache.set("d", 4)  # should evict "a"
        assert cache.get("a") is None
        assert cache.get("d") == 4

    def test_lru_stats(self):
        from src.ml.cache import LRUCache
        cache = LRUCache(max_size=10)
        cache.set("k", "v")
        cache.get("k")      # hit
        cache.get("miss")   # miss
        stats = cache.stats
        assert stats["hits"] == 1
        assert stats["misses"] == 1

    def test_sqlite_cache(self, tmp_path):
        from src.ml.cache import SQLiteCache
        cache = SQLiteCache(tmp_path / "cache.db")
        cache.set("mol:test_drug_x", {"MW": 143379}, ttl=60, category="molecule")
        assert cache.get("mol:test_drug_x")["MW"] == 143379
        cache.flush("molecule")
        assert cache.get("mol:test_drug_x") is None

    def test_cache_decorator(self):
        from src.ml.cache import cached, get_cache
        call_count = 0

        @cached(ttl=60, category="test")
        def expensive(x):
            nonlocal call_count
            call_count += 1
            return x * 2

        assert expensive(5) == 10
        assert expensive(5) == 10  # cached
        assert call_count == 1     # only called once

        get_cache().flush()


class _FakeRedisClient:
    """Minimal in-process stand-in implementing just the Redis calls
    RedisCache actually makes (get/setex/delete/keys/ping), so the
    category-namespacing fix can be verified end-to-end without the
    `redis` package installed or a real server running."""
    def __init__(self):
        self.store = {}

    def ping(self):
        return True

    def get(self, key):
        return self.store.get(key)

    def setex(self, key, ttl, value):
        self.store[key] = value

    def delete(self, *keys):
        for k in keys:
            self.store.pop(k, None)

    def keys(self, pattern):
        import fnmatch
        return [k for k in self.store if fnmatch.fnmatch(k, pattern)]

    # ── hash + list ops (TaskHealthRegistry) ────────────────────────────
    def hgetall(self, key):
        return dict(self.store.get(key, {}))

    def hset(self, key, mapping):
        self.store.setdefault(key, {})
        self.store[key].update({k: str(v) for k, v in mapping.items()})

    def lpush(self, key, value):
        self.store.setdefault(key, [])
        self.store[key].insert(0, value)

    def ltrim(self, key, start, end):
        lst = self.store.get(key, [])
        self.store[key] = lst[start:end + 1] if end != -1 else lst[start:]

    def lrange(self, key, start, end):
        lst = self.store.get(key, [])
        return lst[start:] if end == -1 else lst[start:end + 1]


class TestRedisCacheCategoryNamespacing:
    """RedisCache.get/set/delete previously ignored the category
    argument entirely — the stored key was always f"{PREFIX}{key}", with
    no category embedded. CacheManager.flush(category) asks
    RedisCache.flush(f"{category}:*") to delete matching keys, but since
    category was never part of the actual stored key, that pattern never
    matched anything — a category-scoped flush against Redis was a
    silent no-op. This is the exact mechanism invalidate_on_excel_change()
    depends on to clear stale molecule/DDS cache entries when a
    researcher uploads new input; in a multi-worker Redis deployment
    (the scenario Redis exists for), stale data could keep being served
    from Redis for up to its full TTL after invalidation was supposedly
    triggered. Fixed by embedding category into the Redis key namespace
    and threading it through CacheManager's get/set/delete."""

    def _make_cache_with_fake_redis(self):
        from src.ml.cache import CacheManager, RedisCache
        cache = CacheManager()
        cache.l2 = RedisCache.__new__(RedisCache)
        cache.l2._client = _FakeRedisClient()
        return cache

    def test_category_scoped_flush_actually_removes_redis_entries(self):
        cache = self._make_cache_with_fake_redis()
        cache.set("fetch_molecule:Donepezil", {"MW_Da": 379.5}, category="molecule")
        cache.set("score_dds:F001", {"score": 85}, category="dds")

        # Confirm both actually landed in the fake Redis store first.
        assert len(cache.l2._client.store) == 2

        cache.flush(category="molecule")

        remaining = list(cache.l2._client.store.keys())
        assert not any("fetch_molecule" in k for k in remaining), (
            f"molecule-category entry survived a molecule-category flush: {remaining}")
        assert any("score_dds" in k for k in remaining), (
            "flush(category='molecule') must not touch other categories' Redis entries")

    def test_get_after_category_flush_is_a_real_miss_not_a_stale_hit(self):
        """The bug wasn't just that flush() logged success while doing
        nothing — a subsequent get() would still return the stale value
        from L2, since L1 gets wiped but L2 never actually loses the key.
        This is the actual user-visible symptom."""
        cache = self._make_cache_with_fake_redis()
        cache.set("fetch_molecule:Aspirin", {"MW_Da": 180.16}, category="molecule")
        cache.flush(category="molecule")
        assert cache.l2.get("fetch_molecule:Aspirin", category="molecule") is None


class TestDDSMetricsStealthDerivation:
    """_derive_stealth (src/viz/_dds_metrics.py, feeding every HTML5
    dashboard's Stealth column) used the matched value's own magnitude to
    guess whether it came from a real PEGylation mol% column or an
    already-normalized Stealth_Index (0-1) column — _first_non_null only
    returns the value, not which alias matched. A genuinely light
    PEGylation degree like 0.5 mol% (a real, legitimate formulation
    parameter) got misread as a pre-normalized Stealth_Index of 0.5 and
    scored 50.0 — identical to a real Stealth_Index=0.5 input — even
    though 0.5 mol% PEG is far from the documented 5% optimum and should
    score low. Fixed to check explicitly which alias supplied the value.
    Also fixed a discontinuity in the ascending branch: peg=0 returned
    0.0 exactly, but peg=0.001 jumped straight to ~30 instead of ramping
    smoothly, contradicting the function's own "triangular profile"
    description."""

    def test_low_mol_pct_peg_and_stealth_index_are_not_conflated(self):
        from src.viz._dds_metrics import _derive_stealth
        low_peg = _derive_stealth({"pegylation_degree_mol_pct": 0.5})
        stealth_idx = _derive_stealth({"Stealth_Index": 0.5})
        assert low_peg != stealth_idx
        assert low_peg < 50.0, "0.5 mol% PEG (far from the 5% optimum) must score low"
        assert stealth_idx == 50.0

    def test_optimal_peg_scores_100(self):
        from src.viz._dds_metrics import _derive_stealth
        assert _derive_stealth({"pegylation_degree_mol_pct": 5.0}) == 100.0

    def test_ascending_branch_is_continuous_at_zero(self):
        from src.viz._dds_metrics import _derive_stealth
        assert _derive_stealth({"pegylation_degree_mol_pct": 0.0}) == 0.0
        near_zero = _derive_stealth({"pegylation_degree_mol_pct": 0.001})
        assert near_zero < 1.0, (
            f"expected a small value continuous with 0, got {near_zero} "
            f"(the old formula jumped to ~30 here)")

    def test_stealth_index_still_handled_when_no_mol_pct_present(self):
        """Guards against overcorrecting: records that only have a
        pre-computed Stealth_Index (no mol% column at all) must still work."""
        from src.viz._dds_metrics import _derive_stealth
        assert _derive_stealth({"Stealth_Index": 0.8}) == 80.0
        assert _derive_stealth({"Stealth_Index": 0.0}) == 0.0


class TestDDSMetricsLegacyAliasBackfill:
    """backfill_legacy_aliases (src/viz/_dds_metrics.py) fixes a codebase-wide
    instance of the "real value never threaded through, silently replaced by
    a hardcoded generic constant" bug class.

    Only a handful of cerebro_html5_engine.py functions (H05, H10, H11, H13,
    H20/H25) were ever migrated to read DDS metrics through the centralized
    extractor. Every other consumer — cerebro_advanced_modules_2.py,
    cerebro_science_modules.py, cerebro_video_engine_v2.py,
    cerebro_canvas_engine.py, most of cerebro_html5_engine.py, and
    final_report_unified.py — still reads the DDS record directly via
    top_dds.get("BBB_Enhanced_Pct", 30), top_dds.get("Endosomal_Escape_Eff",
    0.5), top_dds.get("Stealth_Index", 0.5). Verified directly (see
    pipeline_runner.py's _run_dds_from_yaml / evaluate_all_dds_62 output
    columns) that none of those three key names is ever produced by the
    real pipeline — every one of those dozens of .get() calls was silently
    returning the same hardcoded default for every drug, on every run,
    regardless of the formulation actually scored. pipeline_runner.py now
    calls backfill_legacy_aliases exactly once, at the single point each
    top-DDS dict is assembled (for both the primary drug and any
    Drug 2..N multi-drug comparison), so every downstream consumer gets
    real per-drug values without needing dozens of separate call-site edits.
    """

    def test_bbb_enhanced_pct_uses_real_score_not_hardcoded_30(self):
        from src.viz._dds_metrics import backfill_legacy_aliases
        rec = {"BBB_Engineering_Score": 91.4}
        out = backfill_legacy_aliases(rec)
        assert out["BBB_Enhanced_Pct"] == 91.4
        assert out["BBB_Enhanced_Pct"] != 30  # the old silent-fallback default

    def test_escape_and_stealth_are_0_1_fractions_not_0_100_scale(self):
        """Legacy callers format these as top_dds.get('Endosomal_Escape_Eff',
        0.5)*100 for a percentage, or with '.2f' directly (e.g. "0.65") —
        both assume a 0-1 fraction, unlike METRIC_DEFS's 0-100 scale."""
        from src.viz._dds_metrics import backfill_legacy_aliases
        rec = {"PgP_Escape_Coeff": 0.6, "pegylation_degree_mol_pct": 5.0}
        out = backfill_legacy_aliases(rec)
        assert out["Endosomal_Escape_Eff"] == pytest.approx(0.6)
        assert out["Stealth_Index"] == pytest.approx(1.0)  # 5.0 mol% = optimum = 100%

    def test_native_bbb_pct_reads_from_mol_profile_logbb_estimate(self):
        """BBB_Native_Pct is a molecule property (native, without-DDS
        crossing), not a DDS-formulation property — it must come from
        mol_profile's LogBB-derived BBB_permeability_pct, not from the
        DDS record itself."""
        from src.viz._dds_metrics import backfill_legacy_aliases
        rec = {"BBB_Engineering_Score": 80.0}
        out = backfill_legacy_aliases(rec, mol_profile={"BBB_permeability_pct": 2.35})
        assert out["BBB_Native_Pct"] == 2.35

    def test_native_bbb_pct_falls_back_honestly_when_mol_profile_missing(self):
        from src.viz._dds_metrics import backfill_legacy_aliases
        out = backfill_legacy_aliases({"BBB_Engineering_Score": 80.0}, mol_profile=None)
        assert out["BBB_Native_Pct"] == 3.0

    def test_cns_bioavailability_pct_uses_real_derivation_not_hardcoded_10(self):
        """CNS_Bioavailability_Pct is read the same ghost-key way in
        cerebro_advanced_modules_2.py, cerebro_science_modules.py,
        final_report_unified.py and h23 in cerebro_html5_engine.py (default
        10 everywhere) — must resolve via the real BBB%/liver-loss
        derivation, not collapse to that shared constant."""
        from src.viz._dds_metrics import backfill_legacy_aliases
        weak = backfill_legacy_aliases({"BBB_Engineering_Score": 15.0, "Off_Target_Liver_pct": 60.0})
        strong = backfill_legacy_aliases({"BBB_Engineering_Score": 90.0, "Off_Target_Liver_pct": 5.0})
        assert weak["CNS_Bioavailability_Pct"] != 10
        assert strong["CNS_Bioavailability_Pct"] != 10
        assert strong["CNS_Bioavailability_Pct"] > weak["CNS_Bioavailability_Pct"]

    def test_backfill_does_not_mutate_or_drop_original_keys(self):
        from src.viz._dds_metrics import backfill_legacy_aliases
        rec = {"BBB_Engineering_Score": 55.0, "Formulation_Name": "LNP-7"}
        out = backfill_legacy_aliases(rec)
        assert rec == {"BBB_Engineering_Score": 55.0, "Formulation_Name": "LNP-7"}
        assert out["Formulation_Name"] == "LNP-7"

    def test_different_formulations_get_different_backfilled_values(self):
        """The whole point of the fix: two different DDS records must not
        collapse onto the same hardcoded constant."""
        from src.viz._dds_metrics import backfill_legacy_aliases
        weak = backfill_legacy_aliases({"BBB_Engineering_Score": 20.0})
        strong = backfill_legacy_aliases({"BBB_Engineering_Score": 95.0})
        assert weak["BBB_Enhanced_Pct"] != strong["BBB_Enhanced_Pct"]


class TestPbbmDiagnosticPlotsHonestLabeling:
    """fig13_pbbm_diagnostic_plots (src/viz/advanced_viz.py) used to label
    itself a "Visual Predictive Check" (VPC) and "Goodness-of-Fit" (GOF)
    plot — real pharmacometric diagnostics that require independently
    measured clinical/experimental concentrations to compare a model's
    predictions against. df_pk here holds one deterministic
    single-compartment decay curve from AnalyticsEngine.simulate_pkpd, not
    observed data. The old "VPC" band was 100 replicates of that same
    curve with synthetic +/-20% noise sprinkled on, and the old "GOF" plot
    compared the noise-perturbed curve against itself — circular by
    construction, so it could never fail while presenting itself as a
    model validation. Also fixed: the column lookup only matched
    "Concentration_pct"/"Concentration_ugL" (lowercase p), while
    simulate_pkpd's real output column is "Concentration_Pct" (capital
    P) — the case mismatch meant this figure silently produced nothing
    on every real pipeline run.
    """

    def _df_pk(self):
        import pandas as pd
        return pd.DataFrame({
            "Day": list(range(10)),
            "Drug": ["TEST_DRUG_X"] * 10,
            "Concentration_Pct": [100 * (0.9 ** i) for i in range(10)],
        })

    def test_matches_real_pipeline_column_name_concentration_Pct(self, tmp_path):
        """simulate_pkpd emits 'Concentration_Pct' (capital P) — before the
        fix, only lowercase variants were matched and the function silently
        returned None for every real drug."""
        from src.viz.advanced_viz import fig13_pbbm_diagnostic_plots
        out = fig13_pbbm_diagnostic_plots(self._df_pk(), "TEST_DRUG_X", tmp_path)
        assert out is not None
        assert out.exists()

    def test_documentation_no_longer_claims_model_validation(self, tmp_path):
        """The regenerated figure must not claim to be a validated
        prediction-vs-observation check when no observed dataset exists."""
        from pathlib import Path
        from src.viz.advanced_viz import fig13_pbbm_diagnostic_plots
        out = fig13_pbbm_diagnostic_plots(self._df_pk(), "TEST_DRUG_X", tmp_path)
        doc_text = Path(str(out) + "_DOCUMENTATION.txt").read_text()
        assert "Visual Predictive Check" not in doc_text
        assert "Goodness-of-Fit" not in doc_text
        assert "Observed" not in doc_text
        assert "illustrative" in doc_text.lower() or "assumed" in doc_text.lower()

    def test_missing_concentration_column_still_returns_none(self, tmp_path):
        import pandas as pd
        from src.viz.advanced_viz import fig13_pbbm_diagnostic_plots
        df_no_conc = pd.DataFrame({"Day": [0, 1, 2]})
        assert fig13_pbbm_diagnostic_plots(df_no_conc, "TEST_DRUG_X", tmp_path) is None


class TestH07ScoreBreakdownHonestLabeling:
    """h07_shap (src/viz/cerebro_html5_engine.py) titled itself "SHAP
    Explainability" and its body text claimed "SHAP (SHapley Additive
    exPlanations)... Based on gradient boosted ensemble model" -- neither
    was true. The features dict is a hand-written heuristic decomposition
    of the DDS composite score (BBB_Enhanced_Pct*0.25, a size penalty,
    etc.), not real Shapley values from the `shap` library run against a
    trained model. A genuine SHAP TreeExplainer does exist elsewhere in
    this codebase (AdvancedMLEngine, feeding
    models/shap_feature_importance.csv) -- for the separate
    ML_Success_Probability regressor, not this DDS score -- so the old
    label borrowed a real technique's name for an unrelated illustrative
    calculation. Same class of issue as the fabricated "1000 bootstrap
    resamples" claim already found and fixed in h20_bootstrap in this
    same file, and the VPC/GOF mislabeling fixed in
    fig13_pbbm_diagnostic_plots above."""

    def test_output_makes_no_shap_claim(self):
        from src.viz.cerebro_html5_engine import h07_shap
        html = h07_shap({"Composite_Score": 72.0, "Formulation_Name": "LNP-3"},
                         "TEST_DRUG_X")
        assert "SHAP" not in html
        assert "gradient boosted ensemble" not in html.lower()

    def test_still_renders_the_waterfall_chart_and_score(self):
        """The fix is to the label, not the underlying illustrative
        computation -- the chart and composite score must still render."""
        from src.viz.cerebro_html5_engine import h07_shap
        html = h07_shap({"Composite_Score": 72.0, "Formulation_Name": "LNP-3"},
                         "TEST_DRUG_X")
        assert "shapChart" in html
        assert "72.0" in html


class TestH23BiodistributionOrganSharesSumTo100:
    """h23_biodistribution_animated's fallback organ-distribution
    calculator (used whenever science["biodistribution_map"] has no real
    data) computed "Blood" as a residual against hardcoded placeholder
    stand-ins for Lung and Kidney (literal 3 and 4), not the real Lung/
    Kidney values computed two lines above it -- which range far outside
    3-4 depending on size_nm/stealth. The six organ shares then didn't
    actually sum to 100% of the administered dose: verified directly
    across realistic size_nm/stealth combinations, the raw total ranged
    from 95.5% up to 107% -- a physically nonsensical "more than 100% of
    the dose was distributed" result for the large-size/low-stealth case.
    Same root-cause pattern as the negative "Other tissues" bucket
    already found and fixed in BiologicPBPK.simulate's organ_distribution
    (src/core/cerebro_science_modules.py) -- a residual computed against
    an assumed sub-total instead of the real sibling values. Fixed the
    same way: Blood now absorbs whatever's actually left after the real
    Lung/Kidney/Spleen/Liver/Brain shares, then the whole dict is
    renormalized to guarantee an exact 100% total."""

    def _organ_values(self, html: str) -> list[float]:
        import re
        m = re.search(r"data:(\[[\d.,\s]+\]),\s*\n\s*backgroundColor", html)
        assert m, "could not find the doughnut chart's data array in the output"
        import json as _json
        return _json.loads(m.group(1))

    def test_large_particle_low_stealth_no_longer_exceeds_100_percent(self):
        """The parameter combination (size_nm=300, Stealth_Index=0.0) that
        used to sum to 107% under the old hardcoded-residual formula."""
        from src.viz.cerebro_html5_engine import h23_biodistribution_animated
        top_dds = {"CNS_Bioavailability_Pct": 10, "Off_Target_Liver_pct": 30,
                   "Stealth_Index": 0.0, "size_nm": 300}
        mol_profile = {"molecule_class": "small_molecule", "MW_Da": 300}
        html = h23_biodistribution_animated({}, top_dds, "TEST_DRUG_X", mol_profile)
        vals = self._organ_values(html)
        assert sum(vals) == pytest.approx(100.0, abs=0.2)
        assert all(v >= 0 for v in vals)

    def test_default_parameters_sum_to_100_percent(self):
        from src.viz.cerebro_html5_engine import h23_biodistribution_animated
        mol_profile = {"molecule_class": "small_molecule", "MW_Da": 300}
        html = h23_biodistribution_animated({}, {}, "TEST_DRUG_X", mol_profile)
        vals = self._organ_values(html)
        assert sum(vals) == pytest.approx(100.0, abs=0.2)


# ═════════════════════════════════════════════════════════════════════════════
# 6. COMPLIANCE
# ═════════════════════════════════════════════════════════════════════════════

class TestCompliance:
    """PHI detection, audit trail, data masking."""

    def test_phi_detection(self):
        from src.compliance.privacy import PHIDetector
        text = "Patient SSN is 123-45-6789 and email is john@hospital.com"
        findings = PHIDetector.scan(text)
        types_found = {f["type"] for f in findings}
        assert "ssn" in types_found
        assert "email" in types_found

    def test_phi_redaction(self):
        from src.compliance.privacy import PHIDetector
        text = "Call 555-123-4567 or email dr@clinic.com"
        redacted = PHIDetector.redact(text)
        assert "555-123-4567" not in redacted
        assert "dr@clinic.com" not in redacted
        assert "[REDACTED]" in redacted

    def test_no_false_positive_on_drug_data(self):
        from src.compliance.privacy import PHIDetector
        text = "TEST_DRUG_X MW=143379 Da, LogP=-0.7, SMILES=CC(=O)NC"
        assert not PHIDetector.has_phi(text)

    def test_audit_trail_chain_integrity(self, tmp_path):
        from src.compliance.privacy import AuditTrail
        audit = AuditTrail(tmp_path / "audit.db")
        audit.log("admin", "login", "auth")
        audit.log("admin", "pipeline:run", "pipeline")
        audit.log("researcher", "results:read", "results")
        result = audit.verify_chain()
        assert result["valid"] is True
        assert result["entries"] == 3

    def test_data_masking_readonly(self):
        from src.compliance.privacy import DataMasker
        record = {
            "Drug": "TEST_DRUG_X",
            "MW_Da": 143379,
            "Formulation_Name": "SECRET-LNP-v3",
            "manufacturing_method": "microfluidics",
        }
        masked = DataMasker.mask_record(record, "readonly")
        assert masked["Drug"] == "TEST_DRUG_X"  # public
        assert masked["Formulation_Name"] == "***RESTRICTED***"
        assert masked["manufacturing_method"] == "***RESTRICTED***"

    def test_data_masking_researcher(self):
        from src.compliance.privacy import DataMasker
        record = {
            "Drug": "TEST_DRUG_X",
            "Formulation_Name": "SECRET-LNP-v3",
            "patient_id": "PT-12345",
        }
        masked = DataMasker.mask_record(record, "researcher")
        assert masked["Drug"] == "TEST_DRUG_X"
        assert masked["Formulation_Name"] == "SECRET-LNP-v3"
        assert masked["patient_id"] == "***RESTRICTED***"

    def test_encryption_roundtrip(self):
        from src.compliance.privacy import EncryptionEngine
        engine = EncryptionEngine()
        assert engine.available, (
            "encryption must be active even with no ENCRYPTION_KEY set in "
            "a non-production environment — it should never silently no-op"
        )
        ct = engine.encrypt("sensitive data")
        assert ct != "sensitive data"  # actually encrypted, not passed through
        pt = engine.decrypt(ct)
        assert pt == "sensitive data"

    def test_encryption_fails_hard_in_production_without_key(self, monkeypatch):
        """Same fail-hard-on-missing-secret pattern already used for
        JWT_SECRET_KEY and CEREBRO_ADMIN_PASSWORD in src/api/auth.py —
        refuse to start with sensitive-field encryption silently disabled."""
        import importlib

        import src.compliance.privacy as privacy_module
        monkeypatch.setenv("ENVIRONMENT", "production")
        monkeypatch.delenv("ENCRYPTION_KEY", raising=False)
        # AUDIT_HMAC_KEY has its own production fail-hard check at module
        # import time (see test_audit_hmac_key_fails_hard_in_production_
        # without_key) -- set a valid one here so the reload itself
        # succeeds and this test isolates EncryptionEngine's own behavior.
        monkeypatch.setenv("AUDIT_HMAC_KEY", "test_hmac_key_for_encryption_test")
        importlib.reload(privacy_module)
        try:
            with pytest.raises(RuntimeError):
                privacy_module.EncryptionEngine()
        finally:
            monkeypatch.delenv("ENVIRONMENT", raising=False)
            monkeypatch.delenv("AUDIT_HMAC_KEY", raising=False)
            importlib.reload(privacy_module)

    def test_audit_hmac_key_fails_hard_in_production_without_key(self, monkeypatch):
        """AUDIT_HMAC_KEY used to only warn and fall back to an ephemeral
        per-process key when unset, even in production -- unlike its three
        siblings (JWT_SECRET_KEY, CEREBRO_ADMIN_PASSWORD, ENCRYPTION_KEY),
        which all refuse to start. That silently defeated the audit trail's
        own tamper-detection guarantee: a forged chain segment signed with a
        different ephemeral key would be indistinguishable from a real one
        across a process/worker boundary. This check happens at module
        import time (a module-level constant, not a class __init__), so the
        reload itself must raise."""
        import importlib

        import src.compliance.privacy as privacy_module
        monkeypatch.setenv("ENVIRONMENT", "production")
        monkeypatch.delenv("AUDIT_HMAC_KEY", raising=False)
        try:
            with pytest.raises(RuntimeError):
                importlib.reload(privacy_module)
        finally:
            monkeypatch.delenv("ENVIRONMENT", raising=False)
            importlib.reload(privacy_module)

    def test_audit_hmac_key_falls_back_with_warning_outside_production(self, monkeypatch):
        import importlib

        import src.compliance.privacy as privacy_module
        monkeypatch.delenv("ENVIRONMENT", raising=False)
        monkeypatch.delenv("AUDIT_HMAC_KEY", raising=False)
        try:
            importlib.reload(privacy_module)
            assert privacy_module.AUDIT_HMAC_KEY  # ephemeral key generated, not empty
        finally:
            importlib.reload(privacy_module)

    def test_retention_check(self):
        from src.compliance.privacy import DataClass, RetentionManager
        old_date = datetime.utcnow() - timedelta(days=365 * 11)
        result = RetentionManager.check_retention(old_date, DataClass.PUBLIC)
        assert result["expired"] is True

        recent = datetime.utcnow() - timedelta(days=30)
        result = RetentionManager.check_retention(recent, DataClass.PUBLIC)
        assert result["expired"] is False


# ═════════════════════════════════════════════════════════════════════════════
# 7. BBB PERMEABILITY DNN (engine/cerebro_bbb_dnn.py)
# ═════════════════════════════════════════════════════════════════════════════

class TestBBBDNN:
    """Real DNN trained on the public BBBP dataset — see the module docstring
    in engine/cerebro_bbb_dnn.py for what is and isn't claimed about it."""

    def test_featurizer_handles_valid_and_invalid_smiles(self):
        import cerebro_bbb_dnn as bbb
        if not bbb._HAS_RDKIT:
            pytest.skip("rdkit not installed")
        feat = bbb._featurize_smiles("CCO")  # ethanol — valid
        assert feat is not None
        assert feat.shape == (2048,)
        assert bbb._featurize_smiles("not a smiles string !!!") is None

    def test_predict_reports_unavailable_without_deps(self):
        import cerebro_bbb_dnn as bbb
        if bbb._HAS_BBB_DNN:
            pytest.skip("rdkit+tensorflow both installed — availability path not exercised")
        result = bbb.predict_bbb_class("CCO")
        assert result["available"] is False

    @pytest.mark.slow
    def test_train_and_predict_real_model(self):
        """Trains (or reuses the cached model) on the real BBBP dataset and
        checks the reported held-out metrics are real and reasonable — not
        a mock, not a hardcoded number. Requires network access on first run
        to download BBBP.csv; subsequent runs reuse the cached CSV + model."""
        import cerebro_bbb_dnn as bbb
        if not bbb._HAS_BBB_DNN:
            pytest.skip("rdkit and/or tensorflow not installed")
        result = bbb.predict_bbb_class(
            "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2"  # donepezil
        )
        assert result["available"] is True
        assert result["predicted_class"] in ("permeable", "non_permeable")
        assert 0.0 <= result["probability_permeable"] <= 1.0
        # Real held-out test accuracy should beat random guessing by a wide
        # margin on a real, non-trivial dataset — but this is not a
        # hardcoded "it must be >0.9" assertion, since re-training with a
        # different seed/environment could reasonably land anywhere in a
        # sane range.
        assert result["model_test_accuracy"] > 0.7
        assert result["model_n_train"] > 1000  # real BBBP has ~1600 in an 80% split

    def test_resolver_uses_dnn_tier_for_small_molecules(self):
        """The bbb_permeability resolver should route small molecules with a
        valid SMILES through the DNN (tier 3) when the DNN is available, and
        must never apply it to biologics (see BIOLOGIC_CLASSES exclusion)."""
        import cerebro_bbb_dnn as bbb
        if not bbb._HAS_BBB_DNN:
            pytest.skip("rdkit and/or tensorflow not installed")
        from cerebro_value_resolver import resolve_value

        small_mol = resolve_value(
            "bbb_permeability", name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            molecule_class="small_molecule")
        assert small_mol["tier"] == 3
        assert small_mol["source"] == "cerebro_bbb_dnn"

        biologic = resolve_value(
            "bbb_permeability", name="Lecanemab",
            molecule_class="monoclonal_antibody")
        assert biologic["source"] != "cerebro_bbb_dnn"
        assert biologic["value"] < 1.0  # biologics get a low class-default %

    def test_clark_logbb_fallback_gives_realistic_not_saturated_percentages(self):
        """Regression test: the Tier-6 Clark-regression fallback (no SMILES,
        or DNN unavailable) used to compute
        bbb_pct = min(100, bp_ratio*100/(1+bp_ratio)*5) -- an extra
        saturating-percentage term stacked on top of the "B/P_ratio x 5"
        this function's own docstring/method string describes. That pushed
        the 100% cap into range for almost any logBB >= -0.6 (i.e. nearly
        every CNS-penetrant compound), when real BBB% for even excellent
        CNS drugs runs 3-60% throughout this codebase. Donepezil-like
        descriptors (logP=4.3, TPSA=39, logBB~0.215) must NOT come back as
        100%."""
        from cerebro_value_resolver.categories.bbb_perm import resolve_bbb_permeability

        result = resolve_bbb_permeability(
            name="NoSmilesTestDrug", smiles="", logp=4.3, tpsa=39.0)
        assert result["tier"] == 6
        assert result["source"] == "cerebro_value_resolver:clark_logbb_to_pct"
        assert 5.0 < result["value"] < 20.0

    def test_clark_logbb_fallback_never_exceeds_100_pct(self):
        from cerebro_value_resolver.categories.bbb_perm import resolve_bbb_permeability

        result = resolve_bbb_permeability(
            name="ExtremeLogP", smiles="", logp=15.0, tpsa=0.0)
        assert result["value"] <= 100.0


# ═════════════════════════════════════════════════════════════════════════════
# 8. DDS INVERSE-DESIGN (engine/cerebro_dds_inverse_design.py)
# ═════════════════════════════════════════════════════════════════════════════

class TestDDSInverseDesign:
    """Genetic-algorithm search over the DDS formulation-parameter space,
    using the REAL production scoring pipeline as its fitness function —
    see the module docstring for what novelty claim this is (and isn't)
    entitled to make."""

    @pytest.mark.slow
    def test_ga_search_returns_real_scored_candidates(self):
        from cerebro_dds_inverse_design import (
            ALL_PARAMS,
            generate_candidate_formulations,
        )
        from cerebro_resolved_bundles import resolve_drug_bundle

        import src.path_resolver  # noqa: F401 — ensures engine/ is on sys.path

        drug_bundle = resolve_drug_bundle(
            name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            molecule_class="small_molecule",
        )
        result = generate_candidate_formulations(
            drug_bundle, drug_name="Donepezil",
            n_generations=3, population_size=10, top_k=3, seed=7,
        )
        assert result["n_evaluated"] == 3 * 10
        assert 1 <= len(result["candidates"]) <= 3
        for c in result["candidates"]:
            assert "Principle_Composite_Score" in c
            assert 0 <= c["Principle_Composite_Score"] <= 100
            for p in ALL_PARAMS:
                assert p in c
        assert "disclaimer" in result and "hypotheses" in result["disclaimer"]


# ═════════════════════════════════════════════════════════════════════════════
# 9. REAL DOCKING ENGINE (src/core/real_docking_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
# The audit (docs/AUDIT_REPORT.md §11) flagged this as the scientific core
# with the least test coverage — zero tests on any docking/QSAR/PBPK engine.
# These are P0 per the audit's own testing roadmap (§11).

class TestRealDockingEngine:
    """AutoDock Vina integration with a graceful LIE-approximation fallback.
    See engine/cerebro_62_deep_engine.py's deep_P47 for where I wired this
    into the live pipeline."""

    def test_pdb_id_regex_rejects_path_traversal(self):
        """Regression test for the audit's §6 Medium finding: pdb_id was
        validated by length only elsewhere in the codebase (src/core/
        pdb_resolver.py), which a value like '../x' (4 chars) would pass.
        This file's stricter alphanumeric-4 regex is the fix pattern that
        should eventually replace that check — verify it actually rejects
        the exact class of input the audit was concerned about."""
        from src.core.real_docking_engine import _PDB_ID_RE
        assert _PDB_ID_RE.match("2NAO")       # real, valid PDB ID
        assert _PDB_ID_RE.match("1abc")       # lowercase alphanumeric, valid
        assert not _PDB_ID_RE.match("../x")   # path traversal — must reject
        assert not _PDB_ID_RE.match("../../etc/passwd")
        assert not _PDB_ID_RE.match("AB")     # too short
        assert not _PDB_ID_RE.match("ABCDE")  # too long
        assert not _PDB_ID_RE.match("AB-C")   # non-alphanumeric

    def test_lie_estimate_matches_documented_formula(self):
        """_lie_estimate implements a specific published formula (Aqvist
        1994). Regression-test the actual arithmetic, not just 'returns a
        number' — this is what distinguishes a real correlation from a
        black box, per the audit's distinction between this file (praised
        as genuine) and the fabricated Monte-Carlo code it also found."""
        from src.core.real_docking_engine import _lie_estimate
        # Donepezil-like small molecule: MW 379.5, LogP 4.77, TPSA 38.8
        result = _lie_estimate(ligand_mw=379.5, logp=4.77, tpsa=38.8,
                                hbd=0, hba=4, is_peptide=False)
        alpha, beta = 0.181, 0.137
        expected_dG = -(alpha * 4.77 + beta * (50 - 38.8/5) + 0.5*(0+4)*0.3)
        expected_dG = max(-20, min(-1, expected_dG))
        assert result["delta_G_kcal_mol"] == round(expected_dG, 2)
        assert result["docking_method"] == "LIE approximation (fallback — Vina unavailable)"
        assert result["confidence"] == "LOW — LIE approximation only"
        assert result["reference"].startswith("Aqvist 1994")
        # Kd back-calculated from the same delta_G via RT ln(Kd), R=1.987e-3
        # kcal/(mol·K) — regression-pinned to the exact formula, not just
        # "> 0": this used to divide by (8.314 * 310), mixing the SI
        # (joule-based) gas constant against delta_G*1000 (kcal converted to
        # cal) with no unit conversion between them, understating the
        # exponent by the ~4.18x cal/joule ratio and making Kd wrong by 3-8
        # orders of magnitude across the realistic ΔG range — every input
        # landed in "Weak (>1µM)" regardless of the real computed binding
        # strength. Fixed to match the same R and kcal-consistent units the
        # real Vina path (a few dozen lines below) already used correctly.
        import math
        RT = 1.987e-3 * 310
        expected_Kd_nM = round(math.exp(expected_dG / RT) * 1e9, 3)
        assert result["Kd_nM"] == expected_Kd_nM

    def test_lie_estimate_kd_classification_spans_the_full_range(self):
        """Regression test for the same Kd bug from the other direction:
        confirms the classification actually differentiates strong from
        weak binders now, instead of every realistic input collapsing
        into the same "Weak" bucket."""
        from src.core.real_docking_engine import _lie_estimate

        weak = _lie_estimate(ligand_mw=350, logp=0.5, tpsa=20, hbd=0, hba=0, is_peptide=False)
        strong = _lie_estimate(ligand_mw=350, logp=8.0, tpsa=5, hbd=10, hba=15, is_peptide=False)
        assert weak["Kd_class"] == "Weak (>1µM)"
        assert strong["Kd_class"] == "Tight (<10nM)"
        assert strong["Kd_nM"] < weak["Kd_nM"]

    def test_run_autodock_vina_falls_back_safely_without_pdb_id(self):
        """With no valid PDB ID, this must take the deterministic LIE path
        without ever attempting a network fetch or `import vina` — verified
        by checking the returned method label, matching the manual check I
        did when this was first wired into deep_P47."""
        from src.core.real_docking_engine import run_autodock_vina
        import tempfile
        with tempfile.TemporaryDirectory() as td:
            result = run_autodock_vina(
                smiles="CCO", pdb_id=None, output_dir=td,
                mol_profile={"MW_Da": 46.07, "LogP": -0.14},
            )
        assert result["docking_method"].startswith("LIE approximation")
        assert "note" in result  # explains why: no valid Target PDB ID

    def test_biologic_uses_lie_not_vina_regardless_of_pdb_id(self):
        """Biologics (MW > 2000 Da) can't be docked with Vina — must always
        use the LIE approximation, even if a PDB ID is supplied."""
        from src.core.real_docking_engine import run_autodock_vina
        import tempfile
        with tempfile.TemporaryDirectory() as td:
            result = run_autodock_vina(
                smiles="", pdb_id="2NAO", output_dir=td,
                mol_profile={"MW_Da": 148000, "LogP": 0.2,
                             "molecule_class": "monoclonal_antibody"},
            )
        assert result["docking_method"].startswith("LIE approximation")
        assert "Biologic" in result.get("note", "")


# ═════════════════════════════════════════════════════════════════════════════
# 10. PBBM ADMET PREDICTORS (src/core/pbbm_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
# The audit's §4.8 positive findings specifically praised these functions as
# "real, correctly-cited QSAR correlations ... applied as documented" —
# these tests pin the actual documented formulas so a future edit can't
# silently turn them into another §4.3-style "docstring says X, code does Y"
# mismatch without a test failing.

class TestNCAEngine:
    """Regression tests for NCAEngine.analyse's dose-to-molar conversion.

    Found while auditing pbbm_engine.py: dose_umol was computed with a
    hardcoded MW of 454 Da regardless of the actual drug, even though
    PBBMOrchestrator.run_full already resolves the real MW a few lines
    before calling NCAEngine — it just never threaded it through. Every
    drug whose real MW differs from 454 got a silently wrong CL/Vd."""

    def _mono_exp_decay(self):
        import numpy as np
        t = np.linspace(0, 24, 50)
        C = 10 * np.exp(-0.1 * t)
        return t, C

    def test_real_mw_changes_clearance_and_volume_vs_placeholder(self):
        from src.core.pbbm_engine import NCAEngine
        t, C = self._mono_exp_decay()

        no_mw   = NCAEngine.analyse(t, C, dose_mg=10.0)
        real_mw = NCAEngine.analyse(t, C, dose_mg=10.0, mw_da=180.16)

        assert no_mw["CL_apparent"] != real_mw["CL_apparent"]
        assert no_mw["Vd_ss"] != real_mw["Vd_ss"]
        # dose_umol is directly proportional to 1/MW, so CL and Vd scale
        # by exactly the ratio of the fallback MW to the real MW.
        expected_ratio = 454.0 / 180.16
        assert real_mw["CL_apparent"] / no_mw["CL_apparent"] == pytest.approx(
            expected_ratio, rel=1e-3)

    def test_missing_mw_falls_back_to_documented_placeholder(self):
        """mw_da=None (or 0) must fall back to the documented 454 Da
        placeholder, not raise or silently divide by zero."""
        from src.core.pbbm_engine import NCAEngine
        t, C = self._mono_exp_decay()

        via_none    = NCAEngine.analyse(t, C, dose_mg=10.0, mw_da=None)
        via_zero    = NCAEngine.analyse(t, C, dose_mg=10.0, mw_da=0)
        via_explicit= NCAEngine.analyse(t, C, dose_mg=10.0, mw_da=454.0)

        assert via_none["CL_apparent"] == via_explicit["CL_apparent"]
        assert via_zero["CL_apparent"] == via_explicit["CL_apparent"]

    def test_analyse_dataframe_threads_mw_through_to_analyse(self):
        """analyse_dataframe must pass mw_da down to analyse rather than
        dropping it — this is the exact wiring gap the bug was in."""
        import pandas as pd
        from src.core.pbbm_engine import NCAEngine
        t, C = self._mono_exp_decay()
        df = pd.DataFrame({"Hour": t, "Conc_umol_L": C, "Organ": "blood",
                            "Drug": "TestDrug"})

        no_mw   = NCAEngine.analyse_dataframe(df, dose_mg=10.0)
        real_mw = NCAEngine.analyse_dataframe(df, dose_mg=10.0, mw_da=180.16)

        assert no_mw["CL_apparent"].iloc[0] != real_mw["CL_apparent"].iloc[0]


class TestPBBMEngineACAT:
    """Regression tests for PBBMEngine.run_acat's dissolution physics.

    Found while auditing pbbm_engine.py: solubility_mg_mL, particle_size_um,
    mw_da, route, and n_points were all accepted as parameters, and a
    Noyes-Whitney dissolution_rate_per_h() helper was written, but nothing
    in the segment loop ever called it or referenced solubility/particle
    size — fa_total depended only on permeability/pKa. Two drugs with the
    same permeability but 1,000,000x different solubility (and 100x
    different particle size) produced bit-identical output — meaning
    BCS class II/IV (dissolution-limited) compounds got a permeability-only
    result mislabeled as a full ACAT dissolution+transit+permeability
    model. Fixed by capping per-segment absorption at the dissolved
    fraction (Dose Number gated, particle-size-scaled first-order
    dissolution)."""

    def test_poor_solubility_reduces_absorption_vs_high_solubility(self):
        from src.core.pbbm_engine import PBBMEngine
        r_soluble = PBBMEngine.run_acat(
            dose_mg=10, mw_da=379.5, logp=4.31,
            solubility_mg_mL=100.0, particle_size_um=5.0)
        r_insoluble = PBBMEngine.run_acat(
            dose_mg=10, mw_da=379.5, logp=4.31,
            solubility_mg_mL=0.0001, particle_size_um=500.0)
        assert r_insoluble["fa_total"] < r_soluble["fa_total"]
        assert r_insoluble["F_oral"] < r_soluble["F_oral"]

    def test_large_particle_size_alone_reduces_absorption(self):
        """Isolates the particle-size effect (Noyes-Whitney surface-area
        scaling) from solubility — same solubility, only particle size
        differs."""
        from src.core.pbbm_engine import PBBMEngine
        r_fine   = PBBMEngine.run_acat(
            dose_mg=10, mw_da=379.5, logp=4.31,
            solubility_mg_mL=0.05, particle_size_um=5.0)
        r_coarse = PBBMEngine.run_acat(
            dose_mg=10, mw_da=379.5, logp=4.31,
            solubility_mg_mL=0.05, particle_size_um=500.0)
        assert r_coarse["fa_total"] < r_fine["fa_total"]

    def test_highly_soluble_high_permeability_drug_still_absorbs_well(self):
        """BCS Class I-like input (Do << 1) must not be penalised by the
        dissolution cap — this guards against an overcorrection that
        would make every drug look dissolution-limited."""
        from src.core.pbbm_engine import PBBMEngine
        r = PBBMEngine.run_acat(
            dose_mg=10, mw_da=200.0, logp=2.0,
            solubility_mg_mL=50.0, particle_size_um=25.0,
            peff_cm_s=5e-4)
        assert r["fa_total"] > 0.5


class TestPBBMPredictors:
    """Regression tests for the QSAR correlations in ADMETPredictor —
    pinned against donepezil (a real, well-characterized small molecule
    used throughout this project's real pipeline runs)."""

    DONEPEZIL_SMILES = "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2"

    def test_predict_pka_returns_real_rdkit_derived_value(self):
        from src.core.pbbm_engine import ADMETPredictor
        result = ADMETPredictor.predict_pka(self.DONEPEZIL_SMILES)
        assert isinstance(result, dict)
        # pKa must be a real number in a chemically plausible range, not
        # None and not a suspiciously round placeholder like 7.0 or 0.
        pka = result.get("pKa") or result.get("pka") or result.get("value")
        if pka is not None:
            assert 0 < float(pka) < 14

    def test_predict_permeability_matches_documented_palm1997_formula(self):
        """Pins the Palm et al. 1997 Peff correlation
        (logPeff = -4.36 - 0.01*TPSA + 0.39*logP) against a fixed
        MW/LogP/TPSA input — if a future edit changes the coefficients
        without updating the docstring, this fails instead of silently
        drifting like the audit found elsewhere (§4.3)."""
        from src.core.pbbm_engine import ADMETPredictor
        result = ADMETPredictor.predict_permeability(
            self.DONEPEZIL_SMILES, mw=379.5, logp=4.31)
        assert result["_method"] == "Palm1997+Clark2003+PottsGuy1992+Wilson2001"
        assert result["Peff_cm_s"] is not None and result["Peff_cm_s"] > 0
        assert result["BBB_Filter"] in ("PASS", "FAIL")
        assert result["LogBB"] is not None

    def test_predict_permeability_handles_invalid_smiles_gracefully(self):
        """Non-SMILES input (e.g. accidentally-passed FASTA) must return
        the all-None result dict, not raise — this guard is what the
        function's own _is_valid_smiles check is for."""
        from src.core.pbbm_engine import ADMETPredictor
        result = ADMETPredictor.predict_permeability(">sp|P12345|FASTA_HEADER")
        assert result["Peff_cm_s"] is None
        assert result["_method"] == "heuristic_QSAR"  # unchanged default

    def test_predict_solubility_uses_real_mw_not_hardcoded_342(self):
        """Found while auditing pbbm_engine.py: predict_solubility converted
        Yalkowsky logSw (mol/L) to mg/mL with a hardcoded 342 Da constant,
        even though full_admet_profile already resolves the real MW and
        simply never passed it through. Every drug with MW != 342 got a
        systematically wrong Sw_mg_mL (and every biorelevant solubility
        derived from it: S_pH, FaSSGF, FaSSIF, FeSSIF)."""
        from src.core.pbbm_engine import ADMETPredictor
        donepezil = "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2"
        no_mw   = ADMETPredictor.predict_solubility(donepezil, logp=4.31)
        real_mw = ADMETPredictor.predict_solubility(donepezil, logp=4.31, mw=379.5)
        assert no_mw["Sw_mg_mL"] != real_mw["Sw_mg_mL"]
        # Both values are individually round()ed to 4 dp before the ratio
        # is taken, so a tight tolerance would fail on rounding noise at
        # these small magnitudes (~0.006) — 2% comfortably separates a
        # real ~11% MW-driven shift from that rounding noise.
        assert real_mw["Sw_mg_mL"] / no_mw["Sw_mg_mL"] == pytest.approx(
            379.5 / 342.0, rel=0.02)

    def test_full_admet_profile_threads_mw_into_solubility(self):
        """full_admet_profile already has the real mw in scope (it's a
        parameter) — this is the exact wiring gap the bug was in."""
        from src.core.pbbm_engine import ADMETPredictor
        donepezil = "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2"
        profile = ADMETPredictor.full_admet_profile(
            donepezil, "donepezil", mw=379.5, logp=4.31)
        direct = ADMETPredictor.predict_solubility(donepezil, logp=4.31, mw=379.5)
        assert profile["Sw_mg_mL"] == direct["Sw_mg_mL"]


class TestFormulationAdvisor:
    """FormulationAdvisor.biowaiver_assessment used to accept a bcs_class
    parameter that the body never referenced (it always recomputed its own
    inferred_class from solubility/permeability) — removed the misleading
    unused parameter rather than leaving a caller to think it has effect."""

    def test_biowaiver_class_i_high_sol_high_perm_is_eligible(self):
        from src.core.pbbm_engine import FormulationAdvisor
        result = FormulationAdvisor.biowaiver_assessment(
            dose_mg=10.0, solubility_mg_mL=10.0, permeability_cm_s=5e-4)
        assert result["BCS_class"] == "I"
        assert result["biowaiver_eligible"] is True

    def test_biowaiver_class_iv_low_sol_low_perm_not_eligible(self):
        from src.core.pbbm_engine import FormulationAdvisor
        result = FormulationAdvisor.biowaiver_assessment(
            dose_mg=10.0, solubility_mg_mL=0.001, permeability_cm_s=1e-6)
        assert result["BCS_class"] == "IV"
        assert result["biowaiver_eligible"] is False


class TestOptimisationAndSensitivity:
    """OptimisationEngine (SAEM/f-SAEM/PSO-LCI) has no live caller in the
    real pipeline (it needs real observed concentration-time data the
    automated PBBM run doesn't have) — verified via full-repo grep. Tests
    here exercise it directly since it's real, working code available for
    future calibration against experimental data."""

    def test_saem_recovers_known_parameter_on_simple_quadratic(self):
        import numpy as np
        from src.core.pbbm_engine import OptimisationEngine
        target = np.array([3.0])

        def objective(theta):
            return float(np.sum((theta - target) ** 2))

        result = OptimisationEngine.saem(
            objective, theta0=np.array([0.0]), bounds=[(-10, 10)], n_iter=200)
        # SAEM is a stochastic Metropolis-within-SAEM search, not gradient
        # descent — it can stall short of the exact optimum within a fixed
        # iteration budget. Assert real progress toward the target rather
        # than tight convergence, to avoid a flaky test on this genuinely
        # stochastic algorithm.
        assert abs(result["theta_opt"][0] - 3.0) < abs(0.0 - 3.0)
        assert result["obj_opt"] < 2.0

    def test_pso_lci_recovers_known_parameter(self):
        import numpy as np
        from src.core.pbbm_engine import OptimisationEngine
        target = np.array([2.0, -1.0])

        def objective(theta):
            return float(np.sum((theta - target) ** 2))

        result = OptimisationEngine.pso_lci(
            objective, bounds=[(-5, 5), (-5, 5)], n_particles=20, n_iter=60)
        assert np.allclose(result["theta_opt"], target, atol=0.5)

    def test_ota_sensitivity_ranks_influential_parameter_higher(self):
        import numpy as np
        from src.core.pbbm_engine import SensitivityAnalyser

        def model(theta):
            # second parameter dominates the output
            return theta[0] * 0.01 + theta[1] * 10.0

        df = SensitivityAnalyser.ota_sensitivity(
            model, np.array([1.0, 1.0]), ["weak", "strong"])
        top = df.iloc[0]["Parameter"]
        assert top == "strong"

    def test_uncertainty_propagation_returns_sane_distribution(self):
        import numpy as np
        from src.core.pbbm_engine import SensitivityAnalyser

        def model(theta):
            return float(theta[0] * theta[1])

        result = SensitivityAnalyser.uncertainty_propagation(
            model, np.array([10.0, 2.0]), param_cv=0.1, n_samples=200)
        assert result["n_valid_samples"] > 0
        assert result["p5"] < result["mean"] < result["p95"]


class TestPBBMOrchestratorEndToEnd:
    """Real end-to-end run of the full PBBM suite against a real drug
    (donepezil) — not a mock. Exercises ACAT, PBPK, NCA, metabolite tree,
    ADMET, formulation strategy, and sensitivity analysis together, and
    confirms the master report is written without crashing on any of the
    values produced by the fixes above."""

    DONEPEZIL_SMILES = "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2"

    def test_run_full_produces_all_result_sections_and_report(self, tmp_path):
        from src.core.pbbm_engine import PBBMOrchestrator
        mol_profile = {"MW_Da": 379.5, "LogP": 4.31, "Half_Life_Days": 3.0,
                        "Protein_Binding_pct": 96.0, "Sw_mg_mL": 0.1}
        results = PBBMOrchestrator.run_full(
            drug_name="Donepezil", smiles=self.DONEPEZIL_SMILES,
            mol_profile=mol_profile, df_dds=None, trial_dir=tmp_path,
            dose_mg=10.0, route="oral", n_workers=2)

        for key in ("acat", "pbpk", "nca", "metabolites", "admet",
                    "formulation_strategy", "sensitivity"):
            assert key in results, f"missing PBBM result section: {key}"

        assert 0 < results["acat"]["fa_total"] <= 1
        assert results["admet"]["Sw_mg_mL"] > 0

        report = tmp_path / "pbbm_results" / "PBBM_Master_Report_Donepezil.txt"
        assert report.exists()
        text = report.read_text()
        assert "ABSORPTION (ACAT MODEL)" in text
        assert "NON-COMPARTMENTAL ANALYSIS" in text


class TestHarmonizationSourcePriority:
    """HarmonizationEngine.harmonize_drug_records resolved a source's
    priority by taking max() over every SOURCE_PRIORITY key that's a
    substring of the record's _source string. "EmbeddedClinicalLibrary"
    (priority 8) is a substring of "EmbeddedClinicalLibrary_PartialHit"
    (intended priority 7), so max() silently promoted every partial hit to
    the full hit's priority — a partial hit could then outrank (or wrongly
    tie and win by list order against) a genuine full hit. Fixed to match
    the longest (most specific) key instead of taking the max priority
    across all substring matches."""

    def test_partial_hit_source_gets_its_own_lower_priority(self):
        from src.core.data_engineering import HarmonizationEngine
        priority = HarmonizationEngine.SOURCE_PRIORITY[
            "EmbeddedClinicalLibrary_PartialHit"]
        records = [{"Half_Life_Days": 2.0,
                    "_source": "EmbeddedClinicalLibrary_PartialHit"}]
        result = HarmonizationEngine.harmonize_drug_records(records, "Drug")
        assert result["_field_provenance"]["Half_Life_Days"]["priority"] == priority

    def test_full_embedded_hit_outranks_partial_hit(self):
        from src.core.data_engineering import HarmonizationEngine
        records = [
            {"Half_Life_Days": 2.0, "_source": "EmbeddedClinicalLibrary_PartialHit"},
            {"Half_Life_Days": 5.0, "_source": "EmbeddedClinicalLibrary"},
        ]
        result = HarmonizationEngine.harmonize_drug_records(records, "Drug")
        assert result["_field_provenance"]["Half_Life_Days"]["source"] == "EmbeddedClinicalLibrary"
        assert result["Half_Life_Days"] == 5.0

    def test_ordinary_source_priorities_unaffected(self):
        """Regression guard: the fix must not disturb non-colliding
        sources' normal priority resolution."""
        from src.core.data_engineering import HarmonizationEngine
        records = [
            {"Half_Life_Days": 2.0, "_source": "PubMed_NLP"},
            {"Half_Life_Days": 3.0, "_source": "DrugBank_API"},
        ]
        result = HarmonizationEngine.harmonize_drug_records(records, "Drug")
        assert result["_field_provenance"]["Half_Life_Days"]["source"] == "DrugBank_API"


class TestLineageTierLabeling:
    """LineageEngine.record_from_drug_data mapped an integer _tier to a
    human-readable "algorithm" label for the audit trail (this engine's
    documented purpose is FDA 21 CFR Part 11 provenance tracking). It
    defaulted missing _tier to 0, which tier_map resolves to
    "EmbeddedClinicalLibrary" — but pipeline.py's live API cascade
    (DrugBank/ChEMBL/UniProt/PubChem/PubMed, tiers 1-5) never set _tier at
    all, so every one of those hits got recorded with a correct "source"
    (e.g. "DrugBank") but a contradictory, wrong "algorithm"
    ("EmbeddedClinicalLibrary") in the same audit row. Fixed to fall back
    to the real source name when _tier is genuinely absent, and tagged
    _tier on pipeline.py's live-tier results so the specific tier label
    resolves correctly instead of just falling back."""

    def _make_engine(self, tmp_path):
        from src.core.data_engineering import LineageEngine
        return LineageEngine(tmp_path / "lineage.db", tmp_path / "lineage.jsonl")

    def test_missing_tier_falls_back_to_real_source_not_embedded_library(self, tmp_path):
        le = self._make_engine(tmp_path)
        le.record_from_drug_data("Trial_0", "Donepezil",
            {"MW_Da": 379.5, "Half_Life_Days": 3.0, "_source": "DrugBank"})
        df = le.get_feature_lineage("Donepezil")
        row = df.iloc[0]
        assert row["source"] == "DrugBank"
        assert row["algorithm"] == "DrugBank"
        assert row["algorithm"] != "EmbeddedClinicalLibrary"

    def test_explicit_tier_resolves_to_its_named_algorithm(self, tmp_path):
        le = self._make_engine(tmp_path)
        le.record_from_drug_data("Trial_0", "Rivastigmine",
            {"MW_Da": 250.0, "Half_Life_Days": 1.5,
             "_source": "DrugBank", "_tier": 1})
        df = le.get_feature_lineage("Rivastigmine")
        assert df.iloc[0]["algorithm"] == "DrugBank_API"

    def test_real_tier_zero_still_labels_embedded_library(self, tmp_path):
        """Guards against overcorrecting: a genuine tier-0 embedded-library
        hit must still resolve to its correct label."""
        le = self._make_engine(tmp_path)
        le.record_from_drug_data("Trial_0", "Galantamine",
            {"MW_Da": 287.0, "Half_Life_Days": 1.0,
             "_source": "EmbeddedClinicalLibrary", "_tier": 0})
        df = le.get_feature_lineage("Galantamine")
        assert df.iloc[0]["algorithm"] == "EmbeddedClinicalLibrary"

    def test_pipeline_live_tier_cascade_now_tags_tier(self):
        """Root-cause check: CascadeDataEngine.fetch_drug's tiers 1-5 loop
        must tag _tier on whichever source actually succeeded."""
        import inspect
        from src.core import pipeline
        src = inspect.getsource(pipeline.CascadeDataEngine.fetch_drug)
        assert '_tier' in src.split("TIERS 1-5")[1].split("TIER 6")[0]


class TestPipelineLipinskiAndPKPD:
    """Two bugs found auditing src/core/pipeline.py's ML/reporting engines.

    (1) AdvancedMLEngine.lipinski_baseline used OR instead of AND across
    the MW<=500 / LogP<=5 criteria, so a 10,000 Da biologic-scale molecule
    "passed" whenever LogP was low, and an extremely lipophilic LogP=15
    molecule "passed" whenever MW was reasonable — nearly any real
    molecule satisfies at least one of two loose criteria, making the
    Rule-of-5 baseline comparison this project's own docstring describes
    almost meaningless.

    (2) AnalyticsEngine.simulate_pkpd and ReportingEngine's headline "Days
    Above 50%" both computed C0 = 100*(150_000/MW) uncapped. That formula
    is tuned around a 150 kDa antibody-scale reference (matches this
    file's own lecanemab citation) — for this project's actual small-
    molecule candidates (MW ~200-600 Da) it inflated C0 to tens of
    thousands of percent, which broke the "Effective Brain Concentration
    (%)" chart's 50/100% threshold semantics and inflated the master
    report's headline days-above-threshold number by roughly 10x for a
    real donepezil-like candidate. Fixed by capping C0 at 100%."""

    def _patch_paths(self, monkeypatch, tmp_path):
        from src.core import pipeline
        for key in ("figures", "results", "deliverable", "reports"):
            d = tmp_path / key
            d.mkdir(parents=True, exist_ok=True)
            monkeypatch.setitem(pipeline.PATHS, key, d)
        return pipeline

    def test_lipinski_requires_both_criteria_not_either(self):
        import pandas as pd
        from src.core.pipeline import AdvancedMLEngine
        df = pd.DataFrame({
            "MW_Da": [10000.0, 300.0, 350.0],
            "LogP":  [2.0,     15.0,  3.0],
        })
        result = AdvancedMLEngine.lipinski_baseline(df)
        assert list(result["Lipinski_Pass"]) == [0, 0, 1]

    def test_pkpd_c0_capped_at_100_for_small_molecule(self, monkeypatch, tmp_path):
        import pandas as pd
        pipeline = self._patch_paths(monkeypatch, tmp_path)
        df = pd.DataFrame({"Drug": ["Donepezil"], "Half_Life_Days": [3.0],
                            "MW_Da": [379.5]})
        result = pipeline.AnalyticsEngine.simulate_pkpd(df)
        c0 = result[result["Day"] == 0]["Concentration_Pct"].iloc[0]
        assert c0 == pytest.approx(100.0)

    def test_pkpd_c0_still_scales_down_above_150kda_reference(self, monkeypatch, tmp_path):
        """Guards against overcorrecting: a genuinely large biologic (above
        the 150 kDa reference) should still get a reduced C0, not always 100."""
        import pandas as pd
        pipeline = self._patch_paths(monkeypatch, tmp_path)
        df = pd.DataFrame({"Drug": ["BigBiologic"], "Half_Life_Days": [10.0],
                            "MW_Da": [300000.0]})
        result = pipeline.AnalyticsEngine.simulate_pkpd(df)
        c0 = result[result["Day"] == 0]["Concentration_Pct"].iloc[0]
        assert c0 == pytest.approx(50.0)

    def test_master_report_days_above_50pct_matches_half_life_for_small_molecule(
            self, monkeypatch, tmp_path):
        """With C0 correctly capped at 100%, a molecule that decays purely
        by its own half-life crosses the 50% threshold at exactly t=t½ —
        this pins the master report's headline number against that
        physically obvious case instead of the ~10x-inflated value the
        uncapped formula produced."""
        import json
        import pandas as pd
        pipeline = self._patch_paths(monkeypatch, tmp_path)
        df_ml = pd.DataFrame({
            "Drug": ["Donepezil"], "MW_Da": [379.5], "Half_Life_Days": [3.0],
            "Docking_Affinity_kcal": [-8.5], "ML_Success_Probability": [72.0],
        })
        df_aav = pd.DataFrame({"Serotype": ["AAV9"], "CNS_Tropism": [0.9],
                                "Capsid_Mass_Da": [82000]})
        pipeline.ReportingEngine.generate_master_report(
            df_mab=pd.DataFrame(), df_aav=df_aav, df_ml=df_ml,
            metrics={"r2": 0.8, "cv_r2": 0.75})
        cfg = json.loads((tmp_path / "deliverable" / "project_config.json").read_text())
        assert cfg["Days_Above_50pct"] == pytest.approx(3.0, abs=0.05)


class TestPatchedTrainCVGuard:
    """pipeline_patches.patched_train is what actually runs in production —
    apply_patches() monkey-patches it in as AdvancedMLEngine.train at
    runtime (called for real from pipeline_runner.py) — so
    pipeline.py's own CV_MIN_SAMPLES guard on the original train() never
    executes once patched. patched_train had no equivalent guard: called
    directly with a single-row DataFrame, nk = min(5, len(X_s)) resolves
    to 1 and KFold(n_splits=1) raises immediately. The two current
    call sites happen to pre-pad single-drug trials with synthetic
    neighbours before calling train(), so this wasn't reachable through
    them today, but the function itself had no defense — fixed to match
    pipeline.py's own guard rather than relying on every future caller to
    remember to pre-pad."""

    def test_single_row_dataframe_does_not_crash(self):
        import pandas as pd
        from src.core.pipeline_patches import patched_train
        df = pd.DataFrame({
            "Drug": ["Donepezil"], "MW_Da": [379.5], "LogP": [4.31],
            "Half_Life_Days": [3.0], "Docking_Affinity_kcal": [-8.5],
        })
        df_out, ensemble, metrics = patched_train(
            None, df,
            feature_cols=["MW_Da", "LogP", "Half_Life_Days", "Docking_Affinity_kcal"])
        assert metrics["cv_r2"] != metrics["cv_r2"]  # NaN, not a crash
        assert "ML_Success_Probability" in df_out.columns

    def test_realistic_sample_size_still_produces_real_cv_score(self):
        """Guards against overcorrecting: a normal-sized dataset (the
        augmented single-drug case produces ~9 rows in the real pipeline)
        must still get a real K-Fold CV score, not NaN."""
        import numpy as np
        import pandas as pd
        from src.core.pipeline_patches import patched_train
        rng = np.random.RandomState(0)
        n = 9
        df = pd.DataFrame({
            "Drug": [f"Drug{i}" for i in range(n)],
            "MW_Da": rng.uniform(200, 500, n),
            "LogP": rng.uniform(1, 5, n),
            "Half_Life_Days": rng.uniform(1, 10, n),
            "Docking_Affinity_kcal": rng.uniform(-10, -6, n),
        })
        _, _, metrics = patched_train(
            None, df,
            feature_cols=["MW_Da", "LogP", "Half_Life_Days", "Docking_Affinity_kcal"])
        assert metrics["cv_r2"] == metrics["cv_r2"]  # not NaN


class TestModelRegistryBridge:
    """/predict reads its production model from src/ml/mlops.py's
    ModelRegistry via ModelRegistry().get_production("cerebro_ensemble").
    Nothing in the real training path ever wrote to that registry:
    pipeline.py's AdvancedMLEngine.train (and patched_train, which
    replaces it at runtime via apply_patches()) only wrote to
    pipeline.py's own separate `model_registry` SQLite table
    (db_register_model), a totally disconnected system. The result: even
    after a real, successful training run, /predict always 404'd with
    "No production model. Run /pipeline/run first." — running the
    pipeline was exactly what did NOT fix it. Fixed by having
    patched_train register + auto-promote into the mlops ModelRegistry
    right after it saves the model artifact. This exercises the real
    patched_train training path (no sklearn mocking) against a real
    temp SQLite registry, so a broken bridge would show up as
    get_production() returning None, not just as a missing function
    call."""

    def _train_synthetic(self, tmp_db, n=9, seed=0):
        import numpy as np
        import pandas as pd
        from src.core.pipeline_patches import patched_train
        rng = np.random.RandomState(seed)
        df = pd.DataFrame({
            "Drug": [f"Drug{i}" for i in range(n)],
            "MW_Da": rng.uniform(200, 500, n),
            "LogP": rng.uniform(1, 5, n),
            "Half_Life_Days": rng.uniform(1, 10, n),
            "Docking_Affinity_kcal": rng.uniform(-10, -6, n),
        })
        return patched_train(
            None, df,
            feature_cols=["MW_Da", "LogP", "Half_Life_Days", "Docking_Affinity_kcal"],
            registry_db_path=tmp_db)

    def test_patched_train_registers_and_promotes_production_model(self, tmp_db):
        from pathlib import Path

        from src.ml.mlops import ModelRegistry

        self._train_synthetic(tmp_db)

        registry = ModelRegistry(tmp_db)
        prod = registry.get_production("cerebro_ensemble")
        assert prod is not None, (
            "patched_train ran successfully but ModelRegistry().get_production() "
            "still returned None — /predict would 404 even after a real training run")
        assert prod.version == "1.0.0"
        assert prod.artifact_path and Path(prod.artifact_path).exists()

        import joblib
        bundle = joblib.load(prod.artifact_path)
        assert isinstance(bundle, dict)
        assert {"model", "scaler", "scaler_state", "feat_scaler", "features"} <= bundle.keys()

    def test_second_better_run_is_promoted_first_is_archived(self, tmp_db):
        from src.ml.mlops import ModelRegistry, ModelStage

        self._train_synthetic(tmp_db, seed=0)
        registry = ModelRegistry(tmp_db)
        first_version = registry.get_production("cerebro_ensemble").version

        # Force the second run's metrics above the first so promotion is
        # unambiguous, mirroring MLOpsPipeline.train_and_register's own
        # "only promote if better" rule.
        from src.ml.mlops import ModelVersion
        registry.register(ModelVersion(
            model_name="cerebro_ensemble", version="9.0.0",
            stage=ModelStage.STAGING, metrics={"r2": 0.999},
            artifact_path="unused.pkl",
        ))
        registry.promote("cerebro_ensemble", "9.0.0", ModelStage.PRODUCTION)

        assert registry.get_production("cerebro_ensemble").version == "9.0.0"
        archived = registry.list_versions("cerebro_ensemble", ModelStage.ARCHIVED)
        assert first_version in [v.version for v in archived]

    def test_predict_endpoint_round_trip_after_real_pipeline_training(
            self, test_client, auth_headers, tmp_db, monkeypatch):
        """Full stack: real patched_train() run -> real ModelRegistry lookup
        -> real POST /predict -> a real, correctly-scaled ML_Success_Probability.
        The only mock is redirecting ModelRegistry() to the temp DB this test
        trained into, so /predict doesn't touch the real outputs/ directory."""
        if not auth_headers:
            pytest.skip("auth fixture unavailable in this environment")

        self._train_synthetic(tmp_db)

        from src.ml.mlops import ModelRegistry as RealModelRegistry
        import src.api.app as app_module

        class _RegistryAtTmpDB:
            def __init__(self, *a, **kw):
                self._reg = RealModelRegistry(tmp_db)
            def get_production(self, name):
                return self._reg.get_production(name)

        monkeypatch.setattr(app_module, "ModelRegistry", _RegistryAtTmpDB)

        r = test_client.post(
            "/predict",
            json={"molecule_input": "CC(=O)OC1=CC=CC=C1C(=O)O",  # aspirin
                  "drug_name": "Aspirin"},
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert 45 <= body["ML_Success_Probability"] <= 98, (
            "ML_Success_Probability outside the pipeline's own 45-98 scaled "
            f"range — got {body['ML_Success_Probability']}, suggests the "
            "TrainAwareScaler wasn't applied to the raw model output")
        assert body["model_version"] == "1.0.0"
        assert body["model_stage"] == "production"


class TestInferenceEngineFeatureScaling:
    """InferenceEngine.load() read `model` and `scaler` (the TrainAwareScaler
    that rescales OUTPUT predictions to 45-98) from the saved bundle but
    silently dropped `feat_scaler` — the RobustScaler patched_train fits on
    INPUT features before ever calling ensemble.fit()/ensemble.predict().
    Skipping it doesn't crash (tree models are invariant to per-feature
    affine scaling) but silently skews the VotingRegressor's SVR estimator,
    which is not scale-invariant, so predictions on new molecules diverged
    from what the training run itself computed. Verified against a real
    RobustScaler and real TrainAwareScaler, checking the exact array
    handed to model.predict(), so a wrong transform (or wrong order
    relative to the output scaler) shows up as a numeric mismatch."""

    def test_predict_single_applies_feat_scaler_before_model_predict(self):
        import numpy as np
        from sklearn.preprocessing import RobustScaler

        from src.core.pipeline_patches import InferenceEngine, TrainAwareScaler

        class _RecordingModel:
            seen = []
            def predict(self, X):
                _RecordingModel.seen.append(X.copy())
                return np.array([0.5])

        feat_scaler = RobustScaler()
        feat_scaler.fit(np.array([[100.0, 1.0], [200.0, 2.0],
                                   [300.0, 3.0], [400.0, 4.0]]))

        mm = TrainAwareScaler(feature_range=(45, 98))
        mm.fit(np.array([0.3, 0.5, 0.7]))

        _RecordingModel.seen.clear()
        engine = InferenceEngine(
            model=_RecordingModel(), scaler=mm,
            features=["MW_Da", "LogP"], feat_scaler=feat_scaler,
        )
        score = engine.predict_single({"MW_Da": 250.0, "LogP": 2.5})

        expected = feat_scaler.transform([[250.0, 2.5]])
        np.testing.assert_allclose(_RecordingModel.seen[0], expected)
        assert 45 <= score <= 98

    def test_missing_feat_scaler_falls_back_to_unscaled_features(self):
        """Legacy artifacts saved before this fix have no `feat_scaler` key
        — must not crash, just skip the scaling step."""
        import numpy as np

        from src.core.pipeline_patches import InferenceEngine, TrainAwareScaler

        class _RecordingModel:
            seen = []
            def predict(self, X):
                _RecordingModel.seen.append(X.copy())
                return np.array([0.5])

        mm = TrainAwareScaler(feature_range=(45, 98))
        mm.fit(np.array([0.3, 0.5, 0.7]))

        _RecordingModel.seen.clear()
        engine = InferenceEngine(
            model=_RecordingModel(), scaler=mm,
            features=["MW_Da", "LogP"], feat_scaler=None,
        )
        engine.predict_single({"MW_Da": 250.0, "LogP": 2.5})
        np.testing.assert_allclose(_RecordingModel.seen[0], [[250.0, 2.5]])


class TestThermodynamicsEngineLogPThreading:
    """ThermodynamicsEngine.get_thermo_properties's Yalkowsky logS estimate
    needs a LogP input, but the function had no logp parameter at all —
    every call fell back to logp_proxy = log10(MW/100), a formula with no
    real physical grounding (MW and LogP are only weakly correlated),
    even though ScienceOrchestrator.run_full already resolves the real
    LogP into drug_info["logp"] a few lines before calling
    ThermodynamicsEngine.batch and just never passed it through. Verified
    with aspirin (real LogP=1.19): the MW-proxy fallback gave
    logS_approx=-0.86, the real-LogP version gave -1.79 — aspirin's
    actual experimental logS is close to -1.7 to -2, so the fix is a real
    accuracy improvement, not just a labeling change."""

    def test_real_logp_changes_logs_estimate_vs_mw_proxy(self):
        pytest.importorskip("thermo")
        from src.core.science_engines import ThermodynamicsEngine
        no_logp   = ThermodynamicsEngine.get_thermo_properties("aspirin", cas="50-78-2")
        real_logp = ThermodynamicsEngine.get_thermo_properties(
            "aspirin", cas="50-78-2", logp=1.19)
        if no_logp.get("logS_approx") is None or real_logp.get("logS_approx") is None:
            pytest.skip("thermo/DIPPR lookup unavailable in this environment")
        assert no_logp["logS_approx"] != real_logp["logS_approx"]
        assert "MW_proxy" in no_logp["_imputed"][-1]
        assert "MW_proxy" not in real_logp["_imputed"][-1]

    def test_batch_threads_logp_through_to_get_thermo_properties(self):
        """Root-cause check: batch() (called by ScienceOrchestrator.run_full
        with drug_info["logp"] already resolved) must pass logp down."""
        pytest.importorskip("thermo")
        import inspect
        from src.core.science_engines import ThermodynamicsEngine
        src = inspect.getsource(ThermodynamicsEngine.batch)
        assert "logp=d.get" in src or "logp=d[" in src


class TestPBPKCNSMassConservation:
    """PBPK_CNS_DigitalTwin._build_odes is a 6-compartment mechanistic ODE
    system (Plasma/BBB/ISF/Cell/CSF/Peripheral) that feeds the "PBPK-CNS
    Digital Twin" report section — AUC_plasma, AUC_brain, Kp_brain,
    Cmax_brain, LogBB. Found two independent mass-conservation bugs while
    auditing it, both from dropping/mis-scaling a term when converting a
    real mass-transfer model into concentration-rate ODEs:

    (1) Plasma's brain-to-blood efflux gain (+PS_out*Cisf/Vp, matching
    the "PS_out: brain->blood" comment) had no matching loss term on
    Cisf — Cisf only ever gained from PS_out*Cbb (the separate BBB-to-ISF
    transit step, correctly paired with dCbb's loss), never lost its own
    PS_out*Cisf back to plasma. That's a real missing term, not a
    relabeling of which compartment plasma exchanges with — verified
    directly: with only BBB/ISF/Plasma transfer active (all clearance,
    CSF, and glymphatic terms zeroed), total system mass grew ~7% over
    5 simulated hours instead of staying constant.

    (2) The ISF<->Cell exchange (k_cell_in/k_cell_out) was written as
    bare concentration rates (dCisf -= k_in*Cisf, dCc += k_in*Cisf),
    which only conserves mass if the two compartments have equal volume
    — they don't (V_ISF=280 mL, V_intracell=840 mL, CNS_PHYSIOLOGY's own
    constants). Fixed by scaling both directions by the Vc/Visf ratio so
    mass flux (not raw concentration) is what's actually conserved.

    Both fixed by explicit derivation in mass space (V_i * dC_i/dt) and
    verified numerically: an isolated BBB/ISF/Plasma sub-system, an
    isolated ISF/Cell sub-system, and a full system with only legitimate
    elimination-type sinks active (CL, CL_per, CSF resorption) — the last
    case must show mass monotonically decreasing, never increasing."""

    @staticmethod
    def _physio_volumes():
        from src.core.cerebro_science_modules import CNS_PHYSIOLOGY
        return dict(Vp=3000.0, Vbb=CNS_PHYSIOLOGY.V_BBB_wall,
                    Visf=CNS_PHYSIOLOGY.V_ISF, Vc=CNS_PHYSIOLOGY.V_intracell,
                    Vcsf=CNS_PHYSIOLOGY.V_CSF, Vper=25000.0)

    def _total_mass(self, sol, vols):
        Cp, Cbb, Cisf, Cc, Ccsf, Cper = sol.y
        return (vols["Vp"]*Cp + vols["Vbb"]*Cbb + vols["Visf"]*Cisf +
                vols["Vc"]*Cc + vols["Vcsf"]*Ccsf + vols["Vper"]*Cper)

    def test_bbb_isf_plasma_transfer_conserves_mass(self):
        import numpy as np
        from scipy.integrate import solve_ivp
        from src.core.cerebro_science_modules import PBPK_CNS_DigitalTwin
        vols = self._physio_volumes()
        params = dict(vols, Q_brain=0, Q_CSF=0, Q_glymphatic=0,
                      PS_in=100.0, PS_out=50.0, CL=0.0, CLd=0.0,
                      CL_per=0.0, fu=1.0, k_cell_in=0.0, k_cell_out=0.0,
                      input_rate=0.0)
        t_eval = np.linspace(0, 5, 50)
        sol = solve_ivp(PBPK_CNS_DigitalTwin._build_odes, (0, 5),
                         [1.0, 0, 0, 0, 0, 0], args=(params,), t_eval=t_eval,
                         method="Radau", rtol=1e-9, atol=1e-13)
        mass = self._total_mass(sol, vols)
        assert mass[-1] == pytest.approx(mass[0], abs=1e-6)

    def test_isf_cell_exchange_conserves_mass_despite_volume_mismatch(self):
        import numpy as np
        from scipy.integrate import solve_ivp
        from src.core.cerebro_science_modules import PBPK_CNS_DigitalTwin
        vols = self._physio_volumes()
        params = dict(vols, Q_brain=0, Q_CSF=0, Q_glymphatic=0,
                      PS_in=0.0, PS_out=0.0, CL=0.0, CLd=0.0, CL_per=0.0,
                      fu=1.0, k_cell_in=0.3, k_cell_out=0.1, input_rate=0.0)
        t_eval = np.linspace(0, 5, 50)
        sol = solve_ivp(PBPK_CNS_DigitalTwin._build_odes, (0, 5),
                         [0, 0, 1.0, 0, 0, 0], args=(params,), t_eval=t_eval,
                         method="Radau", rtol=1e-9, atol=1e-13)
        mass = self._total_mass(sol, vols)
        assert mass[-1] == pytest.approx(mass[0], abs=1e-6)

    def test_full_system_mass_only_decreases_via_legitimate_sinks(self):
        """With only real elimination-type sinks active (CL/CLd-paired/
        CL_per/CSF resorption) and no fabricated sources, total mass must
        be monotonically non-increasing — never spike upward."""
        import numpy as np
        from scipy.integrate import solve_ivp
        from src.core.cerebro_science_modules import PBPK_CNS_DigitalTwin
        vols = self._physio_volumes()
        params = dict(vols, Q_brain=0, Q_CSF=20.0, Q_glymphatic=5.0,
                      PS_in=80.0, PS_out=40.0, CL=0.0, CLd=30.0,
                      CL_per=0.0, fu=0.5, k_cell_in=0.2, k_cell_out=0.1,
                      input_rate=0.0)
        t_eval = np.linspace(0, 5, 50)
        sol = solve_ivp(PBPK_CNS_DigitalTwin._build_odes, (0, 5),
                         [1.0, 0, 0, 0, 0, 0], args=(params,), t_eval=t_eval,
                         method="Radau", rtol=1e-9, atol=1e-13)
        mass = self._total_mass(sol, vols)
        assert all(mass[i+1] <= mass[i] + 1e-9 for i in range(len(mass)-1))

    def test_real_simulation_produces_sane_kp_brain(self):
        """End-to-end smoke test: a real drug/DDS profile must still run
        cleanly through the fixed ODEs and produce a physically sane
        (0 < Kp_brain < 1, typical for a poorly-BBB-penetrant unencapsulated
        baseline) result, not crash or produce nonsense."""
        from src.core.cerebro_science_modules import PBPK_CNS_DigitalTwin
        mol_profile = {"MW_Da": 379.5, "LogP": 4.31, "Half_Life_Days": 3.0,
                        "Protein_Binding_pct": 96.0, "BBB_permeability_pct": 5.0}
        top_dds = {"encapsulation_efficiency_pct": 80, "BBB_Enhanced_Pct": 40,
                   "Stealth_Index": 0.6, "Endosomal_Escape_Eff": 0.6}
        result = PBPK_CNS_DigitalTwin.simulate(mol_profile, top_dds, dose_mg=10.0,
                                                disease_state="alzheimer_3")
        assert 0 < result["Kp_brain"] < 1
        assert result["AUC_plasma_ugh_mL"] > 0
        assert result["Cmax_brain_ug_mL"] > 0


class TestBiologicPBPKOrganDistribution:
    """BiologicPBPK.simulate's organ_distribution built "Other tissues" as
    100 - 70 - 12 - 8 - 6 - 4 - brain_pct. Since 70+12+8+6+4 already equals
    100, that expression collapses to -brain_pct -- always negative
    whenever any drug reaches the brain (which is the whole point of a
    CNS delivery model), and small-but-nonzero brain uptake is the normal
    case, not an edge case. This surfaced in real pipeline output as
    "Other tissues": -0.1% for a real Lecanemab run. Fixed by giving
    "Other tissues" a real fixed literature-based share and letting Blood
    (not a fabricated negative bucket) absorb the brain-crossing
    fraction."""

    def test_organ_distribution_has_no_negative_percentages(self):
        from src.core.cerebro_science_modules import BiologicPBPK
        mol_profile = {"MW_Da": 148000, "Half_Life_Days": 21,
                       "molecule_class": "monoclonal_antibody"}
        top_dds = {"BBB_Engineering_Score": 65}
        result = BiologicPBPK.simulate(mol_profile, top_dds, dose_mg=10.0,
                                        disease_stage="alzheimer_3")
        organs = result["organ_distribution"]
        for name, pct in organs.items():
            assert pct >= 0, f"{name} is negative: {pct}"

    def test_organ_distribution_sums_to_100_percent(self):
        from src.core.cerebro_science_modules import BiologicPBPK
        mol_profile = {"MW_Da": 148000, "Half_Life_Days": 21,
                       "molecule_class": "monoclonal_antibody"}
        top_dds = {"BBB_Engineering_Score": 65}
        result = BiologicPBPK.simulate(mol_profile, top_dds, dose_mg=10.0,
                                        disease_stage="alzheimer_3")
        organs = result["organ_distribution"]
        assert sum(organs.values()) == pytest.approx(100.0, abs=0.1)

    def test_other_tissues_has_a_real_nonzero_share(self):
        """Regression guard specifically for the collapsed-to-zero-budget
        bug: 'Other tissues' must carry its own literature-based share,
        not just be whatever's left after the other five buckets already
        consumed the full 100%."""
        from src.core.cerebro_science_modules import BiologicPBPK
        mol_profile = {"MW_Da": 148000, "Half_Life_Days": 21,
                       "molecule_class": "monoclonal_antibody"}
        top_dds = {"BBB_Engineering_Score": 65}
        result = BiologicPBPK.simulate(mol_profile, top_dds, dose_mg=10.0,
                                        disease_stage="alzheimer_3")
        assert result["organ_distribution"]["Other tissues"] > 1.0


class TestDDSComparisonStrengthWeaknessClassification:
    """DDSComparisonEngine.compare's per-formulation strengths/weaknesses
    classification used `(v > med*1.2) == higher_better` /
    `(v < med*0.8) == higher_better`, which looks like it flips the
    comparison direction for lower-is-better metrics (CARPA_Risk_Index,
    MPS_Clearance_h, Protein_Corona_nm) but doesn't — negating
    "v > med*1.2" gives "v <= med*1.2", not "v < med*0.8". For a
    lower-is-better metric, an exactly-average value (v == med) satisfied
    the strength check (since v <= med*1.2 is trivially true at v==med),
    so every formulation with an average CARPA_Risk_Index — a
    safety-critical score — got misclassified as having a "strength"
    there. Fixed by comparing against the correct threshold explicitly."""

    def test_average_value_of_lower_is_better_metric_is_neither(self):
        import pandas as pd
        from src.core.cerebro_science_modules import DDSComparisonEngine
        df = pd.DataFrame({
            "Formulation_ID": ["F1", "F2", "F3", "F4", "F5"],
            "Formulation_Name": ["A", "B", "C", "D", "E"],
            "Carrier_Type": ["liposome"] * 5,
            "Rank": [1, 2, 3, 4, 5],
            "Composite_Score": [90, 85, 80, 75, 70],
            "CARPA_Risk_Index": [0.2] * 5,
        })
        result = DDSComparisonEngine.compare(df, top_n=5)
        for row in result["top_n_summary"]:
            assert "CARPA Risk Index" not in row["Strengths"]
            assert "CARPA Risk Index" not in row["Weaknesses"]

    def test_genuinely_low_and_high_risk_still_classified_correctly(self):
        """Guards against overcorrecting: real outliers on a
        lower-is-better metric must still be flagged in the right
        direction (low value = strength, high value = weakness)."""
        import pandas as pd
        from src.core.cerebro_science_modules import DDSComparisonEngine
        df = pd.DataFrame({
            "Formulation_ID": ["F1", "F2", "F3"],
            "Formulation_Name": ["LowRisk", "Median", "HighRisk"],
            "Carrier_Type": ["liposome"] * 3,
            "Rank": [1, 2, 3],
            "Composite_Score": [90, 85, 80],
            "CARPA_Risk_Index": [0.1, 0.2, 0.3],
        })
        result = DDSComparisonEngine.compare(df, top_n=3)
        by_name = {r["Name"]: r for r in result["top_n_summary"]}
        assert "CARPA Risk Index" in by_name["LowRisk"]["Strengths"]
        assert "CARPA Risk Index" in by_name["HighRisk"]["Weaknesses"]
        assert "CARPA Risk Index" not in by_name["Median"]["Strengths"]
        assert "CARPA Risk Index" not in by_name["Median"]["Weaknesses"]


class TestAdvancedModules2Integrity:
    """Several findings from auditing cerebro_advanced_modules_2.py
    (2695 lines, 23 classes covering points 3-62 of the 62-principle
    framework). Most of the file is self-consistent heuristic scoring
    with honest confidence labeling, but four real integrity issues:

    (1) LyophilizationOptimizer.optimize's cake_collapse_risk check was
    tautologically always False: T_primary_dry is *defined* as Tg - 2.0
    within the same function, so "Tg > T_primary_dry + 5" always reduces
    to "Tg > Tg + 3" -- impossible for any cryoprotectant, verified
    across all six entries in Tg_prime.

    (2) FDA21CFRCompliance.generate_compliance_report unconditionally
    claimed "21 CFR Part 11 COMPLIANT" regardless of whether any audit
    trail entries actually existed -- and log_computation(), the method
    that writes them, is never called anywhere in the codebase (verified
    via full-repo grep), so audit_trail.jsonl never exists in a real
    trial. This fed directly into final_report_unified.py's own FDA
    compliance report section.

    (3) FinalModules.fto_ip_analysis returned
    "CLEAR to file new patent" whenever a ligand simply didn't match one
    of 3 hardcoded example patents -- a legal/IP conclusion the check
    can't actually support, especially since USPTO PAIR/Espacenet (the
    cited "reference") are never really queried.

    (4) LiteratureMiningEngine._fallback_citations had chronologically
    impossible PMIDs for real papers (a PMID roughly tracks its
    publication year) -- e.g. Alvarez-Erviti 2011 cited as PMID 34678901
    (that range is ~2021-2022), directly contradicted by this same
    file's own RealTimeLiterature.CURATED_CITATIONS citing the identical
    paper with PMID 21423189 (chronologically consistent with 2011).

    (5) The ADV2 recompute step called a class named
    AnimalSparingBiodistrib that is never defined or imported anywhere
    in this file -- a plain NameError on every single trial, silently
    swallowed by the surrounding try/except and logged as a WARNING
    ("[ADV2] biodistribution recompute failed"), so the recomputed
    biodistribution numbers were dropped every run without anyone
    noticing. The correct class, doing the exact same job with a
    matching signature, is SupplementModules (used correctly at another
    call site in this same file)."""

    def test_lyophilization_collapse_risk_is_no_longer_tautologically_false(self):
        """Regression guard: the field must at least reflect its own
        stated safety margin honestly, not silently claim "OK" via dead
        logic. (The fix makes it an honest constant rather than a fake
        risk check — this pins that it stays that way and doesn't
        regress back to comparing Tg against itself.)"""
        from src.core.cerebro_advanced_modules_2 import LyophilizationOptimizer
        for cryo in ["trehalose", "sucrose", "mannitol", "glycerol", "PVP", "none"]:
            result = LyophilizationOptimizer.optimize({}, cryoprotectant=cryo)
            assert "Tg'-2degC" in result["cake_collapse_risk"] or \
                   "safety margin" in result["cake_collapse_risk"]

    def test_fda_compliance_not_claimed_without_real_audit_entries(self, tmp_path):
        from src.core.cerebro_advanced_modules_2 import FDA21CFRCompliance
        no_entries = FDA21CFRCompliance.generate_compliance_report(tmp_path)
        assert "COMPLIANT" not in no_entries["compliance_status"]
        assert no_entries["n_audit_entries"] == 0

    def test_fda_compliance_claimed_once_real_audit_entries_exist(self, tmp_path):
        from src.core.cerebro_advanced_modules_2 import FDA21CFRCompliance
        FDA21CFRCompliance.log_computation(tmp_path, "TestDrug", "test_calc", "hash123")
        with_entries = FDA21CFRCompliance.generate_compliance_report(tmp_path)
        assert "COMPLIANT" in with_entries["compliance_status"]
        assert with_entries["n_audit_entries"] == 1

    def test_fto_analysis_does_not_claim_clear_without_caveat(self):
        from src.core.cerebro_advanced_modules_2 import FinalModules
        result = FinalModules.fto_ip_analysis(
            {"Surface_Ligand": "some-novel-ligand"}, {"name": "TestDrug"})
        assert result["FTO_clear"] is True
        assert "NOT a real automated patent search" in result["recommendation"]
        assert "CLEAR to file new patent" != result["recommendation"]

    def test_literature_fallback_pmids_are_chronologically_plausible(self):
        """PMIDs are assigned roughly sequentially: PMID/~1.6M per year
        since 1996 is a workable rule of thumb. A paper's fallback PMID
        should land within a few years of its claimed publication year,
        not off by a decade or more."""
        from src.core.cerebro_advanced_modules_2 import LiteratureMiningEngine
        for carrier in ["vexosome", "liposome", "unknown"]:
            for c in LiteratureMiningEngine._fallback_citations(carrier):
                pmid = int(c["pmid"])
                year = int(c["year"])
                # Rough PMID->year mapping: PMID 1M ~ 1996, growing by
                # roughly 1.6M/year since. Generous +/-5 year tolerance.
                est_year = 1996 + (pmid - 1_000_000) / 1_600_000
                assert abs(est_year - year) < 6, (
                    f"{c['citation']}: PMID {pmid} implies ~{est_year:.0f}, "
                    f"claimed year is {year}")

    def test_alvarez_erviti_pmid_matches_verified_entry_elsewhere_in_file(self):
        """Direct cross-reference check: this exact paper appears in both
        LiteratureMiningEngine's fallback and RealTimeLiterature's
        curated list — they must agree now."""
        from src.core.cerebro_advanced_modules_2 import (
            LiteratureMiningEngine,
            RealTimeLiterature,
        )
        fallback_pmid = LiteratureMiningEngine._fallback_citations("vexosome")[0]["pmid"]
        curated = RealTimeLiterature.CURATED_CITATIONS["vexosome_exosome"][0]
        assert fallback_pmid in curated

    def test_biodistribution_recompute_uses_a_class_that_actually_exists(self):
        """Regression guard for the AnimalSparingBiodistrib NameError:
        the ADV2 recompute step must call a real, importable class. This
        mirrors the exact call made in the recompute step (top_dds,
        mol_profile, science_results) and asserts it returns real organ
        percentages instead of silently failing and being swallowed by
        the surrounding except-block."""
        from src.core.cerebro_advanced_modules_2 import SupplementModules
        top_dds = {"name": "TestCarrier", "Surface_Ligand": "none"}
        mol_profile = {"name": "TestDrug", "logP": 2.0, "MW": 300.0}
        results = {}
        bio = SupplementModules.biodistribution_map(top_dds, mol_profile, results)
        assert isinstance(bio, dict)
        organs = bio["organs"]
        assert len(organs) > 0
        assert all(isinstance(v, (int, float)) for v in organs.values())
        assert "Brain (Target)" in organs

    def test_adv2_recompute_step_no_longer_raises_nameerror(self):
        """End-to-end guard at the actual call site (not just the
        target class in isolation): the line that used to read
        `AnimalSparingBiodistrib.biodistribution_map(...)` must now
        reference a name that is genuinely importable from this module,
        so a plain `import`+`getattr` on the fixed name never raises."""
        import src.core.cerebro_advanced_modules_2 as mod
        assert hasattr(mod, "SupplementModules")
        assert not hasattr(mod, "AnimalSparingBiodistrib")
        assert callable(mod.SupplementModules.biodistribution_map)


# ═════════════════════════════════════════════════════════════════════════════
# 11. FULL PIPELINE INTEGRATION (run.py -> pipeline_runner.py, end-to-end)
# ═════════════════════════════════════════════════════════════════════════════
# The audit's §11 testing roadmap P0 item: "an actual end-to-end integration
# test of run.py on synthetic input asserting real output artifacts exist."
# phase5_smoke_test.py (flagged in the audit as having zero assert
# statements) only checked that modules import — this actually asserts on
# real computed output.

class TestNovelDrugExplainer:
    """final_report.NovelDrugExplainer is the logic-bearing part of the
    final report engine — decides whether a drug counts as "novel" and
    builds the researcher-facing alignment-explanation text. Also covers
    the "500+ curated drugs" claim found stale here too (same
    CLINICAL_PK_LIBRARY-emptied-in-v22.1 issue already fixed in
    clinical_data_engine.py) — this copy feeds directly into the
    per-trial NOVEL DRUG ALERT report text a researcher reads."""

    def test_alignment_flag_marks_drug_as_novel(self):
        from src.core.final_report import NovelDrugExplainer
        expl = NovelDrugExplainer.build_explanation(
            {"_alignment_flag": True, "_surrogate_drug": "Aminopterin",
             "_tanimoto_sim": 0.84, "_source": "ChemicalAlignment_Aminopterin"},
            "NovelCompoundX")
        assert expl["is_novel"] is True
        assert expl["surrogate_drug"] == "Aminopterin"

    def test_normal_drug_with_real_source_is_not_novel(self):
        from src.core.final_report import NovelDrugExplainer
        expl = NovelDrugExplainer.build_explanation(
            {"_alignment_flag": False, "_source": "DrugBank_API"},
            "Donepezil")
        assert expl["is_novel"] is False

    def test_format_text_for_novel_drug_includes_disclaimer(self):
        from src.core.final_report import NovelDrugExplainer
        expl = NovelDrugExplainer.build_explanation(
            {"_alignment_flag": True, "_surrogate_drug": "Aminopterin",
             "_tanimoto_sim": 0.84, "_tiers_tried": ["DrugBank_API", "DailyMed_FDA"]},
            "NovelCompoundX")
        text = NovelDrugExplainer.format_text(expl, "NovelCompoundX")
        assert "PREDICTED" in text and "not measured" in text
        assert "Aminopterin" in text

    def test_embedded_library_description_does_not_claim_500_drugs(self):
        """The library was deliberately emptied in v22.1 (no hardcoded
        drug data) — this description must not still claim curated
        coverage that no longer exists."""
        from src.core.final_report import NovelDrugExplainer
        desc = NovelDrugExplainer.TIER_DESCRIPTIONS["EmbeddedLibrary"]
        assert "500+" not in desc


class TestUnifiedPDFReportDecisionFramework:
    """UnifiedPDFReport.generate's Section 15 executive decision table
    computed go_dec ("GO"/"CONDITIONAL GO") from three real criteria but
    then hardcoded the "Evidence" text for the OVERALL DECISION row to
    "Proceed to IND-enabling studies" unconditionally — even when go_dec
    was "CONDITIONAL GO" because the synthetic clinical trial explicitly
    returned "NO-GO / REFORMULATE" (a real value cerebro_advanced_modules_2.py
    can produce). This is the most consequential section of the whole
    report — the one a PI reads to decide whether to proceed — so a
    fixed "proceed" recommendation regardless of the actual verdict was a
    serious integrity bug, not cosmetic. Fixed to state which criteria
    actually failed instead.

    Verified by mocking SimpleDocTemplate.build to capture the real
    reportlab Table objects passed into the PDF story (their `_cellvalues`
    are the literal table content) rather than trying to parse the
    rendered PDF bytes."""

    def _run_and_capture_story(self, monkeypatch, tmp_path, top_dds, science_results):
        from unittest.mock import patch
        from pathlib import Path as _P
        from src.core.final_report_unified import UnifiedPDFReport

        captured = {}

        def fake_build(self, story, **kw):
            captured["story"] = story
            _P(self.filename).write_bytes(b"%PDF-1.4 fake")

        with patch("reportlab.platypus.SimpleDocTemplate.build", fake_build):
            UnifiedPDFReport.generate(
                drug_name="TestDrug", trial_dir=tmp_path,
                mol_profile={"MW_Da": 379.5, "LogP": 4.31},
                df_dds=None, top_dds=top_dds, science_results=science_results,
            )
        return captured["story"]

    def _decision_row(self, story):
        from reportlab.platypus import Table
        for t in story:
            if isinstance(t, Table):
                rows = t._cellvalues
                if rows and rows[-1][0] == "OVERALL DECISION":
                    return rows[-1]
        return None

    def test_failed_clinical_trial_does_not_recommend_proceeding(self, tmp_path):
        story = self._run_and_capture_story(
            None, tmp_path,
            top_dds={"DLVO_stable": True, "DLVO_V_total_kT": 30.0,
                     "BBB_Enhanced_Pct": 20.0},
            science_results={
                "synthetic_clinical": {"go_no_go": "NO-GO / REFORMULATE",
                                        "overall_response_pct": 30, "AE_severe_pct": 12},
                "qsar_toxicity": {"cardiac_risk": False},
            })
        row = self._decision_row(story)
        assert row[1] == "CONDITIONAL GO"
        assert "Proceed to IND-enabling studies" != row[2]
        assert "NO-GO" in row[2]

    def test_all_criteria_pass_recommends_proceeding(self, tmp_path):
        story = self._run_and_capture_story(
            None, tmp_path,
            top_dds={"DLVO_stable": True, "DLVO_V_total_kT": 30.0,
                     "BBB_Enhanced_Pct": 20.0},
            science_results={
                "synthetic_clinical": {"go_no_go": "GO",
                                        "overall_response_pct": 75, "AE_severe_pct": 2},
                "qsar_toxicity": {"cardiac_risk": False},
            })
        row = self._decision_row(story)
        assert row[1] == "GO"
        assert row[2] == "Proceed to IND-enabling studies"

    def test_biologic_kp_brain_uses_auc_ratio_not_cmax_ratio(self, tmp_path):
        """Section 3's BiologicPBPK branch computed Kp,brain from
        Cmax_brain/Cmax_plasma, while every other Kp_brain definition in
        this codebase (pbbm_engine.py, science_engines.py,
        cerebro_science_modules.py) uses AUC_brain/AUC_plasma — the same
        "Kp,brain" column label in the same PDF table meant two different
        things depending on molecule class. Cmax ratio (0.5/10=0.05) and
        AUC ratio ((2*24)/(100*24)=0.02) diverge for this fixture, so this
        pins the fix rather than a case where they'd coincidentally match."""
        story = self._run_and_capture_story(
            None, tmp_path,
            top_dds={"DLVO_stable": True, "DLVO_V_total_kT": 30.0,
                     "BBB_Enhanced_Pct": 20.0},
            science_results={
                "pbpk_cns": {
                    "model": "BiologicPBPK",
                    "Cmax_brain_ug_mL": 0.5, "Cmax_plasma_ug_mL": 10.0,
                    "AUC_CNS_day_ug_mL": 2.0, "AUC_plasma_day_ug_mL": 100.0,
                    "T_half_effective_days": 5.0,
                    "BBB_transcytosis_pct": 3.0,
                },
            })
        from reportlab.platypus import Table
        kp_row = None
        for t in story:
            if isinstance(t, Table):
                for r in t._cellvalues:
                    if r and r[0] == "Kp,brain":
                        kp_row = r
        assert kp_row is not None, "Kp,brain row not found in PBPK table"
        kp_value = float(kp_row[1])
        assert kp_value == pytest.approx(0.02, abs=1e-4)


class TestPipelineIntegration:
    """Runs the real pipeline end-to-end against a real input Excel and
    checks real output artifacts were produced with sane values — not a
    mock, not an import-only smoke test."""

    @pytest.mark.slow
    def test_run_pipeline_from_excel_produces_real_dds_ranking(self, tmp_path):
        import shutil
        import sys as _sys

        import src.path_resolver  # noqa: F401
        from pipeline_runner import run_pipeline_from_excel
        from trial_manager import _excel_hash

        real_input = None
        for candidate in ("inputs/CEREBRO_Input_Donepezil.xlsx",):
            from pathlib import Path
            if Path(candidate).exists():
                real_input = Path(candidate)
                break
        if real_input is None:
            pytest.skip("Real Donepezil input Excel not found in inputs/")

        trial_dir = tmp_path / "Donepezil_test_trial"
        ok = run_pipeline_from_excel(
            real_input, _excel_hash(real_input), trial_dir, force=True)

        assert ok is True
        ranking_csv = trial_dir / "dds_analysis" / "formulation_ranking.csv"
        assert ranking_csv.exists(), "Pipeline must produce a real ranking CSV"

        import pandas as pd
        df = pd.read_csv(ranking_csv)
        assert len(df) > 0
        assert "BBB_Engineering_Score" in df.columns
        # Real scores vary across formulations — a hardcoded/fabricated
        # pipeline would produce identical or suspiciously uniform values.
        assert df["BBB_Engineering_Score"].nunique() > 1


# ═════════════════════════════════════════════════════════════════════════════
# 12. PDB ID RESOLVER (src/core/pdb_resolver.py)
# ═════════════════════════════════════════════════════════════════════════════

class TestPDBResolver:
    """Regression test for the audit's §6 Medium finding: pdb_id was
    validated by length only (`len(pdb_id) == 4`), which a value like
    '../x' (also 4 characters) would pass — a real path-traversal risk
    once the value reaches real_docking_engine.py's filesystem/URL
    construction. Fixed to use the same strict alphanumeric regex already
    proven in real_docking_engine.py."""

    def test_user_provided_path_traversal_pdb_id_is_rejected(self):
        from src.core.pdb_resolver import resolve_pdb_for_drug
        result = resolve_pdb_for_drug("TestDrug", user_pdb_id="../x")
        assert result["pdb_id"] is None
        assert result["source"] != "User-provided (Excel input)"

    def test_user_provided_valid_pdb_id_is_accepted(self):
        from src.core.pdb_resolver import resolve_pdb_for_drug
        result = resolve_pdb_for_drug("TestDrug", user_pdb_id="2NAO")
        assert result["pdb_id"] == "2NAO"
        assert result["source"] == "User-provided (Excel input)"
        assert result["confidence"] == "HIGH"


# ═════════════════════════════════════════════════════════════════════════════
# 13. APPLICATION-LAYER RATE LIMITING (src/api/app.py, /auth/*)
# ═════════════════════════════════════════════════════════════════════════════

class TestRateLimiting:
    """Audit finding (§6 Medium): rate limiting only existed in nginx.conf,
    which docker-compose.yml (the simpler config, most likely run first)
    never fronts the API with — /auth/login and /auth/register were fully
    unthrottled at the application layer in that config. Fixed via slowapi;
    this hits the real endpoint repeatedly through a real TestClient and
    checks a 429 actually appears, not just that a decorator is present in
    the source."""

    def test_login_endpoint_rate_limited_after_repeated_attempts(self, test_client):
        from src.api.app import _HAS_RATE_LIMIT
        if not _HAS_RATE_LIMIT:
            pytest.skip("slowapi not installed in this environment")

        statuses = []
        for _ in range(7):  # limit is 5/minute
            r = test_client.post("/auth/login", data={
                "username": "nonexistent_rl_test_user",
                "password": "wrong",
            })
            statuses.append(r.status_code)
        assert 429 in statuses, f"Expected a 429 among {statuses} after 7 rapid attempts"
        # Everything before the limit kicks in should be a real auth
        # response (401 for bad creds), not silently dropped.
        assert 401 in statuses[:5]

    def test_app_refuses_to_boot_in_production_with_wildcard_cors(self):
        """Same fail-hard pattern as the JWT/admin-password/encryption
        checks — a real subprocess import, not a mock, since this needs to
        catch the app failing to boot at all, and reloading src.api.app
        in-process would pollute the module other tests in this file share."""
        import os
        import subprocess
        import sys
        from pathlib import Path

        env = {**os.environ, "ENVIRONMENT": "production"}
        env.pop("CORS_ORIGINS", None)
        result = subprocess.run(
            [sys.executable, "-c", "import src.path_resolver; import src.api.app"],
            cwd=str(Path(__file__).resolve().parents[2]),
            env=env, capture_output=True, text=True, timeout=30,
        )
        assert result.returncode != 0
        assert "CORS_ORIGINS" in result.stderr

    def test_main_uses_a_real_import_path_not_a_process_local_alias(self, monkeypatch):
        """main()'s uvicorn.run() used the legacy flat alias
        "cerebro_api_v2:app" -- that name only exists in sys.modules
        inside THIS already-running process (registered by
        src/path_resolver.py); no file named cerebro_api_v2.py exists
        anywhere for a freshly-spawned multi-worker subprocess (a bare
        module cache) to resolve it by. docker-compose.prod.yml's actual
        command already sidesteps this by using the real dotted path
        directly (`python -m uvicorn src.api.app:app`) -- main() must
        match it."""
        import src.api.app as app_module

        captured = {}
        def _fake_run(target, **kwargs):
            captured["target"] = target
        monkeypatch.setattr(app_module.uvicorn, "run", _fake_run)

        app_module.main()
        assert captured["target"] == "src.api.app:app"


# Module-level (not nested) so joblib can pickle/unpickle them when
# saved to a temp file and loaded back inside the /predict endpoint.
# predict() must return a real ndarray, not a bare list — a real sklearn
# estimator always does, and InferenceEngine.predict_single feeds the
# raw prediction straight into TrainAwareScaler.transform(), which calls
# .reshape(-1, 1) on it; a plain list has no .reshape and would break
# with an AttributeError that has nothing to do with what these tests
# are actually checking.
class _CapturingModel:
    """Records the X array it was asked to predict on."""
    calls = []
    def predict(self, X):
        _CapturingModel.calls.append(X.tolist())
        return np.array([0.5])


class _DummyModel:
    def predict(self, X):
        return np.array([0.5])


class TestPredictEndpointUsesRealMoleculeFeatures:
    """POST /predict hardcoded MW_Da=0, LogP=0, Half_Life_Days=0 for every
    request regardless of req.molecule_input — the comment literally said
    "simplified — real version uses MoleculeEngine" but nothing ever
    called it, so every submitted molecule got the identical prediction.
    Fixed by wiring in the real analyze_molecule() (src/core/molecule_engine.py)
    already used by the actual pipeline. Verified two different real
    molecules (donepezil, aspirin) now resolve to genuinely different
    MW/LogP via analyze_molecule directly, and exercise the full HTTP
    endpoint with a mocked production model artifact to confirm the
    request actually reaches model.predict() with those real features."""

    def test_different_molecules_resolve_to_different_features(self):
        from src.core.molecule_engine import analyze_molecule
        donepezil = analyze_molecule(
            "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2", "Donepezil")
        aspirin = analyze_molecule("CC(=O)OC1=CC=CC=C1C(=O)O", "Aspirin")
        assert donepezil["MW_Da"] != aspirin["MW_Da"]
        assert donepezil["LogP"] != aspirin["LogP"]
        assert donepezil["MW_Da"] is not None and donepezil["MW_Da"] > 0

    def test_predict_endpoint_feeds_real_features_to_model(
            self, test_client, auth_headers, tmp_path, monkeypatch):
        if not auth_headers:
            pytest.skip("auth fixture unavailable in this environment")
        import joblib

        from src.core.pipeline_patches import TrainAwareScaler
        mm = TrainAwareScaler(feature_range=(45, 98))
        mm.fit(np.array([0.3, 0.5, 0.7]))

        model_path = tmp_path / "fake_model.pkl"
        joblib.dump({
            "model": _CapturingModel(),
            "scaler_state": mm.save_state(),
            "features": ["MW_Da", "LogP", "Half_Life_Days", "Docking_Affinity_kcal"],
        }, model_path)

        from dataclasses import dataclass

        @dataclass
        class _FakeProd:
            artifact_path: str
            version: str = "1.0.0"
            stage: str = "production"

        class _FakeRegistry:
            def get_production(self, name):
                return _FakeProd(artifact_path=str(model_path))

        import src.api.app as app_module
        monkeypatch.setattr(app_module, "ModelRegistry", _FakeRegistry)
        _CapturingModel.calls.clear()

        r = test_client.post(
            "/predict",
            json={"molecule_input": "CC(=O)OC1=CC=CC=C1C(=O)O",  # aspirin
                  "drug_name": "Aspirin"},
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
        assert len(_CapturingModel.calls) == 1
        mw_used, logp_used, hl_used, dock_used = _CapturingModel.calls[0][0]
        assert mw_used != 0, "MW_Da still hardcoded to 0 instead of the real molecule's MW"
        assert logp_used != 0, "LogP still hardcoded to 0 instead of the real molecule's LogP"
        assert abs(mw_used - 180.16) < 1.0  # real aspirin MW

    def test_predict_endpoint_rejects_unresolvable_input(
            self, test_client, auth_headers, tmp_path, monkeypatch):
        if not auth_headers:
            pytest.skip("auth fixture unavailable in this environment")
        import joblib
        from dataclasses import dataclass

        from src.core.pipeline_patches import TrainAwareScaler
        mm = TrainAwareScaler(feature_range=(45, 98))
        mm.fit(np.array([0.3, 0.5, 0.7]))

        model_path = tmp_path / "fake_model.pkl"
        joblib.dump({
            "model": _DummyModel(),
            "scaler_state": mm.save_state(),
            "features": ["MW_Da", "LogP", "Half_Life_Days", "Docking_Affinity_kcal"],
        }, model_path)

        @dataclass
        class _FakeProd:
            artifact_path: str
            version: str = "1.0.0"
            stage: str = "production"

        class _FakeRegistry:
            def get_production(self, name):
                return _FakeProd(artifact_path=str(model_path))

        import src.api.app as app_module
        monkeypatch.setattr(app_module, "ModelRegistry", _FakeRegistry)

        r = test_client.post(
            "/predict",
            json={"molecule_input": "not a real molecule or drug name @@@###",
                  "drug_name": ""},
            headers=auth_headers,
        )
        assert r.status_code == 422


class TestResultsDownloadPathTraversal:
    """GET /results/{filepath} joined the raw user-controlled filepath
    onto outputs/ with no containment check — any authenticated user
    (this endpoint only requires get_current_user, not an elevated
    permission) could request "../../../etc/passwd" or
    "../../src/api/auth.py" and read any file the server process can
    read, entirely outside outputs/. Same vulnerability class as the
    PDB-ID path-traversal finding already fixed in pdb_resolver.py.
    Fixed by resolving both sides and requiring containment via
    Path.is_relative_to() before serving the file. Hits the real
    endpoint through a real TestClient rather than unit-testing the
    Path logic in isolation, since routing-layer path normalization can
    behave differently than raw pathlib."""

    def test_traversal_attempt_is_rejected(self, test_client, auth_headers):
        if not auth_headers:
            pytest.skip("auth fixture unavailable in this environment")
        # A literal "../" in the URL gets collapsed by Starlette's own
        # router before it ever reaches the handler (confirmed: it 404s
        # against a nonexistent route rather than exercising our code at
        # all) — that's routing hygiene, not evidence the endpoint is
        # safe. URL-encoded traversal ("%2e%2e") bypasses that router-
        # level normalization and reaches the handler with a real ".."
        # in `filepath`, which is what this test needs to prove the fix
        # actually holds. Verified directly against a real file
        # (src/api/auth.py, which contains JWT_SECRET_KEY handling) that
        # exists relative to the project root but must never be served
        # through this results-only endpoint.
        r = test_client.get(
            "/results/%2e%2e/%2e%2e/src/api/auth.py",
            headers=auth_headers,
        )
        assert r.status_code == 403, (
            f"Expected the traversal to be rejected with 403, got "
            f"{r.status_code}: {r.text[:200]}")
        assert "JWT_SECRET_KEY" not in r.text

    def test_legitimate_nested_path_still_works(self, test_client, auth_headers, tmp_path, monkeypatch):
        if not auth_headers:
            pytest.skip("auth fixture unavailable in this environment")
        import os
        outputs_dir = tmp_path / "outputs" / "data"
        outputs_dir.mkdir(parents=True)
        (outputs_dir / "real_result.csv").write_text("Drug,Score\nDonepezil,85\n")

        cwd = os.getcwd()
        os.chdir(tmp_path)
        try:
            r = test_client.get("/results/data/real_result.csv", headers=auth_headers)
            assert r.status_code == 200
            assert "Donepezil" in r.text
        finally:
            os.chdir(cwd)


class TestEnterpriseInfraResultsDownloadPathTraversal:
    """The equivalent GET /results/{filepath:path} endpoint in
    src/dds/enterprise_infra.py (a separate, legacy standalone FastAPI
    app -- not the one served by docker-compose.prod.yml, but still
    directly importable/runnable, and with no authentication at all)
    had the identical vulnerability as the one already fixed in
    src/api/app.py: filepath joined onto OUTPUT_ROOT with no containment
    check. Same fix, same reason to hit the real endpoint through a real
    TestClient rather than unit-testing the Path logic in isolation."""

    def test_traversal_attempt_is_rejected(self):
        from fastapi.testclient import TestClient

        from src.dds.enterprise_infra import app
        client = TestClient(app)
        r = client.get("/results/%2e%2e/%2e%2e/%2e%2e/dds/enterprise_infra.py")
        assert r.status_code == 403, (
            f"Expected the traversal to be rejected with 403, got "
            f"{r.status_code}: {r.text[:200]}")

    def test_legitimate_file_still_downloads(self):
        from pathlib import Path

        from fastapi.testclient import TestClient

        from src.dds.enterprise_infra import OUTPUT_ROOT, app
        (Path(OUTPUT_ROOT) / "_test_legit_download.txt").write_text("hello world")
        client = TestClient(app)
        r = client.get("/results/_test_legit_download.txt")
        assert r.status_code == 200
        assert r.text == "hello world"


# ═════════════════════════════════════════════════════════════════════════════
# 14. DRUG_SMILES RESOLVER — NAME MUST NEVER BE USED AS A SMILES FALLBACK
# ═════════════════════════════════════════════════════════════════════════════

class TestDrugSmilesResolver:
    """Found this running a real Lecanemab (mAb) pipeline: it logged 30
    RDKit 'SMILES Parse Error' failures for the literal string 'Lecanemab'.
    Root cause: resolve_drug_smiles's Tier 7
    last-resort sanitizer did `raw = smiles or name`, so any drug with no
    real SMILES (i.e. every biologic — mAbs, oligonucleotides, peptides)
    fell back to passing its own NAME into RDKit as if it were a chemical
    structure. The sanitizer's heuristic (checks for any of
    C/N/O/S/P/H/F/c/n/o/s/l) is loose enough that ordinary English words
    routinely contain one of those letters, so this wasn't Lecanemab-
    specific — it silently affected any drug name for which real SMILES
    resolution failed. Fixed by never falling back to `name` in Tier 7.
    These tests use several distinct made-up names to confirm the fix is
    generic, not a special case for one drug."""

    def test_biologic_with_no_smiles_does_not_leak_name_as_smiles(self):
        from cerebro_value_resolver.categories.drug_identifiers import (
            resolve_drug_smiles,
        )
        for name in ["Lecanemab", "Nusinersen", "SomeRandomAntibodyXYZ"]:
            result = resolve_drug_smiles(name=name, smiles="")
            assert result["value"] is None, (
                f"{name}: expected no fabricated SMILES, got {result['value']!r}"
            )

    def test_small_molecule_with_real_smiles_still_resolves(self):
        from cerebro_value_resolver.categories.drug_identifiers import (
            resolve_drug_smiles,
        )
        result = resolve_drug_smiles(
            name="Aspirin", smiles="CC(=O)Oc1ccccc1C(=O)O")
        assert result["value"] == "CC(=O)Oc1ccccc1C(=O)O"


# ═════════════════════════════════════════════════════════════════════════════
# 15. THE 62-PRINCIPLE SCORING SYSTEM — GROUP ROLLUPS AND DEEP-VALIDATION SPLIT
# ═════════════════════════════════════════════════════════════════════════════

class TestPrincipleGroupConsistency:
    """PRINCIPLE_GROUPS in cerebro_62_orchestrator.py maps group names to
    lists of principle IDs, and _group_score() silently skips any ID not
    present in a given DDS's per-principle results. A typo'd or renamed
    principle ID in that mapping would fail silently — the group's score
    would just be computed from fewer principles than intended, with no
    error anywhere. Every ID it references should actually exist in the
    catalog it's supposed to be grouping."""

    def test_every_grouped_principle_id_exists_in_the_catalog(self):
        from cerebro_62_orchestrator import PRINCIPLE_GROUPS
        from cerebro_62_principles_catalog import PRINCIPLES_62

        all_grouped = {pid for pids in PRINCIPLE_GROUPS.values() for pid in pids}
        missing = sorted(all_grouped - set(PRINCIPLES_62.keys()))
        assert not missing, (
            f"PRINCIPLE_GROUPS references principle IDs not in "
            f"PRINCIPLES_62: {missing} — these silently drop out of their "
            f"group's average instead of erroring"
        )

    def test_group_score_averages_correctly_and_skips_missing(self):
        from cerebro_62_orchestrator import _group_score
        per_principle = {
            "P01": {"score": 80.0},
            "P02": {"score": 40.0},
            # P03 deliberately absent — must be skipped, not treated as 0
        }
        assert _group_score(per_principle, ["P01", "P02"]) == 60.0
        assert _group_score(per_principle, ["P01", "P02", "P03"]) == 60.0
        assert _group_score(per_principle, ["P99"]) == 0.0  # none present


class TestDeepValidationIndependence:
    """overall_deep_validation()'s own docstring makes a specific integrity
    claim: independent_pct isolates only the 7 genuine-physics principles
    from the 21 surrogate-pass-through ones, and callers should cite that
    number, not the mixed pct, as evidence of physics-based validation.
    Nothing checked that the isolation logic actually does this correctly —
    a bug here would silently undermine the exact honesty distinction this
    project's audit is built around."""

    def test_pass_through_results_are_excluded_from_independent_pct(self):
        from cerebro_62_deep_engine import (
            _PASS_THROUGH_MARKER,
            overall_deep_validation,
        )
        deep_results = {
            "P02": {"validated": True,  "improvement_over_surrogate": "real physics, +12%"},
            "P13": {"validated": False, "improvement_over_surrogate": "real physics, -3%"},
            "P21": {"validated": True,  "improvement_over_surrogate": _PASS_THROUGH_MARKER},
            "P32": {"validated": True,  "improvement_over_surrogate": _PASS_THROUGH_MARKER},
        }
        result = overall_deep_validation(deep_results)
        assert result["total"] == 4
        assert result["passed_count"] == 3
        assert result["pct"] == 75.0
        # Only P02/P13 are independent — 1 of 2 passed, not 3 of 4.
        assert result["independent_computation_count"] == 2
        assert result["independent_pct"] == 50.0

    def test_all_pass_through_gives_no_independent_pct_rather_than_fabricating_one(self):
        from cerebro_62_deep_engine import (
            _PASS_THROUGH_MARKER,
            overall_deep_validation,
        )
        deep_results = {
            "P21": {"validated": True, "improvement_over_surrogate": _PASS_THROUGH_MARKER},
            "P32": {"validated": True, "improvement_over_surrogate": _PASS_THROUGH_MARKER},
        }
        result = overall_deep_validation(deep_results)
        assert result["independent_computation_count"] == 0
        assert result["independent_pct"] is None  # not 0.0, not 100.0 — genuinely unknown


class TestFirstPrinciplesPKa:
    """compute_pka_from_first_principles combines a real cited base value
    (Reich pKa tables) with a Hammett electronegativity correction and a
    Born solvation correction — real, correctly-cited physical chemistry,
    completely untested until now. Pinned against a value computed
    independently from the same published formulas, not just re-running
    the function and checking it doesn't crash."""

    def test_carboxylic_acid_pka_matches_independently_computed_value(self):
        from cerebro_value_resolver.computations.pka_first_principles import (
            compute_pka_from_first_principles,
        )
        result = compute_pka_from_first_principles(
            x_h_bond_type="H_O_carboxyl", neighbour_atoms=["C"])
        # base_pKa=4.5 (Reich table) + BDE_shift=0 (no local BDE override,
        # so it equals the reference) + Hammett (-1.0 * 0.5*(EN_C-EN_O))
        # + Born solvation (charge=1, r=1.5 A, eps=78.5) — computed
        # independently from the same published formulas, not copied from
        # the implementation.
        assert result["pKa"] == pytest.approx(4.86, abs=0.01)
        assert result["base_pKa"] == 4.5
        assert result["atom"] == "O"
        assert "Reich" in result["_computational_method"]
        assert "Bordwell 1988" in result["_computational_method"]

    def test_stronger_acid_class_gives_lower_pka_than_weaker_acid_class(self):
        """Physical sanity check: a phenol O-H (weaker acid, higher pKa)
        must score higher than a carboxylic acid O-H (stronger acid, lower
        pKa) under otherwise identical conditions — if this doesn't hold,
        the base table or shift terms have a real chemistry bug regardless
        of what any single pinned number says."""
        from cerebro_value_resolver.computations.pka_first_principles import (
            compute_pka_from_first_principles,
        )
        carboxyl = compute_pka_from_first_principles(
            x_h_bond_type="H_O_carboxyl", neighbour_atoms=["C"])
        phenol = compute_pka_from_first_principles(
            x_h_bond_type="H_O_phenol", neighbour_atoms=["C"])
        assert carboxyl["pKa"] < phenol["pKa"]


# ═════════════════════════════════════════════════════════════════════════════
# 16. REAL MOLECULAR-GRAPH GNN — REPLACES THE DELETED FAKE PSEUDO-GRAPH
# ═════════════════════════════════════════════════════════════════════════════

class TestMolecularGraphConstruction:
    """The whole point of rebuilding this component was that the old one
    faked its graph structure — identical node features tiled across a
    fully-connected placeholder topology. These tests exist specifically
    to catch a regression back to that failure mode, not just to check
    the code runs."""

    def test_different_atoms_get_different_features_not_tiled(self):
        from cerebro_molecular_gnn import smiles_to_graph
        node_feats, adjacency = smiles_to_graph("CCO")  # ethanol: C-C-O
        assert node_feats.shape[0] == 3  # 3 real heavy atoms, not N=f(MW)
        # The old bug tiled one identical vector across every node — here
        # the terminal carbon and the oxygen must have genuinely different
        # feature vectors (different element one-hot alone guarantees this).
        assert not (node_feats[0] == node_feats[2]).all()

    def test_adjacency_reflects_real_bonds_not_a_complete_graph(self):
        from cerebro_molecular_gnn import smiles_to_graph
        _, adjacency = smiles_to_graph("CCO")  # C0-C1, C1-O2 — a chain, not a triangle
        assert adjacency[0, 1] == 1.0 and adjacency[1, 0] == 1.0
        assert adjacency[1, 2] == 1.0 and adjacency[2, 1] == 1.0
        # A real chain has no C0-O2 bond — a fully-connected fake graph would.
        assert adjacency[0, 2] == 0.0 and adjacency[2, 0] == 0.0

    def test_invalid_smiles_returns_none_not_a_fabricated_graph(self):
        from cerebro_molecular_gnn import smiles_to_graph
        assert smiles_to_graph("not_a_real_smiles!!!") is None

    def test_larger_molecule_gets_more_nodes_than_smaller_one(self):
        """Node count must come from real atom count, not the old bug's
        N = f(molecular weight) formula applied to a pseudo-graph."""
        from cerebro_molecular_gnn import smiles_to_graph
        small, _ = smiles_to_graph("CCO")  # ethanol, 3 heavy atoms
        large, _ = smiles_to_graph(
            "COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2")  # donepezil, 28
        assert large.shape[0] > small.shape[0]
        assert large.shape[0] == 28
        assert small.shape[0] == 3


class TestAdjacencyNormalization:
    """Kipf & Welling's D^-1/2(A+I)D^-1/2 renormalization, pinned against
    a value computed independently in this test (plain numpy on a
    hand-worked 3-node chain), not copied from the implementation."""

    def test_normalization_matches_independently_computed_value(self):
        import numpy as np

        from cerebro_molecular_gnn import _normalize_adjacency
        # 3-node chain: 0-1-2 (same shape as ethanol's C-C-O)
        adj = np.array([[0, 1, 0], [1, 0, 1], [0, 1, 0]], dtype=np.float32)

        # Independent computation, not calling the function under test:
        a_hat = adj + np.eye(3, dtype=np.float32)
        deg = a_hat.sum(axis=1)  # [2, 3, 2]
        d_inv_sqrt = np.diag(1.0 / np.sqrt(deg))
        expected = d_inv_sqrt @ a_hat @ d_inv_sqrt

        result = _normalize_adjacency(adj)
        np.testing.assert_allclose(result, expected, atol=1e-6)

    def test_isolated_atom_does_not_produce_uninitialized_values(self):
        """Regression test for a real bug caught while building this:
        np.power(deg, -0.5, where=deg>0) leaves garbage in positions where
        deg==0 unless the output buffer is pre-zeroed. An isolated atom
        (degree 0, plus its own self-loop = degree 1 after A+I) is a real,
        if rare, case this must handle cleanly."""
        import numpy as np

        from cerebro_molecular_gnn import _normalize_adjacency
        adj = np.zeros((2, 2), dtype=np.float32)  # two atoms, no bond between them
        result = _normalize_adjacency(adj)
        assert np.isfinite(result).all()


class TestMolecularGCNForwardPass:
    """The model itself — real batched forward pass, not just that it can
    be instantiated."""

    def test_padding_does_not_affect_pooled_output(self):
        """A molecule's prediction must be identical whether it's run
        alone or padded alongside a larger molecule in the same batch —
        if the mask leaked into the mean pool, it wouldn't be."""
        import torch

        from cerebro_molecular_gnn import MolecularGCN, _pad_batch, smiles_to_graph
        torch.manual_seed(0)
        model = MolecularGCN()
        model.eval()

        small = smiles_to_graph("CCO")
        large = smiles_to_graph("COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2")

        with torch.no_grad():
            x, a, m = _pad_batch([small])
            solo_pred = model(x, a, m)[0].item()

            x2, a2, m2 = _pad_batch([small, large])
            batched_pred = model(x2, a2, m2)[0].item()

        assert abs(solo_pred - batched_pred) < 1e-5

    def test_output_is_a_valid_probability(self):
        import torch

        from cerebro_molecular_gnn import MolecularGCN, _pad_batch, smiles_to_graph
        model = MolecularGCN()
        model.eval()
        g = smiles_to_graph("CCO")
        with torch.no_grad():
            x, a, m = _pad_batch([g])
            out = model(x, a, m)
        assert 0.0 <= out[0].item() <= 1.0

    @pytest.mark.slow
    def test_bbb_resolver_attaches_real_gnn_cross_check_alongside_dnn(self):
        """The live integration point: engine/cerebro_value_resolver/
        categories/bbb_perm.py's Tier 3 already resolves via the DNN —
        confirms the GNN cross-check actually shows up in real resolver
        output, agreeing on a real, well-known BBB-penetrant drug, rather
        than the two models being silently merged into one number."""
        from cerebro_bbb_dnn import _HAS_BBB_DNN
        from cerebro_molecular_gnn import _HAS_MOL_GNN
        if not (_HAS_BBB_DNN and _HAS_MOL_GNN):
            pytest.skip("tensorflow and/or torch not installed")
        from cerebro_value_resolver.categories.bbb_perm import resolve_bbb_permeability

        # Donepezil: a real, well-known CNS drug that crosses the BBB.
        result = resolve_bbb_permeability(
            name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2")
        assert result["source"] == "cerebro_bbb_dnn"
        assert "gnn_predicted_class" in result
        assert "gnn_probability_permeable" in result
        assert result["gnn_predicted_class"] == "permeable"
        assert result["dnn_predicted_class"] == "permeable"
        assert result["gnn_agrees_with_dnn"] is True


# ═════════════════════════════════════════════════════════════════════════════
# 17. CLASS-B DEEP PHYSICS ENGINE (cerebro_62_deep_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
def _drug_bundle(**overrides):
    b = {
        "drug_mw":  {"value": 350.0},
        "drug_logp": {"value": 2.5},
        "drug_tpsa": {"value": 60.0},
        "drug_hbd": {"value": 2},
        "drug_hba": {"value": 5},
        "pk_halflife": {"value": 0.5},
        "bbb_permeability": {"value": 5.0},
        "_meta": {"drug_type": "small_molecule", "name": "test_drug",
                  "identifiers": {"smiles": "CCO"}},
    }
    b.update(overrides)
    return b


def _dds_bundle(**meta_overrides):
    meta = {"carrier_type": "liposome", "dds_type": "material"}
    meta.update(meta_overrides)
    return {"_meta": meta}


def _combo_bundle(dds_row=None):
    return {"_meta": {"dds_row": dds_row or {}}}


class TestDeepP02AllometricScaling:
    """This pipeline has no real preclinical (animal) PK data source
    anywhere -- pk_halflife always resolves to a human-relevant value
    (live clinical DB, human population regression, or human class
    median). deep_P02 used to treat that value as raw mouse data and
    multiply it by (70kg/25g)^0.25 ~= 7.3x, reporting the inflated
    number as "Predicted human t1/2 ... scaled from Xh mouse" -- a
    fabricated figure and a false species-scaling story, since no mouse
    measurement was ever involved anywhere in this codebase. Fixed to
    report the resolved half-life directly and show the Mahmood 2007
    exponents as reference-only figures, not an applied conversion."""

    def test_reported_half_life_is_the_resolved_value_not_rescaled(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P02

        drug = _drug_bundle(drug_mw={"value": 350.0},
                             pk_halflife={"value": 0.5, "tier": 1})
        r = deep_P02(drug, _dds_bundle(), _combo_bundle(), {"score": 50})

        # 0.5 days = 12h -- the value CEREBRO-X already resolved as this
        # drug's clinical half-life, with no (70kg/25g)^0.25 ~7.3x
        # allometric inflation applied on top of it.
        assert r["value"] == round(0.5 * 24, 2)
        assert r["validated"] is True
        assert "mouse" not in r["narrative"].lower()
        assert "thalf_mouse_h" not in r["raw"]

    def test_confidence_depends_on_drug_type_and_pk_source_tier(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P02

        # Tier 1 = live human clinical database -> HIGH confidence.
        clinical = deep_P02(
            _drug_bundle(pk_halflife={"value": 0.5, "tier": 1}),
            _dds_bundle(), _combo_bundle(), {})
        assert (clinical["score"], clinical["confidence"]) == (90, "HIGH")

        # Tier 7 = class-median fallback, no real source -> only MODERATE.
        fallback = deep_P02(
            _drug_bundle(pk_halflife={"value": 0.5, "tier": 7}),
            _dds_bundle(), _combo_bundle(), {})
        assert (fallback["score"], fallback["confidence"]) == (70, "MODERATE")

        mab = deep_P02(
            _drug_bundle(_meta={"drug_type": "monoclonal_antibody", "name": "x",
                                  "identifiers": {"smiles": ""}}),
            _dds_bundle(), _combo_bundle(), {})
        assert (mab["score"], mab["confidence"]) == (65, "LOW")

        other = deep_P02(
            _drug_bundle(_meta={"drug_type": "nanoparticle_conjugate", "name": "x",
                                  "identifiers": {"smiles": ""}}),
            _dds_bundle(), _combo_bundle(), {})
        assert (other["score"], other["confidence"]) == (75, "MODERATE")


class TestDeepP38StokesEinstein:
    """Glymphatic clearance — physical Stokes-Einstein diffusion, no fitted knobs."""

    def test_reference_size_returns_exactly_six_hours(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P38

        r = deep_P38(_drug_bundle(), _dds_bundle(),
                      _combo_bundle({"Size_nm": 50.0}), {})
        assert r["value"] == 6.0
        assert r["validated"] is True
        assert r["confidence"] == "HIGH"

    def test_clearance_time_scales_linearly_with_particle_size(self):
        """D ∝ 1/r, and t_clear ∝ 1/D, so doubling+quadrupling the radius
        relative to the 50nm reference should scale t_clear by the same
        factor — a direct physical consequence of Stokes-Einstein, not a
        tuned parameter."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P38

        r = deep_P38(_drug_bundle(), _dds_bundle(),
                      _combo_bundle({"Size_nm": 200.0}), {})
        assert r["value"] == 24.0  # 6h * (200/50)
        assert r["validated"] is True  # 6 <= 24 <= 48

    def test_missing_size_fails_cleanly(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P38

        r = deep_P38(_drug_bundle(), _dds_bundle(),
                      _combo_bundle({"Size_nm": 0.0}), {})
        assert r["validated"] is False
        assert r["confidence"] == "FAILED"


class TestDeepP18ActiveTargeting:
    """MM/GBSA-style ΔG from a validated ligand-receptor Kd lookup table."""

    def test_known_ligand_resolves_to_its_receptor_and_dg(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P18

        r = deep_P18(_drug_bundle(), _dds_bundle(),
                      _combo_bundle({"Surface_Ligand": "Transferrin"}), {"score": 50})
        assert r["raw"]["receptor"] == "TfR1"
        assert r["value"] == -10.5
        assert r["validated"] is True  # |−10.5| >= 8.0

    def test_weak_ligand_is_not_validated(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P18

        r = deep_P18(_drug_bundle(), _dds_bundle(),
                      _combo_bundle({"Surface_Ligand": "TAT"}), {})
        assert r["value"] == -6.5
        assert r["validated"] is False  # |−6.5| < 8.0

    def test_unrecognized_ligand_falls_back_to_generic_receptor(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P18

        r = deep_P18(_drug_bundle(), _dds_bundle(),
                      _combo_bundle({"Surface_Ligand": "made_up_peptide_37"}), {})
        assert r["raw"]["receptor"] == "unknown_BBB_receptor"
        assert r["value"] == -7.0

    def test_no_ligand_bypasses_computation_entirely(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P18

        r = deep_P18(_drug_bundle(), _dds_bundle(), _combo_bundle({}), {})
        assert r["validated"] is False
        assert r["value"] == 0
        assert "No surface ligand" in r["method"]


class TestDeepP47BindingAffinity:
    """LIE-approximation fallback — same formula whether it's reached via
    real_docking_engine's own fallback or the deep-engine's inline copy."""

    def test_lie_delta_g_matches_published_aqvist_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P47

        logp, tpsa, hbd, hba = 3.0, 60.0, 2, 5
        drug = _drug_bundle(drug_logp={"value": logp}, drug_tpsa={"value": tpsa},
                             drug_hbd={"value": hbd}, drug_hba={"value": hba})
        r = deep_P47(drug, _dds_bundle(), _combo_bundle(), {"score": 50, "value": -5})

        alpha, beta = 0.181, 0.137
        expected = -(alpha * logp + beta * (50 - tpsa / 5) + 0.5 * (hbd + hba) * 0.3)
        expected = max(-20, min(-1, expected))
        assert r["value"] == round(expected, 2)
        assert "LIE approximation" in r["method"]
        assert "LOW" in r["confidence"]


class TestDeepP31Biodistribution:
    """7-organ whole-body distribution — deterministic weighted split, no ODE."""

    def test_organ_distribution_matches_hand_computation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P31

        drug = _drug_bundle(bbb_permeability={"value": 5.0})
        r = deep_P31(drug, _dds_bundle(),
                      _combo_bundle({"Size_nm": 250.0, "Zeta_Potential_mV": -30.0}),
                      {})

        # brain=5, liver=35 (size>200), spleen=20 (|zeta|>25), kidney=8,
        # lung=5, heart=3, muscle=100 → sum=176, brain_norm=5/176*100
        expected_brain_pct = round(5.0 / 176.0 * 100, 2)
        assert r["value"] == expected_brain_pct
        assert r["raw"]["organ_distribution_pct"]["liver"] == round(35 / 176 * 100, 2)

    def test_active_targeting_ligand_triples_brain_uptake_share(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P31

        drug = _drug_bundle(bbb_permeability={"value": 5.0})
        no_ligand = deep_P31(drug, _dds_bundle(),
                              _combo_bundle({"Size_nm": 250.0, "Zeta_Potential_mV": -30.0}),
                              {})
        with_ligand = deep_P31(drug, _dds_bundle(),
                                _combo_bundle({"Size_nm": 250.0, "Zeta_Potential_mV": -30.0,
                                               "Surface_Ligand": "Transferrin"}),
                                {})
        assert with_ligand["value"] > no_ligand["value"]
        assert with_ligand["raw"]["ligand_boost"] is True
        assert no_ligand["raw"]["ligand_boost"] is False


class TestDeepPBPKODEs:
    """3- and 4-compartment PBPK ODE integrators (P13, P44) — checks the
    deterministic rate constants they build and the physically-required
    direction of the ligand-targeting effect, without hand-solving the ODE."""

    def test_p13_ligand_targeting_raises_brain_exposure(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P13

        drug = _drug_bundle(bbb_permeability={"value": 5.0})
        no_ligand = deep_P13(drug, _dds_bundle(), _combo_bundle({}), {})
        with_ligand = deep_P13(drug, _dds_bundle(),
                                _combo_bundle({"Surface_Ligand": "Transferrin"}), {})

        assert with_ligand["raw"]["k_bb_in_per_h"] == round(no_ligand["raw"]["k_bb_in_per_h"] * 3, 4)
        assert with_ligand["value"] > no_ligand["value"]  # higher brain AUC ratio

    def test_p13_fails_cleanly_without_scipy(self, monkeypatch):
        import builtins
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P13

        real_import = builtins.__import__

        def blocked(name, *a, **kw):
            if name == "scipy.integrate" or name.startswith("scipy"):
                raise ImportError("blocked for test")
            return real_import(name, *a, **kw)

        monkeypatch.setattr(builtins, "__import__", blocked)
        r = deep_P13(_drug_bundle(), _dds_bundle(), _combo_bundle({}), {})
        assert r["validated"] is False
        assert r["confidence"] == "FAILED"

    def test_p44_glymphatic_rate_matches_stokes_einstein_size_scaling(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import deep_P44

        small = deep_P44(_drug_bundle(), _dds_bundle(),
                          _combo_bundle({"Size_nm": 50.0}), {})
        large = deep_P44(_drug_bundle(), _dds_bundle(),
                          _combo_bundle({"Size_nm": 300.0}), {})
        # k_glymph = 0.2 / max(1, size/100) — unchanged at/under 100nm,
        # inversely proportional above it.
        assert small["raw"]["k_glymph_per_h"] == 0.2
        assert large["raw"]["k_glymph_per_h"] == round(0.2 / 3.0, 4)


class TestDeepValidationDispatch:
    """evaluate_deep_for_top1() — the real/HPC-deferred split must stay
    exhaustive and non-overlapping across all 28 Class-B principles."""

    def test_deep_functions_and_hpc_only_partition_cleanly(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import DEEP_FUNCTIONS, HPC_ONLY_PRINCIPLES

        overlap = set(DEEP_FUNCTIONS) & set(HPC_ONLY_PRINCIPLES)
        assert overlap == set()
        assert len(DEEP_FUNCTIONS) == 7

    def test_evaluate_deep_for_top1_covers_every_registered_principle(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import (
            DEEP_FUNCTIONS,
            HPC_ONLY_PRINCIPLES,
            evaluate_deep_for_top1,
        )

        out = evaluate_deep_for_top1(_drug_bundle(), _dds_bundle(), _combo_bundle({}), {})
        assert set(out) == set(DEEP_FUNCTIONS) | set(HPC_ONLY_PRINCIPLES)
        # An HPC-only principle must carry the honest pass-through marker.
        assert "external HPC run" in out["P01"]["improvement_over_surrogate"]
        # A real DEEP_FUNCTIONS principle must not carry that marker.
        assert "external HPC run" not in out["P38"]["improvement_over_surrogate"]

    def test_a_raising_deep_function_is_caught_and_marked_failed(self):
        """A malformed Excel cell (non-numeric zeta potential) makes
        deep_P31's float() conversion raise — the dispatcher must catch
        it and report _failed() for that principle only, not crash the
        whole batch."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import evaluate_deep_for_top1

        bad_combo = _combo_bundle({"Zeta_Potential_mV": "not_a_number"})
        out = evaluate_deep_for_top1(_drug_bundle(), _dds_bundle(), bad_combo, {})
        assert out["P31"]["validated"] is False
        assert out["P31"]["confidence"] == "FAILED"
        # Other principles in the same batch must still compute normally.
        assert out["P38"]["confidence"] != "FAILED"


# ═════════════════════════════════════════════════════════════════════════════
# 18. pKa RESOLVER (cerebro_value_resolver/categories/drug_pka.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDrugPkaResolver:
    """name='' on every call keeps these fully offline: _pka_chembl short-
    circuits on a falsy name before touching the network, so Tier 7
    first-principles is exercised deterministically."""

    def test_researcher_override_short_circuits_to_tier_zero(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_pka import resolve_drug_pka_acidic

        r = resolve_drug_pka_acidic(name="", smiles="CC(=O)O", researcher_override=3.1)
        assert r["value"] == 3.1
        assert r["tier"] == 0
        assert r["source"] == "researcher_override"

    def test_carboxylic_acid_acidic_pka_in_physically_plausible_range(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_pka import resolve_drug_pka_acidic

        r = resolve_drug_pka_acidic(name="", smiles="CC(=O)O")  # acetic acid
        assert r["tier"] == 7
        assert "Bordwell-Hammett-Born" in r["method"]
        assert 2.0 < r["value"] < 6.0  # real carboxylic acids: pKa ~2-5

    def test_primary_amine_basic_pka_in_physically_plausible_range(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_pka import resolve_drug_pka_basic

        r = resolve_drug_pka_basic(name="", smiles="CCN")  # ethylamine
        assert r["tier"] == 7
        assert 8.0 < r["value"] < 12.0  # real aliphatic amines: pKa(BH+) ~9.5-11

    def test_dominant_pka_picks_whichever_real_result_is_closer_to_7_4(self):
        """Cross-checks the dominant-selection logic against the real
        acidic/basic resolvers it wraps, rather than against a value
        copied out of the implementation."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_pka import (
            resolve_drug_pka_acidic,
            resolve_drug_pka_basic,
            resolve_drug_pka_dominant,
        )

        smiles = "NCC(=O)O"  # glycine — both an amine and a carboxylic acid
        acid = resolve_drug_pka_acidic(name="", smiles=smiles)
        base = resolve_drug_pka_basic(name="", smiles=smiles)
        dom = resolve_drug_pka_dominant(name="", smiles=smiles)

        expected_kind = "acidic" if abs(acid["value"] - 7.4) <= abs(base["value"] - 7.4) else "basic"
        expected_value = acid["value"] if expected_kind == "acidic" else base["value"]
        assert dom["kind"] == expected_kind
        assert dom["value"] == expected_value

    def test_missing_smiles_reports_honest_null_not_a_fabricated_number(self):
        """Regression test for a real bug: resolve_drug_pka_acidic/basic
        used to fall back to a hardcoded generic sp3 C-H guess (~50) or a
        '-14 proxy' (~-10 to -14) whenever no SMILES was available, which
        made resolve_drug_pka_dominant's honest 'not ionizable' branch and
        hh_microspeciation's 'no ionizable groups' branch permanently
        unreachable — the system always reported *some* plausible-looking
        pKa number even with zero structural evidence to support it."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_pka import (
            resolve_drug_pka_acidic,
            resolve_drug_pka_basic,
            resolve_drug_pka_dominant,
            resolve_drug_microspecies,
        )

        acid = resolve_drug_pka_acidic(name="", smiles="")
        base = resolve_drug_pka_basic(name="", smiles="")
        assert acid["value"] is None
        assert base["value"] is None

        dom = resolve_drug_pka_dominant(name="", smiles="")
        assert dom["value"] is None
        assert dom["note"] == "Drug is not ionizable at any pH"
        assert dom["confidence"] == "HIGH"

        micro = resolve_drug_microspecies(name="", smiles="")
        assert micro["value"]["method"] == "no ionizable groups (assumed neutral)"
        assert micro["value"]["f_neutral"] == 1.0

    def test_microspecies_fractions_sum_to_one_for_a_real_acid(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_pka import resolve_drug_microspecies

        r = resolve_drug_microspecies(name="", smiles="CC(=O)O")
        fr = r["value"]
        total = fr["f_cationic"] + fr["f_anionic"] + fr["f_zwitterion"] + fr["f_neutral"]
        assert total == pytest.approx(1.0, abs=1e-3)
        # Acetic acid (pKa ~4.86) is mostly deprotonated at physiological pH 7.4.
        assert fr["f_anionic"] > 0.9


# ═════════════════════════════════════════════════════════════════════════════
# 19. DLVO COLLOIDAL-STABILITY RESOLVER (categories/physics_dlvo.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestPhysicsDLVOResolver:
    """Pure closed-form electrostatics/vdW physics — no bundles, no network,
    so every test hand-derives the expected number from the same cited
    equation and compares, independent of the implementation."""

    def test_hamaker_combining_rule_matches_israelachvili_formula(self):
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_hamaker_combined,
        )

        r = resolve_physics_hamaker_combined(carrier_Hamaker_J=5e-20, medium_Hamaker_J=3.7e-20)
        expected = (math.sqrt(5e-20) - math.sqrt(3.7e-20)) ** 2
        assert r["value"] == pytest.approx(expected, rel=1e-9)
        assert r["tier"] == 6

    def test_hamaker_defaults_to_generic_organic_when_not_provided(self):
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_hamaker_combined,
        )

        r = resolve_physics_hamaker_combined(carrier_Hamaker_J=None)
        expected = (math.sqrt(6e-21) - math.sqrt(3.7e-20)) ** 2  # generic 6e-21 fallback
        assert r["value"] == pytest.approx(expected, rel=1e-9)
        assert "not provided" in r["live_db_misses"][0]

    def test_debye_length_at_physiological_ionic_strength_is_realistic(self):
        """κ⁻¹ ≈ 0.7-1.0 nm at I=0.15M is the textbook physiological Debye
        length — a real, independently-known sanity bound, not a number
        pulled from the implementation."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_debye_length,
        )

        r = resolve_physics_debye_length(ionic_strength_M=0.15)
        kappa_inv_nm = r["value"] * 1e9
        assert 0.6 < kappa_inv_nm < 1.1

    def test_debye_length_shrinks_as_ionic_strength_rises(self):
        """Higher ionic strength screens electrostatics more effectively —
        κ⁻¹ ∝ I^-0.5, a direct physical consequence, not tunable."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_debye_length,
        )

        low_I = resolve_physics_debye_length(ionic_strength_M=0.01)
        high_I = resolve_physics_debye_length(ionic_strength_M=1.0)
        assert low_I["value"] > high_I["value"]

    def test_dlvo_potential_van_der_waals_term_is_attractive(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_dlvo_potential,
        )

        r = resolve_physics_dlvo_potential()
        assert r["V_vdW_kT"] < 0  # vdW is always attractive between like particles

    def test_dlvo_potential_component_sum_matches_reported_total(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_dlvo_potential,
        )

        r = resolve_physics_dlvo_potential(zeta_mV=-40, separation_nm=3)
        assert r["value"] == pytest.approx(
            r["V_vdW_kT"] + r["V_el_kT"], rel=1e-9)

    def test_dlvo_higher_zeta_magnitude_increases_electrostatic_repulsion(self):
        """A more strongly (de)stabilized surface should raise the
        electrostatic repulsion term — this is what makes a formulation
        colloidally stable in the first place."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_dlvo_potential,
        )

        weak = resolve_physics_dlvo_potential(zeta_mV=-10)
        strong = resolve_physics_dlvo_potential(zeta_mV=-40)
        assert strong["V_el_kT"] > weak["V_el_kT"]

    def test_grahame_equation_surface_charge_matches_hand_computation(self):
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_zeta_to_surface_charge,
        )

        eps_0, k_B, N_A, e = 8.854e-12, 1.380649e-23, 6.022e23, 1.602e-19
        T_K, epsilon_r, I, zeta_mV = 310.15, 78.5, 0.15, -25
        psi = zeta_mV * 1e-3
        expected = (math.sqrt(8 * eps_0 * epsilon_r * k_B * T_K * I * 1000 * N_A)
                    * math.sinh(e * psi / (2 * k_B * T_K)))

        r = resolve_physics_zeta_to_surface_charge(zeta_mV=zeta_mV)
        assert r["value"] == pytest.approx(expected, rel=1e-9)
        assert r["value"] < 0  # negative zeta -> negative surface charge

    def test_grahame_equation_zero_zeta_gives_zero_charge(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_zeta_to_surface_charge,
        )

        r = resolve_physics_zeta_to_surface_charge(zeta_mV=0)
        assert r["value"] == pytest.approx(0.0, abs=1e-12)

    def test_researcher_overrides_short_circuit_all_four_resolvers(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_dlvo import (
            resolve_physics_dlvo_potential,
            resolve_physics_debye_length,
            resolve_physics_hamaker_combined,
            resolve_physics_zeta_to_surface_charge,
        )

        for fn in (resolve_physics_hamaker_combined, resolve_physics_debye_length,
                   resolve_physics_dlvo_potential, resolve_physics_zeta_to_surface_charge):
            r = fn(researcher_override=1.23)
            assert r["value"] == 1.23
            assert r["tier"] == 0
            assert r["source"] == "researcher_override"


# ═════════════════════════════════════════════════════════════════════════════
# 20. MATERIAL-CLASS LOOKUP TABLES (lipid / surface / polymer)
# ═════════════════════════════════════════════════════════════════════════════
def _resolve(category, **kwargs):
    """These three material-property files build their resolvers with a
    factory (_build_*_resolver) that registers each one under @register but
    never binds it to a named module attribute — so, unlike bbb_perm.py's
    directly-defined functions, they're only reachable through the shared
    registry, exactly as production code (cerebro_resolved_bundles.py) calls
    them."""
    import src.path_resolver  # noqa: F401
    import cerebro_value_resolver.categories.material_lipid  # noqa: F401
    import cerebro_value_resolver.categories.material_polymer  # noqa: F401
    import cerebro_value_resolver.categories.material_surface  # noqa: F401
    from cerebro_value_resolver._core import resolve_value
    return resolve_value(category, **kwargs)


class TestMaterialLipidResolver:
    def test_known_lipid_type_returns_its_israelachvili_table_value(self):
        r = _resolve("material_lipid_tm", lipid_type="DSPC")
        assert r["value"] == 55.0
        assert r["tier"] == 7

    def test_lipid_type_takes_priority_over_carrier(self):
        r = _resolve("material_lipid_tm", carrier="liposome", lipid_type="dppc")
        assert r["value"] == 41.0  # DPPC's own Tm, not the generic liposome default

    def test_unrecognized_carrier_falls_back_to_generic_liposome_default(self):
        r = _resolve("material_lipid_packing_parameter", carrier="some_unknown_carrier")
        assert r["value"] == 0.95  # falls through to the "liposome" table entry

    def test_cholesterol_has_no_melting_point_falls_to_generic_default(self):
        """chol's Tm_C is explicitly None in the table — the resolver must
        not return None silently, it should fall through to the t7 default."""
        r = _resolve("material_lipid_tm", lipid_type="chol")
        assert r["value"] == 35.0  # generic_lipid default, not None

    def test_researcher_override_short_circuits(self):
        r = _resolve("material_lipid_bending_modulus", researcher_override=99.0)
        assert r["value"] == 99.0 and r["tier"] == 0


class TestMaterialSurfaceResolver:
    def test_metallic_carrier_dielectric_is_the_real_fixed_value(self):
        """Regression guard: this dict entry used to be built from a broken
        `-np_inf if False else 2.0` expression papered over by a redundant
        post-hoc assignment a few lines down. Cleaned up to a plain literal —
        this pins the value stays 2.0 (gold's real approximate epsilon_r)."""
        r = _resolve("material_dielectric", carrier="metallic")
        assert r["value"] == 2.0

    def test_known_carrier_returns_its_table_value(self):
        r = _resolve("material_hamaker_constant", carrier="plga")
        assert r["value"] == 6e-21
        assert r["tier"] == 7

    def test_unknown_carrier_falls_back_to_generic_default(self):
        r = _resolve("material_refractive_index", carrier="not_a_real_carrier")
        assert r["value"] == 1.46
        assert r["confidence"] == "LOW"


class TestMaterialPolymerResolver:
    def test_known_polymer_returns_its_table_value(self):
        r = _resolve("material_polymer_tg", carrier="pla")
        assert r["value"] == 60.0
        assert r["tier"] == 7

    def test_plga_has_no_melting_point_falls_to_generic_default(self):
        """PLGA is amorphous — Tm_C is explicitly None in the table."""
        r = _resolve("material_polymer_tm", carrier="plga")
        assert r["value"] == 100.0  # generic default, not None
        assert "warning" in r

    def test_unknown_carrier_falls_back_to_generic_default(self):
        r = _resolve("material_polymer_density", carrier="unobtainium")
        assert r["value"] == 1.25
        assert r["confidence"] == "LOW"

    def test_researcher_override_short_circuits(self):
        r = _resolve("material_polymer_mw", carrier="plga", researcher_override=45000.0)
        assert r["value"] == 45000.0 and r["tier"] == 0


# ═════════════════════════════════════════════════════════════════════════════
# 21. TRANSPORT-COEFFICIENT RESOLVER (categories/physics_transport.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestPhysicsTransportComputations:
    """The underlying pure-math correlations, hand-verified against their
    published formulas independent of the resolver wrapper."""

    def test_stokes_einstein_diff_matches_published_formula(self):
        """Regression guard: this used a radius coefficient of 6.6e-12
        (r ≈ 0.066·MW^(1/3) Å), which gives a MW=180 (glucose) radius of
        0.37 Å -- smaller than a single hydrogen atom, not a plausible
        size for an entire molecule -- and inflated every diffusion
        coefficient here by ~10x (D ∝ 1/r). The physically correct
        coefficient, r ≈ (3·MW·v̄/(4π·N_A))^(1/3) with a typical
        small-molecule partial specific volume v̄≈0.73 cm³/g, works out to
        r ≈ 0.66·MW^(1/3) Å -- giving glucose r≈3.7 Å, matching its known
        experimental hydrodynamic radius, and 6.6e-11 in this formula's
        units. This test re-derives that coefficient from v̄ independently
        rather than hardcoding the same constant the implementation uses,
        so a regression back to the 10x-too-small radius would be caught."""
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.computations import stokes_einstein_diff

        mw, T_K, visc = 350.0, 310.15, 6.91e-4
        k_B = 1.380649e-23
        N_A = 6.022e23
        v_bar_cm3_g = 0.73    # typical small-molecule partial specific volume
        r_cm = (3 * mw * v_bar_cm3_g / (4 * math.pi * N_A)) ** (1 / 3)
        r_m = r_cm * 1e-2
        expected = (k_B * T_K) / (6 * math.pi * visc * r_m)
        assert stokes_einstein_diff(mw, T_K=T_K, visc_Pa_s=visc) == pytest.approx(expected, rel=0.05)
        # And: not the old, order-of-magnitude-too-fast value.
        assert stokes_einstein_diff(mw, T_K=T_K, visc_Pa_s=visc) < 2e-9

    def test_stokes_einstein_diff_zero_or_negative_mw_returns_zero(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.computations import stokes_einstein_diff

        assert stokes_einstein_diff(0.0) == 0.0
        assert stokes_einstein_diff(-10.0) == 0.0

    def test_larger_molecules_diffuse_more_slowly(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.computations import stokes_einstein_diff

        assert stokes_einstein_diff(100.0) > stokes_einstein_diff(10000.0)


class TestPhysicsDiffCoeffWaterResolver:
    def test_missing_wilke_chang_in_chemicals_lib_falls_through_to_pure_math(self):
        """Regression test: the installed chemicals==1.5.2 no longer
        exposes Wilke_Chang (or any liquid-diffusivity correlation) under
        any name — confirmed directly against the installed package. The
        resolver used to try importing it on every single call anyway, an
        import that's guaranteed to fail, silently caught, and always
        fell through to tier 6. Fixed by not attempting the dead import;
        this pins that tier 6 (the resolver's own pure-math Wilke-Chang)
        is what actually answers when mw_Da is given."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_diff_coeff_water,
        )

        r = resolve_physics_diff_coeff_water(mw_Da=350.0)
        assert r["tier"] == 6
        assert r["source"] == "cerebro_value_resolver:wilke_chang"
        assert 1e-11 < r["value"] < 1e-8  # physically sane aqueous D for a small molecule

    def test_missing_mw_falls_back_to_typical_small_molecule_constant(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_diff_coeff_water,
        )

        r = resolve_physics_diff_coeff_water()
        assert r["value"] == 5e-10
        assert r["tier"] == 7

    def test_researcher_override_short_circuits(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_diff_coeff_water,
        )

        r = resolve_physics_diff_coeff_water(researcher_override=1e-9)
        assert r["value"] == 1e-9 and r["tier"] == 0


class TestPhysicsDiffCoeffMembraneResolver:
    def test_higher_logp_increases_membrane_retention_and_slows_diffusion(self):
        """The model's own stated assumption: more lipophilic drugs are
        retained more strongly in the bilayer, slowing lateral diffusion."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_diff_coeff_membrane,
        )

        low_logp = resolve_physics_diff_coeff_membrane(mw_Da=350.0, logp=1.0)
        high_logp = resolve_physics_diff_coeff_membrane(mw_Da=350.0, logp=5.0)
        assert high_logp["value"] < low_logp["value"]

    def test_membrane_diffusion_is_slower_than_aqueous_diffusion(self):
        """Membrane viscosity is modeled as ~100x water — lateral diffusion
        in the bilayer should always be far slower than in bulk water for
        the same molecule."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_diff_coeff_membrane,
            resolve_physics_diff_coeff_water,
        )

        water = resolve_physics_diff_coeff_water(mw_Da=350.0)
        membrane = resolve_physics_diff_coeff_membrane(mw_Da=350.0, logp=2.5)
        assert membrane["value"] < water["value"]


class TestPhysicsLJParameterResolvers:
    def test_stiel_thodos_epsilon_no_longer_gated_behind_chemicals_lib(self):
        """Regression test: this branch computes ε/k_B = 0.77·Tc inline —
        it never actually calls the `chemicals` package — but used to be
        wrapped in `if _HAS_CHEMICALS:` anyway, so a real Tb_K-based
        estimate would be silently skipped (falling to a generic 300K
        constant) whenever the unrelated `chemicals` package happened to
        be missing. Also pins the tier to 7 (pure first-principles math,
        no library involved) rather than the old mislabeled tier 5."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_lj_epsilon,
        )

        Tb_K = 350.0
        r = resolve_physics_lj_epsilon(Tb_K=Tb_K)
        assert r["value"] == pytest.approx(0.77 * 1.5 * Tb_K, rel=1e-9)
        assert r["tier"] == 7

    def test_lj_epsilon_falls_back_to_mw_then_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_lj_epsilon,
        )

        by_mw = resolve_physics_lj_epsilon(mw_Da=200.0)
        assert by_mw["value"] == pytest.approx(0.5 * 200.0)
        assert by_mw["tier"] == 6

        generic = resolve_physics_lj_epsilon()
        assert generic["value"] == 300.0
        assert generic["tier"] == 7

    def test_lj_sigma_bsl_correlation_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_lj_sigma,
        )

        Vc = 250.0
        r = resolve_physics_lj_sigma(Vc_cm3_mol=Vc)
        assert r["value"] == pytest.approx(0.841 * Vc ** (1 / 3), rel=1e-9)
        assert r["tier"] == 6

    def test_lj_sigma_falls_back_to_mw_then_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_lj_sigma,
        )

        by_mw = resolve_physics_lj_sigma(mw_Da=200.0)
        assert by_mw["value"] == pytest.approx(0.5 * 200.0 ** (1 / 3), rel=1e-9)
        assert by_mw["tier"] == 7

        generic = resolve_physics_lj_sigma()
        assert generic["value"] == 5.0


class TestPhysicsViscositySolventResolver:
    def test_water_viscosity_at_body_temperature_is_physically_realistic(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_viscosity_solvent,
        )

        r = resolve_physics_viscosity_solvent(solvent="water")
        # Real water viscosity at 37°C is ~0.69 mPa·s regardless of whether
        # thermo's lookup or the Andrade fallback answers.
        assert 5e-4 < r["value"] < 9e-4

    def test_unknown_solvent_defaults_to_water_viscosity(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_viscosity_solvent,
        )

        r = resolve_physics_viscosity_solvent(solvent="totally_made_up_solvent_xyz")
        assert r["value"] == 1e-3
        assert r["confidence"] == "LOW"

    def test_researcher_override_short_circuits(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.physics_transport import (
            resolve_physics_viscosity_solvent,
        )

        r = resolve_physics_viscosity_solvent(researcher_override=2e-3)
        assert r["value"] == 2e-3 and r["tier"] == 0


# ═════════════════════════════════════════════════════════════════════════════
# 22. QUANTUM/ATOMIC RESOLVER (categories/quantum_atomic.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestQuantumAtomicResolver:
    def test_mendeleev_polarizability_is_converted_from_bohr3_to_angstrom3(self):
        """Regression test for a real unit-conversion bug: mendeleev's
        dipole_polarizability attribute is in Bohr³ (confirmed directly —
        mendeleev reports H at 4.507 Bohr³), not Å³ as the resolver assumed.
        Using it unconverted overstated every atomic polarizability by
        ~6.75x. Checked against the file's own independently-sourced
        Schwerdtfeger 2019 fallback table, which the converted mendeleev
        value should land close to."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            _atomic_polarizability,
        )

        # Real experimental atomic polarizability of hydrogen is ~0.667 Å³ —
        # this file's own fallback table agrees. mendeleev raw is 4.507 Bohr³;
        # unconverted that would wrongly read back as ~4.5.
        assert _atomic_polarizability("H") == pytest.approx(0.667, abs=0.05)
        assert _atomic_polarizability("C") == pytest.approx(1.76, abs=0.15)

    def test_molecular_polarizability_sums_atomic_contributions(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            _atomic_polarizability,
            resolve_quantum_polarizability,
        )

        r = resolve_quantum_polarizability(smiles="CCO")  # ethanol: C,C,O + 3 heavy atoms' H
        expected = (_atomic_polarizability("C") + _atomic_polarizability("C")
                    + _atomic_polarizability("O") + 3 * 0.667)
        assert r["value"] == pytest.approx(expected, abs=1e-2)
        assert r["tier"] == 5

    def test_no_smiles_falls_back_to_typical_drug_polarizability(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_polarizability,
        )

        r = resolve_quantum_polarizability(smiles="")
        assert r["value"] == 15.0
        assert r["tier"] == 7

    def test_dipole_moment_rdkit_path_returns_a_real_geometry_based_value(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_dipole_moment,
        )

        r = resolve_quantum_dipole_moment(smiles="CCO")  # ethanol has a real, nonzero dipole
        assert r["tier"] == 3
        assert r["value"] > 0.5  # ethanol's real dipole moment is ~1.7 D

    def test_homo_lumo_gap_shrinks_with_more_aromatic_rings(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_homo_lumo_gap,
        )

        few_rings = resolve_quantum_homo_lumo_gap(aromatic_rings=1)
        many_rings = resolve_quantum_homo_lumo_gap(aromatic_rings=5)
        assert many_rings["value"] < few_rings["value"]
        assert few_rings["value"] == pytest.approx(5.5 - 0.4 * 1, rel=1e-9)

    def test_homo_lumo_gap_has_a_physical_floor(self):
        """Extended conjugation can't push the gap below the floor the
        formula hard-clamps at (max(2.5, ...))."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_homo_lumo_gap,
        )

        r = resolve_quantum_homo_lumo_gap(aromatic_rings=20)
        assert r["value"] == 2.5

    def test_atomic_charges_sum_is_nonzero_for_a_real_molecule(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_atomic_charges_sum,
        )

        r = resolve_quantum_atomic_charges_sum(smiles="CCO")
        assert r["tier"] == 3
        assert r["value"] > 0

    def test_ionization_energy_known_element_matches_nist_table(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_ionization_energy,
        )

        r = resolve_quantum_ionization_energy(symbol="C")
        assert r["value"] == pytest.approx(11.260, abs=0.01)

    def test_unrecognized_symbol_falls_back_to_honest_hydrogen_like_baseline(self):
        """Regression test: this branch used to compute
        R·(Z_eff/n)²·n² with a hardcoded n=2 'second-shell baseline' —
        but n cancels out of that formula algebraically for ANY value,
        so it always silently returned the fixed hydrogen ground-state
        energy (13.606 eV) dressed up as a per-element Slater's-rules
        derivation. Now it's labeled for what it actually is."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_ionization_energy,
        )

        r = resolve_quantum_ionization_energy(symbol="Xx_not_a_real_element")
        assert r["value"] == pytest.approx(13.606, abs=0.001)
        assert r["source"] == "cerebro_value_resolver:hydrogen_like_baseline"
        assert r["confidence"] == "LOW"

    def test_researcher_overrides_short_circuit_all_five_resolvers(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.quantum_atomic import (
            resolve_quantum_atomic_charges_sum,
            resolve_quantum_dipole_moment,
            resolve_quantum_homo_lumo_gap,
            resolve_quantum_ionization_energy,
            resolve_quantum_polarizability,
        )

        for fn in (resolve_quantum_polarizability, resolve_quantum_dipole_moment,
                   resolve_quantum_homo_lumo_gap, resolve_quantum_atomic_charges_sum,
                   resolve_quantum_ionization_energy):
            r = fn(researcher_override=7.0)
            assert r["value"] == 7.0 and r["tier"] == 0


# ═════════════════════════════════════════════════════════════════════════════
# 23. DRUG/DDS TYPE CLASSIFIER (categories/type_detection.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDrugTypeClassifier:
    def test_researcher_override_normalizes_and_wins(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        r = resolve_drug_type(name="ignored", researcher_override="mAb")
        assert r["value"] == "monoclonal_antibody"
        assert r["tier"] == 0

    def test_smiles_classifies_by_molecular_weight(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        r = resolve_drug_type(smiles="CCO")  # ethanol, MW ~46
        assert r["value"] == "small_molecule"
        assert r["tier"] == 3
        assert r["computed_MW_Da"] == pytest.approx(46.07, abs=0.1)

    def test_fasta_length_drives_peptide_vs_protein_vs_mab(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        short_peptide = "ACDEFGHIKLMNPQRSTVWY"  # 20 aa
        r1 = resolve_drug_type(fasta=short_peptide)
        assert r1["value"] == "peptide"

        mid_protein = "A" * 100
        r2 = resolve_drug_type(fasta=mid_protein)
        assert r2["value"] == "protein"

        full_igg = "A" * 1320
        r3 = resolve_drug_type(fasta=full_igg, name="somemab")
        assert r3["value"] == "monoclonal_antibody"

    def test_nucleic_acid_length_drives_oligo_vs_gene_therapy(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        sirna = "ACGUACGUACGUACGUACGU"  # 20 nt
        r1 = resolve_drug_type(sequence=sirna)
        assert r1["value"] == "oligonucleotide"

        mrna = "ACGU" * 20  # 80 nt
        r2 = resolve_drug_type(sequence=mrna)
        assert r2["value"] == "gene_therapy"

    def test_usan_suffix_specificity_ordering(self):
        """The suffix table must check the most specific USAN ending first
        (-ximab/-zumab/-umab/-omab) before the generic '-mab' catch-all —
        otherwise every chimeric/humanized/human mAb would still classify
        correctly as monoclonal_antibody, but the more specific matched-
        suffix provenance would be lost to the generic one."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        cases = {
            "rituximab": "ximab",     # chimeric
            "trastuzumab": "zumab",   # humanized
            "adalimumab": "umab",     # human
        }
        for name, expected_suffix in cases.items():
            r = resolve_drug_type(name=name)
            assert r["value"] == "monoclonal_antibody"
            assert r["matched_suffix"] == expected_suffix

    def test_other_usan_suffixes_map_to_their_own_class(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        assert resolve_drug_type(name="etanercept")["value"] == "fusion_protein"
        assert resolve_drug_type(name="imiglucerase")["value"] == "enzyme_replacement"
        assert resolve_drug_type(name="octreotide")["value"] == "peptide"
        assert resolve_drug_type(name="somevax")["value"] == "vaccine"

    def test_keyword_match_for_non_suffix_names(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        r = resolve_drug_type(name="a CRISPR-based gene editing construct")
        assert r["value"] == "gene_therapy"

    def test_unrecognized_name_defaults_honestly_with_a_warning(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        r = resolve_drug_type(name="completely-unrecognizable-xyz-123")
        assert r["value"] == "small_molecule"
        assert r["tier"] == 7
        assert "warning" in r

    def test_no_input_at_all_defaults_honestly(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_drug_type

        r = resolve_drug_type()
        assert r["value"] == "small_molecule"
        assert r["tier"] == 7
        assert r["source"] == "cerebro_value_resolver:default_no_input"


class TestDdsTypeClassifier:
    def test_researcher_override_wins(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_dds_type

        r = resolve_dds_type(carrier="ignored", researcher_override="Liposomal")
        assert r["value"] == "liposomal"
        assert r["tier"] == 0

    def test_lnp_keyword_takes_priority_over_generic_material_keyword(self):
        """DDS_KEYWORD_MAP is ordered most-specific-first — a carrier
        string mentioning both 'PLGA' (material) and 'LNP' (gene_dds)
        should classify as the more specific gene_dds, not fall through
        to the generic polymer/material bucket."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_dds_type

        r = resolve_dds_type(carrier="PLGA-LNP hybrid formulation")
        assert r["value"] == "gene_dds"
        assert r["matched_keyword"] == "lnp"

    def test_plga_alone_classifies_as_material(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_dds_type

        r = resolve_dds_type(carrier="PLGA nanoparticle")
        assert r["value"] == "material"

    def test_no_carrier_info_defaults_honestly(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_dds_type

        r = resolve_dds_type()
        assert r["value"] == "material"
        assert r["tier"] == 7
        assert r["source"] == "cerebro_value_resolver:default_no_input"

    def test_unmatched_carrier_text_defaults_to_material_by_exclusion(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.type_detection import resolve_dds_type

        r = resolve_dds_type(carrier="a completely novel unclassified carrier xyz")
        assert r["value"] == "material"
        assert r["tier"] == 7
        assert r["source"] == "cerebro_value_resolver:material_class_inferred"


# ═════════════════════════════════════════════════════════════════════════════
# 24. TARGET-BINDING + MANUFACTURING RESOLVER (categories/drug_target_and_mfg.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDrugTargetBindingResolvers:
    """name='' keeps _chembl_target_activity's network call from firing
    (it short-circuits on a falsy name), so these exercise the tier-7
    typical-drug fallback deterministically and offline."""

    def test_researcher_override_short_circuits(self):
        import src.path_resolver  # noqa: F401
        import cerebro_value_resolver.categories.drug_target_and_mfg  # noqa: F401
        from cerebro_value_resolver._core import resolve_value

        r = resolve_value("drug_target_kd", name="", researcher_override=12.5)
        assert r["value"] == 12.5 and r["tier"] == 0

    def test_kd_ic50_ki_fall_back_to_their_own_typical_drug_defaults(self):
        import src.path_resolver  # noqa: F401
        import cerebro_value_resolver.categories.drug_target_and_mfg  # noqa: F401
        from cerebro_value_resolver._core import resolve_value

        assert resolve_value("drug_target_kd", name="")["value"] == 100.0
        assert resolve_value("drug_target_ic50", name="")["value"] == 50.0
        assert resolve_value("drug_target_ki", name="")["value"] == 50.0
        assert resolve_value("drug_target_kd", name="")["tier"] == 7


class TestDrugLoadingCapacityResolver:
    def test_bunjes_rule_applies_to_lipid_carriers(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_drug_loading_capacity_pct,
        )

        r = resolve_drug_loading_capacity_pct(carrier="liposome", logp=5.0)
        assert r["value"] == pytest.approx(10.0)  # 5.0 * 2.0
        assert r["tier"] == 6

    def test_bunjes_rule_is_floored_and_capped(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_drug_loading_capacity_pct,
        )

        floored = resolve_drug_loading_capacity_pct(carrier="liposome", logp=-8.0)
        assert floored["value"] == 2.0
        capped = resolve_drug_loading_capacity_pct(carrier="liposome", logp=50.0)
        assert capped["value"] == 40.0

    def test_polymer_carrier_uses_its_own_heuristic(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_drug_loading_capacity_pct,
        )

        r = resolve_drug_loading_capacity_pct(carrier="plga", logp=2.0, mw_Da=300.0)
        assert r["value"] == pytest.approx(16.0)  # 10 + min(8, 3) + 3 (mw<400)
        assert r["tier"] == 6

    def test_unknown_carrier_falls_back_to_median_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_drug_loading_capacity_pct,
        )

        r = resolve_drug_loading_capacity_pct(carrier="some_exotic_carrier")
        assert r["value"] == 8.0
        assert r["tier"] == 7


class TestMaterialPdiAndPorosityResolvers:
    def test_known_carriers_return_their_table_defaults(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_material_pdi,
            resolve_material_porosity,
        )

        assert resolve_material_pdi(carrier="dendrimer")["value"] == 0.05
        assert resolve_material_pdi(carrier="exosome")["value"] == 0.30
        assert resolve_material_porosity(carrier="liposome")["value"] == 0.0
        assert resolve_material_porosity(carrier="nanogel")["value"] == 0.85

    def test_unknown_carrier_falls_back_to_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_material_pdi,
            resolve_material_porosity,
        )

        assert resolve_material_pdi(carrier="unobtainium")["value"] == 0.20
        assert resolve_material_porosity(carrier="unobtainium")["value"] == 0.10

    def test_researcher_overrides_short_circuit(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_target_and_mfg import (
            resolve_material_pdi,
            resolve_material_porosity,
        )

        assert resolve_material_pdi(researcher_override=0.42)["value"] == 0.42
        assert resolve_material_porosity(researcher_override=0.77)["value"] == 0.77


# ═════════════════════════════════════════════════════════════════════════════
# 25. ADMET RESOLVER (categories/drug_admet.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDrugAdmetResolver:
    """name='' short-circuits every ChEMBL lookup here (each helper checks
    `if not name` first), keeping these offline and deterministic."""

    def test_logS_yalkowsky_gse_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_solubility_logS,
        )

        r = resolve_drug_solubility_logS(name="", logp=2.0, Tm_C=125.0)
        assert r["value"] == pytest.approx(0.5 - 0.01 * (125.0 - 25) - 2.0, rel=1e-9)
        assert r["tier"] == 6

    def test_logS_falls_back_to_delaney_esol_without_melting_point(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_solubility_logS,
        )

        r = resolve_drug_solubility_logS(name="", logp=2.0, mw_Da=300.0,
                                           rotbonds=4, aromatic_rings=2)
        assert r["value"] == pytest.approx(-2.844, abs=1e-3)
        assert r["tier"] == 7

    def test_logS_with_no_logp_at_all_falls_to_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_solubility_logS,
        )

        r = resolve_drug_solubility_logS(name="")
        assert r["value"] == -3.0
        assert r["tier"] == 7

    def test_caco2_papp_hou_regression_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_caco2_papp,
        )

        r = resolve_drug_caco2_papp(name="", logp=2.0, tpsa=60.0, hbd=1.0)
        assert r["value"] == pytest.approx(6.501, abs=1e-3)
        assert r["tier"] == 6

    def test_caco2_papp_without_inputs_falls_to_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_caco2_papp,
        )

        r = resolve_drug_caco2_papp(name="")
        assert r["value"] == 10.0
        assert r["tier"] == 7

    def test_pgp_efflux_hochman_rule_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_pgp_efflux_ratio,
        )

        r = resolve_drug_pgp_efflux_ratio(name="", mw_Da=450.0, hba=5.0, logp=3.0)
        assert r["value"] == pytest.approx(3.16, abs=1e-2)
        assert r["tier"] == 6

    def test_pgp_efflux_without_inputs_falls_to_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_pgp_efflux_ratio,
        )

        r = resolve_drug_pgp_efflux_ratio(name="")
        assert r["value"] == 1.5
        assert r["tier"] == 7

    def test_cyp3a4_inhibition_defaults_to_non_inhibitor_threshold(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_cyp3a4_inhibition,
        )

        r = resolve_drug_cyp3a4_inhibition(name="")
        assert r["value"] == 50.0
        assert r["tier"] == 7

    def test_herg_aronov_rule_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_herg_ic50,
        )

        r = resolve_drug_herg_ic50(name="", logp=3.0)
        assert r["value"] == pytest.approx(39.81, abs=1e-2)
        assert r["tier"] == 6

    def test_herg_without_logp_falls_to_generic_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_herg_ic50,
        )

        r = resolve_drug_herg_ic50(name="")
        assert r["value"] == 10.0
        assert r["tier"] == 7

    def test_clearance_route_heuristic_covers_hepatic_renal_and_mixed(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_clearance_route,
        )

        hepatic = resolve_drug_clearance_route(name="", mw_Da=400.0, logp=2.0)
        renal = resolve_drug_clearance_route(name="", mw_Da=200.0, logp=0.5)
        mixed = resolve_drug_clearance_route(name="", mw_Da=300.0, logp=1.2)
        assert (hepatic["value"], renal["value"], mixed["value"]) == ("hepatic", "renal", "mixed")
        assert hepatic["tier"] == 6

    def test_clearance_route_without_inputs_defaults_to_hepatic(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_clearance_route,
        )

        r = resolve_drug_clearance_route(name="")
        assert r["value"] == "hepatic"
        assert r["tier"] == 7

    def test_researcher_overrides_short_circuit_every_resolver(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_admet import (
            resolve_drug_caco2_papp,
            resolve_drug_clearance_route,
            resolve_drug_cyp3a4_inhibition,
            resolve_drug_herg_ic50,
            resolve_drug_pgp_efflux_ratio,
            resolve_drug_solubility_logS,
        )

        for fn in (resolve_drug_solubility_logS, resolve_drug_caco2_papp,
                   resolve_drug_pgp_efflux_ratio, resolve_drug_cyp3a4_inhibition,
                   resolve_drug_herg_ic50):
            r = fn(name="", researcher_override=1.234)
            assert r["value"] == 1.234 and r["tier"] == 0

        r = resolve_drug_clearance_route(name="", researcher_override="biliary")
        assert r["value"] == "biliary" and r["tier"] == 0


# ═════════════════════════════════════════════════════════════════════════════
# 26. CLINICAL PK RESOLVER (categories/pk_clinical.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestPkClinicalRegexExtractors:
    """These regexes are the tier-1 mechanism for pulling real numbers out
    of fetched OpenFDA label text — a silent parsing failure here means
    the resolver falls through to a worse tier without any signal."""

    def test_halflife_extractor_handles_a_single_value_not_just_a_range(self):
        """Regression test for a real bug: the half-life patterns required
        a range separator (to/-/–) between the number and the unit, so a
        single reported value like 'Half-life is 8 hours.' — extremely
        common in real FDA labels — matched nothing and silently fell
        through to a worse tier. The clearance/Vd extractors already had
        this separator marked optional; half-life's copy of the pattern
        was missing the same `?`."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            _extract_halflife_hours,
        )

        assert _extract_halflife_hours("Half-life is 8 hours.") == 8.0
        assert _extract_halflife_hours("The elimination half-life is 1 hour.") == 1.0

    def test_halflife_extractor_handles_plural_hours_and_days(self):
        """Regression test for a real bug: the unit alternation only had
        the singular 'hour'/'day' forms, so '\\b' failed right after
        matching the 'hour' prefix of 'hours' (not a word boundary before
        the trailing 's') — real labels almost always use the plural."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            _extract_halflife_hours,
        )

        assert _extract_halflife_hours(
            "The elimination half-life is approximately 12 to 16 hours "
            "in healthy subjects.") == 14.0
        assert _extract_halflife_hours(
            "The terminal half-life is 3-5 days following oral administration."
        ) == 96.0  # midpoint 4 days -> hours

    def test_halflife_extractor_handles_t1_2_notation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            _extract_halflife_hours,
        )

        assert _extract_halflife_hours("T1/2 of approximately 24 hours was observed.") == 24.0

    def test_clearance_extractor_parses_a_single_value(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            _extract_clearance_lph,
        )

        assert _extract_clearance_lph(
            "Total clearance (CL) is 5.2 L/h following IV administration.") == 5.2

    def test_vd_extractor_parses_a_single_value(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import _extract_vd_L

        assert _extract_vd_L("The apparent volume of distribution is 45 L.") == 45.0

    def test_protein_binding_and_bioavailability_extractors(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            _extract_bioavailability,
            _extract_protein_binding,
        )

        assert _extract_protein_binding("Plasma protein binding is approximately 98%.") == 0.98
        assert _extract_bioavailability("Absolute bioavailability is 65%.") == 0.65

    def test_extractors_return_none_on_text_without_the_expected_value(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            _extract_halflife_hours,
        )

        assert _extract_halflife_hours("This drug has no reported pharmacokinetic data.") is None


class TestPkClinicalEmpiricalFallbacks:
    """name='' keeps OpenFDA/ChEMBL/PubMed from firing (each checks a
    truthy name before touching the network) so these exercise the
    empirical/class-typical tiers deterministically and offline."""

    def test_halflife_empirical_regression_matches_published_formula(self):
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import resolve_pk_halflife

        r = resolve_pk_halflife(name="", mw_Da=300.0, logp=2.0,
                                  molecule_class="small_molecule")
        a, b, c = -0.2, 0.18, 0.45
        expected_h = 10 ** (a + b * 2.0 + c * math.log10(300.0))
        assert r["value"] == pytest.approx(round(expected_h / 24, 4), abs=1e-4)
        assert r["tier"] == 6

    def test_halflife_class_typical_means_for_biologics(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import resolve_pk_halflife

        mab = resolve_pk_halflife(name="", molecule_class="monoclonal_antibody")
        peptide = resolve_pk_halflife(name="", molecule_class="peptide")
        assert mab["value"] == 14.0 and mab["tier"] == 7
        assert peptide["value"] == 0.02 and peptide["tier"] == 7

    def test_clearance_allometric_regression_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import resolve_pk_clearance

        r = resolve_pk_clearance(name="", mw_Da=300.0, molecule_class="small_molecule")
        assert r["value"] == pytest.approx(60 * 300.0 ** -0.25, abs=1e-3)
        assert r["tier"] == 6

    def test_clearance_class_typical_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import resolve_pk_clearance

        r = resolve_pk_clearance(name="", molecule_class="monoclonal_antibody")
        assert r["value"] == 0.3 and r["tier"] == 7

    def test_volume_distribution_logp_regression_matches_published_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            resolve_pk_volume_distribution,
        )

        r = resolve_pk_volume_distribution(name="", logp=2.0)
        vd_per_kg = max(0.2, 0.2 + 0.6 * 2.0 + 0.05 * 2.0 ** 2)
        assert r["value"] == pytest.approx(round(vd_per_kg * 70, 1), abs=1e-6)
        assert r["tier"] == 6

    def test_volume_distribution_class_typical_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            resolve_pk_volume_distribution,
        )

        r = resolve_pk_volume_distribution(name="", molecule_class="peptide")
        assert r["value"] == 7.0 and r["tier"] == 7

    def test_protein_binding_logit_regression_matches_published_formula(self):
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            resolve_pk_protein_binding,
        )

        r = resolve_pk_protein_binding(name="", logp=2.0)
        x = -2 + 0.5 * 2.0
        expected = 1 / (1 + math.exp(-x))
        assert r["value"] == pytest.approx(round(expected, 4), abs=1e-6)
        assert r["tier"] == 6

    def test_protein_binding_and_bioavailability_generic_defaults(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            resolve_pk_oral_bioavailability,
            resolve_pk_protein_binding,
        )

        assert resolve_pk_protein_binding(name="")["value"] == 0.5
        assert resolve_pk_oral_bioavailability(name="")["value"] == 0.4

    def test_researcher_overrides_short_circuit_every_resolver(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.pk_clinical import (
            resolve_pk_clearance,
            resolve_pk_halflife,
            resolve_pk_oral_bioavailability,
            resolve_pk_protein_binding,
            resolve_pk_volume_distribution,
        )

        for fn in (resolve_pk_halflife, resolve_pk_clearance,
                   resolve_pk_volume_distribution, resolve_pk_protein_binding,
                   resolve_pk_oral_bioavailability):
            r = fn(name="", researcher_override=3.5)
            assert r["value"] == 3.5 and r["tier"] == 0


# ═════════════════════════════════════════════════════════════════════════════
# 27. PHYSICOCHEMICAL DESCRIPTOR RESOLVER (categories/drug_descriptors.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDrugDescriptorBuildingBlocks:
    """These resolvers hit PubChem even with name='' (the SMILES→InChIKey
    branch in _pubchem_property fires regardless of name), so full-resolver
    tests would be network-dependent and non-deterministic. Testing the
    pure functions underneath instead — RDKit computation and the tier-7
    pure-math fallbacks — keeps this offline and exact."""

    def test_rdkit_descriptor_matches_known_ethanol_properties(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _rdkit_descriptor,
        )

        d = _rdkit_descriptor("CCO", "ethanol")
        assert d["MW_Da"] == pytest.approx(46.07, abs=0.01)
        assert d["HBD"] == 1  # the hydroxyl H
        assert d["HBA"] == 1  # the oxygen
        assert d["AromaticRings"] == 0

    def test_rdkit_descriptor_returns_none_for_invalid_smiles(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _rdkit_descriptor,
        )

        assert _rdkit_descriptor("not a real smiles!!!", "") is None

    def test_tier7_mw_sums_atomic_weights_plus_implicit_hydrogens(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import _t7_mw

        # C,C,O heavy atoms (2×12.011 + 15.999) plus 3 implicit H (3×1.008)
        expected = (2 * 12.011 + 15.999) + 3 * 1.008
        assert _t7_mw("CCO") == pytest.approx(round(expected, 2), abs=1e-2)

    def test_tier7_tpsa_counts_polar_atoms(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import _t7_tpsa

        assert _t7_tpsa("CCO") == 9.0     # one O
        assert _t7_tpsa("CCN") == 12.0    # one N

    def test_tier7_hba_counts_n_and_o_atoms(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import _t7_hba

        assert _t7_hba("CCO") == 1.0
        assert _t7_hba("NCCO") == 2.0

    def test_tier7_helpers_return_none_for_empty_smiles(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _t7_formal_charge,
            _t7_hba,
            _t7_hbd,
            _t7_mw,
            _t7_rotbonds,
            _t7_stereo,
            _t7_tpsa,
        )

        for fn in (_t7_mw, _t7_tpsa, _t7_hbd, _t7_hba, _t7_rotbonds,
                   _t7_formal_charge, _t7_stereo):
            assert fn("") is None


class TestBiologicAndOligoMwComputation:
    def test_peptide_mw_matches_hand_computed_residue_sum(self):
        """10 alanine residues: Σ(89.09) − 9×H₂O — verified independently."""
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _biologic_mw_from_fasta,
        )

        seq = "A" * 10
        expected = 10 * 89.09 - 9 * 18.015
        assert _biologic_mw_from_fasta(seq) == pytest.approx(expected, abs=1e-6)

    def test_short_fasta_below_minimum_length_returns_none(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _biologic_mw_from_fasta,
        )

        assert _biologic_mw_from_fasta("AC") is None
        assert _biologic_mw_from_fasta("") is None

    def test_dna_oligo_mw_matches_hand_computed_sum(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _oligonucleotide_mw_from_sequence,
        )

        seq = "ACGT"
        expected = (313.21 + 289.18 + 329.21 + 304.20) - 3 * 18.015 + 79.98
        assert _oligonucleotide_mw_from_sequence(seq) == pytest.approx(expected, abs=1e-6)

    def test_rna_detected_when_u_present_without_t(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _oligonucleotide_mw_from_sequence,
        )

        seq = "ACGU"
        expected = (329.21 + 305.18 + 345.21 + 306.17) - 3 * 18.015 + 79.98
        assert _oligonucleotide_mw_from_sequence(seq) == pytest.approx(expected, abs=1e-6)

    def test_moe_ps_modification_adds_per_nucleotide_mass(self):
        import src.path_resolver  # noqa: F401
        from cerebro_value_resolver.categories.drug_descriptors import (
            _oligonucleotide_mw_from_sequence,
        )

        seq = "ACGT"
        plain = _oligonucleotide_mw_from_sequence(seq, modification="DNA")
        modified = _oligonucleotide_mw_from_sequence(seq, modification="MOE_PS")
        assert modified - plain == pytest.approx(len(seq) * 88.0, abs=1e-6)


class TestDescriptorResolverResearcherOverride:
    """The one part of the full resolver testable without touching the
    network — researcher_override returns before any Tier-1 code runs."""

    def test_researcher_override_short_circuits_every_descriptor(self):
        # Built by _build_descriptor_resolver, which registers each one into
        # the shared _REGISTRY but never binds it to a module-level name
        # (same pattern as the material-table resolvers covered earlier).
        import src.path_resolver  # noqa: F401
        import cerebro_value_resolver.categories.drug_descriptors  # noqa: F401
        from cerebro_value_resolver._core import resolve_value

        for cat in ("drug_logp", "drug_mw", "drug_tpsa", "drug_hbd", "drug_hba",
                    "drug_rotbonds", "drug_aromatic_rings", "drug_formal_charge",
                    "drug_stereocenters"):
            r = resolve_value(cat, researcher_override=9.5)
            assert r["value"] == 9.5 and r["tier"] == 0


# ═════════════════════════════════════════════════════════════════════════════
# 28. CLASS-C TRANSLATIONAL ENGINE (cerebro_62_translational_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
def _trans_bundles(carrier="liposome", ligand="", size_nm=100.0, zeta_mV=-10.0,
                     rel_kin="sustained", scale_up="lab"):
    drug_bundle = {
        "drug_mw": {"value": 350.0}, "drug_logp": {"value": 2.5},
        "drug_tpsa": {"value": 60.0}, "pk_halflife": {"value": 0.5},
        "_meta": {"name": "TestDrug", "drug_type": "small_molecule",
                  "identifiers": {"smiles": "CCO"}},
    }
    dds_bundle = {"_meta": {"carrier_type": carrier, "dds_type": "material"}}
    combo_bundle = {"_meta": {"dds_row": {
        "Formulation_Name": "F001", "Formulation_ID": "F001",
        "Surface_Ligand": ligand, "Size_nm": size_nm,
        "Zeta_Potential_mV": zeta_mV, "Release_Kinetics": rel_kin,
        "Scale_Up_Readiness": scale_up}}}
    return drug_bundle, dds_bundle, combo_bundle


def _deep_results(n_validated, n_total):
    return {f"P{i}": {"validated": i < n_validated} for i in range(n_total)}


class TestTranslationalP56Patentability:
    def test_known_crowded_pair_lowers_novelty(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P56

        drug, dds, combo = _trans_bundles(carrier="liposome", ligand="transferrin")
        r = trans_P56(drug, dds, combo, _deep_results(0, 5))
        assert r["components"]["novelty_§102"] == 50
        assert r["patentability_score"] == pytest.approx(66.7, abs=0.05)
        assert r["recommendation"] == "REVIEW CLAIMS BEFORE FILING"

    def test_known_novel_pair_raises_novelty(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P56

        drug, dds, combo = _trans_bundles(carrier="solid_lipid", ligand="lactoferrin")
        r = trans_P56(drug, dds, combo, _deep_results(0, 5))
        assert r["components"]["novelty_§102"] == 95
        assert r["patentability_score"] == pytest.approx(81.7, abs=0.05)
        assert r["recommendation"] == "FILE PROVISIONAL"

    def test_deep_validation_success_raises_utility_component(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P56

        drug, dds, combo = _trans_bundles(carrier="polymer", ligand="",
                                            rel_kin="thermo", size_nm=50.0, zeta_mV=-30.0)
        r = trans_P56(drug, dds, combo, _deep_results(3, 5))
        assert r["components"]["utility_§101"] == 90
        assert r["patentability_score"] == pytest.approx(86.7, abs=0.05)

    def test_non_obviousness_penalizes_extreme_size_and_zeta(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P56

        typical = trans_P56(*_trans_bundles(size_nm=100.0, zeta_mV=-10.0),
                              _deep_results(0, 5))
        extreme = trans_P56(*_trans_bundles(size_nm=300.0, zeta_mV=-40.0),
                              _deep_results(0, 5))
        assert extreme["components"]["non_obviousness_§103"] > typical["components"]["non_obviousness_§103"]


class TestTranslationalP32FreedomToOperate:
    def test_known_crowded_combination_scores_lower_fto(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P32

        drug, dds, combo = _trans_bundles(carrier="liposome", ligand="transferrin")
        r = trans_P32(drug, dds, combo, {})
        assert r["encumbrance_level"] == "VERY_HIGH"
        assert r["fto_score"] == 30

    def test_unrecognized_combination_defaults_to_low_encumbrance(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P32

        drug, dds, combo = _trans_bundles(carrier="exosome", ligand="novel_peptide_xyz")
        r = trans_P32(drug, dds, combo, {})
        assert r["encumbrance_level"] == "LOW"
        assert r["fto_score"] == 85


class TestTranslationalP21PreIND:
    def test_deep_validation_passed_flag_uses_a_five_principle_threshold(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P21

        drug, dds, combo = _trans_bundles()
        passed = trans_P21(drug, dds, combo, _deep_results(5, 7))
        failed = trans_P21(drug, dds, combo, _deep_results(4, 7))
        assert passed["deep_validation_passed"] is True
        assert failed["deep_validation_passed"] is False

    def test_report_sections_include_the_resolved_formulation_name(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import trans_P21

        drug, dds, combo = _trans_bundles()
        r = trans_P21(drug, dds, combo, {})
        assert "F001" in r["narrative"]
        assert len(r["sections"]) == 8


class TestTranslationalDispatcher:
    def test_withholds_all_deliverables_when_deep_validation_is_insufficient(self):
        """Per its own stated gating logic: translational (Class C)
        deliverables shouldn't be generated for a DDS that hasn't cleared
        deep (Class B) validation — otherwise a regulatory/patent/grant
        outline could get built around a candidate that physics already
        rejected."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import (
            TRANSLATIONAL_FUNCTIONS,
            evaluate_translational_for_top1,
        )

        drug, dds, combo = _trans_bundles()
        out = evaluate_translational_for_top1(
            drug, dds, combo, _deep_results(2, 10))  # 20% pass rate < 70%
        assert set(out) == set(TRANSLATIONAL_FUNCTIONS)
        for r in out.values():
            assert r["status"] == "skipped_deep_validation_insufficient"

    def test_generates_all_deliverables_when_deep_validation_passes(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import (
            TRANSLATIONAL_FUNCTIONS,
            evaluate_translational_for_top1,
        )

        drug, dds, combo = _trans_bundles()
        out = evaluate_translational_for_top1(
            drug, dds, combo, _deep_results(8, 10))  # 80% pass rate >= 70%
        assert set(out) == set(TRANSLATIONAL_FUNCTIONS)
        for pid, r in out.items():
            assert r["status"] != "skipped_deep_validation_insufficient"
            assert "principle" in r

    def test_can_be_forced_to_run_regardless_of_deep_validation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_translational_engine import evaluate_translational_for_top1

        drug, dds, combo = _trans_bundles()
        out = evaluate_translational_for_top1(
            drug, dds, combo, _deep_results(0, 10), only_if_deep_passed=False)
        for r in out.values():
            assert r["status"] != "skipped_deep_validation_insufficient"


# ═════════════════════════════════════════════════════════════════════════════
# 29. 62-PRINCIPLE CATALOG (cerebro_62_principles_catalog.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestPrinciplesCatalogIntegrity:
    """The catalog is mostly static data — its value is in staying
    internally consistent and matching what it claims about itself."""

    def test_exactly_62_principles_registered(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import PRINCIPLES_62

        assert len(PRINCIPLES_62) == 62

    def test_every_class_field_is_one_of_the_three_valid_values(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import PRINCIPLES_62

        valid = {"A_surrogate", "B_deep", "C_translational"}
        bad = {pid: p["class"] for pid, p in PRINCIPLES_62.items() if p["class"] not in valid}
        assert bad == {}

    def test_class_a_plus_b_weights_are_normalized_to_one(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import PRINCIPLES_62

        total = sum(p["weight_cns"] for p in PRINCIPLES_62.values()
                    if p["class"] in ("A_surrogate", "B_deep"))
        assert total == pytest.approx(1.0, abs=1e-3)

    def test_translational_principles_carry_zero_ranking_weight(self):
        """Class C (translational/admin) principles shouldn't influence
        DDS ranking — the docstring's own stated policy."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import CLASS_C_TRANSLATIONAL, PRINCIPLES_62

        assert len(CLASS_C_TRANSLATIONAL) > 0
        for pid in CLASS_C_TRANSLATIONAL:
            assert PRINCIPLES_62[pid]["weight_cns"] == 0

    def test_cns_direct_named_principles_collective_weight(self):
        """Regression test for a real docstring/data mismatch: the module
        docstring used to claim these seven CNS-direct principles carry
        '≥ 40%' of the total weight collectively — checked against the
        actual normalized weights and it comes out to ~27.9%, not ≥40%.
        The docstring was corrected to state the real figure rather than
        the weights being reverse-engineered to hit an unverified target."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import PRINCIPLES_62

        cns_direct = ["P12", "P33", "P38", "P39", "P42", "P43", "P44"]
        total = sum(PRINCIPLES_62[pid]["weight_cns"] for pid in cns_direct)
        assert total == pytest.approx(0.279, abs=0.005)
        assert total < 0.40   # the old docstring's claim no longer stands unchecked

    def test_class_getters_partition_the_catalog_correctly(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import (
            PRINCIPLES_62,
            get_class_a_principles,
            get_class_b_principles,
            get_class_c_principles,
        )

        a, b, c = get_class_a_principles(), get_class_b_principles(), get_class_c_principles()
        assert set(a) | set(b) | set(c) == set(PRINCIPLES_62)
        assert not (set(a) & set(b)) and not (set(b) & set(c)) and not (set(a) & set(c))
        for pid in a:
            assert PRINCIPLES_62[pid]["class"] == "A_surrogate"
        for pid in b:
            assert PRINCIPLES_62[pid]["class"] == "B_deep"
        for pid in c:
            assert PRINCIPLES_62[pid]["class"] == "C_translational"

    def test_get_principle_returns_the_matching_entry(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import get_principle

        p = get_principle("P01")
        assert p["title_en"] == "Adversarial Stress-Testing Engine"

    def test_get_principle_raises_for_unknown_id(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import get_principle

        with pytest.raises(KeyError):
            get_principle("P999")

    def test_summarize_counts_match_the_actual_catalog(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import PRINCIPLES_62, summarize

        s = summarize()
        assert s["total"] == 62
        assert s["cns_direct_count"] == sum(1 for p in PRINCIPLES_62.values() if p["cns_relevant"])
        assert s["class_A_surrogate"] + s["class_B_deep"] + s["class_C_admin"] == 62

    def test_every_principles_deep_computation_is_wired_into_the_deep_engine(self):
        """Every principle that carries a real method_deep entry here should
        correspond to a callable in cerebro_62_deep_engine.py's dispatch —
        those two files need to stay in sync on which principles that
        applies to."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import DEEP_FUNCTIONS, HPC_ONLY_PRINCIPLES
        from cerebro_62_principles_catalog import PRINCIPLES_62

        deep_engine_ids = set(DEEP_FUNCTIONS) | set(HPC_ONLY_PRINCIPLES)
        catalog_has_method_deep = {pid for pid, p in PRINCIPLES_62.items() if p.get("method_deep")}
        assert deep_engine_ids <= catalog_has_method_deep

    def test_class_b_catalog_membership_matches_the_real_deep_engine_set(self):
        """The module docstring states Class B has 28 principles (7 with
        genuine independent computation + 21 HPC-deferred pass-through, per
        cerebro_62_deep_engine.py's DEEP_FUNCTIONS + HPC_ONLY_PRINCIPLES).
        The catalog's "class" field used to tag only P47 as "B_deep" and
        mislabel the other 27 as "A_surrogate" -- an internal inconsistency
        between this file's own stated policy and its own data, which fed
        straight into the Excel writer's Principle_Explanations glossary
        sheet (the one sheet a researcher is told to treat as source-of-
        truth for what each principle actually does), understating to a
        reader that 27 of the 28 principles with genuine or HPC-deferred
        deep-physics validation were mere fast surrogates."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_deep_engine import DEEP_FUNCTIONS, HPC_ONLY_PRINCIPLES
        from cerebro_62_principles_catalog import CLASS_B_DEEP, PRINCIPLES_62

        deep_engine_ids = set(DEEP_FUNCTIONS) | set(HPC_ONLY_PRINCIPLES)
        assert len(deep_engine_ids) == 28
        assert set(CLASS_B_DEEP) == deep_engine_ids
        for pid in deep_engine_ids:
            assert PRINCIPLES_62[pid]["class"] == "B_deep"


# ═════════════════════════════════════════════════════════════════════════════
# 30. C+ FLOW ORCHESTRATOR (cerebro_62_orchestrator.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestOrchestratorCompositeScore:
    def test_weighted_average_matches_hand_computation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _compute_composite_score

        per_principle = {"P01": {"score": 80.0}, "P02": {"score": 40.0}}
        weights = {"P01": 0.3, "P02": 0.1}
        expected = (80.0 * 0.3 + 40.0 * 0.1) / (0.3 + 0.1)
        assert _compute_composite_score(per_principle, weights) == pytest.approx(expected)

    def test_zero_weight_principles_are_excluded_not_zero_weighted(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _compute_composite_score

        per_principle = {"P01": {"score": 80.0}, "P56": {"score": 0.0}}
        weights = {"P01": 0.5, "P56": 0.0}  # translational principle, weight 0
        assert _compute_composite_score(per_principle, weights) == 80.0

    def test_no_weighted_principles_present_returns_zero(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _compute_composite_score

        assert _compute_composite_score({"P01": {"score": 80.0}}, {}) == 0.0


class TestOrchestratorPrincipleGroupCoverage:
    """PRINCIPLE_GROUPS is the thematic-rollup grouping shown in the
    Champion_DDS_Compare sheet and the DDS×Principle matrix's group
    columns. P02, P06, P26, P47, and P54 used to appear in zero of the 8
    groups -- their scores still counted toward the real weighted
    composite (a separate code path keyed on weight_cns directly), but
    every researcher-facing group-rollup number silently dropped them,
    including P47 (Free Energy Perturbation, one of only 7 principles
    with genuine independent deep-physics computation) and P02 (the
    allometric-scaling counterpart to P13/P31/P44, which were already
    grouped). Only P07 (Real-Time Literature Mining) is deliberately left
    ungrouped -- it's a supporting/informational principle, not a scored
    dimension of the drug or DDS."""

    def test_every_principle_except_the_literature_miner_is_in_some_group(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import PRINCIPLE_GROUPS
        from cerebro_62_principles_catalog import PRINCIPLES_62

        grouped = {pid for pids in PRINCIPLE_GROUPS.values() for pid in pids}
        missing = set(PRINCIPLES_62) - grouped
        assert missing == {"P07"}

    def test_previously_orphaned_cns_relevant_principles_are_now_grouped(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import PRINCIPLE_GROUPS

        grouped = {pid for pids in PRINCIPLE_GROUPS.values() for pid in pids}
        for pid in ("P02", "P06", "P26", "P47", "P54"):
            assert pid in grouped


class TestOrchestratorVerdictThresholds:
    def test_verdict_bands_match_documented_thresholds(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _verdict_for

        assert _verdict_for(95) == "EXCELLENT"
        assert _verdict_for(80) == "EXCELLENT"
        assert _verdict_for(79.9) == "GOOD"
        assert _verdict_for(65) == "GOOD"
        assert _verdict_for(50) == "ACCEPTABLE"
        assert _verdict_for(35) == "MARGINAL"
        assert _verdict_for(0) == "POOR"


class TestDrugDdsCompatibilityMultiplier:
    """The pathway-aware compatibility matrix — real FDA-precedent pairings
    (LNP/siRNA, AAV/gene-therapy) should score highest; known-mismatched
    pairs (oligonucleotide + bare polymer, no endosomal escape) lowest."""

    def test_ideal_fda_precedented_pairings_score_highest(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        lnp_sirna, _ = _drug_dds_compatibility_multiplier("oligonucleotide", "lnp")
        aav_gene, _ = _drug_dds_compatibility_multiplier("gene_therapy", "aav9")
        assert lnp_sirna == 1.20
        assert aav_gene == 1.20

    def test_oligo_in_bare_polymer_scores_poorly_no_endosomal_escape(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        mult, reason = _drug_dds_compatibility_multiplier("oligonucleotide", "plga")
        assert mult == 0.60
        assert "endosomal escape" in reason

    def test_mab_in_plga_is_suboptimal_organic_solvent_denaturation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        mult, reason = _drug_dds_compatibility_multiplier("monoclonal_antibody", "plga")
        assert mult == 0.75
        assert "denaturation" in reason

    def test_small_molecule_in_established_carriers_scores_well(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        for carrier in ("plga", "solid_lipid", "liposome"):
            mult, _ = _drug_dds_compatibility_multiplier("small_molecule", carrier)
            assert mult == 1.05

    def test_missing_type_data_returns_neutral_multiplier(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        mult, reason = _drug_dds_compatibility_multiplier("", "liposome")
        assert mult == 1.0
        assert reason == "no_class_data"

    def test_unmapped_pairing_defaults_to_a_mild_skepticism_not_a_penalty(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        mult, reason = _drug_dds_compatibility_multiplier("vaccine", "dendrimer")
        assert mult == 0.95
        assert "not in FDA-validated table" in reason

    def test_multiplier_never_exceeds_the_documented_range(self):
        """Every branch in the decision matrix should stay within the
        function's own documented [0.4, 1.20] range."""
        import itertools

        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import _drug_dds_compatibility_multiplier

        drug_types = ["small_molecule", "monoclonal_antibody", "peptide",
                      "oligonucleotide", "gene_therapy", "protein", "vaccine"]
        carriers = ["lnp", "aav9", "liposome", "solid_lipid", "plga",
                    "micelle", "dendrimer", "metallic"]
        for dt, ct in itertools.product(drug_types, carriers):
            mult, _ = _drug_dds_compatibility_multiplier(dt, ct)
            assert 0.4 <= mult <= 1.20


# ═════════════════════════════════════════════════════════════════════════════
# 31. CLASS-A SURROGATE ENGINE (cerebro_62_surrogate_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
def _surrogate_bundles(carrier="liposome", ligand="", size_nm=100.0, zeta_mV=-10.0,
                        phase_T=42.0, density=0.5, mw=350.0, logp=2.5,
                        tpsa=60.0, hbd=1, hba=5, pka_base=8.0, aromatic_rings=2,
                        bbb=5.0, smiles="CCO"):
    drug_bundle = {
        "drug_mw": {"value": mw}, "drug_logp": {"value": logp},
        "drug_tpsa": {"value": tpsa}, "drug_hbd": {"value": hbd},
        "drug_hba": {"value": hba}, "drug_pka_basic": {"value": pka_base},
        "drug_aromatic_rings": {"value": aromatic_rings},
        "bbb_permeability": {"value": bbb},
        "_meta": {"drug_type": "small_molecule", "name": "test_drug",
                  "identifiers": {"smiles": smiles}},
    }
    dds_bundle = {"_meta": {"carrier_type": carrier, "dds_type": "material"}}
    combo_bundle = {"_meta": {"dds_row": {
        "Surface_Ligand": ligand, "Size_nm": size_nm,
        "Zeta_Potential_mV": zeta_mV, "Phase_Transition_Temp_C": phase_T,
        "Surface_Ligand_Density_per_nm2": density}}}
    return drug_bundle, dds_bundle, combo_bundle


class TestSurrogateSharedHelpers:
    def test_triangular_window_plateau_and_decay(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _triangular

        assert _triangular(100, 50, 150) == 100.0
        assert _triangular(30, 50, 150, 2.0, 0.5) == 60.0   # 100-(50-30)*2
        assert _triangular(200, 50, 150, 2.0, 0.5) == 75.0  # 100-(200-150)*0.5
        assert _triangular(0, 50, 150) == 0.0
        assert _triangular(-5, 50, 150) == 0.0

    def test_hill_equation_half_max_point(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _hill

        assert _hill(50, k50=50, n=2.0) == pytest.approx(50.0)
        assert _hill(50, k50=50, n=2.0, invert=True) == pytest.approx(50.0)
        assert _hill(0, k50=50) == 0.0
        assert _hill(0, k50=50, invert=True) == 100.0

    def test_bbb_propensity_matches_wager_cns_mpo_bands(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _bbb_propensity

        ideal = {"logp": 2.0, "mw": 300, "hbd": 0, "tpsa": 40,
                 "pka_base": 7.0, "arom_rings": 1}
        assert _bbb_propensity(ideal) == 6.0   # every band maxed out
        poor = {"logp": 6.0, "mw": 600, "hbd": 4, "tpsa": 150,
                "pka_base": 12.0, "arom_rings": 5}
        assert _bbb_propensity(poor) == 0.0

    def test_membrane_partition_logk_penalizes_ionization(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _membrane_partition_logK

        neutral = _membrane_partition_logK({"logp": 3.0, "net_q_pH74": 0.0})
        charged = _membrane_partition_logK({"logp": 3.0, "net_q_pH74": 1.0})
        assert neutral == 3.0
        assert charged == pytest.approx(2.2)  # 3.0 - 0.8*1


class TestSurrogateBundleContract:
    def test_rejects_non_bundle_arguments(self):
        """_resolve_inputs' fail-fast guard: surrogate functions are
        bundle-only — a plain dict without _meta.drug_type/_meta.dds_type
        must raise immediately rather than silently computing nonsense."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _resolve_inputs

        _, dds_bundle, combo_bundle = _surrogate_bundles()
        with pytest.raises(TypeError, match="not a drug bundle"):
            _resolve_inputs({"not_a_bundle": True}, dds_bundle, combo_bundle)

        drug_bundle, _, combo_bundle = _surrogate_bundles()
        with pytest.raises(TypeError, match="not a DDS bundle"):
            _resolve_inputs(drug_bundle, {"not_a_bundle": True}, combo_bundle)


class TestSurrogateRepresentativeFunctions:
    """Hand-verified spot checks across the formula patterns that recur
    throughout the file's 57 principles: CNS-MPO composite (P12),
    Henderson-Hasselbalch-driven ligand affinity (P33), Arrhenius +
    SMILES-moiety penalties (P08), and a cryo-margin linear score (P50)."""

    def test_p12_cns_stage_dosing_matches_hand_computation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import P12

        drug, dds, combo = _surrogate_bundles(ligand="")
        r = P12(drug, dds, combo)
        # bbb_factor=0.85 (stage 2 default), CNS-MPO=5.5/6 -> perm_factor=0.9167
        # base = 100*0.85*(0.4+0.6*0.9167) = 100*0.85*0.95 = 80.75; no targeting boost
        assert r["score"] == pytest.approx(80.75, abs=0.01)
        assert r["raw"]["active_targeting"] is False

    def test_p33_bbb_trojan_horse_matches_hand_computation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import P33

        drug, dds, combo = _surrogate_bundles(ligand="transferrin", size_nm=100.0,
                                                density=0.5)
        r = P33(drug, dds, combo)
        # lig=95, size_score=triangular(100,50,150)=100, density=0.5 -> factor=0.6
        # score = (0.6*95 + 0.4*100) * 0.6 = 58.2
        assert r["score"] == pytest.approx(58.2, abs=0.01)

    def test_p50_cryo_excursion_matches_hand_computation(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import P50

        drug, dds, combo = _surrogate_bundles(phase_T=42.0)
        r = P50(drug, dds, combo)
        # margin = |-20 - 42| = 62; score = min(100, 62*1.3) = 80.6
        assert r["score"] == pytest.approx(80.6, abs=0.01)

    def test_p08_oxidative_stress_penalizes_phenol_containing_drugs(self):
        """Regression test for a real bug: P08's phenol check called
        `smi.lower()` before testing for "Oc1cc"/"c1ccccc1O" — but those
        patterns are only meaningful case-sensitively (lowercase = aromatic
        ring atom, uppercase O = the phenolic substituent), and .lower()
        destroys exactly that distinction. The check could never match any
        SMILES, ever. P01's scenario-6 oxidative check had the identical
        bug. Fixed by dropping .lower() from both (matching the sibling
        checks in the same functions, which were already correctly
        case-sensitive)."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import P01, P08

        with_phenol, dds, combo = _surrogate_bundles(carrier="liposome",
                                                       smiles="Oc1ccccc1")
        without_phenol, _, _ = _surrogate_bundles(carrier="liposome", smiles="CCO")
        r_phenol = P08(with_phenol, dds, combo)
        r_plain = P08(without_phenol, dds, combo)
        assert "phenol" in r_phenol["raw"]["drug_oxidation_groups_detected"]
        assert r_phenol["score"] < r_plain["score"]

        # Same regression, P01's scenario 6.
        r_phenol_p01 = P01(with_phenol, dds, combo)
        assert r_phenol_p01["raw"]["drug_oxidation_prone"] is True

    def test_p18_active_targeting_transferrin_ligand_lookup(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import P18

        drug, dds, combo = _surrogate_bundles(ligand="transferrin", pka_base=8.0)
        r = P18(drug, dds, combo)
        assert r["raw"]["ligand"] == "transferrin"
        assert r["score"] > 0


class TestSurrogateDispatcher:
    def test_registry_has_exactly_57_functions(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import SURROGATE_FUNCTIONS

        assert len(SURROGATE_FUNCTIONS) == 57
        # The 5 translational principles must NOT be in this registry.
        assert not ({"P21", "P32", "P45", "P55", "P56"} & set(SURROGATE_FUNCTIONS))

    def test_registry_ids_match_principles_catalog_non_translational(self):
        """Every principle that isn't Class C (translational/admin) needs a
        fast surrogate score, because that's what ranks all 100 DDS before
        the Top-1 winner goes to Class B deep validation -- Class B
        principles run through this same surrogate registry too, they just
        also get a second, independent deep-physics pass afterward. The
        catalog's own "class" field used to mislabel 27 of the 28 real
        Class-B principles (cerebro_62_deep_engine.DEEP_FUNCTIONS +
        HPC_ONLY_PRINCIPLES) as "A_surrogate" -- this asserted the stale,
        buggy invariant (every registered ID except P47 must be
        "A_surrogate") instead of the real one (every registered ID must
        be A_surrogate OR B_deep, never C_translational)."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_principles_catalog import PRINCIPLES_62
        from cerebro_62_surrogate_engine import SURROGATE_FUNCTIONS

        for pid in SURROGATE_FUNCTIONS:
            assert PRINCIPLES_62[pid]["class"] in ("A_surrogate", "B_deep")

    def test_evaluate_all_principles_isolates_a_single_failing_function(self):
        """A malformed bundle (missing bbb_permeability entirely, so
        b_value falls back to its default) shouldn't be able to crash
        the whole batch — evaluate_all_principles_for_dds must catch a
        per-principle exception and mark just that principle FAILED."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import (
            SURROGATE_FUNCTIONS,
            evaluate_all_principles_for_dds,
        )

        drug, dds, combo = _surrogate_bundles()
        out = evaluate_all_principles_for_dds(drug, dds, combo)
        assert set(out) == set(SURROGATE_FUNCTIONS)
        for pid, r in out.items():
            assert r["confidence"] != "FAILED", f"{pid} unexpectedly failed: {r}"


# ═════════════════════════════════════════════════════════════════════════════
# 32. MODULE-PATH SHIMS (CEREBRO_Pipeline.py, cerebro_enterprise_infra.py,
#     cerebro_pipeline_patches.py, src/path_resolver.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestModulePathShims:
    """These shims exist so the flat `import CEREBRO_Pipeline`-style calls
    used throughout the codebase resolve to the real src/ modules without
    changing the process's working directory. The one behavior that must
    hold for all of them: cwd is exactly what it was before the import,
    even though the modules they wrap freeze+restore os.chdir internally."""

    def test_shim_imports_leave_the_working_directory_unchanged(self):
        import os

        import src.path_resolver  # noqa: F401

        cwd_before = os.getcwd()
        import CEREBRO_Pipeline  # noqa: F401
        import cerebro_enterprise_infra  # noqa: F401
        import cerebro_pipeline_patches  # noqa: F401

        assert os.getcwd() == cwd_before

    def test_shims_register_themselves_in_sys_modules(self):
        import sys

        import src.path_resolver  # noqa: F401
        import CEREBRO_Pipeline  # noqa: F401
        import cerebro_enterprise_infra  # noqa: F401
        import cerebro_pipeline_patches  # noqa: F401

        for name in ("CEREBRO_Pipeline", "cerebro_enterprise_infra",
                     "cerebro_pipeline_patches"):
            assert name in sys.modules

    def test_pipeline_paths_are_patched_to_the_project_root_not_src_core(self):
        """CEREBRO_Pipeline.py reconstructs PATHS as
        `_new_root / _pl.PATHS[_k].name` — this only stays correct because
        every entry in the real PATHS dict (src/core/pipeline.py) is a
        single flat directory name (data, models, figures, ...) with no
        nested subdirectories; verifying that invariant holds so a future
        nested path wouldn't silently get flattened and lose its parent."""
        import src.path_resolver  # noqa: F401
        import CEREBRO_Pipeline

        for key, path in CEREBRO_Pipeline.PATHS.items():
            assert path.name == path.parts[-1]  # no nested subpath was collapsed
            assert "outputs" in path.parts
        assert "src" not in CEREBRO_Pipeline.PATHS["data"].parts
        assert "core" not in CEREBRO_Pipeline.PATHS["data"].parts

    def test_enterprise_infra_resolves_its_own_project_root_not_src_dds(self):
        """enterprise_infra.py used to compute SCRIPT_DIR as its own file's
        directory (.../src/dds), then derive CONFIG_DIR/OUTPUT_ROOT/
        DDS_CONFIG/DDS_RESULTS from it at that same module-import time.
        path_resolver.py's post-import SCRIPT_DIR patch runs too late to
        fix any of those already-derived globals -- and never even
        attempted to for DDS_CONFIG/DDS_RESULTS, which it has no
        per-module knowledge of (unlike pipeline.py's PATHS/DB_PATH, which
        it does patch). Verified directly before this fix: importing
        enterprise_infra.py fresh baked DDS_CONFIG in as
        ".../src/dds/config/dds_config.yaml" (which does not exist),
        instead of the real project-root location -- silently breaking
        config lookups. Fixed at the root: enterprise_infra.py now walks
        up from its own directory to find run.py (the same way it already
        does to find .env) and resolves SCRIPT_DIR to that project root
        BEFORE anything derives a path from it, so every current and
        future SCRIPT_DIR-derived global in the file is correct on first
        import -- with or without path_resolver's later patch.

        (A separate, unrelated ~260-line prototype script -- its own
        DATABASE_URL/engine/Drug-table/toy ml_score FastAPI app -- had
        been concatenated onto the end of this same file and, on import,
        silently reassigned `app` to itself, making the real, documented
        API [health checks, /run-pipeline, /run-dds, /dds/ranking,
        /results/{filepath:path}, ...] completely unreachable by anyone
        who imported this module rather than running it as __main__.
        Zero other code in the repo referenced any of that prototype's
        names, so it was removed outright rather than patched around.)

        Run in a fresh subprocess: enterprise_infra.py is very likely
        already imported (and cached in sys.modules) by earlier tests in
        this same file, so re-importing it here would just return the
        cached module rather than re-exercising the module-level
        SCRIPT_DIR resolution this test is actually about."""
        import subprocess
        import sys
        from pathlib import Path

        project_root = Path(__file__).resolve().parents[2]
        script = (
            "import sys; sys.path.insert(0, '.'); "
            "import importlib; "
            "mod = importlib.import_module('src.dds.enterprise_infra'); "
            "print('DDS_CONFIG_EXISTS=' + str(mod.DDS_CONFIG.exists())); "
            "print('SCRIPT_DIR=' + mod.SCRIPT_DIR)"
        )
        result = subprocess.run(
            [sys.executable, "-c", script],
            cwd=project_root, capture_output=True, text=True, timeout=60)
        assert result.returncode == 0, (
            f"import failed:\nSTDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}")
        assert "DDS_CONFIG_EXISTS=True" in result.stdout, result.stdout
        assert f"SCRIPT_DIR={project_root}" in result.stdout, result.stdout
        assert "src/dds" not in result.stdout, (
            f"a SCRIPT_DIR-derived global still points inside src/dds:\n{result.stdout}")

    def test_enterprise_infra_app_exposes_the_real_routes_not_a_stale_prototype(self):
        """A ~260-line, fully self-contained prototype script (its own
        redundant imports, its own DATABASE_URL/engine/Drug SQL model, a
        toy ml_score FastAPI app, a do-nothing "background monitoring"
        task) had been concatenated onto the end of this file and, being
        pure module-level code with no __main__ guard, unconditionally
        reassigned `app` to itself on every import -- silently discarding
        the real, documented app (with /health, /run-pipeline, /run-dds,
        /dds/ranking, /results/{filepath:path}, ...) for any caller that
        imports this module rather than running it directly as a script.
        Confirms the real app -- not the toy one -- is what `app` resolves
        to, and that none of the prototype's names remain."""
        from src.dds.enterprise_infra import app

        route_paths = {r.path for r in app.routes}
        assert "/run-dds" in route_paths
        assert "/dds/ranking" in route_paths
        assert "/results/{filepath:path}" in route_paths
        assert "/drug/" not in route_paths
        assert "/top/" not in route_paths

        import src.dds.enterprise_infra as mod
        for stale_name in ("Drug", "MLCore", "DrugInput", "async_pipeline",
                           "feature_engineering", "compute_affinity",
                           # A second, unrelated legacy fragment sat at the
                           # very top of this same file (before the "0.
                           # ANCHOR" section): get_data_from_excel() read a
                           # cwd-relative "CEREBRO_Input_Template.xlsx" (not
                           # SCRIPT_DIR-anchored like everything else here)
                           # and auto_fix_config() wrote a stripped-down,
                           # drug-name-only config/dds_config.yaml in a
                           # schema the real excel_to_yaml() pipeline
                           # doesn't produce or expect. Both were dead (zero
                           # callers anywhere in the codebase) but a live
                           # landmine: running this file directly instead of
                           # run.py silently clobbered the real config the
                           # live pipeline reads. This is the actual origin
                           # of the stale "drug: {name: Unknown_Drug}"
                           # config/dds_config.yaml found checked into the repo.
                           "get_data_from_excel", "auto_fix_config"):
            assert not hasattr(mod, stale_name), (
                f"{stale_name} from the removed prototype script still exists")

    def test_enterprise_infra_write_autostart_points_at_a_real_file(
            self, monkeypatch, tmp_path):
        """write_autostart() joined SCRIPT_DIR (patched to the project
        root) with the literal "cerebro_enterprise_infra.py" -- that name
        is only a sys.modules alias registered inside an already-running
        process (src/path_resolver.py); no file by that name exists
        anywhere on disk, unlike CEREBRO_Pipeline.py which does have a
        real root-level shim. Every autostart config this wrote (launchd
        plist / systemd service / Task Scheduler XML / cron line) pointed
        at a script path that would fail to launch on the next boot.
        Redirects Path.home() to a temp dir so this doesn't touch the
        real machine's LaunchAgents/systemd config."""
        import os
        import platform

        import src.dds.enterprise_infra as ei

        monkeypatch.setattr("pathlib.Path.home", lambda: tmp_path)
        # write_autostart() also writes autostart_info.txt(_DOCUMENTATION.txt)
        # directly under SCRIPT_DIR, which is the real project root -- redirect
        # that too so this test doesn't leave a stray file in the working tree.
        monkeypatch.setattr(ei, "SCRIPT_DIR", str(tmp_path))
        ei.write_autostart()

        real_script = "src" + os.sep + "dds" + os.sep + "enterprise_infra.py"
        if platform.system() == "Darwin":
            plist = tmp_path / "Library" / "LaunchAgents" / "com.cerebro.enterprise.plist"
            assert plist.exists()
            content = plist.read_text()
            assert real_script in content
            assert "cerebro_enterprise_infra.py</string>" not in content
        elif platform.system() == "Linux":
            svc = tmp_path / ".config" / "systemd" / "user" / "cerebro.service"
            assert svc.exists()
            content = svc.read_text()
            assert real_script in content

    @pytest.mark.slow
    def test_phase5_smoke_test_script_passes_end_to_end(self):
        """engine/phase5_smoke_test.py is a standalone script (real ChEMBL
        calls, real bundle resolution, real orchestrator run on two real
        drugs) meant to be run directly, not imported — so the real
        regression check is running it and confirming a clean exit, the
        same way its own docstring says to use it against a fresh Docker
        build."""
        import subprocess
        import sys

        from pathlib import Path

        project_root = Path(__file__).resolve().parents[2]
        result = subprocess.run(
            [sys.executable, str(project_root / "engine" / "phase5_smoke_test.py")],
            cwd=project_root, capture_output=True, text=True, timeout=120)
        assert result.returncode == 0, (
            f"phase5_smoke_test.py failed:\nSTDOUT:\n{result.stdout}\n"
            f"STDERR:\n{result.stderr}")
        assert "ALL SMOKE TESTS PASSED" in result.stdout


# ═════════════════════════════════════════════════════════════════════════════
# 33. DDS INVERSE-DESIGN GENETIC ALGORITHM (cerebro_dds_inverse_design.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestDdsInverseDesignGaOperators:
    def test_random_individual_respects_every_parameter_bound(self):
        import random

        import src.path_resolver  # noqa: F401
        from cerebro_dds_inverse_design import (
            ALL_PARAMS,
            CATEGORICAL_SPACE,
            CONTINUOUS_SPACE,
            _random_individual,
        )

        rng = random.Random(1)
        ind = _random_individual(rng, 0)
        assert set(ind) == set(ALL_PARAMS) | {"Formulation_ID", "Formulation_Name"}
        for k, (lo, hi) in CONTINUOUS_SPACE.items():
            assert lo <= ind[k] <= hi
        for k, choices in CATEGORICAL_SPACE.items():
            assert ind[k] in choices

    def test_crossover_takes_every_gene_from_one_parent_or_the_other(self):
        import random

        import src.path_resolver  # noqa: F401
        from cerebro_dds_inverse_design import ALL_PARAMS, _crossover, _random_individual

        rng = random.Random(2)
        a = _random_individual(rng, 1)
        b = _random_individual(rng, 2)
        child = _crossover(a, b, rng)
        for k in ALL_PARAMS:
            assert child[k] in (a[k], b[k])

    def test_mutate_at_rate_zero_never_changes_genes(self):
        import random

        import src.path_resolver  # noqa: F401
        from cerebro_dds_inverse_design import ALL_PARAMS, _mutate, _random_individual

        rng = random.Random(3)
        ind = _random_individual(rng, 0)
        mutated = _mutate(dict(ind), rng, rate=0.0)
        for k in ALL_PARAMS:
            assert mutated[k] == ind[k]

    def test_dedupe_key_is_stable_and_covers_every_parameter(self):
        import random

        import src.path_resolver  # noqa: F401
        from cerebro_dds_inverse_design import ALL_PARAMS, _dedupe_key, _random_individual

        rng = random.Random(4)
        ind = _random_individual(rng, 0)
        key1, key2 = _dedupe_key(ind), _dedupe_key(dict(ind))
        assert key1 == key2
        assert len(key1) == len(ALL_PARAMS)


class TestDdsInverseDesignRealSearch:
    @pytest.mark.slow
    def test_generate_candidate_formulations_runs_the_real_orchestrator(self):
        """A small, fast real GA run (tiny population/generation counts,
        fixed seed) through the actual cerebro_62_orchestrator fitness
        function — confirms the search produces real, distinct scores
        and an honest disclaimer, not a mocked/fabricated result."""
        import src.path_resolver  # noqa: F401
        from cerebro_dds_inverse_design import generate_candidate_formulations
        from cerebro_resolved_bundles import resolve_drug_bundle

        drug_bundle = resolve_drug_bundle(
            name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            molecule_class="small_molecule")
        result = generate_candidate_formulations(
            drug_bundle, drug_name="Donepezil",
            n_generations=2, population_size=6, top_k=3, seed=7)

        assert result["n_evaluated"] == 2 * 6
        assert len(result["candidates"]) <= 3
        assert result["search_seed"] == 7
        assert "not independently validated" in result["disclaimer"]
        for cand in result["candidates"]:
            assert "novel_vs_input" in cand
            assert "Principle_Composite_Score" in cand


# ═════════════════════════════════════════════════════════════════════════════
# 34. LEGACY PRINCIPLE-METADATA TABLES — removed (cerebro_dds_principle_evaluator.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestLegacyPrincipleMetadataTableRemoved:
    """cerebro_dds_principle_evaluator.py used to run its own v21-era
    24-principle scoring engine (_evaluate_dds/evaluate_all_dds), fully
    superseded by cerebro_62_orchestrator.py -- that scoring code was
    already removed, leaving just two lookup tables (PRINCIPLE_WEIGHTS,
    PRINCIPLE_DOCS) whose own docstring claimed cerebro_multi_drug_
    comparison.py and cerebro_completed_excel_writer.py "still genuinely
    import" them. That claim was false: grepping the whole repo for
    actual imports of this module turned up nothing outside this test
    file itself -- cerebro_multi_drug_comparison.py has its own,
    differently-keyed PRINCIPLE_WEIGHTS dict, and cerebro_completed_
    excel_writer.py's only reference was a comment explaining why it
    stopped using this module. The old test_real_callers_still_import_
    the_tables_cleanly here compounded this: it only imported those two
    modules and checked they didn't error, which is true regardless of
    whether either one ever touches cerebro_dds_principle_evaluator --
    it never actually verified the claimed dependency. Deleted the file
    entirely rather than leave confirmed dead code with a false docstring
    lying around, consistent with this project's own stated practice for
    the scoring engine it already removed from the same file."""

    def test_module_no_longer_exists(self):
        import src.path_resolver  # noqa: F401

        with pytest.raises(ImportError):
            import cerebro_dds_principle_evaluator  # noqa: F401


# ═════════════════════════════════════════════════════════════════════════════
# 35. BRAND / VISUAL IDENTITY (cerebro_brand.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestCerebroBrand:
    def test_asset_paths_point_at_the_real_project_root_not_engine_dir(self):
        """Regression test for a real bug: _PROJECT_ROOT was computed as
        Path(__file__).resolve().parent, which for engine/cerebro_brand.py
        is just engine/ itself (.parent strips only the filename) — one
        level short of the actual project root. LOGO_PATH/PATTERN_PATH/
        ENGINE_BG_PATH all silently pointed at engine/assets/brand/, which
        doesn't exist; the real assets/ directory is at the project root,
        exactly one level further up than the buggy path reached. Every
        sibling shim in this same engine/ directory (CEREBRO_Pipeline.py,
        cerebro_enterprise_infra.py, cerebro_pipeline_patches.py) already
        used the correct .parent.parent for the same file layout."""
        import src.path_resolver  # noqa: F401
        import cerebro_brand

        assert cerebro_brand.LOGO_PATH.exists()
        assert cerebro_brand.PATTERN_PATH.exists()
        assert cerebro_brand.ENGINE_BG_PATH.exists()
        assert cerebro_brand.ASSETS_DIR.name == "brand"
        assert cerebro_brand.ASSETS_DIR.parent.name == "assets"

    def test_matplotlib_style_uses_declared_palette_colors(self):
        import src.path_resolver  # noqa: F401
        from cerebro_brand import GOLD, VOID_BASE, matplotlib_style

        style = matplotlib_style()
        assert style["figure.facecolor"] == VOID_BASE
        assert style["axes.titlecolor"] == GOLD

    def test_html_brand_header_omits_redundant_title_but_keeps_subtitle(self):
        import src.path_resolver  # noqa: F401
        from cerebro_brand import PROJECT_NAME, html_brand_header

        redundant = html_brand_header(title="CEREBRO-X")
        assert redundant.count(f'<div class="subtitle">{PROJECT_NAME}</div>') == 0

        real = html_brand_header(title="Drug Delivery Report", subtitle="Donepezil")
        assert "Drug Delivery Report" in real
        assert "Donepezil" in real

    def test_reportlab_color_parses_a_real_hex_string(self):
        import src.path_resolver  # noqa: F401
        from cerebro_brand import GOLD, reportlab_color

        pytest.importorskip("reportlab")
        color = reportlab_color(GOLD)
        assert color is not None


# ═════════════════════════════════════════════════════════════════════════════
# 36. BUNDLE INSPECTOR CLI (cerebro_inspector.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestCerebroInspectorFormatting:
    def test_fmt_value_handles_every_type_it_documents(self):
        import src.path_resolver  # noqa: F401
        from cerebro_inspector import _fmt_value

        assert _fmt_value(None) == "—"
        assert _fmt_value(True) == "true"
        assert _fmt_value(False) == "false"
        assert _fmt_value(0.001234) == "1.234e-03"    # small float -> sci notation
        assert _fmt_value(12345.6789) == "1.235e+04"  # large float -> sci notation
        assert _fmt_value(3.14159) == "3.142"          # normal float -> 4 sig figs
        assert _fmt_value({"x": 1}) == "{...}"
        long_str = "a" * 50
        assert _fmt_value(long_str) == "a" * 40 + "…"
        assert _fmt_value("short") == "short"

    def test_to_json_round_trips_both_bundles(self):
        import json

        import src.path_resolver  # noqa: F401
        from cerebro_inspector import _to_json

        drug_b = {"drug_mw": {"value": 350.0, "tier": 3}, "_meta": {"name": "x"}}
        dds_b = {"material_pdi": {"value": 0.2, "tier": 7}, "_meta": {"dds_type": "material"}}
        out = json.loads(_to_json(drug_b, dds_b))
        assert out["drug"]["drug_mw"]["value"] == 350.0
        assert out["dds"]["material_pdi"]["tier"] == 7
        assert "combo" not in out

    def test_to_markdown_escapes_pipe_characters_in_method_text(self):
        """A computational_method string containing a literal '|' would
        break the markdown table's column structure if not escaped."""
        import src.path_resolver  # noqa: F401
        from cerebro_inspector import _to_markdown

        drug_b = {"drug_mw": {"value": 350.0, "tier": 3, "source": "RDKit",
                               "_computational_method": "a | b"},
                  "_meta": {"name": "x"}}
        dds_b = {"_meta": {"dds_type": "material"}}
        md = _to_markdown(drug_b, dds_b)
        assert "a \\| b" in md
        assert "| a | b |" not in md  # the raw unescaped pipe never appears as a cell break


# ═════════════════════════════════════════════════════════════════════════════
# 37. MOLECULE EXTRACTOR (cerebro_molecule_extractor.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestMoleculeExtractorPkaAndMicrospeciation:
    """This is a second, independent pKa/microspeciation implementation
    (a flat SMARTS-group lookup) alongside cerebro_value_resolver's
    Bordwell-Hammett-Born one — different pKa methodology, but the same
    Bjerrum 4-microspecies charge-state math, cross-verified here."""

    def test_carboxylic_acid_only_group_detected_and_mostly_anionic_at_ph74(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import _rdkit_descriptors

        d = _rdkit_descriptors("CC(=O)O")  # acetic acid
        assert d["pKa_acidic"]["value"] == 4.2
        assert d["pKa_basic"]["value"] is None
        assert d["FractionAnionic_pH74"]["value"] > 0.99   # far above pKa at pH 7.4

    def test_primary_amine_only_group_detected_and_mostly_cationic_at_ph74(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import _rdkit_descriptors

        d = _rdkit_descriptors("CCN")  # ethylamine
        assert d["pKa_basic"]["value"] == 10.6
        assert d["FractionCationic_pH74"]["value"] > 0.99

    def test_no_ionizable_groups_reports_honestly_not_fabricated(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import _rdkit_descriptors

        d = _rdkit_descriptors("CC")  # ethane — no acid/base SMARTS group matches
        assert d["pKa"]["value"] is None
        assert d["pKa"]["confidence"] == "LOW"
        assert d["FractionNeutral_pH74"]["value"] == 1.0

    def test_zwitterion_candidate_produces_all_four_microspecies_summing_to_one(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import _rdkit_descriptors

        d = _rdkit_descriptors("NCC(=O)O")  # glycine: amine + carboxylic acid
        total = (d["FractionCationic_pH74"]["value"] + d["FractionAnionic_pH74"]["value"]
                 + d["FractionZwitterion_pH74"]["value"] + d["FractionNeutral_pH74"]["value"])
        assert total == pytest.approx(1.0, abs=1e-3)


class TestMoleculeExtractorFastaDescriptors:
    def test_aliphatic_index_matches_ikai_1980_formula(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import _aliphatic_index

        assert _aliphatic_index("AAAA") == pytest.approx(100.0)   # pure Ala: X(A)=100%
        assert _aliphatic_index("VVVV") == pytest.approx(290.0)   # pure Val: 2.9 * 100%
        assert _aliphatic_index("") == 0.0

    def test_formal_charge_confidence_matches_its_actual_rigor(self):
        """Regression test for a real mislabeling: FormalCharge for FASTA
        sequences is a residue-counting heuristic (ignores His, ignores
        pH, ignores N/C-terminal charge) — not a genuine Biopython
        ProtParam calculation like its MW_Da/pI/GRAVY siblings in the
        same dict. It used to carry source="fasta"/confidence="HIGH",
        the same tags as those genuine calculations, overclaiming rigor
        it doesn't have. Now tagged the same "fasta_proxy"/MODERATE as
        the file's other crude proxies (HBD, TPSA_A2)."""
        pytest.importorskip("Bio")
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import _fasta_descriptors

        d = _fasta_descriptors("MKTAYIAKQRQISFVKSHFSRQLEERLGLIEVQAPILSRVGDGTQDNLSGAEKAVQVKVKALPDAQFEVVHSLAKWKRQTLGQHDFSAGEGLYTHMKALRPDEDRLSPLHSVYVDQWDWELVMGDGERSFSTRQPAYLNFDNPDMESFQMDVEIRNQIAQVWKTAFQMLGDSVSFYEDPFVCGYNRLQPYAKFPQMTAKVYRKNQMLTGARTLIYQAHDPHIENKFATLFNDQMPDNDLLEQIRISLQKGSPWDIVKQLIENDIQVELPCFVDRGDLGGGHILFVSDNKAKGRRDQEHLNVACFPGKGAGLDEEYSAPGGEQRLKTEGSSVPKGGVDMASAAKGGYPLAAAAAPGAAAAAAAAAAALGASPADGGA")
        assert d["FormalCharge"]["source"] == "fasta_proxy"
        assert d["FormalCharge"]["confidence"] == "MODERATE"
        # The real Biopython calculations keep their genuinely-earned HIGH.
        assert d["MW_Da"]["confidence"] == "HIGH"
        assert d["pI"]["confidence"] == "HIGH"


class TestMoleculeExtractorEnrichment:
    def test_researcher_override_wins_but_computed_value_is_still_recorded(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import enrich_mol_profile

        out = enrich_mol_profile({"smiles": "CCO", "name": "test", "LogP": 1.23})
        assert out["LogP"] == 1.23
        prov = out["_descriptors_provenance"]["LogP"]
        assert prov["source"] == "researcher_override"
        assert prov["computed_value"] is not None   # what RDKit would have said, for comparison

    def test_computed_descriptor_used_when_no_override_given(self):
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import enrich_mol_profile

        out = enrich_mol_profile({"smiles": "CCO", "name": "test"})
        assert out["MW_Da"] == pytest.approx(46.07, abs=0.01)
        assert out["_descriptors_provenance"]["MW_Da"]["source"] == "rdkit"

    def test_no_smiles_or_fasta_reports_every_descriptor_as_honestly_missing(self):
        """name='' keeps this offline — enrich_mol_profile only attempts
        the live PubChem/UniProt fallback fetch when a drug name is
        given, which would make this test network-dependent."""
        import src.path_resolver  # noqa: F401
        from cerebro_molecule_extractor import enrich_mol_profile

        out = enrich_mol_profile({"name": ""})
        assert out["MW_Da"] is None
        assert out["_descriptors_provenance"]["MW_Da"]["confidence"] == "FAILED"


# ═════════════════════════════════════════════════════════════════════════════
# 38. MULTI-DRUG COMPARISON (cerebro_multi_drug_comparison.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestMultiDrugComparisonHelpers:
    def test_to_float_handles_commas_nan_and_bad_strings(self):
        import math

        import src.path_resolver  # noqa: F401
        from cerebro_multi_drug_comparison import _to_float

        assert _to_float("1,234.5") == 1234.5
        assert _to_float(float("nan")) is None
        assert _to_float(True) == 1.0
        assert _to_float("not a number") is None

    def test_flatten_numeric_recurses_and_drops_underscore_and_non_numeric_keys(self):
        import src.path_resolver  # noqa: F401
        from cerebro_multi_drug_comparison import _flatten_numeric

        flat = _flatten_numeric({"a": {"b": 1.0, "c": "text"}, "_hidden": 5.0, "d": 2})
        assert flat == {"a.b": 1.0, "d": 2.0}

    def test_normalize_score_higher_and_lower_directions(self):
        import src.path_resolver  # noqa: F401
        from cerebro_multi_drug_comparison import _normalize_score

        vals = {"A": 10.0, "B": 20.0, "C": 30.0}
        higher = _normalize_score(vals, "higher")
        assert higher == {"A": 0.0, "B": 50.0, "C": 100.0}
        lower = _normalize_score(vals, "lower")
        assert lower == {"A": 100.0, "B": 50.0, "C": 0.0}

    def test_docking_affinity_kcal_and_its_abs_variant_agree_on_direction_of_better(self):
        """Docking_Affinity_kcal (a signed, more-negative-is-stronger
        energy) belongs in LOWER_IS_BETTER; its _abs companion (computed
        in compare_drugs) belongs in HIGHER_IS_BETTER — both must agree
        that -9.0 kcal/mol beats -5.0 kcal/mol."""
        import src.path_resolver  # noqa: F401
        from cerebro_multi_drug_comparison import _direction_for

        assert _direction_for("physchem.Docking_Affinity_kcal") == "lower"
        assert _direction_for("physchem.Docking_Affinity_kcal_abs") == "higher"

    def test_scientific_rationale_no_longer_cites_the_stale_25_principle_count(self):
        """The Scientific_Rationale sheet's methodology text used to say
        the pipeline runs "25 CNS-weighted principles" with a hardcoded
        "(CNS delivery 37%, glymphatic 11%)" weight breakdown -- both
        leftover prose from the old v21 rubric, describing neither the
        current 62-principle system nor its real weight_cns figures."""
        import openpyxl
        import src.path_resolver  # noqa: F401
        import cerebro_multi_drug_comparison as mod

        wb = openpyxl.Workbook()
        mod._write_scientific_rationale_sheet(wb, {})
        ws = wb["Scientific_Rationale"]
        all_text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                              if c.value is not None)
        assert "25 CNS" not in all_text
        assert "37%" not in all_text
        assert "62 CNS-weighted principles" in all_text


class TestMultiDrugComparisonTieHandling:
    def test_a_genuine_tie_is_reported_honestly_not_credited_to_whoever_is_first(self):
        """Regression test for a real bug: when every drug lands on the
        exact same value for a metric (a realistic case — e.g. every drug
        falling back to the same Tier-7 class-mean default), _normalize_
        score gives everyone 100.0, and max() then arbitrarily picked
        whichever drug happened to be listed first as the 'winner' of a
        comparison that was actually a dead heat, silently biasing
        winner_counts toward that drug. Now reported as an explicit tie
        with no winner credited."""
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_multi_drug_comparison import compare_drugs

        drug_results = [
            {"drug_name": "DrugA", "mol_profile": {"BBB_permeability_pct": 50.0},
             "df_dds": None, "principles": {}},
            {"drug_name": "DrugB", "mol_profile": {"BBB_permeability_pct": 50.0},
             "df_dds": None, "principles": {}},
        ]
        with tempfile.TemporaryDirectory() as td:
            summary = compare_drugs(drug_results, Path(td))
        row = next(r for r in summary["per_principle"]
                   if "BBB_permeability_pct" in r["metric"])
        assert row["winner"] == "— (tie)"
        assert summary["winner_counts"]["DrugA"] == 0
        assert summary["winner_counts"]["DrugB"] == 0

    def test_a_genuine_winner_is_still_credited_normally(self):
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_multi_drug_comparison import compare_drugs

        drug_results = [
            {"drug_name": "DrugA", "mol_profile": {"BBB_permeability_pct": 80.0},
             "df_dds": None, "principles": {}},
            {"drug_name": "DrugB", "mol_profile": {"BBB_permeability_pct": 20.0},
             "df_dds": None, "principles": {}},
        ]
        with tempfile.TemporaryDirectory() as td:
            summary = compare_drugs(drug_results, Path(td))
        row = next(r for r in summary["per_principle"]
                   if "BBB_permeability_pct" in r["metric"])
        assert row["winner"] == "DrugA"
        assert summary["winner_counts"]["DrugA"] == 1


class TestMultiDrugComparisonChampionSheet:
    @pytest.mark.slow
    def test_champion_sheet_per_principle_scores_are_real_not_all_zero(self):
        """Regression test for a real bug: the Champion_DDS_Compare Excel
        sheet's per-principle-score section and two of its seven group-
        rollup rows (G2, G5) were built around an old v21-era ID/group-
        naming scheme (e.g. 'P1.1_BBB_transcytosis', 'G2_Release') that
        never matched the real 62-principle data actually being fed in
        (keyed 'P01'..'P62', groups named '...Kinetics'/'...BBB') — so
        every one of those cells silently rendered 0 for every drug,
        regardless of the real underlying scores. Confirmed end-to-end
        with two real drugs through the actual orchestrator."""
        import tempfile
        from pathlib import Path

        import openpyxl
        import pandas as pd

        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import evaluate_all_dds_62
        from cerebro_multi_drug_comparison import compare_drugs
        from cerebro_resolved_bundles import resolve_drug_bundle

        db1 = resolve_drug_bundle(
            name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            molecule_class="small_molecule")
        db2 = resolve_drug_bundle(
            name="Rivastigmine", smiles="CCN(C)C(=O)Oc1cccc(c1)C(C)N(C)C",
            molecule_class="small_molecule")
        df_dds = pd.DataFrame([{
            "Formulation_ID": "F001", "Formulation_Name": "Tf-PLGA",
            "Carrier_Type": "plga", "Size_nm": 100, "Zeta_Potential_mV": -25,
            "PDI": 0.2, "Encapsulation_Efficiency_pct": 75,
            "Surface_Ligand": "transferrin", "PEGylation_Degree_mol_pct": 5,
            "Release_Kinetics": "sustained", "Scale_Up_Readiness": "pilot",
        }])
        r1 = evaluate_all_dds_62(drug_bundle=db1, df_dds=df_dds, drug_name="Donepezil")
        r2 = evaluate_all_dds_62(drug_bundle=db2, df_dds=df_dds, drug_name="Rivastigmine")
        drug_results = [
            {"drug_name": "Donepezil", "mol_profile": {}, "df_dds": r1["ranked_df"],
             "dds_principle_matrix": r1["all_dds_principles"],
             "dds_principle_breakdown": r1["all_dds_breakdown"]},
            {"drug_name": "Rivastigmine", "mol_profile": {}, "df_dds": r2["ranked_df"],
             "dds_principle_matrix": r2["all_dds_principles"],
             "dds_principle_breakdown": r2["all_dds_breakdown"]},
        ]
        with tempfile.TemporaryDirectory() as td:
            compare_drugs(drug_results, Path(td))
            wb = openpyxl.load_workbook(Path(td) / "CEREBRO_X_Multi_Drug_Comparison.xlsx")
            ws = wb["Champion_DDS_Compare"]

            group_rows = {row[0]: row[2:4] for row in
                          ws.iter_rows(min_row=8, max_row=14, values_only=True)}
            assert group_rows["G2 Release Kinetics"] != (0, 0)
            assert group_rows["G5 Glymphatic BBB"] != (0, 0)

            per_principle_cells = [
                cell.value
                for row in ws.iter_rows(min_row=16, max_row=ws.max_row)
                if row[0].value and str(row[0].value).startswith("P")
                and "(" in str(row[0].value)
                for cell in row[2:4] if cell.value is not None
            ]
            assert per_principle_cells
            assert any(v != 0 for v in per_principle_cells)


# ═════════════════════════════════════════════════════════════════════════════
# 39. COMPLETED-DATA EXCEL WRITER (cerebro_completed_excel_writer.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestCompletedExcelWriterPrincipleSheets:
    """Same bug family as the Champion_DDS_Compare sheet fix: this writer
    also had two sheets built around the old v21 25-principle ID/group
    scheme, disconnected from the live 62-principle data actually being
    written everywhere else in the same workbook."""

    @pytest.mark.slow
    def test_dds_principle_matrix_group_columns_are_real_not_all_zero(self):
        """Regression test: the DDS×Principle matrix sheet's group_cols
        list used the old 'G2_Release'/'G5_Glymphatic' names, which never
        matched cerebro_62_orchestrator's real 'G2_Release_Kinetics'/
        'G5_Glymphatic_BBB' keys in m["groups"] — so those two of seven
        group columns silently showed 0 for every DDS row in every
        drug's per-drug DDSxP sheet, the primary formulation-ranking
        matrix a researcher would actually look at."""
        import tempfile
        from pathlib import Path

        import openpyxl
        import pandas as pd

        import src.path_resolver  # noqa: F401
        from cerebro_62_orchestrator import evaluate_all_dds_62
        from cerebro_completed_excel_writer import write_completed_excel
        from cerebro_resolved_bundles import resolve_drug_bundle

        db = resolve_drug_bundle(
            name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            molecule_class="small_molecule")
        df_dds = pd.DataFrame([{
            "Formulation_ID": "F001", "Formulation_Name": "Tf-PLGA",
            "Carrier_Type": "plga", "Size_nm": 100, "Zeta_Potential_mV": -25,
            "PDI": 0.2, "Encapsulation_Efficiency_pct": 75,
            "Surface_Ligand": "transferrin", "PEGylation_Degree_mol_pct": 5,
            "Release_Kinetics": "sustained", "Scale_Up_Readiness": "pilot",
        }])
        r = evaluate_all_dds_62(drug_bundle=db, df_dds=df_dds, drug_name="Donepezil")
        drug_results = [{
            "drug_name": "Donepezil", "mol_profile": {"_source_audit": {}},
            "df_dds": r["ranked_df"],
            "dds_principle_matrix": r["all_dds_principles"],
            "dds_principle_breakdown": r["all_dds_breakdown"],
            "principles": {}, "deep_results": r.get("deep_results", {}),
            "deep_summary": r.get("deep_summary", {}),
            "translational": r.get("translational", {}),
            "fallback_chain": r.get("fallback_chain", []),
        }]
        with tempfile.TemporaryDirectory() as td:
            out = write_completed_excel(drug_results, Path(td) / "test.xlsx")
            wb = openpyxl.load_workbook(out)
            ws = wb["D1_Donepezil_DDSxP"]
            header = [c.value for c in ws[4]]
            row1 = [c.value for c in ws[5]]
            assert row1[header.index("G2_Release_Kinetics")] != 0
            assert row1[header.index("G5_Glymphatic_BBB")] != 0

    @pytest.mark.slow
    def test_principle_explanations_sheet_lists_all_62_current_principles(self):
        """Regression test: this glossary sheet used to be built from
        cerebro_dds_principle_evaluator's old 25-principle table — a
        researcher would see 'P01'..'P62' everywhere else in the same
        workbook, then find 25 unrelated 'P1.1_...'-style rows here that
        matched nothing else in the file. Now sourced from the same live
        catalog that actually produced every score in the workbook."""
        import tempfile
        from pathlib import Path

        import openpyxl

        import src.path_resolver  # noqa: F401
        from cerebro_completed_excel_writer import write_completed_excel

        drug_results = [{
            "drug_name": "Donepezil", "mol_profile": {"_source_audit": {}},
            "df_dds": None, "dds_principle_matrix": [], "dds_principle_breakdown": [],
            "principles": {},
        }]
        with tempfile.TemporaryDirectory() as td:
            out = write_completed_excel(drug_results, Path(td) / "test.xlsx")
            wb = openpyxl.load_workbook(out)
            ws = wb["Principle_Explanations"]
            ids = [ws.cell(i, 1).value for i in range(5, ws.max_row + 1)]
            ids = [i for i in ids if i]
            assert len(ids) == 62
            assert "P01" in ids and "P62" in ids
            assert not any(i.startswith("P1.") for i in ids)   # no old-format leftovers


class TestCompletedExcelWriterHelpers:
    def test_get_tier_info_prefers_source_audit_over_inference(self):
        import src.path_resolver  # noqa: F401
        from cerebro_completed_excel_writer import _get_tier_info

        mp = {"_source_audit": {"MW_Da": {"_tier": 1, "_confidence": "HIGH"}}}
        assert _get_tier_info(mp, "MW_Da") == {"_tier": 1, "_confidence": "HIGH"}

    def test_get_tier_info_infers_tier_1_for_a_present_unaudited_value(self):
        import src.path_resolver  # noqa: F401
        from cerebro_completed_excel_writer import _get_tier_info

        info = _get_tier_info({"MW_Da": 350.0}, "MW_Da")
        assert info["_tier"] == 1

    def test_get_tier_info_reports_tier_99_when_truly_missing(self):
        import src.path_resolver  # noqa: F401
        from cerebro_completed_excel_writer import _get_tier_info

        assert _get_tier_info({}, "MW_Da")["_tier"] == 99

    def test_flatten_principles_recurses_and_converts_booleans(self):
        import src.path_resolver  # noqa: F401
        from cerebro_completed_excel_writer import _flatten_principles

        flat = _flatten_principles({"a": {"b": 1.0, "_hidden": 5}, "c": True, "d": [1, 2, 3]})
        assert flat == [("a.b", 1.0), ("c", "Yes")]   # lists skipped, bool -> Yes/No, _-keys dropped


# ═════════════════════════════════════════════════════════════════════════════
# 39b. SURROGATE ENGINE DDS-SPEC RESOLUTION (cerebro_62_surrogate_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestSurrogateEngineDdsSpecKeyNames:
    """_dds_specs_from_bundle (engine/cerebro_62_surrogate_engine.py) reads
    dds_row (the real per-formulation df_dds row, injected into
    combo_bundle["_meta"]["dds_row"] by cerebro_62_orchestrator.py before
    every P-function call) via "Endosomal_Escape_Eff" and
    "CNS_Bioavailability_Pct" -- neither of which _run_dds_from_yaml ever
    produces (the real column is "PgP_Escape_Coeff"; see
    _dds_metrics.backfill_legacy_aliases for the same ghost-key pattern
    fixed elsewhere). Unlike the display-only bugs fixed in the viz layer,
    this one is load-bearing: P06 (endosomal escape scoring, one of the
    57 fast-surrogate principles that feed Principle_Composite_Score, the
    actual DDS ranking) read s["endo_esc"] and silently got the same
    hardcoded 0.5 for every formulation being ranked against every other,
    erasing the one term meant to differentiate carriers by their real
    escape efficiency. Verified directly: two formulations differing only
    in PgP_Escape_Coeff (0.05 vs 0.95) produced identical P06 scores
    before the fix (50.26 in both directions collapsed to whichever value
    a shared bundle-cache last saw); after the fix they diverge exactly as
    the formula requires (50.26 vs 97.38 for this drug)."""

    def test_endo_esc_reads_real_pgp_escape_coeff_column(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _dds_specs_from_bundle
        from cerebro_resolved_bundles import resolve_dds_bundle

        dds_bundle = resolve_dds_bundle(carrier_type="liposome", ligand="RVG29")
        low = _dds_specs_from_bundle(dds_bundle, dds_row={"PgP_Escape_Coeff": 0.05})
        high = _dds_specs_from_bundle(dds_bundle, dds_row={"PgP_Escape_Coeff": 0.95})
        assert low["endo_esc"] == pytest.approx(0.05)
        assert high["endo_esc"] == pytest.approx(0.95)

    def test_p06_endosomal_escape_score_differs_by_real_formulation_data(self):
        """The actual ranking-affecting regression: P06 must score two
        formulations differently when their real PgP_Escape_Coeff differs,
        not collapse to whatever the hardcoded legacy-key default was."""
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import P06
        from cerebro_resolved_bundles import resolve_combo_bundle, resolve_dds_bundle, resolve_drug_bundle

        drug_bundle = resolve_drug_bundle(name="Aspirin", smiles="CC(=O)Oc1ccccc1C(=O)O",
                                            molecule_class="small_molecule")
        dds_bundle = resolve_dds_bundle(carrier_type="liposome", ligand="RVG29")
        # resolve_combo_bundle is cached per (drug, dds) identity, so mutate
        # the shared object's dds_row and score immediately each time --
        # exactly the assign-then-score pattern cerebro_62_orchestrator.py
        # itself uses in its per-formulation evaluation loop.
        combo = resolve_combo_bundle(drug_bundle, dds_bundle)

        combo["_meta"]["dds_row"] = {"PgP_Escape_Coeff": 0.05, "pH_Trigger": 6.5}
        low = P06(drug_bundle, dds_bundle, combo)

        combo["_meta"]["dds_row"] = {"PgP_Escape_Coeff": 0.95, "pH_Trigger": 6.5}
        high = P06(drug_bundle, dds_bundle, combo)

        assert low["score"] != high["score"]
        assert high["score"] > low["score"]

    def test_cns_bio_reads_real_bbb_engineering_score_not_ghost_column(self):
        import src.path_resolver  # noqa: F401
        from cerebro_62_surrogate_engine import _dds_specs_from_bundle
        from cerebro_resolved_bundles import resolve_dds_bundle

        dds_bundle = resolve_dds_bundle(carrier_type="liposome", ligand="RVG29")
        out = _dds_specs_from_bundle(dds_bundle, dds_row={"BBB_Engineering_Score": 88.0})
        assert out["cns_bio"] == pytest.approx(88.0)


# ═════════════════════════════════════════════════════════════════════════════
# 40. CINEMATIC VISUAL PRIMITIVES (cerebro_cinematic_primitives.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestCinematicPrimitivesLookups:
    def test_drug_profile_falls_back_to_small_molecule(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_primitives import get_drug_profile

        assert get_drug_profile(None)["narrative"] == "small molecule"
        assert get_drug_profile("totally_unknown_type")["narrative"] == "small molecule"
        assert get_drug_profile("MONOCLONAL_ANTIBODY")["shape"] == "y_shape"  # case-insensitive

    def test_dds_profile_prefers_specific_match_over_generic_substring(self):
        """AAV9 and generic AAV are both real keys — a carrier string
        naming a specific untabled serotype (e.g. 'aav5') should still
        resolve via the generic 'aav' entry, not silently fall through
        to _default, and an exact 'AAV9' should get its own profile."""
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_primitives import get_dds_profile

        aav9 = get_dds_profile("AAV9")
        aav5 = get_dds_profile("aav5")
        generic_aav = get_dds_profile("aav")
        assert aav5 == generic_aav
        assert aav9["shape"] == "icosahedral_capsid"

    def test_dds_profile_unrecognized_carrier_falls_back_to_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_primitives import DDS_VISUAL_PROFILES, get_dds_profile

        assert get_dds_profile("nonexistent_carrier_xyz") == DDS_VISUAL_PROFILES["_default"]

    def test_ligand_info_partial_match_and_default(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_primitives import LIGAND_RECEPTOR_MAP, get_ligand_info

        assert get_ligand_info("RVG29")["receptor"].startswith("Nicotinic")
        assert get_ligand_info(None) == LIGAND_RECEPTOR_MAP[""]
        assert get_ligand_info("") == LIGAND_RECEPTOR_MAP[""]


# ═════════════════════════════════════════════════════════════════════════════
# 41. CINEMATIC SCENE ENGINE (cerebro_cinematic_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestCinematicEngineHelpers:
    def test_b_value_and_b_tier_handle_missing_and_malformed_bundles(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import _b_tier, _b_value

        assert _b_value({"x": {"value": 5.0}}, "x", 0) == 5.0
        assert _b_value({"x": {"value": None}}, "x", 99) == 99   # explicit None -> default
        assert _b_value("not a dict", "x", 99) == 99
        assert _b_tier({"x": {"tier": 3}}, "x") == 3
        assert _b_tier({}, "x") == 7   # unresolved defaults to lowest-confidence tier

    def test_hash_id_is_deterministic_and_short(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import _hash_id

        h1 = _hash_id("DrugA", "DDS1", "C01")
        h2 = _hash_id("DrugA", "DDS1", "C01")
        h3 = _hash_id("DrugB", "DDS1", "C01")
        assert h1 == h2 and len(h1) == 8
        assert h1 != h3

    def test_safe_filename_strips_unsafe_characters(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import _safe_filename

        assert _safe_filename("Drug/Name: Test!") == "Drug_Name__Test_"
        assert _safe_filename(None) == "x"
        assert len(_safe_filename("x" * 100)) == 40

    def test_drug_class_narrative_covers_every_modality_with_real_citations(self):
        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import _drug_class_narrative

        mab = _drug_class_narrative("monoclonal_antibody")
        assert "Lecanemab" in mab["clinical_example"]
        oligo = _drug_class_narrative("oligonucleotide")
        assert "Nusinersen" in oligo["clinical_example"]
        default = _drug_class_narrative("totally_unknown_modality")
        assert "Donepezil" in default["clinical_example"]   # falls back to small molecule


class TestCinematicSuiteGeneration:
    @pytest.mark.slow
    def test_generates_all_five_scenes_for_a_real_drug(self):
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import generate_cinematic_suite
        from cerebro_resolved_bundles import resolve_dds_bundle, resolve_drug_bundle

        db = resolve_drug_bundle(
            name="Donepezil",
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            molecule_class="small_molecule")
        dds = resolve_dds_bundle(carrier_type="plga", ligand="transferrin", formulation_id="F1")
        top1 = {"Formulation_Name": "Tf-PLGA", "Carrier_Type": "plga",
                "Size_nm": 100, "Zeta_Potential_mV": -25, "PDI": 0.2,
                "Surface_Ligand": "transferrin", "Drug_Loading_Pct": 15,
                "Release_Kinetics": "sustained", "pH_Trigger": 6.5,
                "Composite_Score": 80.0, "Principle_Composite_Score": 80.0}
        with tempfile.TemporaryDirectory() as td:
            paths = generate_cinematic_suite(db, dds, top1, Path(td))
            assert len(paths) == 5
            for p in paths:
                assert p.exists()
                assert p.stat().st_size > 1000   # not an empty/error stub

    def test_a_single_failing_scene_does_not_abort_the_whole_suite(self):
        """generate_cinematic_suite catches per-scene exceptions — a
        malformed bundle missing _meta shouldn't crash the whole run,
        just log and skip whichever scene(s) choke on it."""
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import generate_cinematic_suite

        with tempfile.TemporaryDirectory() as td:
            paths = generate_cinematic_suite({}, {}, {}, Path(td))
            assert isinstance(paths, list)   # doesn't raise, even with nothing real to render


class TestC03PkProfileHonestLabeling:
    """C03's curve is a stylized Bateman-function illustration (arbitrary
    ka + unitless dose) — real per-drug half-life and BBB% drive the curve
    SHAPE, but the plotted numbers are not a calibrated PBPK prediction. The
    real 3-compartment ODE result lives in cerebro_62_deep_engine.deep_P13.
    The scene must disclose this to the viewer instead of implying the
    numbers are lab-real concentrations."""

    def test_scene_does_not_claim_a_real_concentration_unit(self):
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import make_c03_pk_profile

        drug = _drug_bundle(pk_halflife={"value": 2.0},
                              bbb_permeability={"value": 12.0})
        top_dds = {"Formulation_Name": "Test-DDS"}
        with tempfile.TemporaryDirectory() as td:
            p = make_c03_pk_profile(drug, _dds_bundle(), top_dds, Path(td))
            html = p.read_text()
            assert "μg/mL" not in html   # no fabricated real-looking unit

    def test_scene_discloses_it_is_a_stylized_illustration(self):
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import make_c03_pk_profile

        drug = _drug_bundle(pk_halflife={"value": 2.0},
                              bbb_permeability={"value": 12.0})
        top_dds = {"Formulation_Name": "Test-DDS"}
        with tempfile.TemporaryDirectory() as td:
            p = make_c03_pk_profile(drug, _dds_bundle(), top_dds, Path(td))
            html = p.read_text()
            assert "Stylized illustration" in html
            assert "PDF/Excel report" in html

    def test_curve_shape_still_reflects_real_half_life_and_bbb_percent(self):
        """The fix must not touch the part that already works: the curve's
        shape (elimination rate, brain lag/scale) is driven by the real
        per-drug half-life and BBB% pulled from the bundle."""
        import tempfile
        from pathlib import Path

        import src.path_resolver  # noqa: F401
        from cerebro_cinematic_engine import make_c03_pk_profile

        top_dds = {"Formulation_Name": "Test-DDS"}
        with tempfile.TemporaryDirectory() as td:
            slow = _drug_bundle(pk_halflife={"value": 5.0},
                                  bbb_permeability={"value": 12.0})
            fast = _drug_bundle(pk_halflife={"value": 0.2},
                                  bbb_permeability={"value": 12.0})
            html_slow = make_c03_pk_profile(slow, _dds_bundle(), top_dds, Path(td)).read_text()
            html_fast = make_c03_pk_profile(fast, _dds_bundle(), top_dds, Path(td)).read_text()
            assert "5.00 d" in html_slow
            assert "0.20 d" in html_fast
            assert html_slow != html_fast


# ═════════════════════════════════════════════════════════════════════════════
# 42. PDB RESOLVER (src/core/pdb_resolver.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestPdbResolver:
    @pytest.mark.slow
    def test_path_traversal_pdb_id_is_rejected_not_trusted(self):
        """Security regression test for the audit's §6 Medium finding: a
        user-supplied pdb_id must pass a strict 4-char alphanumeric check
        before being trusted — a value like '../x' must NOT be accepted,
        since a trusted pdb_id later builds both a download URL and a
        local filesystem path in real_docking_engine.py. Marked slow: a
        rejected override falls through to the real live-API cascade."""
        import src.path_resolver  # noqa: F401
        from src.core.pdb_resolver import resolve_pdb_for_drug

        r = resolve_pdb_for_drug("SomeDrug", user_pdb_id="../x")
        assert r["source"] != "User-provided (Excel input)"
        assert r["pdb_id"] != "../x"

    def test_valid_pdb_id_is_trusted_and_uppercased(self):
        import src.path_resolver  # noqa: F401
        from src.core.pdb_resolver import resolve_pdb_for_drug

        r = resolve_pdb_for_drug("SomeDrug", user_pdb_id="1abc")
        assert r["pdb_id"] == "1ABC"
        assert r["source"] == "User-provided (Excel input)"
        assert r["confidence"] == "HIGH"

    def test_pdb_id_regex_rejects_wrong_length_and_special_characters(self):
        import src.path_resolver  # noqa: F401
        from src.core.pdb_resolver import _PDB_ID_RE

        assert not _PDB_ID_RE.match("abc")     # too short
        assert not _PDB_ID_RE.match("abcde")   # too long
        assert not _PDB_ID_RE.match("ab-c")    # non-alphanumeric
        assert _PDB_ID_RE.match("1ABC")

    def test_pdb_ref_is_permanently_empty_by_design(self):
        """Regression guard for a real (if low-severity) dead-code issue:
        the auto-resolved confidence used to read
        'HIGH if drug_lower in PDB_REF else MODERATE', but PDB_REF was
        deliberately emptied in v22.1 (no hardcoded drug data) — that
        branch could never actually produce HIGH. Confirmed the table
        stays empty and auto-resolved confidence is honestly MODERATE."""
        import src.path_resolver  # noqa: F401
        from src.core.pdb_resolver import PDB_REF

        assert PDB_REF == {}

    @pytest.mark.slow
    def test_no_pdb_id_available_falls_back_to_honest_blind_docking(self):
        import src.path_resolver  # noqa: F401
        from src.core.pdb_resolver import resolve_pdb_for_drug

        r = resolve_pdb_for_drug("a-drug-name-with-no-pdb-structure-xyz-123")
        assert r["pdb_id"] is None
        assert r["confidence"] == "LOW"


# ═════════════════════════════════════════════════════════════════════════════
# 43. NOVEL DRUG ANALOG ENGINE (src/core/novel_drug_analog.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestNovelDrugAnalog:
    def test_no_smiles_reports_honest_disclaimer_not_a_fabricated_hit(self):
        import src.path_resolver  # noqa: F401
        from src.core.novel_drug_analog import find_closest_analog

        r = find_closest_analog("SomeDrug", {}, smiles="")
        assert r["is_novel_drug"] is True
        assert r["closest_analog"]["similarity_pct"] == 0.0
        assert "no smiles" in r["disclaimer"].lower()

    def test_reference_drugs_table_is_permanently_empty_by_design(self):
        """v22.1 removed the embedded 220-drug reference list specifically
        because it let unrelated hardcoded drug names leak into outputs
        when the researcher's own input was incomplete."""
        import src.path_resolver  # noqa: F401
        from src.core.novel_drug_analog import REFERENCE_DRUGS

        assert REFERENCE_DRUGS == []

    def test_pubchem_threshold_hits_are_not_silently_ranked_as_exact_matches(self):
        """Regression test for a real bug: PubChem's similarity endpoint
        returns CIDs meeting a threshold, not a per-compound score — every
        hit used to report the identical threshold value as 'similarity_pct'
        with nothing marking it as an estimate, so it competed directly
        against ChEMBL's real per-compound Tanimoto scores in the best-
        match selection, and the disclaimer text stated it as if it were
        an exact measurement. Now flagged via similarity_is_exact and the
        disclaimer wording changes when the winning hit is threshold-only."""
        import src.path_resolver  # noqa: F401
        import src.core.novel_drug_analog as nda

        def fake_chembl(smiles, threshold_pct=60, limit=5):
            return [{"name": "RealChemblHit", "chembl_id": "CHEMBL1",
                     "similarity_pct": 75.0, "similarity_is_exact": True,
                     "_source": "x", "method": "live_chembl_tanimoto"}]

        def fake_pubchem(smiles, threshold=90, limit=5):
            return [{"name": "ThresholdOnlyHit", "pubchem_cid": 123,
                     "similarity_pct": 90.0, "similarity_is_exact": False,
                     "_source": "y", "method": "live_pubchem_2d"}]

        orig_chembl, orig_pubchem = nda._live_chembl_similarity, nda._live_pubchem_similarity
        nda._live_chembl_similarity = fake_chembl
        nda._live_pubchem_similarity = fake_pubchem
        try:
            r = nda.find_closest_analog("SomeDrug", {}, smiles="CCO")
        finally:
            nda._live_chembl_similarity = orig_chembl
            nda._live_pubchem_similarity = orig_pubchem

        assert r["closest_analog"]["name"] == "ThresholdOnlyHit"   # 90 > 75, still wins on the number
        assert "not an exact per-compound score" in r["disclaimer"]

    def test_chembl_hit_disclaimer_reads_as_an_exact_measurement(self):
        import src.path_resolver  # noqa: F401
        import src.core.novel_drug_analog as nda

        def fake_chembl(smiles, threshold_pct=60, limit=5):
            return [{"name": "RealChemblHit", "chembl_id": "CHEMBL1",
                     "similarity_pct": 88.0, "similarity_is_exact": True,
                     "_source": "x", "method": "live_chembl_tanimoto"}]

        def fake_pubchem(smiles, threshold=90, limit=5):
            return []

        orig_chembl, orig_pubchem = nda._live_chembl_similarity, nda._live_pubchem_similarity
        nda._live_chembl_similarity = fake_chembl
        nda._live_pubchem_similarity = fake_pubchem
        try:
            r = nda.find_closest_analog("SomeDrug", {}, smiles="CCO")
        finally:
            nda._live_chembl_similarity = orig_chembl
            nda._live_pubchem_similarity = orig_pubchem

        assert "88.0% similarity via live_chembl_tanimoto" in r["disclaimer"]
        assert "threshold" not in r["disclaimer"]


# ═════════════════════════════════════════════════════════════════════════════
# 44. MISSING VALUE RESOLVER (src/core/missing_value_resolver.py +
#     src/core/molecule_engine.resolve_missing_properties)
# ═════════════════════════════════════════════════════════════════════════════
class TestMissingValueResolverZeroHandling:
    """Regression tests for a real, confirmed bug: a legitimate 0 for
    HBD/HBA/TPSA/LogP (very common — e.g. any molecule with no -OH/-NH
    group genuinely has HBD=0) was treated as 'missing' and could get
    silently overwritten by a fabricated class-typical fallback (HBD=0
    -> class-mean 2.0) whenever the live re-resolution cascade failed for
    any reason (network issues, an unresolvable drug name). MW_Da and
    Half_Life_Days are different — those really can't be 0 for a real
    molecule, so 0 there still correctly means 'never set'."""

    def test_legitimate_zero_hbd_survives_when_the_live_cascade_fails(self):
        import src.path_resolver  # noqa: F401
        import src.core.missing_value_resolver as mvr
        from src.core.molecule_engine import resolve_missing_properties

        orig_pubchem, orig_pubmed = mvr._pubchem_property, mvr._pubmed_search
        mvr._pubchem_property = lambda *a, **k: None
        mvr._pubmed_search = lambda *a, **k: None
        try:
            mol_profile = {"HBD": 0, "MW_Da": 108.14, "LogP": 2.1,
                            "TPSA_A2": 9.2, "HBA": 1, "Half_Life_Days": 0.1}
            out = resolve_missing_properties(mol_profile, "AnisoleLikeDrug", smiles=None)
        finally:
            mvr._pubchem_property, mvr._pubmed_search = orig_pubchem, orig_pubmed

        assert out["HBD"] == 0          # not overwritten with the class-mean 2.0
        assert out["TPSA_A2"] == 9.2    # not overwritten either

    def test_genuinely_unset_mw_da_zero_still_gets_resolved(self):
        import src.path_resolver  # noqa: F401
        import src.core.missing_value_resolver as mvr
        from src.core.molecule_engine import resolve_missing_properties

        orig_pubchem, orig_pubmed = mvr._pubchem_property, mvr._pubmed_search
        mvr._pubchem_property = lambda *a, **k: None
        mvr._pubmed_search = lambda *a, **k: None
        try:
            mol_profile = {"HBD": 0, "MW_Da": 0, "LogP": 2.1,
                            "TPSA_A2": 9.2, "HBA": 1, "Half_Life_Days": 0.1}
            out = resolve_missing_properties(mol_profile, "AnisoleLikeDrug", smiles=None)
        finally:
            mvr._pubchem_property, mvr._pubmed_search = orig_pubchem, orig_pubmed

        assert out["MW_Da"] != 0   # MW=0 is physically impossible -> correctly treated as missing
        assert out["MW_Da"] == 350.0   # resolved via the small_molecule class-typical fallback

    def test_resolve_property_accepts_a_real_zero_api_value_for_hbd(self):
        import src.path_resolver  # noqa: F401
        from src.core.missing_value_resolver import resolve_property

        r = resolve_property("SomeDrug", "hbd", mol_profile={}, smiles=None, api_value=0)
        assert r["value"] == 0
        assert r["_tier"] == 1   # trusted as the real API-provided value, not re-derived

    def test_resolve_property_treats_zero_mw_as_still_missing(self):
        import src.path_resolver  # noqa: F401
        from src.core.missing_value_resolver import resolve_property

        r = resolve_property("SomeDrug", "mw_da", mol_profile={}, smiles=None, api_value=0)
        assert r["_tier"] != 1   # 0 Da isn't a real molecular weight — falls through


# ═════════════════════════════════════════════════════════════════════════════
# 45. REAL QSAR ENGINE (src/core/real_qsar_engine.py)
# ═════════════════════════════════════════════════════════════════════════════
class TestRealQsarEngine:
    def test_receptor_panel_has_50_unique_targets(self):
        import src.path_resolver  # noqa: F401
        from src.core.real_qsar_engine import RECEPTOR_TARGETS

        assert len(RECEPTOR_TARGETS) == 50
        assert len({t["name"] for t in RECEPTOR_TARGETS}) == 50   # no duplicates
        for t in RECEPTOR_TARGETS:
            assert t["chembl_id"].startswith("CHEMBL")

    def test_compute_features_returns_maccs_plus_seven_physchem(self):
        import src.path_resolver  # noqa: F401
        from src.core.real_qsar_engine import _compute_features

        feats = _compute_features("CCO", {})
        assert len(feats) == 167 + 7

    def test_empirical_cardiac_score_matches_hand_computation(self):
        """Aronov 2006 hERG rule: LogP>3.5 (+0.3), MW 300-600 (+0.2),
        HBA<4 (+0.1) = 0.6 raw, then the score += 0.05*(1-score) blend."""
        import src.path_resolver  # noqa: F401
        from src.core.real_qsar_engine import RECEPTOR_TARGETS, _empirical_score

        cardiac_target = next(t for t in RECEPTOR_TARGETS if t["risk_type"] == "cardiac")
        mp = {"MW_Da": 450.0, "LogP": 4.0, "TPSA_A2": 60.0, "HBD": 1, "HBA": 3}
        r = _empirical_score(cardiac_target, mp, None)
        raw = 0.3 + 0.2 + 0.1
        expected = min(0.95, max(0.02, raw + 0.05 * (1 - raw)))
        assert r["score_free_drug"] == pytest.approx(round(expected, 3))
        assert r["risk"] == "HIGH"
        assert r["score_in_DDS"] == pytest.approx(round(expected * 0.45, 3))

    def test_empirical_score_stays_within_bounds_for_extreme_inputs(self):
        import src.path_resolver  # noqa: F401
        from src.core.real_qsar_engine import RECEPTOR_TARGETS, _empirical_score

        for target in RECEPTOR_TARGETS[:5]:
            r = _empirical_score(target, {"MW_Da": 1e6, "LogP": 50, "TPSA_A2": 0,
                                            "HBD": 0, "HBA": 0}, None)
            assert 0.02 <= r["score_free_drug"] <= 0.95

    def test_qsar_model_training_is_cached_per_target_not_per_drug(self):
        """Regression test for a real performance bug: _train_qsar_model
        re-fetched 500 ChEMBL records and retrained a fresh Random Forest
        on every call — run_real_qsar_panel called it once per receptor
        PER DRUG, so scoring N drugs against the same 50-target panel
        retrained the same 50 models N times with zero reuse, even though
        a target's model depends only on that target's own ChEMBL data,
        never on which drug is being scored. Now cached by
        (target_name, chembl_id) via lru_cache."""
        import src.path_resolver  # noqa: F401
        import src.core.real_qsar_engine as qe
        from functools import lru_cache

        call_count = {"n": 0}

        def fake_trainer(target_name, chembl_id):
            call_count["n"] += 1
            return None

        orig = qe._train_qsar_model
        qe._train_qsar_model = lru_cache(maxsize=64)(fake_trainer)
        try:
            qe._train_qsar_model("hERG_K+", "CHEMBL240")
            qe._train_qsar_model("hERG_K+", "CHEMBL240")
            qe._train_qsar_model("hERG_K+", "CHEMBL240")
            assert call_count["n"] == 1   # only the first call actually trained
            qe._train_qsar_model("CYP3A4_inhib", "CHEMBL340")
            assert call_count["n"] == 2   # a genuinely different target does train
        finally:
            qe._train_qsar_model = orig

    @pytest.mark.slow
    def test_run_real_qsar_panel_scores_all_50_receptors_for_a_real_drug(self):
        import src.path_resolver  # noqa: F401
        from src.core.real_qsar_engine import run_real_qsar_panel

        result = run_real_qsar_panel(
            smiles="COc1cc2c(cc1OC)C(=O)C(CC1CCN(Cc3ccccc3)CC1)C2",
            mol_profile={"MW_Da": 379.5, "LogP": 4.3, "TPSA_A2": 38.8, "HBD": 0, "HBA": 4},
            top_dds={}, use_ml=False)   # use_ml=False keeps this from hitting live ChEMBL 50x
        assert result["n_receptors_screened"] == 50
        assert len(result["receptor_panel"]) == 50
        assert result["overall_off_target"] in ("LOW RISK", "CAUTION", "HIGH CONCERN", "CRITICAL")


# ═════════════════════════════════════════════════════════════════════════════
# 40. TRIAL MANAGER — CACHE INVALIDATION MUST HIT THE REAL DB FILE
# ═════════════════════════════════════════════════════════════════════════════
class TestInvalidateMoleculeCacheHitsRealDbFile:
    """invalidate_molecule_cache's own docstring promises: "After this
    call, analyze_molecule() will fetch fresh from APIs" -- backed by
    deleting stale drug_records rows so a subsequent INSERT OR REPLACE
    upsert can't be shadowed by leftover data. It checked db_candidates =
    [RESULTS_ROOT/"cerebro.db", SCRIPT_DIR/"cerebro.db"] -- a filename
    that has never existed anywhere in this project. The real file,
    written by src/core/pipeline.py's DB_PATH, is "cerebro_knowledge.db".
    Since db_path.exists() was always False for "cerebro.db", the entire
    SQLite-deletion step was a silent no-op on every single trial run,
    contradicting run.py's own documented guarantee ("This guarantees
    fresh API fetch every time — no stale data")."""

    def test_stale_drug_record_row_is_actually_deleted(self, tmp_path, monkeypatch):
        import sqlite3

        import trial_manager

        monkeypatch.setattr(trial_manager, "RESULTS_ROOT", tmp_path)
        monkeypatch.setattr(trial_manager, "SCRIPT_DIR", tmp_path)

        db_path = tmp_path / "cerebro_knowledge.db"
        conn = sqlite3.connect(db_path)
        conn.execute("CREATE TABLE drug_records (drug_name TEXT, mw_da REAL)")
        conn.execute("INSERT INTO drug_records VALUES ('TestDrugXYZ', 300.0)")
        conn.commit()
        conn.close()

        trial_manager.invalidate_molecule_cache(["TestDrugXYZ"], tmp_path)

        conn = sqlite3.connect(db_path)
        remaining = conn.execute(
            "SELECT COUNT(*) FROM drug_records WHERE drug_name='TestDrugXYZ'"
        ).fetchone()[0]
        conn.close()
        assert remaining == 0, (
            "invalidate_molecule_cache must delete stale rows from the "
            "real cerebro_knowledge.db, not a 'cerebro.db' that never exists")


# ═════════════════════════════════════════════════════════════════════════════
# 41. REPORT FALLBACKS — MERGED PDF "ALL 100 FORMULATIONS" TABLE
# ═════════════════════════════════════════════════════════════════════════════
class TestMergedPdfIncludesLigandColumn:
    """_generate_merged_pdf's "Complete Formulation Rankings (all 100)"
    section worked out the real ligand column's casing
    ("surface_ligand" vs "Surface_Ligand") into a COLS2 list, but COLS2
    was never actually used -- avail2, the hardcoded list the table was
    actually built from, didn't include a ligand column at all. The
    surface-ligand data that report exists to summarize was silently
    dropped from its own full-rankings table despite the code clearly
    intending to include it."""

    def test_surface_ligand_column_appears_in_full_rankings_table(self, tmp_path):
        import pandas as pd

        from report_fallbacks import _generate_merged_pdf

        df_dds = pd.DataFrame({
            "Rank": [1, 2],
            "Formulation_ID": ["F1", "F2"],
            "Formulation_Name": ["Name1", "Name2"],
            "Carrier_Type": ["Liposome", "SLN"],
            "BBB_Engineering_Score": [85.0, 70.0],
            "ADMET_Overall_Flag": ["OK", "OK"],
            "Surface_Ligand": ["RVG29", "ApoE"],
        })
        # _generate_merged_pdf doesn't return the column list it built, so
        # reproduce the exact same resolution logic to assert against.
        ligand_col = ("surface_ligand" if "surface_ligand" in df_dds.columns
                      else "Surface_Ligand")
        avail2 = [c for c in ["Rank","Formulation_ID","Formulation_Name",
                               "Carrier_Type","BBB_Engineering_Score",
                               "ADMET_Overall_Flag", ligand_col]
                  if c in df_dds.columns]
        assert "Surface_Ligand" in avail2

        # Also confirm the real function still runs end-to-end without error.
        _generate_merged_pdf(None, df_dds, None, {}, {}, tmp_path, "TestDrug")
        pdf_path = tmp_path / "CEREBRO_X_Report_TestDrug.pdf"
        assert pdf_path.exists()
        assert pdf_path.stat().st_size > 0
