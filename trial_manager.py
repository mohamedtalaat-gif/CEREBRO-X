"""
================================================================================
CEREBRO-X — Trial Index, Excel-to-YAML Conversion, and Cache Invalidation
================================================================================
File: trial_manager.py

Extracted from run.py (was Sections 4, 5, 6: "TRIAL INDEX", "EXCEL -> YAML
CONVERTER", "CACHE INVALIDATOR") as part of splitting run.py's mixed
responsibilities — see docs/AUDIT_REPORT.md section 13.

Path constants here mirror run.py's Section 2 exactly (same SCRIPT_DIR
derivation, since this file lives in the same project root) rather than
importing them back from run.py, to avoid a circular import between the
two modules.
================================================================================
"""
from __future__ import annotations

import hashlib
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

log = logging.getLogger("CEREBRO-TRIALS")

try:
    SCRIPT_DIR = Path(os.path.abspath(__file__)).parent
except NameError:
    SCRIPT_DIR = Path(os.path.abspath(sys.argv[0])).parent

EXCEL_GLOB_PATTERNS = [
    "CEREBRO_Input*.xlsx",
    "CEREBRO_Input*.xls",
    "cerebro_input*.xlsx",
]
INPUTS_DIR     = SCRIPT_DIR / "inputs"
RESULTS_ROOT   = SCRIPT_DIR / "outputs"
TRIAL_INDEX_DB = RESULTS_ROOT / "trial_index.db"

INPUTS_DIR.mkdir(parents=True, exist_ok=True)
RESULTS_ROOT.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TRIAL INDEX  (SQLite — tracks which Excel files have been processed)
# ─────────────────────────────────────────────────────────────────────────────

def _init_trial_db() -> None:
    """Create trial index table if it doesn't exist."""
    conn = sqlite3.connect(TRIAL_INDEX_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS trials (
            trial_id    INTEGER PRIMARY KEY AUTOINCREMENT,
            excel_path  TEXT    NOT NULL,
            excel_hash  TEXT    NOT NULL UNIQUE,
            drug_name   TEXT,
            n_forms     INTEGER,
            run_at      TEXT,
            output_dir  TEXT,
            status      TEXT DEFAULT 'pending'
        )""")
    conn.commit()
    conn.close()


def _excel_hash(xlsx_path: Path) -> str:
    """SHA-256 of the Excel file bytes — changes if ANY cell changes."""
    h = hashlib.sha256()
    with open(xlsx_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def find_new_excel_files() -> list:
    """
    Scan INPUTS_DIR for Excel files matching CEREBRO_Input*.xlsx.
    Return only files whose hash is NOT in trial_index.db.
    These are files we haven't processed yet.
    """
    _init_trial_db()
    conn = sqlite3.connect(TRIAL_INDEX_DB)
    known_hashes = {row[0] for row in
                    conn.execute("SELECT excel_hash FROM trials").fetchall()}
    conn.close()

    new_files = []
    for pattern in EXCEL_GLOB_PATTERNS:
        for xlsx in sorted(INPUTS_DIR.glob(pattern)):
            if xlsx.stem.endswith("_Template"):
                continue  # blank template, not a real trial input
            h = _excel_hash(xlsx)
            if h not in known_hashes:
                new_files.append((xlsx, h))

    if new_files:
        log.info(f"[WATCHER] {len(new_files)} new/changed Excel file(s) found")
    else:
        log.info("[WATCHER] No new Excel files — all already processed")
    return new_files


def register_trial(excel_path: Path, excel_hash: str,
                   drug_name: str, n_forms: int,
                   output_dir: Path, status: str = "complete") -> int:
    """Record a completed trial in the index database."""
    _init_trial_db()
    conn = sqlite3.connect(TRIAL_INDEX_DB)
    conn.execute("""
        INSERT OR REPLACE INTO trials
            (excel_path, excel_hash, drug_name, n_forms, run_at, output_dir, status)
        VALUES (?,?,?,?,?,?,?)""",
        (str(excel_path), excel_hash, drug_name, n_forms,
         datetime.utcnow().isoformat(), str(output_dir), status))
    conn.commit()
    trial_id = conn.execute(
        "SELECT trial_id FROM trials WHERE excel_hash=?", (excel_hash,)
    ).fetchone()[0]
    conn.close()
    return trial_id


def next_trial_dir(excel_path: "Path | None" = None) -> Path:
    """
    Return the output directory for a trial, named after the drug rather
    than an auto-incremented Trial_N — outputs/Donepezil/ instead of
    outputs/Trial_0/, so results are discoverable by name instead of by
    an opaque counter. Derived from the input Excel's filename
    (CEREBRO_Input_<DrugName>.xlsx -> <DrugName>) since the drug name
    itself isn't parsed yet at the point this is called. Falls back to
    Trial_N only if no usable name can be derived (e.g. no excel_path,
    or a filename that doesn't match the expected pattern).
    """
    _init_trial_db()
    name = None
    if excel_path is not None:
        stem = Path(excel_path).stem
        for prefix in ("CEREBRO_Input_", "cerebro_input_"):
            if stem.startswith(prefix):
                stem = stem[len(prefix):]
                break
        safe = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in stem).strip("_")
        if safe:
            name = safe
    if name is None:
        conn = sqlite3.connect(TRIAL_INDEX_DB)
        n = conn.execute("SELECT COUNT(*) FROM trials").fetchone()[0]
        conn.close()
        name = f"Trial_{n}"
    d = RESULTS_ROOT / name
    d.mkdir(parents=True, exist_ok=True)
    return d


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL -> YAML CONVERTER  (the single source of truth)
# ─────────────────────────────────────────────────────────────────────────────

def excel_to_yaml(xlsx_path: "Path", output_yaml: "Path",
                  force_refresh: bool = True) -> dict:
    """
    Read CEREBRO_Input_*.xlsx -> write dds_config.yaml.

    REWRITTEN v19 — fully robust, dynamic-N drug parser.
    ─────────────────────────────────────────────────────
    Design principles:
      • NO hardcoded row positions — pure label-based scan
      • Dynamic drug count: 1, 2, 3, …, N (no upper limit)
      • Fuzzy label matching: handles whitespace, embedded \\n,
        parenthetical suffixes, ordering variations
      • Researcher override capture: any value typed into an
        "(auto-fetched)" cell is stored with provenance=researcher_input
      • Markers (▶, bullets, help text, section headers) auto-skipped
      • Returns unified `drugs` list — single drug is just N=1
      • Backward compatible: also exposes `drug` (=drugs[0])
        and `additional_drugs` (=drugs[1:])

    Output cfg structure:
      {
        "drugs":            [drug1, drug2, ..., drugN],   # primary container
        "drug":             drugs[0],                      # backward compat
        "additional_drugs": drugs[1:],                     # backward compat
        "formulations":     [...],
        "pipeline_config":  {...},
      }
    """
    import openpyxl
    import yaml

    log.info(f"[EXCEL→YAML] Reading: {xlsx_path.name}")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # ─── Sheet detection (auto-handles legacy 'Drugs' format too) ────────
    if "1_Drug_Input" in wb.sheetnames:
        ws1 = wb["1_Drug_Input"]
    elif "Drugs" in wb.sheetnames:
        ws1 = wb["Drugs"]
        log.info("[EXCEL] Legacy 'Drugs' sheet detected — parsing as flat rows")
    else:
        raise ValueError(
            f"Excel format not recognized. Need '1_Drug_Input' or 'Drugs' sheet. "
            f"Found: {', '.join(wb.sheetnames)}")

    # ═══════════════════════════════════════════════════════════════════
    #  ROBUST LABEL NORMALIZATION
    # ═══════════════════════════════════════════════════════════════════
    _LBL_NL_RE  = re.compile(r"[\r\n\t]+")          # collapse newlines/tabs
    _LBL_WS_RE  = re.compile(r"\s+")                # collapse whitespace
    _LBL_PAREN  = re.compile(r"\s*\([^)]*\)\s*$")   # strip trailing (…)
    _LBL_PREFIX = re.compile(r"^[\s\d\.▶►●•◾◽\-–—]+")  # strip bullets/numbers
    def _norm(label: str) -> str:
        if not label: return ""
        s = str(label)
        s = _LBL_NL_RE.sub(" ", s)
        s = _LBL_PREFIX.sub("", s)
        s = _LBL_PAREN.sub("", s)
        s = _LBL_WS_RE.sub(" ", s).strip().lower()
        return s

    # ═══════════════════════════════════════════════════════════════════
    #  CANONICAL FIELD MAP — normalized label → canonical key
    # ═══════════════════════════════════════════════════════════════════
    _FIELD_MAP = {
        # Identity
        "drug name":              "name",
        "name":                   "name",
        "molecule class":         "molecule_class",
        "class":                  "molecule_class",
        "molecule input":         "molecule_input",
        "smiles":                 "smiles",
        "fasta":                  "fasta",
        "indication":             "indication",
        "indication disease target": "indication",
        "disease target":         "indication",
        "target protein":         "target_protein",
        "target":                 "target_protein",
        "target pdb id":          "target_pdb_id",
        "pdb id":                 "target_pdb_id",
        "native bbb penetration %": "bbb_native_pct",
        "native bbb %":           "bbb_native_pct",
        "native bbb":             "bbb_native_pct",
        "clinical phase":         "clinical_phase",
        "phase":                  "clinical_phase",
        "fda approval date":      "fda_approval_date",
        "approval date":          "fda_approval_date",
        # Properties (researcher CAN override these by typing values)
        "mw da":                  "mw_da",
        "mw":                     "mw_da",
        "molecular weight":       "mw_da",
        "logp":                   "logp",
        "half-life days":         "half_life_days",
        "half life days":         "half_life_days",
        "half-life":              "half_life_days",
        "half life":              "half_life_days",
        "h-bond donors":          "hbd",
        "hbd":                    "hbd",
        "h-bond acceptors":       "hba",
        "hba":                    "hba",
        "tpsa å²":                "tpsa",
        "tpsa":                   "tpsa",
        "pi":                     "pi",
        "instability index":      "instability_index",
        "uniprot id":             "uniprot_id",
        "uniprot":                "uniprot_id",
        "logbb":                  "logbb",
        "bbb penetration %":      "bbb_pct",
        "bbb penetration":        "bbb_pct",
    }

    # Properties that are normally auto-fetched — if researcher fills them,
    # capture as an override with HIGH confidence
    _OVERRIDABLE_KEYS = {
        "mw_da", "logp", "half_life_days", "hbd", "hba", "tpsa",
        "pi", "instability_index", "uniprot_id", "logbb", "bbb_pct",
        "bbb_native_pct",
    }

    # Sentinel patterns that mean "leave blank — pipeline will fetch"
    _AUTO_PATTERNS = re.compile(
        r"^\s*\(?\s*(fetched\s+automatically|auto[\-\s]?fetch(?:ed)?|"
        r"auto|tbd|n[/]?a)\s*\)?\s*$",
        re.IGNORECASE)

    # Section markers to skip silently (help text, layout, headers)
    _SKIP_PREFIXES = (
        "fill ", "note", "cerebro", "format", "brand",
        "if you", "field", "your input", "for drug",
        "drug identity", "molecule inputs", "optional",
    )
    def _is_skippable_label(lbl_norm: str, raw: str) -> bool:
        # Single Unicode bullet/icon-only rows
        if not lbl_norm: return True
        if lbl_norm.startswith(_SKIP_PREFIXES): return True
        # Help text starting with ⚡ / icon
        if any(ch in raw for ch in "⚡✅❌🔬"): return True
        return False

    # Section header regex: detects "Drug N" ONLY at start of normalized label.
    # This rejects help text like "If you enter Drug 2 …" or "For Drug 2 & 3:".
    _DRUG_SECTION_RE = re.compile(r"^drug\s+(\d+)\b", re.IGNORECASE)

    # ═══════════════════════════════════════════════════════════════════
    #  SCAN ROW-BY-ROW — DYNAMIC-N DRUG EXTRACTION
    # ═══════════════════════════════════════════════════════════════════
    drugs_raw: list = [{}]   # Drug 1 implicitly starts at index 0
    current_idx = 0
    overrides_log: list = []

    for row_n, row in enumerate(ws1.iter_rows(min_row=1, values_only=True), 1):
        if not row: continue
        raw_label = row[0]
        raw_value = row[1] if len(row) > 1 else None

        if raw_label is None and raw_value is None:
            continue

        label_str = str(raw_label or "").strip()
        lbl_norm  = _norm(label_str)

        # Detect "Drug N" section marker — strict start-of-label match
        m = _DRUG_SECTION_RE.match(lbl_norm)
        if m:
            n = int(m.group(1))
            # Pad drugs_raw so index n-1 exists
            while len(drugs_raw) < n:
                drugs_raw.append({})
            current_idx = n - 1
            log.info(f"[EXCEL→YAML]  ↳ Section detected: Drug {n} (row {row_n})")
            continue

        # Skip pure layout / help / header rows
        if _is_skippable_label(lbl_norm, label_str):
            continue

        # Empty value? skip — but ONLY after section/marker checks above
        if raw_value is None:
            continue
        val_s = str(raw_value).strip()
        if not val_s: continue
        # Auto-fetch sentinel? skip — pipeline will fetch live
        if _AUTO_PATTERNS.match(val_s):
            continue

        # Resolve canonical key from normalized label
        key = _FIELD_MAP.get(lbl_norm)
        if key is None:
            # Try partial match: does any canonical label contain lbl_norm?
            for canon, k in _FIELD_MAP.items():
                if canon and (canon in lbl_norm or lbl_norm in canon):
                    key = k
                    break
        if key is None:
            # Unknown field — preserve as snake_case for downstream debugging
            key = re.sub(r"[^a-z0-9_]+", "_", lbl_norm).strip("_")
            if not key: continue

        # Type-coerce numeric properties when the key expects a number
        _NUMERIC_KEYS = {"mw_da", "logp", "half_life_days", "hbd", "hba",
                         "tpsa", "pi", "instability_index", "logbb",
                         "bbb_pct", "bbb_native_pct"}
        if key in _NUMERIC_KEYS:
            try:    val_out = float(val_s)
            except (ValueError, TypeError): val_out = val_s
        else:
            val_out = val_s

        # Make sure target drug index exists
        while len(drugs_raw) <= current_idx:
            drugs_raw.append({})

        drugs_raw[current_idx][key] = val_out

        # Log researcher overrides (auto-fetch keys with real values)
        if key in _OVERRIDABLE_KEYS:
            overrides_log.append(
                f"Drug {current_idx+1}.{key} = {val_out} (researcher override)")

    # ═══════════════════════════════════════════════════════════════════
    #  FILTER + ENRICH EACH DRUG
    # ═══════════════════════════════════════════════════════════════════
    def _detect_molecule_type(mol_in: str) -> dict:
        """Classify molecule_input string → smiles/fasta/helm/pdb_id/name."""
        out = {"smiles": None, "fasta": None, "helm": None, "pdb_id": None}
        if not mol_in: return out
        s = str(mol_in).strip()
        if s.startswith(">"):
            out["fasta"] = s
        elif s.upper().startswith("PEPTIDE"):
            out["helm"] = s
        elif len(s) == 4 and s.isalnum() and not s.isdigit():
            out["pdb_id"] = s
        elif (5 < len(s) < 600 and
              any(c in s for c in "=#()[]@+/\\.-")):
            out["smiles"] = s
        # else: treat as plain name — leave all four None
        return out

    drugs_clean: list = []
    for idx, d in enumerate(drugs_raw, 1):
        if not d.get("name"):
            log.debug(f"[EXCEL→YAML]  Drug {idx} has no name — skipped (empty slot)")
            continue
        if d["name"].strip().lower() == "example":
            # "example" is the literal Excel placeholder text — skip it.
            log.debug(f"[EXCEL→YAML]  Drug {idx} 'example' placeholder skipped")
            continue
        # Detect molecule type from molecule_input
        mol_in = d.get("molecule_input", "") or d.get("smiles", "") or d.get("fasta","")
        mol_types = _detect_molecule_type(mol_in)
        for k, v in mol_types.items():
            if v is not None and not d.get(k):
                d[k] = v
        # Always force fresh fetch for new runs
        d["force_refresh"] = force_refresh
        # Default molecule_class if missing
        d.setdefault("molecule_class", "small_molecule")

        drugs_clean.append(d)

    if not drugs_clean:
        raise ValueError(
            "Excel Sheet 1 must have at least one drug with a Drug Name. "
            "Check that cell B5 (or the corresponding 'Drug Name' row) is filled.")

    log.info(f"[EXCEL→YAML] Drugs detected: {len(drugs_clean)} → "
             f"{[d['name'] for d in drugs_clean]}")
    if overrides_log:
        log.info(f"[EXCEL→YAML] Researcher overrides ({len(overrides_log)}):")
        for ov in overrides_log:
            log.info(f"  • {ov}")

    # ═══════════════════════════════════════════════════════════════════
    #  Sheet 2 — DDS FORMULATIONS  (researcher fills, pipeline scores)
    # ═══════════════════════════════════════════════════════════════════
    if "2_DDS_Formulations" not in wb.sheetnames:
        raise ValueError("Excel must contain '2_DDS_Formulations' sheet")
    ws2 = wb["2_DDS_Formulations"]
    headers = [str(c.value or "").strip() for c in ws2[3]]
    forms: list = []
    RESEARCHER_COLS = set(headers[:21])

    for row in ws2.iter_rows(min_row=4, values_only=True):
        fid = row[0]
        if not fid or str(fid).strip() in ("", "(auto)"):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if i >= len(row): break
            v = row[i]
            if h not in RESEARCHER_COLS: continue
            if v is None or str(v).strip() in ("", "(auto)"):
                rec[h] = None
                continue
            try:    rec[h] = float(v)
            except (ValueError, TypeError): rec[h] = str(v).strip()
        forms.append(rec)

    if not forms:
        raise ValueError(
            "No formulations found in Sheet 2. "
            "Fill at least one row in the yellow columns.")

    # ═══════════════════════════════════════════════════════════════════
    #  Sheet 5 — PIPELINE CONFIG (optional)
    # ═══════════════════════════════════════════════════════════════════
    pipeline_cfg = {
        "run_multi_drug_comparison": True,
        "n_clinical_patients": 500,
        "excursion_temp_C": -20.0,
        "excursion_h": 4.0,
        "generate_html5": True,
        "generate_canvas_videos": True,
        "generate_pdf": True,
    }
    if "5_Pipeline_Config" in wb.sheetnames:
        wsc = wb["5_Pipeline_Config"]
        for row in wsc.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]: continue
            k = str(row[0] or "").strip().lower().replace(" ","_")
            v = str(row[1] or "").strip() if len(row) > 1 else ""
            if not k or not v: continue
            if v.upper() in ("YES","TRUE","1"):  pipeline_cfg[k] = True
            elif v.upper() in ("NO","FALSE","0"): pipeline_cfg[k] = False
            else:
                try:    pipeline_cfg[k] = float(v)
                except: pipeline_cfg[k] = v

    # ═══════════════════════════════════════════════════════════════════
    #  Build FINAL cfg — unified `drugs` + backward-compat keys
    # ═══════════════════════════════════════════════════════════════════
    cfg = {
        "drugs":            drugs_clean,                  # primary container
        "drug":             drugs_clean[0],               # backward compat
        "additional_drugs": drugs_clean[1:],              # backward compat
        "formulations":     forms,
        "pipeline_config":  pipeline_cfg,
    }

    log.info(f"[EXCEL→YAML] Drug 1: {drugs_clean[0]['name']} | "
             f"Formulations: {len(forms)} | "
             f"SMILES: {'yes' if drugs_clean[0].get('smiles') else 'no'}")

    # Write YAML
    output_yaml.parent.mkdir(parents=True, exist_ok=True)
    with open(output_yaml, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True,
                  default_flow_style=False, sort_keys=False)

    log.info(f"[EXCEL→YAML] Wrote: {output_yaml}")
    return cfg


# ─────────────────────────────────────────────────────────────────────────────
# CACHE INVALIDATOR  (wipes molecule engine cache for new drug names)
# ─────────────────────────────────────────────────────────────────────────────

def invalidate_molecule_cache(drug_names: list, trial_dir: Path) -> None:
    """
    Wipe any cached molecule profiles for the given drug names.
    The molecule engine caches to:
      - In-memory dict (cleared automatically per process)
      - outputs/molecule_cache/*.json  (persistent)
      - SQLite drug_records table (force upsert, not read)

    After this call, analyze_molecule() will fetch fresh from APIs.
    """
    # 1. Persistent JSON cache
    for cache_dir in [RESULTS_ROOT / "molecule_cache",
                      SCRIPT_DIR / "molecule_cache"]:
        if cache_dir.exists():
            for drug in drug_names:
                safe = drug.lower().replace(" ", "_").replace("/", "_")
                for f in cache_dir.glob(f"{safe}*.json"):
                    f.unlink()
                    log.info(f"  [CACHE] Deleted: {f.name}")

    # 2. SQLite drug_records — delete old rows for these drugs so Upsert is fresh.
    # The real DB file is "cerebro_knowledge.db" (src/core/pipeline.py's
    # DB_PATH = OUTPUT_ROOT / "cerebro_knowledge.db") -- this used to check
    # for a "cerebro.db" that has never existed anywhere in this project, so
    # db_path.exists() was always False and this whole deletion step was a
    # silent no-op on every trial, contradicting run.py's own documented
    # guarantee ("This guarantees fresh API fetch every time — no stale
    # data"). Verified directly: outputs/cerebro_knowledge.db exists;
    # outputs/cerebro.db does not.
    db_candidates = [
        RESULTS_ROOT / "cerebro_knowledge.db",
        SCRIPT_DIR / "cerebro_knowledge.db",
    ]
    for db_path in db_candidates:
        if db_path.exists():
            try:
                conn = sqlite3.connect(db_path)
                for drug in drug_names:
                    n = conn.execute(
                        "DELETE FROM drug_records WHERE LOWER(drug_name)=LOWER(?)",
                        (drug,)).rowcount
                    if n:
                        log.info(f"  [CACHE] DB: deleted {n} old row(s) for '{drug}'")
                conn.commit()
                conn.close()
            except Exception as e:
                log.debug(f"  [CACHE] DB cleanup: {e}")

    log.info(f"  [CACHE] Invalidated for: {drug_names}")
