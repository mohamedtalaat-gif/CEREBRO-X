"""
================================================================================
CEREBRO-X |  cerebro_completed_excel_writer.py
================================================================================
Created by: Muhammad Talaat (BPharm, R&D Computational Lead) — CEREBRO-X

PURPOSE
═══════
After the pipeline finishes processing a drug (single or multi-drug run),
this engine emits a "completed" Excel workbook that contains ALL data the
researcher left blank — auto-fetched, computed, predicted, or class-typical —
with full provenance for every value.

The researcher submits a sparse Excel (just SMILES + DDS list).
The pipeline returns a fully populated Excel they can:
  • Inspect to verify what was found vs. predicted
  • Override any Tier-6 (class-mean) value with an in-vitro measurement
  • Re-feed into the pipeline (overrides become Tier-0 highest confidence)
  • Cite in publications (every value has source + DOI when available)

OUTPUT STRUCTURE (per pipeline run)
═══════════════════════════════════
Workbook: CEREBRO_X_Completed_Data_<DrugName>.xlsx  (per drug)
   OR     CEREBRO_X_Completed_Data_All_Drugs.xlsx   (combined for multi-drug)

Sheets:
  1. Overview         — summary table: drugs × tier-coverage counts
  2. Drug_<N>_Properties — one row per resolved property:
         Property | Value | Unit | Tier | Confidence% | Source | Reference |
         DOI | Disclaimer | Overridable | Researcher_Override_Cell
  3. Drug_<N>_Principles — flat key:value table for all 62-principle results
  4. Drug_<N>_DDS_Top10 — top 10 DDS formulations ranked by BBB score
  5. Audit_Trail      — full provenance log (which tier resolved what, when)

Cell colour-coding by Tier:
  • Green   (#C6EFCE) — Tier 0 (researcher) / Tier 1 (live API) / Tier 2 (library)
  • Yellow  (#FFEB9C) — Tier 3 (PubMed) / Tier 4 (RDKit)
  • Orange  (#FFC7CE) — Tier 5 (analog) / Tier 6 (class-mean — needs override)
  • Grey    (#A0A0A0) — Tier 99 (truly unknown)

DESIGN PRINCIPLES
═════════════════
  • NEVER hides Tier-6 estimates — they're flagged loudly with a disclaimer
  • Researcher overrides preserved as the highest-confidence column
  • Output is editable: any Tier-6 cell can be replaced with measured value
  • Single-drug run still emits this file (with one Drug_1 sheet)
================================================================================
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

log = logging.getLogger("CEREBRO-COMPLETED-WRITER")

# ──────────────────────────────────────────────────────────────────────────
# Tier presentation tables
# ──────────────────────────────────────────────────────────────────────────
_TIER_FILL = {
    0:  "C6EFCE",   # Researcher override
    1:  "C6EFCE",   # Live API
    2:  "C6EFCE",   # Embedded library
    3:  "FFEB9C",   # PubMed
    4:  "FFEB9C",   # RDKit
    5:  "FFC7CE",   # Analog matching
    6:  "FFC7CE",   # Class-mean fallback (needs override)
    98: "D9D9D9",   # Present but unaudited -- provenance genuinely unknown
    99: "A0A0A0",   # Unknown
}
_TIER_LABELS = {
    0: "Researcher",  1: "Live API",   2: "Library",
    3: "PubMed",      4: "RDKit",      5: "Analog",
    6: "Class-mean",  98: "Unaudited", 99: "Unknown",
}

# Property catalog: every field the pipeline tries to resolve.
# Order is preservation-friendly for the Excel sheet.
# (canonical_key, display_name, unit, expected_class)   class=None means any
_PROPERTY_CATALOG = [
    # ── Identity ──────────────────────────────────────────────────
    ("name",              "Drug Name",                 "",          None),
    ("molecule_class",    "Molecule Class",            "",          None),
    ("molecule_input",    "Molecule Input (raw)",      "",          None),
    ("smiles",            "SMILES",                    "",          None),
    ("fasta",             "FASTA sequence",            "",          "biologic"),
    ("helm",              "HELM notation",             "",          "peptide"),
    ("pdb_id",            "Reference PDB ID",          "",          None),
    ("indication",        "Indication / Disease",      "",          None),
    ("target_protein",    "Target Protein",            "",          None),
    ("target_pdb_id",     "Target PDB ID",             "",          None),
    ("clinical_phase",    "Clinical Phase",            "",          None),
    ("fda_approval_date", "FDA Approval Date",         "",          None),
    # ── Physicochemical (resolver-managed) ────────────────────────
    ("MW_Da",             "Molecular Weight",          "Da",        None),
    ("LogP",              "LogP",                      "",          None),
    ("TPSA_A2",           "TPSA",                      "Å²",        "small_molecule"),
    ("HBD",               "H-Bond Donors",             "count",     "small_molecule"),
    ("HBA",               "H-Bond Acceptors",          "count",     "small_molecule"),
    ("RotBonds",          "Rotatable Bonds",           "count",     "small_molecule"),
    ("pI",                "Isoelectric Point",         "pH",        "biologic"),
    ("Instability_Index", "Instability Index",         "",          "biologic"),
    ("UniProt_ID",        "UniProt ID",                "",          "biologic"),
    # ── PK / CNS ───────────────────────────────────────────────────
    ("Half_Life_Days",    "Half-life",                 "days",      None),
    ("BBB_permeability_pct", "BBB Penetration (native)","%",        None),
    ("LogBB",             "LogBB",                     "",          None),
    ("Docking_Affinity_kcal","Docking Affinity",       "kcal/mol",  None),
]

# ──────────────────────────────────────────────────────────────────────────
# Data extraction helpers
# ──────────────────────────────────────────────────────────────────────────
def _property_applies(expected_class: str | None, molecule_class: str | None) -> bool:
    """True if a _PROPERTY_CATALOG row's expected_class matches this drug's
    molecule_class (None means the property applies to every class). Shared
    by the Overview tier-coverage table and the per-drug Properties sheet so
    both count the same "applicable" property set -- they previously used
    two independent copies of this check that could (and did) drift apart,
    with the Overview counting inapplicable properties as "T99 Unknown"
    that the per-drug sheet correctly never showed in the first place.
    """
    if not expected_class:
        return True
    mc = (molecule_class or "small_molecule").lower()
    return expected_class == mc or (
        expected_class == "biologic" and mc in ("biologic", "monoclonal_antibody", "antibody", "mab"))


def _get_tier_info(mol_profile: dict, prop_key: str) -> dict:
    """
    Extract tier / confidence / source for a property from mol_profile.

    The molecule engine populates `_source_audit` with one entry per
    property the resolver touched. For properties resolved at Tier 1
    by the engine itself (no resolver call), we infer Tier=1 if the value
    exists and isn't None/0.
    """
    audit = mol_profile.get("_source_audit", {}) or {}
    if prop_key in audit and isinstance(audit[prop_key], dict):
        return audit[prop_key]
    # A value present with no _source_audit entry does NOT mean it was
    # resolved via a live API call -- _source_audit is only populated for
    # the handful of properties the resolver actually touches (MW_Da,
    # LogP, Half_Life_Days, TPSA_A2, HBD, HBA -- see molecule_engine.py).
    # Every other catalog property that happens to have a value here
    # could have come from anywhere (Excel input, an upstream default,
    # a computation with no audit hook). Claiming "Tier 1 / Live API /
    # 90% confidence" for it is a fabricated provenance claim -- exactly
    # what this workbook's own stated mission (full, citable provenance)
    # forbids. Label it honestly as present-but-unaudited instead.
    val = mol_profile.get(prop_key)
    if val is not None and val != 0 and val != "":
        return {
            "_tier": 98,
            "_confidence_score": None,
            "_confidence": "UNKNOWN — value present, provenance not audited",
            "_source": "unaudited (no _source_audit entry)",
            "_reference": "",
            "_doi": None, "_disclaimer": None, "_overridable": False,
        }
    return {
        "_tier": 99,
        "_confidence_score": 0,
        "_confidence": "NONE — not resolved",
        "_source": "not_attempted_or_failed",
        "_reference": "", "_doi": None, "_disclaimer": None, "_overridable": False,
    }


def _flatten_principles(principles: dict, prefix: str = "") -> list[tuple]:
    """
    Flatten a nested principles dict into [(key_path, value), ...] pairs.
    Skip non-numeric, non-string scalars (lists, complex objects → skipped).
    """
    out: list[tuple] = []
    if not isinstance(principles, dict):
        return out
    for k, v in principles.items():
        if isinstance(k, str) and k.startswith("_"):
            continue   # internal metadata fields
        path = f"{prefix}.{k}" if prefix else str(k)
        if isinstance(v, dict):
            out.extend(_flatten_principles(v, prefix=path))
        elif isinstance(v, (int, float, str, bool)) and not isinstance(v, bool):
            out.append((path, v))
        elif isinstance(v, bool):
            out.append((path, "Yes" if v else "No"))
        # skip lists, None, complex objects
    return out


# ──────────────────────────────────────────────────────────────────────────
# Excel writer
# ──────────────────────────────────────────────────────────────────────────
def write_completed_excel(drug_results: list[dict],
                           output_path: Path,
                           pipeline_metadata: dict | None = None) -> Path:
    """
    Emit a completed-data Excel workbook for a list of processed drug results.

    Each entry in `drug_results` should be a dict with keys:
      drug_name   (str)
      mol_profile (dict)        — molecule engine output, with _source_audit
      df_dds      (pd.DataFrame|None)
      principles  (dict)        — science_modules + advanced_modules merged
      trial_dir   (Path)        — per-drug trial dir (informational only)

    Returns: the output Path written.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    pipeline_metadata = pipeline_metadata or {}

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 1 — OVERVIEW (summary across drugs)
    # ═══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Overview"

    ws["A1"] = "CEREBRO-X — Completed Data Workbook"
    ws["A1"].font = Font(bold=True, size=16, color="0f2040")
    ws["A2"] = (f"Generated: {datetime.utcnow().isoformat(timespec='seconds')}Z  "
                f"| Pipeline: {pipeline_metadata.get('version','?')}  "
                f"| Source: {Path(pipeline_metadata.get('source_excel','?')).name}")
    ws["A3"] = (f"Drugs processed: {len(drug_results)}  "
                f"| Class-mean (Tier 6) fallbacks should be reviewed and "
                f"overridden with in-vitro values where available.")
    ws["A3"].font = Font(italic=True, color="9CA3AF")

    # Drug × Tier-coverage table
    ws["A5"] = "Drug Tier-Coverage Summary"
    ws["A5"].font = Font(bold=True, size=12)
    hdrs = ["#", "Drug", "Total props",
            "T0 Researcher", "T1 API", "T2 Library", "T3 PubMed",
            "T4 RDKit", "T5 Analog", "T6 Class-mean", "T98 Unaudited",
            "T99 Unknown", "Top DDS", "Top BBB Score"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(6, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, dr in enumerate(drug_results, 1):
        mp = dr.get("mol_profile", {}) or {}
        audit = mp.get("_source_audit", {}) or {}
        tier_counts = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0, 6:0, 98:0, 99:0}
        # Only count properties actually applicable to this drug's molecule
        # class -- matching the same filter _write_drug_properties_sheet
        # applies, so "Total props" and the T* breakdown here describe the
        # same property set the per-drug sheet shows, instead of always
        # showing the full 25-row catalog and inflating T99/T98 with
        # properties (e.g. FASTA, pI for a small molecule) that were never
        # applicable in the first place.
        applicable = [pk for pk, _, _, ec in _PROPERTY_CATALOG
                      if _property_applies(ec, mp.get("molecule_class"))]
        for prop_key in applicable:
            info = _get_tier_info(mp, prop_key)
            tier_counts[info.get("_tier", 99)] = tier_counts.get(info.get("_tier", 99), 0) + 1
        # Top DDS info
        df_dds = dr.get("df_dds")
        top_name = top_score = "—"
        if df_dds is not None and not df_dds.empty:
            top = df_dds.iloc[0]
            top_name = str(top.get("Formulation_Name") or top.get("Formulation_ID") or "—")
            top_score = round(float(top.get("BBB_Engineering_Score", 0) or 0), 2)

        row = [i, dr["drug_name"], len(applicable),
               tier_counts[0], tier_counts[1], tier_counts[2], tier_counts[3],
               tier_counts[4], tier_counts[5], tier_counts[6], tier_counts[98],
               tier_counts[99], top_name, top_score]
        for j, v in enumerate(row, 1):
            c = ws.cell(6+i, j, v)
            c.alignment = Alignment(horizontal="center")
        # Highlight any drug with high Tier-6 reliance (>3 class-mean values)
        if tier_counts[6] > 3:
            ws.cell(6+i, 10).fill = PatternFill("solid", fgColor="FFC7CE")
            ws.cell(6+i, 10).font = Font(bold=True)

    # Column widths
    widths = [4, 22, 10, 12, 10, 12, 12, 10, 10, 14, 12, 22, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Tier legend
    legend_row = 6 + len(drug_results) + 3
    ws.cell(legend_row, 1, "Tier Legend").font = Font(bold=True, size=11)
    for ti, (tier, label) in enumerate(_TIER_LABELS.items(), 1):
        c = ws.cell(legend_row+ti, 1, f"Tier {tier}")
        c.fill = PatternFill("solid", fgColor=_TIER_FILL[tier])
        c.alignment = Alignment(horizontal="center")
        ws.cell(legend_row+ti, 2, label)
        ws.cell(legend_row+ti, 3, _TIER_DESCRIPTIONS.get(tier,""))

    # ═══════════════════════════════════════════════════════════════════
    # PER-DRUG SHEETS — Properties + Principles + DDS + DDS-Principle-Matrix
    #                  + v22: Deep Validation + Translational + Fallback
    # ═══════════════════════════════════════════════════════════════════
    for idx, dr in enumerate(drug_results, 1):
        _write_drug_properties_sheet(wb, idx, dr)
        _write_drug_principles_sheet(wb, idx, dr)
        _write_drug_dds_sheet(wb, idx, dr)
        _write_drug_dds_principle_matrix_sheet(wb, idx, dr)
        _write_drug_dds_breakdown_sheet(wb, idx, dr)
        # v22 C+ Flow additions — surface in EVERY output (Muhammad's mandate)
        _write_drug_surrogate_detail_sheet(wb, idx, dr)
        _write_drug_deep_validation_sheet(wb, idx, dr)
        _write_drug_translational_sheet(wb, idx, dr)
        _write_drug_fallback_chain_sheet(wb, idx, dr)

    # ═══════════════════════════════════════════════════════════════════
    # GLOBAL — Principle Explanations (the "62-principle textbook" sheet)
    # ═══════════════════════════════════════════════════════════════════
    _write_principle_explanations_sheet(wb)

    # ═══════════════════════════════════════════════════════════════════
    # Sheet — AUDIT TRAIL (full provenance log)
    # ═══════════════════════════════════════════════════════════════════
    _write_audit_trail_sheet(wb, drug_results)

    wb.save(str(output_path))
    log.info(f"[COMPLETED-WRITER] ✅ {output_path.name} "
             f"({len(drug_results)} drug(s), {len(wb.sheetnames)} sheets)")
    return output_path


_TIER_DESCRIPTIONS = {
    0:  "Researcher in-vitro override (highest confidence)",
    1:  "Live API fetch from ChEMBL/PubChem/UniProt/DrugBank",
    2:  "Embedded clinical literature library (FDA labels + DOIs)",
    3:  "PubMed E-utilities search (citation only — manual extraction)",
    4:  "RDKit physicochemical computation from SMILES",
    5:  "Analog matching to nearest reference drug",
    6:  "CLASS-TYPICAL ESTIMATE — review & override before publication",
    99: "Truly unknown — not used in calculations",
}

# ──────────────────────────────────────────────────────────────────────────
# Per-drug sheet writers
# ──────────────────────────────────────────────────────────────────────────
def _write_drug_properties_sheet(wb, idx: int, dr: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:25]
    sheet_name = f"D{idx}_{safe_name}_Props"[:31]
    ws = wb.create_sheet(sheet_name)
    mp = dr.get("mol_profile", {}) or {}

    ws["A1"] = f"Drug {idx}: {drug_name} — Resolved Properties"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Every value below carries full provenance. Tier 6 (orange) "
                "values are class-mean estimates — replace with in-vitro "
                "measurements where possible.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    hdrs = ["Property", "Value", "Unit", "Tier", "Confidence %",
            "Source", "Reference / DOI", "Disclaimer", "Overridable"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    row = 5
    for prop_key, display_name, unit, expected_class in _PROPERTY_CATALOG:
        # Skip class-specific properties for wrong class
        if not _property_applies(expected_class, mp.get("molecule_class")):
            continue

        val = mp.get(prop_key)
        info = _get_tier_info(mp, prop_key)
        tier = info.get("_tier", 99)
        conf = info.get("_confidence_score", 0)
        source = info.get("_source", "")
        ref = info.get("_reference", "") or ""
        doi = info.get("_doi", "") or ""
        disclaimer = info.get("_disclaimer", "") or info.get("_warning", "") or ""
        overridable = "Yes" if info.get("_overridable") else "No"

        ref_combined = f"{ref}\n{doi}" if doi and doi not in ref else ref

        cells = [
            display_name,
            val if val is not None else "—",
            unit,
            f"T{tier} ({_TIER_LABELS.get(tier,'?')})",
            conf,
            source,
            ref_combined,
            disclaimer,
            overridable,
        ]
        fill = PatternFill("solid", fgColor=_TIER_FILL.get(tier, "FFFFFF"))
        for j, cv in enumerate(cells, 1):
            c = ws.cell(row, j, cv)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if j in (4, 5, 9):
                c.alignment = Alignment(horizontal="center", vertical="top")
            if tier == 6:
                c.font = Font(bold=(j == 1), color="C62828" if j == 8 else "000000")
        row += 1

    # Column widths
    widths = [26, 18, 10, 18, 12, 26, 36, 40, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28


def _write_drug_principles_sheet(wb, idx: int, dr: dict) -> None:
    """
    Flat 62-principle view (v22): every principle from every C+ Flow class
    listed in one sheet — Class A surrogate (all 57 for Top-1), Class B deep
    (all 28 for Top-1), Class C translational (all 5 for Top-1).
    Each row carries: principle ID, class, score, value, method, reference,
    confidence, narrative.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:25]
    sheet_name = f"D{idx}_{safe_name}_Princ"[:31]
    ws = wb.create_sheet(sheet_name)

    matrix    = dr.get("dds_principle_matrix") or []
    deep      = dr.get("deep_results") or {}
    trans     = dr.get("translational") or {}

    ws["A1"] = f"Drug {idx}: {drug_name} — 62-Principle C+ Flow Flat View (Top-1 DDS)"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Every principle from every C+ Flow class for the Top-1 DDS, in "
                "one flat table: Class A surrogate (57), Class B deep physics (28), "
                "Class C translational (5). For cross-DDS surrogate scores see the "
                "DDSxP matrix sheet.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 38

    hdrs = ["Class", "Principle", "Title (en)", "Score", "Value",
            "Confidence", "Method (truncated)", "Reference"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 26

    # Pull canonical titles from catalog
    try:
        from cerebro_62_principles_catalog import PRINCIPLES_62
    except Exception:
        PRINCIPLES_62 = {}

    row = 5
    CLASS_FILL = {"A":"E8F4FD", "B":"FFF4E6", "C":"F4E6FF"}

    # ─── Class A — surrogate scores for Top-1 DDS ───────────────────
    if matrix:
        top1_principles = matrix[0].get("principles", {})
        for pid in sorted(top1_principles.keys()):
            r = top1_principles[pid]
            cat = PRINCIPLES_62.get(pid, {})
            row_fill = PatternFill("solid", fgColor=CLASS_FILL["A"])
            ws.cell(row, 1, "A — Surrogate").fill = row_fill
            ws.cell(row, 2, pid).font = Font(bold=True)
            ws.cell(row, 3, cat.get("title_en", "—"))
            score = r.get("score", 0)
            sc_cell = ws.cell(row, 4, round(score, 2))
            # Color code by score
            if score >= 80: sc_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif score >= 60: sc_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            else: sc_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            sc_cell.alignment = Alignment(horizontal="center")
            ws.cell(row, 5, str(r.get("value",""))[:30])
            ws.cell(row, 6, r.get("confidence","—")).alignment = Alignment(horizontal="center")
            ws.cell(row, 7, str(r.get("method","") or "")[:80]).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 8, str(r.get("reference","") or "")[:60])
            row += 1

    # ─── Class B — deep physics for Top-1 DDS ──────────────────────
    if deep:
        # Section separator
        sep = ws.cell(row, 1, "── CLASS B — DEEP PHYSICS (Top-1) ──────────────")
        sep.font = Font(bold=True, color="FFFFFF")
        sep.fill = PatternFill("solid", fgColor="C9A84C")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        for pid in sorted(deep.keys()):
            r = deep[pid]
            cat = PRINCIPLES_62.get(pid, {})
            row_fill = PatternFill("solid", fgColor=CLASS_FILL["B"])
            ws.cell(row, 1, "B — Deep").fill = row_fill
            ws.cell(row, 2, pid).font = Font(bold=True)
            ws.cell(row, 3, cat.get("title_en", "—"))
            score = r.get("score", 0)
            sc_cell = ws.cell(row, 4, round(score, 2))
            sc_cell.alignment = Alignment(horizontal="center")
            if r.get("validated"):
                sc_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            else:
                sc_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            ws.cell(row, 5, str(r.get("value",""))[:30])
            ws.cell(row, 6, r.get("confidence","—")).alignment = Alignment(horizontal="center")
            ws.cell(row, 7, str(r.get("method","") or "")[:80]).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 8, str(r.get("reference","") or "")[:60])
            row += 1

    # ─── Class C — translational for Top-1 DDS ─────────────────────
    if trans:
        sep = ws.cell(row, 1, "── CLASS C — TRANSLATIONAL (Top-1) ────────────")
        sep.font = Font(bold=True, color="FFFFFF")
        sep.fill = PatternFill("solid", fgColor="7C4DFF")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        for pid in sorted(trans.keys()):
            t = trans[pid]
            cat = PRINCIPLES_62.get(pid, {})
            row_fill = PatternFill("solid", fgColor=CLASS_FILL["C"])
            ws.cell(row, 1, "C — Translational").fill = row_fill
            ws.cell(row, 2, pid).font = Font(bold=True)
            ws.cell(row, 3, cat.get("title_en", t.get("title","—")))
            score = (t.get("compliance_score") or t.get("fto_score")
                      or t.get("patentability_score") or "—")
            ws.cell(row, 4, score).alignment = Alignment(horizontal="center")
            ws.cell(row, 5, t.get("status","—"))
            ws.cell(row, 6, "—").alignment = Alignment(horizontal="center")
            ws.cell(row, 7, str(t.get("narrative","") or "")[:80]).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 8, str(cat.get("reference","") or "")[:60])
            row += 1

    if row == 5:
        ws.cell(5, 1, "(no principle results — orchestrator did not run)")
        ws.cell(5, 1).font = Font(italic=True, color="C62828")

    widths = [16, 8, 38, 9, 22, 11, 40, 28]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


def _write_drug_dds_sheet(wb, idx: int, dr: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:25]
    sheet_name = f"D{idx}_{safe_name}_DDS"[:31]
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = f"Drug {idx}: {drug_name} — Top 10 DDS Formulations"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")

    df_dds = dr.get("df_dds")
    if df_dds is None or df_dds.empty:
        ws["A3"] = "(no DDS results)"
        ws["A3"].font = Font(italic=True, color="C62828")
        return

    top10 = df_dds.head(10)
    cols = list(top10.columns)
    # Header
    for j, col in enumerate(cols, 1):
        c = ws.cell(3, j, col)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    # Data rows
    for i, (_, drow) in enumerate(top10.iterrows(), 4):
        for j, col in enumerate(cols, 1):
            v = drow.get(col)
            if v is None:
                v = "—"
            elif isinstance(v, float):
                v = round(v, 4)
            ws.cell(i, j, v)
        if i == 4:   # highlight #1 row
            for j in range(1, len(cols)+1):
                ws.cell(i, j).fill = PatternFill("solid", fgColor="C6EFCE")
                ws.cell(i, j).font = Font(bold=True)

    # Column widths
    for j, col in enumerate(cols, 1):
        w = max(12, min(28, len(str(col)) + 2))
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_audit_trail_sheet(wb, drug_results: list[dict]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Audit_Trail")
    ws["A1"] = "Provenance Audit Trail"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("One row per (drug, property) showing exactly which tier "
                "resolved the value and what source was used.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    hdrs = ["Drug", "Property", "Tier", "Confidence %",
            "Source", "Reference", "DOI / URL"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")

    row = 5
    for dr in drug_results:
        mp = dr.get("mol_profile", {}) or {}
        for prop_key, display_name, _, _ in _PROPERTY_CATALOG:
            info = _get_tier_info(mp, prop_key)
            tier = info.get("_tier", 99)
            ws.cell(row, 1, dr["drug_name"])
            ws.cell(row, 2, display_name)
            t_cell = ws.cell(row, 3, f"T{tier} ({_TIER_LABELS.get(tier,'?')})")
            t_cell.fill = PatternFill("solid", fgColor=_TIER_FILL.get(tier, "FFFFFF"))
            ws.cell(row, 4, info.get("_confidence_score", 0))
            ws.cell(row, 5, info.get("_source", "") or "")
            ws.cell(row, 6, info.get("_reference", "") or "")
            ws.cell(row, 7, info.get("_doi", "") or "")
            row += 1

    from openpyxl.utils import get_column_letter
    widths = [22, 28, 18, 12, 26, 38, 28]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ──────────────────────────────────────────────────────────────────────────
# v21 ADDITIONS — Per-DDS principle matrix + breakdown narrative + textbook
# ──────────────────────────────────────────────────────────────────────────
def _write_drug_dds_principle_matrix_sheet(wb, idx: int, dr: dict) -> None:
    """
    Full DDS × Principle matrix sheet.

    Rows: every DDS in the drug's formulation list (ranked by composite)
    Cols: every principle (P01, P02, …, P62)
    Values: per-principle score (0-100) for that DDS

    Plus a Composite column (weighted total) and group columns G1..G7.
    Cell color-graded: green (>=80), yellow-green (60-80), yellow (40-60),
    orange (20-40), red (<20).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:18]
    sheet_name = f"D{idx}_{safe_name}_DDSxP"[:31]
    ws = wb.create_sheet(sheet_name)

    matrix = dr.get("dds_principle_matrix") or []
    if not matrix:
        ws["A1"] = (f"Drug {idx}: {drug_name} — DDS × Principle Matrix")
        ws["A1"].font = Font(bold=True, size=14, color="0f2040")
        ws["A3"] = ("(no per-DDS principle data — pipeline did not produce "
                    "dds_principle_matrix; see DDS_Top10 sheet for ranking)")
        ws["A3"].font = Font(italic=True, color="C62828")
        return

    ws["A1"] = (f"Drug {idx}: {drug_name} — Full DDS × Principle Score Matrix")
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = (f"All {len(matrix)} formulations from the Excel ranked "
                f"top-to-bottom by CNS-weighted composite (57 Class-A "
                f"principles). Hover any cell for the score; refer to "
                f"Principle_Explanations sheet for what each Pnn means "
                f"and how it's computed.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    # Build column list: sorted principle keys from first DDS
    sample_principles = sorted(matrix[0]["principles"].keys())
    # Must match cerebro_62_orchestrator.PRINCIPLE_GROUPS exactly — these
    # names key into m["groups"] below, which the orchestrator built with
    # its own group names ("..._Kinetics", "..._BBB"). The old names here
    # ("G2_Release", "G5_Glymphatic") never matched, so those two columns
    # silently showed 0 for every DDS row regardless of the real score.
    group_cols = ["G1_CNS_Delivery","G2_Release_Kinetics","G3_Stability","G4_Safety",
                   "G5_Glymphatic_BBB","G6_Manufacturability","G7_DrugDDS_Fit"]
    hdrs = ["Rank", "DDS Name", "Composite"] + group_cols + sample_principles

    # Header row
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True,
                                 vertical="center")
    ws.row_dimensions[4].height = 36

    # Data rows
    def _grade_fill(val: float) -> str:
        if val >= 80: return "C6EFCE"   # green
        if val >= 60: return "DDEBCB"
        if val >= 40: return "FFEB9C"   # yellow
        if val >= 20: return "FFC7CE"   # orange
        return "F2A1A1"                  # red

    for i, m in enumerate(matrix, 1):
        ws.cell(4+i, 1, i)
        ws.cell(4+i, 2, m["dds_name"])
        comp_cell = ws.cell(4+i, 3, m["composite"])
        comp_cell.fill = PatternFill("solid", fgColor=_grade_fill(m["composite"]))
        comp_cell.font = Font(bold=True)
        # Group cols
        for jc, gk in enumerate(group_cols, 4):
            v = m["groups"].get(gk, 0)
            cell = ws.cell(4+i, jc, v)
            cell.fill = PatternFill("solid", fgColor=_grade_fill(v))
        # Per-principle cols
        for jp, pid in enumerate(sample_principles, 4 + len(group_cols)):
            pdata = m["principles"].get(pid, {})
            v = pdata.get("score", 0)
            cell = ws.cell(4+i, jp, v)
            cell.fill = PatternFill("solid", fgColor=_grade_fill(v))
            # Add tooltip via cell comment
            method = pdata.get("method", "")
            ref = pdata.get("reference", "")
            conf = pdata.get("confidence", "")
            if method or ref:
                from openpyxl.comments import Comment
                comment_text = (f"Principle: {pid}\n"
                                f"Method: {method}\n"
                                f"Reference: {ref}\n"
                                f"Confidence: {conf}")
                cell.comment = Comment(comment_text[:1500], "CEREBRO-X")

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 6
    ws.column_dimensions[get_column_letter(2)].width = 24
    ws.column_dimensions[get_column_letter(3)].width = 12
    for j in range(4, 4 + len(group_cols)):
        ws.column_dimensions[get_column_letter(j)].width = 14
    for j in range(4 + len(group_cols), 4 + len(group_cols) + len(sample_principles)):
        ws.column_dimensions[get_column_letter(j)].width = 8

    # Freeze panes: keep header + DDS name column visible while scrolling
    ws.freeze_panes = "C5"


def _write_drug_dds_breakdown_sheet(wb, idx: int, dr: dict) -> None:
    """
    Top-10 DDS narrative breakdown sheet.

    For each of the top-10 DDS, shows:
      • Full reasoning narrative
      • Top 3 strengths with explanation
      • Weak spots with improvement hints
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:18]
    sheet_name = f"D{idx}_{safe_name}_Reasoning"[:31]
    ws = wb.create_sheet(sheet_name)

    breakdowns = dr.get("dds_principle_breakdown") or []
    if not breakdowns:
        ws["A1"] = f"Drug {idx}: {drug_name} — DDS Reasoning"
        ws["A1"].font = Font(bold=True, size=14, color="0f2040")
        ws["A3"] = "(no breakdown data)"
        return

    ws["A1"] = f"Drug {idx}: {drug_name} — Top-10 DDS Reasoning & Provenance"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Why each DDS ranked where it did. Top 3 strengths and "
                "weak spots are listed with the underlying principle "
                "explanation, so any decision based on this ranking is "
                "fully traceable to first-principles reasoning.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    row = 4
    for r, b in enumerate(breakdowns[:10], 1):
        # Section header for this DDS
        c = ws.cell(row, 1, f"#{r}")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center")
        ws.cell(row, 2, b["dds_name"]).font = Font(bold=True, size=11)
        score_cell = ws.cell(row, 3, f"{b['composite']:.1f} / 100")
        score_cell.font = Font(bold=True, size=11)
        verdict_cell = ws.cell(row, 4, b["verdict"])
        verdict_cell.font = Font(bold=True)
        VERDICT_COLOR = {"EXCELLENT":"C6EFCE","GOOD":"DDEBCB",
                          "ACCEPTABLE":"FFEB9C","MARGINAL":"FFC7CE","POOR":"F2A1A1"}
        verdict_cell.fill = PatternFill("solid",
                                          fgColor=VERDICT_COLOR.get(b["verdict"],"FFFFFF"))
        row += 1

        # Narrative
        ws.cell(row, 1, "Narrative:").font = Font(italic=True, color="9CA3AF")
        n_cell = ws.cell(row, 2, b["narrative"])
        n_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 36
        row += 1

        # Group scores
        ws.cell(row, 1, "Group scores:").font = Font(italic=True, color="9CA3AF")
        gtxt = "  ".join(f"{g.replace('_Score','')}: {s:.0f}/100"
                          for g, s in b["group_scores"].items())
        ws.cell(row, 2, gtxt)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        row += 1

        # Top 3 strengths
        ws.cell(row, 1, "Strengths:").font = Font(italic=True, color="2E7D32")
        for s in b["top_strengths"]:
            ws.cell(row, 2, s["principle"])
            ws.cell(row, 3, f"{s['score']:.1f}/100")
            ws.cell(row, 4, (s.get("explanation","") or "")[:200])
            ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
            row += 1

        # Weak spots
        if b["weak_spots"]:
            ws.cell(row, 1, "Weak spots:").font = Font(italic=True, color="C62828")
            for w in b["weak_spots"]:
                ws.cell(row, 2, w["principle"])
                ws.cell(row, 3, f"{w['score']:.1f}/100")
                txt = (w.get("explanation","") or "")[:180]
                if w.get("improvement_hint"):
                    txt += f" — Hint: {w['improvement_hint'][:80]}"
                ws.cell(row, 4, txt)
                ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
                row += 1
        row += 1   # blank separator

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    for col in "DEFGH":
        ws.column_dimensions[col].width = 14


def _write_principle_explanations_sheet(wb) -> None:
    """
    The "62-principle textbook" sheet — one row per principle with:
      ID | Group | Class | Weight | Explanation | Computational Method | Reference

    Researcher can refer to this any time they're confused about a metric.

    Reads from the live cerebro_62_principles_catalog (62 principles) +
    cerebro_62_orchestrator.PRINCIPLE_GROUPS — the same sources that
    actually produce the scores shown in the DDS×Principle matrix sheets
    elsewhere in this workbook. This sheet used to be built from
    cerebro_dds_principle_evaluator's old 25-principle v21 tables, which
    predate the current 62-principle system: a researcher would see
    "P01".."P62" everywhere else in the workbook, then land on this
    glossary and find 25 unrelated "P1.1_..."-style IDs that don't match
    anything else in the file — a source-of-truth sheet documenting a
    different system than the one that actually ran.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    try:
        from cerebro_62_principles_catalog import PRINCIPLES_62
        from cerebro_62_orchestrator import PRINCIPLE_GROUPS
    except ImportError:
        log.debug("[COMPLETED-WRITER] principles catalog not importable — "
                  "skipping principle-explanation sheet")
        return

    pid_to_group = {pid: g for g, pids in PRINCIPLE_GROUPS.items() for pid in pids}

    ws = wb.create_sheet("Principle_Explanations")
    ws["A1"] = "CEREBRO-X — Principle Reference Sheet"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Every metric in the DDS×Principle matrix sheets is defined "
                "below with its computational method, scientific reference, "
                "and CNS weight. Use this as the source-of-truth glossary "
                "for any decision based on CEREBRO-X output.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    hdrs = ["ID", "Group", "Class", "Weight",
            "Explanation", "Computational Method", "Reference"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    # Sort principles by ID for consistency
    for i, pid in enumerate(sorted(PRINCIPLES_62.keys()), 5):
        doc = PRINCIPLES_62[pid]
        group = pid_to_group.get(pid, "")
        weight = doc.get("weight_cns", 0)
        ws.cell(i, 1, pid)
        ws.cell(i, 2, group.replace("_", " "))
        ws.cell(i, 3, doc.get("class", ""))
        ws.cell(i, 4, f"{weight*100:.1f}%")
        e = ws.cell(i, 5, doc.get("title_en", ""))
        e.alignment = Alignment(wrap_text=True, vertical="top")
        m = ws.cell(i, 6, doc.get("method_surrogate", ""))
        m.alignment = Alignment(wrap_text=True, vertical="top")
        r = ws.cell(i, 7, doc.get("reference", ""))
        r.alignment = Alignment(wrap_text=True, vertical="top")
        # Color group rows
        GROUP_COLOR = {
            "G1_CNS_Delivery":      "DDEBCB",
            "G2_Release_Kinetics":  "FFEB9C",
            "G3_Stability":         "D9D9D9",
            "G4_Safety":            "FFD7B5",
            "G5_Glymphatic_BBB":    "C2D6F0",
            "G6_Manufacturability": "F0E5C2",
            "G7_DrugDDS_Fit":       "E5DAF2",
            "G8_Translational":     "F2E5DA",
        }
        clr = GROUP_COLOR.get(group, "FFFFFF")
        for j in (1,2,3,4):
            ws.cell(i, j).fill = PatternFill("solid", fgColor=clr)
        ws.row_dimensions[i].height = 56

    # Column widths
    widths = [8, 22, 16, 8, 50, 46, 38]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ──────────────────────────────────────────────────────────────────────────
# v22 ADDITIONS — Deep Validation + Translational + Fallback Chain sheets
# (Per Muhammad's instruction: ALL results visible in every output)
# ──────────────────────────────────────────────────────────────────────────
def _write_drug_deep_validation_sheet(wb, idx: int, dr: dict) -> None:
    """
    Class B Deep Physics validation results for the Top-1 DDS of this drug.
    Shows surrogate vs deep score per principle, with validation verdict.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:18]
    sheet_name = f"D{idx}_{safe_name}_Deep"[:31]
    ws = wb.create_sheet(sheet_name)

    deep = dr.get("deep_results") or {}
    summary = dr.get("deep_summary") or {}

    ws["A1"] = f"Drug {idx}: {drug_name} — Class B Deep Physics Validation (Top-1)"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Each Class B principle re-runs the Top-1 DDS through full "
                "physics (PBPK ODE / FEP+ / Stokes-Einstein / etc.). "
                "Validated=True means the principle's deep result confirms "
                "the surrogate-stage decision.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    if not deep:
        ws["A4"] = "(no deep validation data)"
        ws["A4"].font = Font(italic=True, color="C62828")
        return

    # Summary box
    ws["A4"] = "DEEP VALIDATION SUMMARY"
    ws["A4"].font = Font(bold=True, size=12)
    SUMMARY_FILL = {"PASSED":"C6EFCE","MARGINAL":"FFEB9C",
                     "FAILED":"FFC7CE","NO DATA":"D9D9D9","NOT RUN":"D9D9D9"}
    verdict = summary.get("verdict","NOT RUN")
    fill = PatternFill("solid", fgColor=SUMMARY_FILL.get(verdict,"FFFFFF"))
    ws["A5"] = "Verdict:"
    ws["B5"] = verdict
    ws["B5"].fill = fill
    ws["B5"].font = Font(bold=True)
    ws["A6"] = "Principles passed:"
    ws["B6"] = f"{summary.get('passed_count',0)}/{summary.get('total',0)} ({summary.get('pct',0)}%)"
    ws["A7"] = "Narrative:"
    ws["B7"] = summary.get("narrative","")
    ws["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=8)
    ws.row_dimensions[7].height = 30

    # Per-principle table
    hdrs = ["Principle ID","Validated","Deep Score","Deep Value",
            "Method","Reference","Confidence","Improvement Over Surrogate","Narrative"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(9, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[9].height = 30

    row = 10
    for pid, r in sorted(deep.items()):
        ws.cell(row, 1, pid).font = Font(bold=True)
        v_cell = ws.cell(row, 2, "✅ Yes" if r.get("validated") else "❌ No")
        v_cell.fill = PatternFill("solid",
            fgColor=("C6EFCE" if r.get("validated") else "FFC7CE"))
        v_cell.alignment = Alignment(horizontal="center")
        ws.cell(row, 3, r.get("score", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row, 4, r.get("value", ""))
        ws.cell(row, 5, (r.get("method","") or "")[:200]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 6, r.get("reference",""))
        ws.cell(row, 7, r.get("confidence",""))
        ws.cell(row, 8, (r.get("improvement_over_surrogate","") or "")[:200]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 9, (r.get("narrative","") or "")[:200]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 50
        row += 1

    widths = [12, 10, 11, 14, 36, 30, 12, 32, 36]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_drug_translational_sheet(wb, idx: int, dr: dict) -> None:
    """
    Class C Translational deliverables (Pre-IND, FTO, Compliance, Grants, Patents).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:18]
    sheet_name = f"D{idx}_{safe_name}_Trans"[:31]
    ws = wb.create_sheet(sheet_name)

    trans = dr.get("translational") or {}

    ws["A1"] = f"Drug {idx}: {drug_name} — Class C Translational Deliverables"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Translational outputs (Pre-IND outline, FTO analysis, "
                "compliance audit, grant outline, patentability score) "
                "for the validated Top-1 DDS. Word/PDF generation deferred "
                "to v23 — current outputs are structured JSON/scores.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    if not trans:
        ws["A4"] = "(no translational data — Top-1 may have failed deep validation)"
        ws["A4"].font = Font(italic=True, color="C62828")
        return

    row = 4
    for pid, t in sorted(trans.items()):
        # Section header per principle
        c = ws.cell(row, 1, f"{pid}: {t.get('title','')}")
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        row += 1

        # Status
        status = t.get("status","")
        STATUS_COLOR = {"structured_outline_ready":"C6EFCE",
                          "search_queries_prepared":"DDEBCB",
                          "self_assessment_completed":"DDEBCB",
                          "scored":"C6EFCE",
                          "skipped_deep_validation_insufficient":"FFC7CE",
                          "failed":"F2A1A1"}
        ws.cell(row, 1, "Status:")
        s_cell = ws.cell(row, 2, status)
        s_cell.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(status,"FFFFFF"))
        s_cell.font = Font(bold=True)
        row += 1

        # Narrative
        ws.cell(row, 1, "Narrative:").font = Font(italic=True, color="9CA3AF")
        n_cell = ws.cell(row, 2, t.get("narrative",""))
        n_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 36
        row += 1

        # Specific score-like fields
        for k in ("compliance_score","fto_score","patentability_score",
                  "encumbrance_level","recommendation"):
            if k in t:
                ws.cell(row, 1, k.replace("_"," ").title()+":").font = Font(italic=True)
                ws.cell(row, 2, t[k])
                row += 1

        # Sections (if Pre-IND or Grant)
        if t.get("sections"):
            ws.cell(row, 1, "Sections:").font = Font(italic=True, color="9CA3AF")
            row += 1
            for sec in t["sections"]:
                ws.cell(row, 2, sec.get("heading",""))
                ready_cell = ws.cell(row, 3, "✅" if sec.get("ready") else "⏳")
                ready_cell.alignment = Alignment(horizontal="center")
                ws.cell(row, 4, (sec.get("content_summary","") or "")[:120]).alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

        # Search queries (FTO)
        if t.get("search_queries"):
            ws.cell(row, 1, "Search Queries:").font = Font(italic=True, color="9CA3AF")
            row += 1
            for q in t["search_queries"]:
                ws.cell(row, 2, q)
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                row += 1

        # Components (Patentability / Compliance)
        if t.get("components"):
            ws.cell(row, 1, "Components:").font = Font(italic=True, color="9CA3AF")
            row += 1
            for comp_k, comp_v in t["components"].items():
                ws.cell(row, 2, comp_k); ws.cell(row, 3, comp_v)
                row += 1
        if t.get("features"):
            ws.cell(row, 1, "Features:").font = Font(italic=True, color="9CA3AF")
            row += 1
            for k, ok in t["features"].items():
                ws.cell(row, 2, k.replace("_"," "))
                ws.cell(row, 3, "✅" if ok else "❌")
                row += 1

        # v23 note
        if t.get("v23_note"):
            note_cell = ws.cell(row, 1, f"📌 {t['v23_note']}")
            note_cell.font = Font(italic=True, size=9, color="C9A84C")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

        row += 1   # blank separator

    widths = [22, 36, 16, 50]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_drug_fallback_chain_sheet(wb, idx: int, dr: dict) -> None:
    """Documents which DDS were tried for deep validation, with reasons."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:18]
    sheet_name = f"D{idx}_{safe_name}_Fallback"[:31]
    ws = wb.create_sheet(sheet_name)

    chain = dr.get("fallback_chain") or []
    ws["A1"] = f"Drug {idx}: {drug_name} — Top-N Deep Validation Fallback Chain"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Per the C+ Flow: if the Top-1 DDS fails Class B deep physics, "
                "the orchestrator falls back to Top-2, Top-3, etc. This sheet "
                "documents the FULL audit trail — for each candidate tried: "
                "the failure reason (which deep principles failed) and the "
                "transition reason (why we moved to the next candidate).")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 50

    if not chain:
        ws["A4"] = "(no fallback data)"
        ws["A4"].font = Font(italic=True, color="C62828")
        return

    # Master summary table
    hdrs = ["Rank Tried", "DDS Name", "Surrogate Score",
            "Deep Pass % (count/total)", "Verdict", "Promoted?",
            "Failure Reason", "Transition Reason"]
    for j, h in enumerate(hdrs, 4):
        c = ws.cell(4, j-3, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[4].height = 32

    VERD_COLOR = {"PASSED":"C6EFCE","MARGINAL":"FFEB9C","FAILED":"FFC7CE"}
    row = 5
    for entry in chain:
        ws.cell(row, 1, f"#{entry['rank']}").alignment = Alignment(horizontal="center")
        ws.cell(row, 2, entry["dds_name"])
        ws.cell(row, 3, entry.get("surrogate_score","?")).alignment = Alignment(horizontal="center")
        ws.cell(row, 4, f"{entry['deep_passed_pct']}% "
                         f"({entry.get('deep_passed_count','?')}/"
                         f"{entry.get('deep_total','?')})").alignment = Alignment(horizontal="center")
        v_cell = ws.cell(row, 5, entry["verdict"])
        v_cell.fill = PatternFill("solid", fgColor=VERD_COLOR.get(entry["verdict"],"FFFFFF"))
        v_cell.alignment = Alignment(horizontal="center")
        v_cell.font = Font(bold=True)
        p_cell = ws.cell(row, 6, "✅ YES" if entry.get("promoted") else "❌ no")
        p_cell.alignment = Alignment(horizontal="center")
        if entry.get("promoted"):
            p_cell.font = Font(bold=True, color="2E7D32")
        ws.cell(row, 7, entry.get("failure_reason","—")).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 8, entry.get("transition_reason","—")).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 56
        row += 1

    # Per-candidate detail blocks (failed principles list)
    row += 2
    ws.cell(row, 1, "DETAILED FAILED-PRINCIPLE DIAGNOSTICS").font = Font(bold=True, size=12)
    row += 1

    for entry in chain:
        ws.cell(row, 1,
                 f"#{entry['rank']} — {entry['dds_name']} "
                 f"({entry['verdict']}, {entry['deep_passed_pct']}%)").font = Font(bold=True, color="0f2040")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        failed = entry.get("failed_principles") or []
        if failed:
            sub_hdr = ["Principle", "Deep Score", "Deep Value",
                       "Confidence", "Method", "Narrative"]
            for j, h in enumerate(sub_hdr, 1):
                c = ws.cell(row, j, h)
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="FFE699")
            row += 1
            for f in failed:
                ws.cell(row, 1, f["principle"]).font = Font(bold=True)
                ws.cell(row, 2, f["deep_score"]).alignment = Alignment(horizontal="center")
                ws.cell(row, 3, f.get("deep_value",""))
                ws.cell(row, 4, f["confidence"]).alignment = Alignment(horizontal="center")
                ws.cell(row, 5, f["method"]).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row, 6, f["narrative"]).alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row].height = 40
                row += 1
        else:
            ws.cell(row, 1, "(no failed principles — ALL PASSED)").font = Font(italic=True, color="2E7D32")
            row += 1
        row += 1   # blank separator

    widths = [11, 26, 13, 17, 11, 11, 38, 38]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ──────────────────────────────────────────────────────────────────────────
# v22 — Class A Surrogate Detail Sheet
# Every (DDS, principle) pair in long format with FULL provenance:
# score, value, method, reference, confidence, narrative.
# This is what the researcher reads when they want to know exactly WHY a
# principle gave a particular score for a particular DDS.
# ──────────────────────────────────────────────────────────────────────────
def _write_drug_surrogate_detail_sheet(wb, idx: int, dr: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    drug_name = dr["drug_name"]
    safe_name = "".join(c for c in drug_name if c.isalnum())[:18]
    sheet_name = f"D{idx}_{safe_name}_SurrDetail"[:31]
    ws = wb.create_sheet(sheet_name)

    matrix = dr.get("dds_principle_matrix") or []
    ws["A1"] = (f"Drug {idx}: {drug_name} — Class A Surrogate Detail "
                f"(every DDS × every principle, full provenance)")
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Long-format table: one row per (DDS, principle) pair, with "
                "score, value, method, reference, confidence, and warnings. "
                "Use auto-filter to drill into any specific principle or DDS.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 36

    if not matrix:
        ws["A4"] = "(no surrogate data)"; return

    # Pull catalog titles
    try:
        from cerebro_62_principles_catalog import PRINCIPLES_62
    except Exception:
        PRINCIPLES_62 = {}

    hdrs = ["DDS Rank", "DDS Name", "Principle ID", "Principle Title",
            "Score (0-100)", "Value", "Confidence", "Method", "Reference"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[4].height = 28

    row = 5
    for dds_rank, m in enumerate(matrix, 1):
        principles = m.get("principles", {})
        for pid in sorted(principles.keys()):
            r = principles[pid]
            cat = PRINCIPLES_62.get(pid, {})
            ws.cell(row, 1, f"#{dds_rank}").alignment = Alignment(horizontal="center")
            ws.cell(row, 2, m.get("dds_name","?"))
            ws.cell(row, 3, pid).font = Font(bold=True)
            ws.cell(row, 4, cat.get("title_en","—"))
            score = r.get("score", 0)
            sc = ws.cell(row, 5, round(score, 2))
            sc.alignment = Alignment(horizontal="center")
            if score >= 80:   sc.fill = PatternFill("solid", fgColor="C6EFCE")
            elif score >= 60: sc.fill = PatternFill("solid", fgColor="DDEBCB")
            elif score >= 40: sc.fill = PatternFill("solid", fgColor="FFEB9C")
            elif score >= 20: sc.fill = PatternFill("solid", fgColor="FFC7CE")
            else:             sc.fill = PatternFill("solid", fgColor="F2A1A1")
            ws.cell(row, 6, str(r.get("value",""))[:30])
            ws.cell(row, 7, r.get("confidence","—")).alignment = Alignment(horizontal="center")
            ws.cell(row, 8, str(r.get("method","") or "")[:160]).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 9, str(r.get("reference","") or "")[:80])
            row += 1

    # Auto-filter
    ws.auto_filter.ref = f"A4:I{row-1}"
    ws.freeze_panes = "D5"

    widths = [9, 24, 11, 32, 11, 22, 11, 50, 32]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
