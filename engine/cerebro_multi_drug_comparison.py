"""
================================================================================
CEREBRO-X |  cerebro_multi_drug_comparison.py
================================================================================
Created by: Muhammad Talaat (BPharm, R&D Computational Lead) — CEREBRO-X

PURPOSE
═══════
After EACH drug has been individually processed against ALL DDS (full 62-
principle pipeline), this engine performs a principle-by-principle COMPARISON
across drugs. Used only when the input Excel contained ≥ 2 drugs.

For every numeric science-module result, the engine:
  • Collects the value from each drug
  • Determines the metric direction (higher_is_better / lower_is_better)
    using a built-in pharmacological convention table
  • Computes a normalized 0–100 score per drug per metric
  • Identifies the winner per metric
  • Aggregates a weighted overall ranking grouped by principle category
  • Flags ties, missing values, and Tier-6 (class-mean) reliance

OUTPUTS
═══════
  • CEREBRO_X_Multi_Drug_Comparison.xlsx  (3 sheets)
       Sheet 1 "Overview":         overall ranking + winner counts per drug
       Sheet 2 "Per_Principle":    full N-drug × M-metric matrix with markers
       Sheet 3 "Tier_Coverage":    per-drug Tier-distribution histogram
  • CEREBRO_X_Multi_Drug_Comparison.json  (machine-readable summary)
  • Returns summary dict to caller for embedding in PDF / HTML5

DESIGN PRINCIPLES
═════════════════
  • Pure-data engine — no opinions on drug efficacy beyond metric direction
  • Conservative: if direction unknown, the metric is reported but not ranked
  • Tier-6 reliance is surfaced but does NOT change ranking (researcher
    overrides handle that — see cerebro_completed_excel_writer)
  • CNS-focused weighting: pbpk_cns and glymphatic carry highest weight
================================================================================
"""
from __future__ import annotations

import json
import logging
import math
from datetime import datetime
from pathlib import Path

log = logging.getLogger("CEREBRO-COMPARISON")

# ─── Metric direction conventions ─────────────────────────────────────────
# higher_is_better: increasing metric → better drug performance
HIGHER_IS_BETTER = {
    # CNS PK / efficacy
    "AUC_brain", "Cmax_brain", "T_half_brain", "brain_plasma_ratio",
    "BBB_permeability_pct", "BBB_Engineering_Score", "BBB_score",
    "Tanimoto_Similarity", "Docking_Affinity_kcal_abs",
    # Stability
    "shelf_life", "shelf_life_months", "days_at_25C", "days_at_4C",
    "stability_index",
    # Safety scores (higher = safer in our convention)
    "safety_score", "selectivity_index", "therapeutic_index",
    # Release
    "release_constant", "release_efficiency",
    # Glymphatic
    "glymphatic_clearance_rate", "csf_distribution_volume",
}
# lower_is_better: increasing metric → worse drug performance
LOWER_IS_BETTER = {
    # Toxicity / risk
    "Toxicity_Index", "nanotoxicity_score", "qsar_toxicity_score",
    "off_target_count", "hepatotoxicity_score", "cardiotoxicity_score",
    # PK negatives
    "first_pass_loss_pct", "clearance_rate", "p_glycoprotein_efflux",
    # Release negatives
    "burst_24h_pct", "burst_release_pct",
    # Docking energy (more negative = stronger binding)
    "Docking_Affinity_kcal",
}

# ─── Principle group weights (sum ≈ 1.0, CNS-focused) ────────────────────
PRINCIPLE_WEIGHTS = {
    "pbpk_cns":        0.20,   # CNS PK — primary CEREBRO-X concern
    "glymphatic":      0.15,   # CNS clearance — second priority
    "qsar_toxicity":   0.10,
    "nanotoxicity":    0.10,
    "release":         0.10,
    "shelf_life":      0.05,
    "drug_problems":   0.05,
    "dds_comparison":  0.05,
    "allometric":      0.05,
    "stress_test":     0.05,
    "physchem":        0.05,
    "top_dds":         0.05,
    "default":         0.02,   # any unrecognized module
}


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────
def _to_float(v) -> float | None:
    if v is None: return None
    if isinstance(v, bool): return float(v)
    if isinstance(v, (int, float)) and not math.isnan(v) and math.isfinite(v):
        return float(v)
    if isinstance(v, str):
        try:    return float(v.replace(",","").strip())
        except (ValueError, TypeError): return None
    return None


def _direction_for(metric: str) -> str:
    """Return 'higher','lower', or 'unknown' for a metric path."""
    leaf = metric.split(".")[-1]
    if leaf in HIGHER_IS_BETTER: return "higher"
    if leaf in LOWER_IS_BETTER:  return "lower"
    return "unknown"


def _flatten_numeric(d: dict, prefix: str = "") -> dict[str, float]:
    """Flatten a nested dict, keeping only numeric leaves."""
    out: dict[str, float] = {}
    if not isinstance(d, dict): return out
    for k, v in d.items():
        if isinstance(k, str) and k.startswith("_"): continue
        path = f"{prefix}.{k}" if prefix else str(k)
        if isinstance(v, dict):
            out.update(_flatten_numeric(v, prefix=path))
        else:
            f = _to_float(v)
            if f is not None:
                out[path] = f
    return out


def _normalize_score(values: dict[str, float], direction: str) -> dict[str, float]:
    """
    Normalize values to 0–100 scores per direction.
    For 'higher': max → 100, min → 0
    For 'lower':  min → 100, max → 0
    Ties keep equal scores.
    """
    if not values: return {}
    vmin, vmax = min(values.values()), max(values.values())
    if vmin == vmax:
        return {k: 100.0 for k in values}    # all tied → all win
    out = {}
    for k, v in values.items():
        if direction == "higher":
            out[k] = round(100.0 * (v - vmin) / (vmax - vmin), 2)
        else:
            out[k] = round(100.0 * (vmax - v) / (vmax - vmin), 2)
    return out


def _tier_distribution(mol_profile: dict) -> dict[int, int]:
    """Count how many properties were resolved at each tier for a drug."""
    audit = (mol_profile or {}).get("_source_audit", {}) or {}
    counts: dict[int, int] = {}
    for prop_key, info in audit.items():
        if isinstance(info, dict):
            t = info.get("_tier", 99)
            counts[t] = counts.get(t, 0) + 1
    return counts


# ──────────────────────────────────────────────────────────────────────────
# Main entry — compare_drugs
# ──────────────────────────────────────────────────────────────────────────
def compare_drugs(drug_results: list[dict],
                   output_dir: Path,
                   pipeline_metadata: dict | None = None) -> dict:
    """
    Run the full N-drug comparison and emit Excel + JSON.

    Each entry in `drug_results` must have:
      drug_name   (str)
      mol_profile (dict)           — molecule engine output
      df_dds      (pd.DataFrame|None)
      principles  (dict)           — flattened or nested science_modules output
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if len(drug_results) < 2:
        log.info(f"[COMPARISON] Only {len(drug_results)} drug — skipped "
                 f"(comparison requires ≥ 2)")
        return {"status": "skipped", "reason": "only_one_drug",
                "drug_count": len(drug_results)}

    # Every downstream dict in this function is keyed by drug_name
    # (per_drug_metrics, winner_counts, score_sum, tier_coverage, ...)
    # with no uniqueness check -- two entries sharing a name (a realistic
    # ingestion mistake given this project's own documented history of
    # Excel input mix-ups) would silently clobber each other: the second
    # entry's data overwrites the first's with no warning, and the
    # output still lists the name twice giving the false impression two
    # independent drugs were compared. Disambiguate here (on a local
    # copy -- the caller's own drug_results dicts are left untouched)
    # rather than let one drug's real numbers silently vanish.
    _seen: dict[str, int] = {}
    _deduped = []
    for d in drug_results:
        name = d["drug_name"]
        _seen[name] = _seen.get(name, 0) + 1
        if _seen[name] > 1:
            new_name = f"{name} (dup#{_seen[name]})"
            log.warning(f"[COMPARISON] Duplicate drug_name '{name}' in "
                        f"drug_results — renaming this entry to "
                        f"'{new_name}' so neither dataset silently "
                        f"overwrites the other")
            d = {**d, "drug_name": new_name}
        _deduped.append(d)
    drug_results = _deduped

    drug_names = [d["drug_name"] for d in drug_results]
    log.info(f"[COMPARISON] Comparing {len(drug_names)} drugs: {drug_names}")

    # ─── Step 1: collect all numeric metrics per drug ──────────────────
    per_drug_metrics: dict[str, dict[str, float]] = {}
    for d in drug_results:
        flat = _flatten_numeric(d.get("principles", {}))
        # Add top-line PK & physchem from mol_profile
        mp = d.get("mol_profile", {}) or {}
        for k in ("MW_Da", "LogP", "TPSA_A2", "HBD", "HBA",
                  "Half_Life_Days", "BBB_permeability_pct",
                  "Docking_Affinity_kcal", "LogBB"):
            v = _to_float(mp.get(k))
            if v is not None:
                flat[f"physchem.{k}"] = v
        # Docking_Affinity_kcal (lower/more-negative = stronger binding =
        # better, LOWER_IS_BETTER above) and its abs() (HIGHER_IS_BETTER)
        # are monotonically equivalent for a negative raw value -- "raw
        # is lower-is-better" and "abs is higher-is-better" agree on
        # every comparison. Feeding both into per_drug_metrics used to
        # double-count docking affinity's influence within the physchem
        # weight bucket relative to every other physchem property (MW,
        # LogP, TPSA, HBD, HBA, Half_Life_Days, LogBB), which are each
        # counted once. _direction_for's own "_abs" handling stays
        # generic (still tested directly), but nothing in this module
        # should compute or feed the derived variant into scoring.
        # Add top-DDS metrics
        df_dds = d.get("df_dds")
        try:
            if df_dds is not None and not df_dds.empty:
                top = df_dds.iloc[0]
                for col in df_dds.columns:
                    v = _to_float(top.get(col))
                    if v is not None:
                        flat[f"top_dds.{col}"] = v
        except Exception as e:
            log.debug(f"[COMPARISON] DDS extraction skipped for {d['drug_name']}: {e}")
        per_drug_metrics[d["drug_name"]] = flat

    all_metrics = sorted({m for d in per_drug_metrics.values() for m in d})
    log.info(f"[COMPARISON] Collected {len(all_metrics)} comparable metrics "
             f"across {len(drug_names)} drugs")

    # ─── Step 2: per-metric ranking + score accumulation ───────────────
    per_principle_table: list[dict] = []
    winner_counts = {n: 0 for n in drug_names}
    score_sum     = {n: 0.0 for n in drug_names}
    weight_sum    = {n: 0.0 for n in drug_names}
    # A drug's weighted average is normalized only against the ranked
    # metrics IT actually has a value for -- a drug that's only
    # comparable on one favorable metric can outscore a thoroughly-
    # characterized drug that would have won convincingly if all its
    # data had been comparable, with nothing in the report disclosing
    # the coverage gap driving that result. Rather than silently
    # changing the ranking math (a real methodology decision this
    # module shouldn't make unilaterally), disclose how many ranked
    # metrics each drug's score actually rests on, so a reader can
    # judge a high score built on 2 metrics differently from one built
    # on 40.
    metrics_compared_count = {n: 0 for n in drug_names}
    ranked_count  = 0
    unranked_count = 0
    # Metrics only one drug has a value for can't be *compared*, but
    # silently `continue`-ing past them with no counter meant
    # metrics_total never reconciled against metrics_ranked +
    # metrics_unranked -- confirmed live: production JSON snapshots
    # showed 115-131 of ~480 collected metrics vanishing with no trace
    # anywhere in the report (not in per_principle, not counted anywhere),
    # so a reader had no way to know they were ever computed at all.
    single_drug_count = 0

    for metric in all_metrics:
        values = {n: per_drug_metrics[n][metric]
                  for n in drug_names if metric in per_drug_metrics[n]}
        if len(values) < 2:
            single_drug_count += 1
            continue   # need ≥ 2 drugs with this metric to compare

        direction = _direction_for(metric)
        if direction == "unknown":
            row = {"metric": metric, "direction": "unranked", "winner": "—"}
            for n in drug_names:
                row[n] = values.get(n, "—")
            per_principle_table.append(row)
            unranked_count += 1
            continue

        scores = _normalize_score(values, direction)
        # A genuine tie (e.g. every drug fell back to the same Tier-7
        # class-mean default for this metric) normalizes every score to
        # 100.0 — max() would then arbitrarily pick whichever drug happens
        # to be listed first, silently crediting it with a "win" it didn't
        # actually earn over the others. Report it as an explicit tie
        # instead, and don't count it toward anyone's winner_counts.
        is_tie = len(set(scores.values())) == 1
        winner = "— (tie)" if is_tie else max(scores.items(), key=lambda kv: kv[1])[0]
        if not is_tie:
            winner_counts[winner] += 1
        ranked_count += 1

        prefix = metric.split(".")[0]
        weight = PRINCIPLE_WEIGHTS.get(prefix, PRINCIPLE_WEIGHTS["default"])
        for n, s in scores.items():
            score_sum[n]  += s * weight
            weight_sum[n] += weight
            metrics_compared_count[n] += 1

        row = {"metric": metric, "direction": direction, "winner": winner}
        for n in drug_names:
            row[n] = values.get(n, "—")
            row[f"{n}_score"] = scores.get(n, "—")
        per_principle_table.append(row)

    # ─── Step 3: weighted overall ranking ──────────────────────────────
    overall = []
    for n in drug_names:
        ws = weight_sum[n]
        avg = (score_sum[n] / ws) if ws > 0 else 0.0
        overall.append((n, round(avg, 2)))
    overall.sort(key=lambda kv: kv[1], reverse=True)

    # Competition ranking ("1224" scheme): drugs with an identical
    # weighted_score share the same rank instead of getting distinct
    # ranks from a stable sort's arbitrary tie-break order. Without this,
    # two drugs tied at weighted_score=100.0 were shown as an outright
    # #1/#2 win/loss (confirmed live in production: Lecanemab rank 1,
    # Temozolomide rank 2, both scored exactly 100.0) with the Excel
    # writer then painting rank 1 green ("best") and, when n>1, the last
    # rank red ("worst") -- crowning a winner and a loser out of a tie.
    overall_ranks: list[int] = []
    _prev_score = None
    _prev_rank = 0
    for i, (_n, s) in enumerate(overall, 1):
        if s != _prev_score:
            _prev_rank = i
        overall_ranks.append(_prev_rank)
        _prev_score = s

    # ─── Step 4: tier coverage per drug ────────────────────────────────
    tier_coverage = {d["drug_name"]: _tier_distribution(d.get("mol_profile", {}))
                      for d in drug_results}

    # ─── Step 4b [v21]: extract Drug × Best-DDS Champion data ──────────
    # For each drug, capture the principle-ranked top-1 DDS and its full
    # 62-principle score breakdown. This is the "drug + best DDS pair"
    # head-to-head comparison the researcher cares about most.
    champions: list[dict] = []
    for d in drug_results:
        breakdown = d.get("dds_principle_breakdown") or []
        matrix    = d.get("dds_principle_matrix") or []
        if not breakdown:
            champions.append({"drug": d["drug_name"], "available": False,
                              "reason": "no_dds_principle_breakdown"})
            continue
        top = breakdown[0]   # already sorted by composite descending
        top_matrix = matrix[0] if matrix else {}
        champions.append({
            "drug":          d["drug_name"],
            "available":     True,
            "best_dds":      top["dds_name"],
            "composite":     top["composite"],
            "verdict":       top["verdict"],
            "group_scores":  top["group_scores"],
            "principles":    top_matrix.get("principles", {}),
            "narrative":     top["narrative"],
            "top_strengths": top["top_strengths"],
            "weak_spots":    top["weak_spots"],
        })

    # ─── Step 5: build summary ─────────────────────────────────────────
    # v22: capture C+ Flow data (deep validation, translational, fallback)
    cplus_per_drug = []
    for d in drug_results:
        deep_summary = d.get("deep_summary") or {}
        translational = d.get("translational") or {}
        fallback_chain = d.get("fallback_chain") or []
        # Compose snapshot for cross-drug comparison
        cplus_per_drug.append({
            "drug":              d.get("drug_name"),
            "top1_dds":          (d.get("df_dds").iloc[0].get("Formulation_Name")
                                   if d.get("df_dds") is not None
                                       and not d.get("df_dds").empty else None),
            "deep_verdict":      deep_summary.get("verdict","?"),
            "deep_pct":          deep_summary.get("pct", 0),
            "deep_passed_count": deep_summary.get("passed_count", 0),
            "deep_total":        deep_summary.get("total", 0),
            "fallback_attempts": len(fallback_chain),
            "fallback_chain":    fallback_chain,
            "translational_status": {
                pid: t.get("status","?") for pid, t in translational.items()
            },
            "translational_scores": {
                "fto_score":          translational.get("P32",{}).get("fto_score"),
                "compliance_score":   translational.get("P45",{}).get("compliance_score"),
                "patentability_score": translational.get("P56",{}).get("patentability_score"),
            },
        })

    summary = {
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "drug_count":   len(drug_names),
        "drug_names":   drug_names,
        "metrics_total":      len(all_metrics),
        "metrics_ranked":     ranked_count,
        "metrics_unranked":   unranked_count,
        # metrics_total == metrics_ranked + metrics_unranked +
        # metrics_single_drug always -- see single_drug_count comment above.
        "metrics_single_drug": single_drug_count,
        "winner_counts":      winner_counts,
        "overall_ranking":  [
            {"rank": overall_ranks[i], "drug": n, "weighted_score": s,
             "tied": overall_ranks.count(overall_ranks[i]) > 1,
             "metrics_compared": metrics_compared_count[n]}
            for i, (n, s) in enumerate(overall)
        ],
        "per_principle":  per_principle_table,
        "tier_coverage":  tier_coverage,
        "champions":      champions,
        "cplus_flow":     cplus_per_drug,  # v22
        "metadata":       pipeline_metadata or {},
    }

    # ─── Step 6: emit JSON ─────────────────────────────────────────────
    json_path = output_dir / "CEREBRO_X_Multi_Drug_Comparison.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False, default=str)
    log.info(f"[COMPARISON] JSON → {json_path.name}")

    # ─── Step 7: emit Excel ────────────────────────────────────────────
    xlsx_path = output_dir / "CEREBRO_X_Multi_Drug_Comparison.xlsx"
    _write_comparison_excel(summary, xlsx_path)
    log.info(f"[COMPARISON] Excel → {xlsx_path.name}")

    return summary


# ──────────────────────────────────────────────────────────────────────────
# Excel writer
# ──────────────────────────────────────────────────────────────────────────
def _write_comparison_excel(summary: dict, output_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    drug_names = summary["drug_names"]

    # ─── Sheet 1: Overview ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "CEREBRO-X — Multi-Drug Comparison Report"
    ws["A1"].font = Font(bold=True, size=16, color="0f2040")
    ws["A2"] = (f"Generated: {summary['generated_at']}  "
                f"| Drugs: {summary['drug_count']}  "
                f"| Metrics ranked: {summary['metrics_ranked']}  "
                f"| Unranked (no direction): {summary['metrics_unranked']}  "
                f"| Single-drug only (not comparable, excluded): "
                f"{summary.get('metrics_single_drug', 0)}  "
                f"| Collected total: {summary['metrics_total']}")
    ws["A2"].font = Font(italic=True, color="9CA3AF")

    ws["A4"] = "Overall Weighted Ranking"
    ws["A4"].font = Font(bold=True, size=12)
    hdrs = ["Rank", "Drug", "Weighted Score (0–100)", "Metrics Won", "Metrics Compared"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(5, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center")

    n = len(summary["overall_ranking"])
    ranks = [e["rank"] for e in summary["overall_ranking"]]
    best_rank, worst_rank = (min(ranks), max(ranks)) if ranks else (None, None)
    for i, entry in enumerate(summary["overall_ranking"], 1):
        ws.cell(5+i, 1, entry["rank"])
        ws.cell(5+i, 2, entry["drug"])
        ws.cell(5+i, 3, entry["weighted_score"])
        ws.cell(5+i, 4, summary["winner_counts"][entry["drug"]])
        # A high weighted score built on only a couple of comparable
        # metrics isn't directly comparable to one built on dozens --
        # this column discloses that coverage rather than letting the
        # score alone imply a thoroughness it may not have (see the
        # metrics_compared_count comment above).
        ws.cell(5+i, 5, entry.get("metrics_compared", 0))
        # Highlight every row sharing the top/bottom rank, not just the
        # first/last by list position -- a tie at rank 1 must not paint
        # only one of the tied drugs green while its equal-scoring peer
        # gets no highlight at all (or, worse, gets painted red as
        # "worst" purely because it sorted later).
        if entry["rank"] == best_rank:
            for j in range(1, 6):
                ws.cell(5+i, j).fill = PatternFill("solid", fgColor="C6EFCE")
                ws.cell(5+i, j).font = Font(bold=True)
        elif entry["rank"] == worst_rank and worst_rank != best_rank:
            for j in range(1, 6):
                ws.cell(5+i, j).fill = PatternFill("solid", fgColor="FFC7CE")

    for j, w in enumerate([8, 26, 24, 16, 18], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Note on weighting
    note_row = 5 + n + 3
    ws.cell(note_row, 1, "Weighting (CNS-focused):").font = Font(bold=True)
    weight_text = "  ".join(f"{k}={v:.2f}"
                             for k, v in PRINCIPLE_WEIGHTS.items() if k != "default")
    ws.cell(note_row+1, 1, weight_text).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row+1, start_column=1,
                    end_row=note_row+1, end_column=5)

    # ─── Sheet 2: Per-principle matrix ─────────────────────────────────
    ws2 = wb.create_sheet("Per_Principle")
    cols = ["Metric", "Direction"] + drug_names + ["Winner"]
    for j, h in enumerate(cols, 1):
        c = ws2.cell(1, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(summary["per_principle"], 2):
        ws2.cell(i, 1, row["metric"])
        ws2.cell(i, 2, row["direction"])
        for j, name in enumerate(drug_names, 3):
            v = row.get(name, "—")
            if isinstance(v, float):
                v = round(v, 4)
            cell = ws2.cell(i, j, v)
            if row.get("winner") == name:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(bold=True)
        ws2.cell(i, len(cols), row.get("winner", "—"))

    widths = [40, 14] + [22] * len(drug_names) + [18]
    for j, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ─── Sheet 3: Tier coverage ────────────────────────────────────────
    ws3 = wb.create_sheet("Tier_Coverage")
    ws3["A1"] = "Tier Distribution per Drug"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3["A2"] = ("Counts of how many molecular properties were resolved at "
                 "each tier. High Tier-6 counts indicate the drug is poorly "
                 "characterized — researcher overrides recommended.")
    ws3["A2"].font = Font(italic=True, color="9CA3AF")

    all_tiers = sorted({t for cov in summary["tier_coverage"].values() for t in cov})
    if not all_tiers:
        all_tiers = [1, 2, 4, 5, 6, 99]
    hdrs3 = ["Drug"] + [f"Tier {t}" for t in all_tiers] + ["Tier-6 risk"]
    for j, h in enumerate(hdrs3, 1):
        c = ws3.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center")

    for i, (drug, cov) in enumerate(summary["tier_coverage"].items(), 5):
        ws3.cell(i, 1, drug)
        for j, t in enumerate(all_tiers, 2):
            ws3.cell(i, j, cov.get(t, 0))
        t6 = cov.get(6, 0)
        risk_cell = ws3.cell(i, len(all_tiers)+2,
                              "HIGH" if t6 >= 3 else "MED" if t6 >= 1 else "LOW")
        risk_cell.alignment = Alignment(horizontal="center")
        risk_cell.fill = PatternFill("solid",
            fgColor=("FFC7CE" if t6 >= 3 else "FFEB9C" if t6 >= 1 else "C6EFCE"))
        risk_cell.font = Font(bold=True)

    widths3 = [22] + [10] * len(all_tiers) + [14]
    for j, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    # ─── Sheet 4 [v21]: DRUG × BEST-DDS CHAMPION HEAD-TO-HEAD ────────
    # The "drug + its best DDS pair" comparison the researcher cares about most.
    # Each column = one (drug, best-DDS) champion. Rows = the 62 principles.
    _write_champion_comparison_sheet(wb, summary)

    # ─── Sheet 5 [v21]: SCIENTIFIC RATIONALE NARRATIVE ────────────────
    # A plain-language report explaining how every value was computed,
    # what literature supports it, and what the verdict means.
    _write_scientific_rationale_sheet(wb, summary)

    # ─── Sheets 6-8 [v22]: C+ Flow cross-drug comparison ──────────────
    # Per Muhammad's mandate: ALL C+ Flow results must appear in EVERY
    # output, including this multi-drug comparison.
    _write_cplus_deep_compare_sheet(wb, summary)
    _write_cplus_translational_compare_sheet(wb, summary)
    _write_cplus_fallback_compare_sheet(wb, summary)

    wb.save(str(output_path))


def _write_champion_comparison_sheet(wb, summary: dict) -> None:
    """
    Drug × Best-DDS Champion comparison sheet.

    Each column = one drug's principle-ranked best DDS.
    Each row    = one of the 62 principles.
    Cell        = the principle score for that (drug, best-DDS) pair.

    Winner per row is highlighted green. Group sub-headers separate the
    62 principles into the 7 thematic groups (CNS Delivery, Release, …).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    try:
        from cerebro_62_principles_catalog import PRINCIPLES_62
    except ImportError:
        log.debug("[CHAMPION-SHEET] principles catalog unavailable — skipped")
        return

    champions = summary.get("champions") or []
    available = [c for c in champions if c.get("available")]
    if not available:
        log.info("[CHAMPION-SHEET] No champions with principle data — skipping sheet")
        return

    ws = wb.create_sheet("Champion_DDS_Compare")
    ws["A1"] = "Drug × Best-DDS Champion Head-to-Head Comparison"
    ws["A1"].font = Font(bold=True, size=15, color="0f2040")
    ws["A2"] = ("Each column below = ONE drug paired with its principle-ranked "
                "best DDS. Each row = one of the 62 CNS-weighted principles. "
                "Cells show the per-principle score (0-100) the (drug,DDS) pair "
                "achieved. Green = winner per principle. Read the bottom rows "
                "for composite, verdict, and group-rollup scores.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws.row_dimensions[2].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(available))

    # Header row 4
    headers = ["Principle ID", "Group / Description"] + \
              [f"{c['drug']}\n({c['best_dds']})" for c in available]
    for j, h in enumerate(headers, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 50

    # Composite + verdict rows (rows 5-7) — most important info
    ws.cell(5, 1, "COMPOSITE").font = Font(bold=True, size=11)
    ws.cell(5, 2, "Total CNS-weighted score (0-100)").font = Font(italic=True)
    best_composite = max(c["composite"] for c in available)
    for j, c in enumerate(available, 3):
        cell = ws.cell(5, j, round(c["composite"], 2))
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid",
            fgColor=("C6EFCE" if c["composite"] == best_composite else "FFFFFF"))

    ws.cell(6, 1, "VERDICT").font = Font(bold=True, size=11)
    ws.cell(6, 2, "Verdict label based on composite").font = Font(italic=True)
    VERDICT_FILL = {"EXCELLENT":"C6EFCE","GOOD":"DDEBCB","ACCEPTABLE":"FFEB9C",
                     "MARGINAL":"FFC7CE","POOR":"F2A1A1"}
    for j, c in enumerate(available, 3):
        cell = ws.cell(6, j, c["verdict"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor=VERDICT_FILL.get(c["verdict"], "FFFFFF"))

    # Group-rollup rows — names must match cerebro_62_orchestrator.
    # PRINCIPLE_GROUPS exactly, since that's what actually produced
    # champions[i]["group_scores"]'s keys (G2 is "..._Kinetics" and G5 is
    # "..._BBB" there — this used to read "G2_Release"/"G5_Glymphatic",
    # a leftover from an older naming scheme, so those two rows always
    # rendered zero for every drug).
    group_keys = ["G1_CNS_Delivery","G2_Release_Kinetics","G3_Stability","G4_Safety",
                   "G5_Glymphatic_BBB","G6_Manufacturability","G7_DrugDDS_Fit"]
    GROUP_COLOR = {"G1_CNS_Delivery":"DDEBCB","G2_Release_Kinetics":"FFEB9C",
                    "G3_Stability":"D9D9D9","G4_Safety":"FFD7B5",
                    "G5_Glymphatic_BBB":"C2D6F0","G6_Manufacturability":"F0E5C2",
                    "G7_DrugDDS_Fit":"E5DAF2"}

    row = 7
    ws.cell(row, 1, "GROUP ROLLUPS").font = Font(bold=True, size=11, color="0f2040")
    ws.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=2 + len(available))
    row += 1
    for g in group_keys:
        ws.cell(row, 1, g.replace("_"," "))
        ws.cell(row, 2, f"Average score across {g.replace('_',' ')} group")
        ws.cell(row, 2).font = Font(italic=True, color="9CA3AF", size=9)
        # Find best (highest) per group
        vals = [c["group_scores"].get(g, 0) for c in available]
        max_v = max(vals) if vals else 0
        for j, c in enumerate(available, 3):
            v = c["group_scores"].get(g, 0)
            cell = ws.cell(row, j, round(v, 2))
            cell.alignment = Alignment(horizontal="center")
            if v == max_v and len(available) > 1:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(bold=True)
            elif v == min(vals) and len(available) > 1:
                cell.fill = PatternFill("solid", fgColor="FFE0E0")
        # Color group-name cells with the group palette
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=GROUP_COLOR.get(g,"FFFFFF"))
        ws.cell(row, 1).font = Font(bold=True)
        row += 1

    # Per-principle rows — grouped
    row += 1
    ws.cell(row, 1, "PER-PRINCIPLE SCORES").font = Font(bold=True, size=11, color="0f2040")
    ws.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=2 + len(available))
    row += 1

    # Grouped by the real 62-principle catalog — must match
    # cerebro_62_orchestrator.PRINCIPLE_GROUPS exactly, since that's what
    # actually produced champions[i]["principles"]'s keys ("P01".."P62").
    # This used to list old v21-era dotted IDs ("P1.1_BBB_transcytosis"),
    # which never matched anything in the live data, so every cell in
    # this section rendered 0 for every drug regardless of its real score.
    try:
        from cerebro_62_orchestrator import PRINCIPLE_GROUPS
    except ImportError:
        log.debug("[CHAMPION-SHEET] orchestrator groups unavailable — skipping per-principle rows")
        PRINCIPLE_GROUPS = {}
    GROUP_PRINCIPLES = {g: pids for g, pids in PRINCIPLE_GROUPS.items()
                         if g != "G8_Translational"}   # translational shown elsewhere

    for g_name, pids in GROUP_PRINCIPLES.items():
        # Group divider row
        gc = ws.cell(row, 1, g_name.replace("_"," "))
        gc.font = Font(bold=True, color="0f2040")
        gc.fill = PatternFill("solid", fgColor=GROUP_COLOR.get(g_name,"FFFFFF"))
        ws.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=2 + len(available))
        row += 1
        for pid in pids:
            doc = PRINCIPLES_62.get(pid, {})
            weight = doc.get("weight_cns", 0)
            ws.cell(row, 1, f"{pid}  ({weight*100:.1f}%)").font = Font(size=9)
            descr = (doc.get("title_en","") or "")[:90]
            ws.cell(row, 2, descr).alignment = Alignment(wrap_text=True,
                                                            vertical="center")
            ws.cell(row, 2).font = Font(italic=True, size=8, color="9CA3AF")
            # Find values for each champion
            vals = []
            for j, c in enumerate(available, 3):
                pdata = c["principles"].get(pid, {})
                v = pdata.get("score", 0)
                vals.append(v)
            max_v = max(vals) if vals else 0
            min_v = min(vals) if vals else 0
            for j, c in enumerate(available, 3):
                pdata = c["principles"].get(pid, {})
                v = pdata.get("score", 0)
                cell = ws.cell(row, j, round(v, 2))
                cell.alignment = Alignment(horizontal="center")
                if v == max_v and max_v > min_v:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(bold=True)
                elif v == min_v and max_v > min_v and len(available) > 1:
                    cell.fill = PatternFill("solid", fgColor="FFE0E0")
            ws.row_dimensions[row].height = 22
            row += 1

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(2)].width = 56
    for j in range(3, 3 + len(available)):
        ws.column_dimensions[get_column_letter(j)].width = 22

    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False


def _write_scientific_rationale_sheet(wb, summary: dict) -> None:
    """
    Plain-language scientific rationale for the multi-drug comparison.
    Tells the researcher: what was compared, how scores were computed,
    which references back the methods, and what the verdict means.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Scientific_Rationale")
    ws["A1"] = "How CEREBRO-X Reached These Results — Scientific Rationale"
    ws["A1"].font = Font(bold=True, size=15, color="0f2040")
    ws.merge_cells("A1:D1")

    sections = [
        ("1. Pipeline architecture",
            "CEREBRO-X processes each drug independently — full molecular "
            "resolution, ML training, ADMET prediction, and DDS evaluation. "
            "Drugs are processed sequentially: Drug 1 → fully completed → "
            "Drug 2 → fully completed → ... No state from Drug N leaks into "
            "Drug N+1 (caches invalidated, mol_profile reset, separate Trial "
            "directories). Reference: see CHANGELOG_v19.md for the per-drug "
            "isolation fix."),
        ("2. Per-DDS principle evaluation",
            "For EACH DDS in the input Excel, the pipeline runs 62 CNS-"
            "weighted principles (BBB transcytosis, receptor targeting, "
            "Pgp evasion, brain AUC, release kinetics, stability, safety, "
            "glymphatic clearance, manufacturability, drug-DDS fit). "
            "Each principle returns a 0-100 score with a documented method "
            "and literature reference, individually weighted toward CNS-"
            "relevant criteria (see cerebro_62_principles_catalog.py's "
            "weight_cns field for the exact per-principle weight). The "
            "resulting composite weighted score ranks all DDS. The "
            "principle-ranked top-1 DDS is the 'champion' for that drug — "
            "see the DDSxP matrix sheet in each drug's Completed Excel."),
        ("3. Drug × Best-DDS champion comparison",
            "After every drug is processed, the pipeline pairs each drug "
            "with its principle-ranked best DDS. The 'Champion_DDS_Compare' "
            "sheet shows these pairs side-by-side on every principle. "
            "Winner cells per principle are highlighted green; losers red. "
            "The composite, verdict, and group rollups give the researcher "
            "the headline answer: which (drug, DDS) combination wins overall."),
        ("4. Cross-drug per-metric ranking",
            "The 'Per_Principle' sheet flattens every numeric metric the "
            "science modules produced (PBPK CNS, release, shelf-life, "
            "QSAR toxicity, glymphatic, etc.) and ranks all drugs against "
            "each metric using a metric-direction table (some metrics are "
            "higher=better, others lower=better). Metrics with unknown "
            "direction are reported but not ranked, preserving conservatism."),
        ("5. Tier-coverage quality control",
            "The 'Tier_Coverage' sheet shows how each drug's molecular "
            "properties were resolved — Tier 1 (live API) is highest quality; "
            "Tier 6 (class-mean fallback) is lowest. Drugs flagged with "
            "Tier-6 risk = HIGH should have their Tier-6 values overridden "
            "with in-vitro measurements before publication. Every Tier-6 "
            "value is marked overridable and accompanied by a disclaimer."),
        ("6. Validity statement",
            "Every numeric output in CEREBRO-X has full provenance: the "
            "source (API/Library/PubMed/RDKit/Class-mean), the citation "
            "(DOI when available), the confidence percentage, and the "
            "computational method. All principle scores are deterministic "
            "and reproducible from the input Excel. Use the "
            "'Principle_Explanations' sheet in the per-drug Completed Excel "
            "as the glossary for what each metric means and how it was "
            "computed."),
        ("7. How to use this for decisions",
            "(1) Open Champion_DDS_Compare → the green column wins the most. "
            "(2) Open the per-drug DDSxP matrix → see what the runner-up DDS "
            "would have looked like. (3) Open Principle_Explanations → "
            "understand any score you're uncertain about. (4) Override any "
            "Tier-6 values with in-vitro data and re-run for higher "
            "confidence. (5) Cite the references in the Principle_Explanations "
            "sheet when publishing."),
    ]

    row = 3
    for title, body in sections:
        c = ws.cell(row, 1, title)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 24
        row += 1
        b = ws.cell(row, 1, body)
        b.font = Font(size=10, color="0D1B2A")
        b.fill = PatternFill("solid", fgColor="F2F2F2")
        b.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 90
        row += 2

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────
# v22 — C+ Flow cross-drug comparison sheets
# (Per Muhammad's mandate: results visible in every output)
# ──────────────────────────────────────────────────────────────────────────
def _write_cplus_deep_compare_sheet(wb, summary: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("CPlus_Deep_Validation")
    ws["A1"] = "C+ Flow — Class B Deep Physics Validation (cross-drug)"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("For each drug, the Top-1 DDS (after surrogate ranking) was "
                "validated through full physics. This sheet shows whether the "
                "deep validation passed and how many of the 28 deep principles "
                "confirmed the surrogate decision.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 50

    cplus = summary.get("cplus_flow") or []
    if not cplus:
        ws["A4"] = "(no C+ Flow data)"; return

    hdrs = ["Rank", "Drug", "Final Top-1 DDS", "Deep Verdict",
            "Deep Pass %", "Passed/Total"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 24

    VERD = {"PASSED":"C6EFCE","MARGINAL":"FFEB9C","FAILED":"FFC7CE",
             "NOT RUN":"D9D9D9","NO DATA":"D9D9D9"}
    ranked = sorted(cplus, key=lambda c: c.get("deep_pct",0), reverse=True)
    # Same competition-ranking fix as the Overview sheet's overall
    # ranking: a positional "#{i}" label assigns distinct ordinals to
    # drugs tied at the same deep_pct purely by stable-sort position,
    # implying one beat the other when they scored identically.
    _prev_pct, _prev_rank = None, 0
    for i, e in enumerate(ranked, 5):
        pct = e.get("deep_pct", 0)
        if pct != _prev_pct:
            _prev_rank = i - 4
        _prev_pct = pct
        ws.cell(i, 1, f"#{_prev_rank}").alignment = Alignment(horizontal="center")
        ws.cell(i, 2, e["drug"]).font = Font(bold=True)
        ws.cell(i, 3, e.get("top1_dds") or "—")
        v_cell = ws.cell(i, 4, e["deep_verdict"])
        v_cell.fill = PatternFill("solid", fgColor=VERD.get(e["deep_verdict"],"FFFFFF"))
        v_cell.font = Font(bold=True)
        v_cell.alignment = Alignment(horizontal="center")
        ws.cell(i, 5, f"{e['deep_pct']}%").alignment = Alignment(horizontal="center")
        ws.cell(i, 6, f"{e['deep_passed_count']}/{e['deep_total']}").alignment = Alignment(horizontal="center")
    for j, w in enumerate([8, 24, 28, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_cplus_translational_compare_sheet(wb, summary: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("CPlus_Translational")
    ws["A1"] = "C+ Flow — Class C Translational Deliverables (cross-drug)"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("Class C outputs (Pre-IND outline, FTO analysis, compliance audit, "
                "grant outline, patentability score) generated for each drug's "
                "validated Top-1 DDS. v22 returns structured scores; full Word/PDF "
                "deliverables are scheduled for v23.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 50

    cplus = summary.get("cplus_flow") or []
    if not cplus:
        ws["A4"] = "(no translational data)"; return

    hdrs = ["Drug", "P21 Pre-IND", "P32 FTO Status",
            "FTO Score", "P45 Compliance %", "P55 Grant",
            "P56 Patentability", "P56 Score"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 32

    STATUS_COLOR = {"structured_outline_ready":"C6EFCE",
                      "search_queries_prepared":"DDEBCB",
                      "self_assessment_completed":"DDEBCB",
                      "scored":"C6EFCE",
                      "skipped_deep_validation_insufficient":"FFC7CE",
                      "failed":"F2A1A1"}
    for i, e in enumerate(cplus, 5):
        ws.cell(i, 1, e["drug"]).font = Font(bold=True)
        for j, pid in enumerate(["P21","P32","P45","P55","P56"], 0):
            # We map P21→col2, P32→col3, P45→col5, P55→col6, P56→col7 below
            pass
        # P21
        s21 = e["translational_status"].get("P21","—")
        c1 = ws.cell(i, 2, s21)
        c1.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(s21,"FFFFFF"))
        # P32
        s32 = e["translational_status"].get("P32","—")
        c2 = ws.cell(i, 3, s32)
        c2.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(s32,"FFFFFF"))
        ws.cell(i, 4, e["translational_scores"].get("fto_score") or "—").alignment = Alignment(horizontal="center")
        # P45
        ws.cell(i, 5, f"{e['translational_scores'].get('compliance_score','—')}%" if e["translational_scores"].get("compliance_score") else "—").alignment = Alignment(horizontal="center")
        # P55
        s55 = e["translational_status"].get("P55","—")
        c3 = ws.cell(i, 6, s55)
        c3.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(s55,"FFFFFF"))
        # P56
        s56 = e["translational_status"].get("P56","—")
        c4 = ws.cell(i, 7, s56)
        c4.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(s56,"FFFFFF"))
        ws.cell(i, 8, e["translational_scores"].get("patentability_score") or "—").alignment = Alignment(horizontal="center")

    for j, w in enumerate([22, 22, 22, 11, 16, 22, 22, 11], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_cplus_fallback_compare_sheet(wb, summary: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("CPlus_Fallback_Audit")
    ws["A1"] = "C+ Flow — Top-N Fallback Audit (cross-drug)"
    ws["A1"].font = Font(bold=True, size=14, color="0f2040")
    ws["A2"] = ("If a drug's Top-1 DDS fails the 70% deep-validation threshold, "
                "the orchestrator tries Top-2, then Top-3. This sheet shows the "
                "full attempt log per drug — including failure & transition reasons.")
    ws["A2"].font = Font(italic=True, color="9CA3AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 50

    cplus = summary.get("cplus_flow") or []
    if not cplus:
        ws["A4"] = "(no fallback data)"; return

    row = 4
    for e in cplus:
        # Per-drug header
        c = ws.cell(row, 1, f"💊 {e['drug']} — {len(e['fallback_chain'])} candidate(s) tried")
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0f2040")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 22
        row += 1

        # Sub-header
        sub = ["Rank Tried", "DDS Name", "Surrogate Score",
                "Deep Pass %", "Verdict", "Promoted?",
                "Failure Reason", "Transition Reason"]
        for j, h in enumerate(sub, 1):
            c = ws.cell(row, j, h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDEBF7")
        row += 1

        VERD = {"PASSED":"C6EFCE","MARGINAL":"FFEB9C","FAILED":"FFC7CE"}
        for entry in e["fallback_chain"]:
            ws.cell(row, 1, f"#{entry['rank']}").alignment = Alignment(horizontal="center")
            ws.cell(row, 2, entry["dds_name"])
            ws.cell(row, 3, entry.get("surrogate_score","?")).alignment = Alignment(horizontal="center")
            ws.cell(row, 4, f"{entry['deep_passed_pct']}%").alignment = Alignment(horizontal="center")
            v = ws.cell(row, 5, entry["verdict"])
            v.fill = PatternFill("solid", fgColor=VERD.get(entry["verdict"],"FFFFFF"))
            v.alignment = Alignment(horizontal="center")
            v.font = Font(bold=True)
            p = ws.cell(row, 6, "✅ YES" if entry.get("promoted") else "—")
            p.alignment = Alignment(horizontal="center")
            ws.cell(row, 7, entry.get("failure_reason","—")).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 8, entry.get("transition_reason","—")).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 56
            row += 1
        row += 1   # blank separator

    for j, w in enumerate([10, 24, 13, 13, 11, 11, 32, 32], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
