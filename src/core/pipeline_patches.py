"""
================================================================================
CEREBRO-X |  PIPELINE PATCHES & EXTENSIONS
================================================================================
File: cerebro_pipeline_patches.py

Patches applied to CEREBRO_Pipeline.py (non-invasive monkey-patching):
  1. Scaling Leakage Fix  — fit_transform → train-only fit + transform for new data
  2. Inference Engine     — predict() on NEW molecules without re-training
  3. Excel Reader         — CEREBRO_Input_Template.xlsx → YAML/dict → pipeline
  4. GIF / Video Engine  — animated visualisations from matplotlib frames
  5. Base64 Export        — encode PNG/GIF to Base64 for REST API responses

Import AFTER CEREBRO_Pipeline is imported:
    import CEREBRO_Pipeline as cp
    from cerebro_pipeline_patches import *
    apply_patches(cp)
================================================================================
"""

import base64
import io
import json
import logging
import re
import time
import warnings

import matplotlib
import numpy as np
import pandas as pd

matplotlib.use("Agg")
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt

warnings.filterwarnings("ignore")
log = logging.getLogger("CEREBRO-PATCH")

# ─────────────────────────────────────────────────────────────────────────────
# PATCH 1: SCALING LEAKAGE FIX
# Problem: MinMaxScaler.fit_transform() was called on ALL data (train+unseen).
#          This means the scaler "sees" test rows during training → leakage.
# Fix:     Scaler is fitted ONCE on train data only.
#          New data calls .transform() (NOT .fit_transform()).
# ─────────────────────────────────────────────────────────────────────────────

class TrainAwareScaler:
    """
    Wraps MinMaxScaler to enforce strict train/inference separation.

    Usage:
        scaler = TrainAwareScaler(feature_range=(45, 98))
        scaler.fit(train_predictions)          # Called once, on training data only
        proba = scaler.transform(new_preds)    # Called for any new input
    """

    def __init__(self, feature_range=(45, 98)):
        from sklearn.preprocessing import MinMaxScaler
        self._scaler = MinMaxScaler(feature_range=feature_range)
        self._fitted  = False
        self._train_min = None
        self._train_max = None

    def fit(self, y_train: np.ndarray) -> "TrainAwareScaler":
        """Fit on training predictions only. Must be called before transform."""
        self._scaler.fit(y_train.reshape(-1, 1))
        self._fitted    = True
        self._train_min = float(y_train.min())
        self._train_max = float(y_train.max())
        log.debug(f"  Scaler fitted: min={self._train_min:.4f} max={self._train_max:.4f}")
        return self

    def transform(self, y: np.ndarray) -> np.ndarray:
        """Transform new predictions using the TRAINING distribution."""
        if not self._fitted:
            raise RuntimeError(
                "TrainAwareScaler.fit() must be called before transform(). "
                "Never call fit_transform() on unseen data.")
        return self._scaler.transform(y.reshape(-1, 1)).flatten()

    def fit_transform(self, y_train: np.ndarray) -> np.ndarray:
        """Convenience: fit on train data and return transformed values."""
        self.fit(y_train)
        return self.transform(y_train)

    def save_state(self) -> dict:
        """Serialise scaler state for joblib persistence."""
        return {
            "fitted":    self._fitted,
            "train_min": self._train_min,
            "train_max": self._train_max,
            "scale_":    self._scaler.scale_.tolist() if self._fitted else None,
            "min_":      self._scaler.min_.tolist()   if self._fitted else None,
            "feature_range": list(self._scaler.feature_range),
        }

    @classmethod
    def from_state(cls, state: dict) -> "TrainAwareScaler":
        """Restore scaler from saved state (for inference on new data)."""
        obj = cls(feature_range=tuple(state["feature_range"]))
        if state["fitted"]:
            from sklearn.preprocessing import MinMaxScaler
            obj._scaler = MinMaxScaler(feature_range=tuple(state["feature_range"]))
            import numpy as np
            obj._scaler.scale_ = np.array(state["scale_"])
            obj._scaler.min_   = np.array(state["min_"])
            obj._scaler.data_min_ = np.array([state["train_min"]])
            obj._scaler.data_max_ = np.array([state["train_max"]])
            obj._scaler.data_range_= np.array([state["train_max"] - state["train_min"]])
            obj._fitted    = True
            obj._train_min = state["train_min"]
            obj._train_max = state["train_max"]
        return obj


def patched_train(cls, df: pd.DataFrame, feature_cols: list[str],
                  target_formula=None, run_id: str = None):
    """
    Drop-in replacement for AdvancedMLEngine.train().
    Key fix: TrainAwareScaler replaces MinMaxScaler.fit_transform on all_X.

    The original code:
        mm = MinMaxScaler(feature_range=(45,98))
        all_X = df[avail].fillna(0).values          # ← INCLUDES unseen rows
        df["ML_Success_Probability"] = mm.fit_transform(
            ensemble.predict(all_X).reshape(-1,1))   # ← LEAKAGE: scaler sees all

    The fix:
        mm = TrainAwareScaler(feature_range=(45,98))
        train_preds = ensemble.predict(X)             # X = training rows only
        mm.fit(train_preds)                           # Scaler fitted on train only
        all_preds = ensemble.predict(all_X)           # Predict all rows
        df["ML_Success_Probability"] = mm.transform(all_preds) # Transform (not fit)
    """
    from sklearn.covariance import EllipticEnvelope
    from sklearn.decomposition import PCA
    from sklearn.ensemble import (
        GradientBoostingRegressor,
        RandomForestRegressor,
        VotingRegressor,
    )
    from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
    from sklearn.model_selection import GridSearchCV, KFold, cross_val_score
    from sklearn.preprocessing import RobustScaler, StandardScaler
    from sklearn.svm import SVR

    try:
        import xgboost as xgb
        _HAS_XGB = True
    except ImportError:
        _HAS_XGB = False

    try:
        import shap as shap_lib
        _HAS_SHAP = True
    except ImportError:
        _HAS_SHAP = False

    try:
        import joblib
        _HAS_JOBLIB = True
    except ImportError:
        _HAS_JOBLIB = False

    t0 = time.time()
    log.info("patched AdvancedMLEngine.train() — leakage-free scaler …")
    df = df.copy()

    if target_formula:
        df["_target"] = target_formula(df)
    else:
        aff = next((c for c in ["Docking_Affinity_kcal","Binding_Affinity_kcal",
                                  "Estimated_Affinity_kcal"] if c in df.columns), None)
        df["_target"] = (abs(df[aff])*0.6 + df["Half_Life_Days"]*0.4
                         if aff else df["Half_Life_Days"]*0.4)

    avail = [c for c in feature_cols if c in df.columns]

    # Outlier detection
    X_raw_full = df[avail].fillna(0).values
    df["_is_outlier"] = False
    if len(X_raw_full) >= 5:
        try:
            preds = EllipticEnvelope(contamination=0.1,
                                     random_state=42).fit_predict(X_raw_full)
            df["_is_outlier"] = (preds == -1)
        except Exception as _exc_bare:
            pass

    train = df[~df["_is_outlier"]].copy()
    if len(train) < 4:
        train = df.copy()

    X = train[avail].fillna(0).values
    y = train["_target"].values

    # PCA enrichment (fitted on TRAIN only → transform all)
    if X.shape[0] >= 4 and X.shape[1] >= 2:
        sc_pca = StandardScaler()
        pca    = PCA(n_components=min(2, X.shape[1]))
        X_sc   = sc_pca.fit_transform(X)
        pc     = pca.fit_transform(X_sc)
        # Transform ALL rows with the train-fitted PCA
        all_X_sc = sc_pca.transform(df[avail].fillna(0).values)
        all_pc   = pca.transform(all_X_sc)
        df["PCA_1"] = all_pc[:, 0]
        df["PCA_2"] = all_pc[:, 1] if all_pc.shape[1] > 1 else 0.0

    # ── Feature scaling: fit on TRAIN only, transform all ────────────────────
    # Pre-scaling X replaces the broken Pipeline-in-VotingRegressor pattern.
    # VotingRegressor._validate_estimators() rejects sklearn Pipeline objects in
    # sklearn ≥1.4 unless _estimator_type is explicitly set (version-dependent).
    # Scaling once before fitting works correctly with ALL sklearn versions.
    _feat_scaler = RobustScaler()
    X_s     = _feat_scaler.fit_transform(X)                       # train, scaled
    all_X_s = _feat_scaler.transform(df[avail].fillna(0).values)  # all, scaled

    # Raw regressors — no Pipeline wrappers
    estimators = [
        ("rf",  RandomForestRegressor(n_estimators=200, max_depth=8,
                                      min_samples_leaf=2, random_state=42)),
        ("gbr", GradientBoostingRegressor(n_estimators=150, max_depth=4,
                                          learning_rate=0.05, random_state=42)),
        ("svr", SVR(kernel="rbf", C=10, epsilon=0.1)),
    ]
    if _HAS_XGB:
        estimators.append(
            ("xgb", xgb.XGBRegressor(n_estimators=150, max_depth=5,
                                      learning_rate=0.05,
                                      random_state=42, verbosity=0)))

    # Filter estimators: remove any that sklearn's VotingRegressor would reject.
    # XGBRegressor in some versions (2.0+) doesn't properly expose _estimator_type
    # to sklearn's _validate_estimators(), causing ValueError at fit time.
    try:
        from sklearn.base import is_regressor as _is_reg
        _valid = [(n, e) for n, e in estimators if _is_reg(e)]
        if len(_valid) < len(estimators):
            _skipped = [n for n, e in estimators if not _is_reg(e)]
            log.warning(f"  Excluded from VotingRegressor (sklearn type check): {_skipped}")
        estimators = _valid if _valid else estimators
    except Exception as _exc_bare:
        pass

    ensemble = VotingRegressor(estimators)

    # K-Fold CV
    nk = min(5, len(X_s))
    cv = KFold(n_splits=nk, shuffle=True, random_state=42)
    try:
        cvs    = cross_val_score(ensemble, X_s, y, cv=cv, scoring="r2", n_jobs=-1)
        cv_r2  = float(np.mean(cvs))
        cv_std = float(np.std(cvs))
        log.info(f"  K-Fold CV R²={cv_r2:.4f}±{cv_std:.4f}")
    except Exception as e:
        log.warning(f"  CV failed: {e}")
        cv_r2 = cv_std = 0.0

    # HPT — standalone RF on scaled data
    best_rf = RandomForestRegressor(n_estimators=200, max_depth=8,
                                    min_samples_leaf=2, random_state=42)
    try:
        gs = GridSearchCV(
            RandomForestRegressor(random_state=42),
            {"n_estimators": [100, 200], "max_depth": [5, 8, None]},
            cv=min(3, len(X_s)), scoring="r2", n_jobs=-1)
        gs.fit(X_s, y)
        best_rf = gs.best_estimator_
        log.info(f"  Best RF: {gs.best_params_}  R²={gs.best_score_:.4f}")
    except Exception as e:
        log.warning(f"  HPT failed: {e}")

    # Fit ensemble — graceful fallback to RF if VotingRegressor still fails
    try:
        ensemble.fit(X_s, y)
    except Exception as e:
        log.warning(f"  VotingRegressor.fit failed ({e}) — falling back to RandomForest")
        ensemble = best_rf
        ensemble.fit(X_s, y)

    yp   = ensemble.predict(X_s)
    rmse = float(np.sqrt(mean_squared_error(y, yp)))
    r2   = float(r2_score(y, yp))
    mae  = float(mean_absolute_error(y, yp))

    # ── LEAKAGE FIX: fit scaler on TRAINING predictions only ──────────────────
    mm = TrainAwareScaler(feature_range=(45, 98))
    train_preds = ensemble.predict(X_s)           # train rows only (scaled)
    mm.fit(train_preds)                            # scaler fitted on train preds only
    all_preds   = ensemble.predict(all_X_s)        # predict all rows (scaled)
    df["ML_Success_Probability"] = mm.transform(all_preds)  # transform, NOT fit
    log.info("  Scaler: fitted on train preds only → new data uses .transform()")

    # ── Resolve output directory FIRST — used by both SHAP and model save ──────
    run_id = run_id or datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    import sys as _sys
    _model_dir = Path("CEREBRO_RESULTS") / "models"
    try:
        _cp = _sys.modules.get("CEREBRO_Pipeline") or _sys.modules.get("src.core.pipeline")
        if _cp and hasattr(_cp, "PATHS") and "models" in _cp.PATHS:
            _model_dir = _cp.PATHS["models"]
    except Exception as _exc_bare:
        pass
    _model_dir.mkdir(parents=True, exist_ok=True)

    # SHAP — uses _model_dir (defined above)
    if _HAS_SHAP and len(X_s) >= 4:
        try:
            best_rf.fit(X_s, y)
            exp  = shap_lib.TreeExplainer(best_rf)
            vals = exp.shap_values(X_s)
            imp  = np.abs(vals).mean(axis=0)
            pd.DataFrame({"Feature": avail, "SHAP_Importance": imp}).sort_values(
                "SHAP_Importance", ascending=False).to_csv(
                    str(_model_dir / "shap_feature_importance.csv"), index=False)
        except Exception as e:
            log.warning(f"  SHAP: {e}")

    # Lipinski baseline
    if "LogP" in df.columns and "MW_Da" in df.columns:
        df["Lipinski_BBB_Pred"] = (
            df["LogP"].between(1, 3) & (df["MW_Da"] < 500)).astype(int)

    # Model persistence
    mpath = str(_model_dir / f"ensemble_{run_id}.pkl")
    if _HAS_JOBLIB:
        try:
            joblib.dump({
                "model":         ensemble,
                "scaler":        mm,           # TrainAwareScaler (not raw MM)
                "scaler_state":  mm.save_state(),
                "feat_scaler":   _feat_scaler,  # RobustScaler for feature pre-scaling
                "features":      avail,
                "run_id":        run_id,
                "r2":            r2,
                "rmse":          rmse,
                "train_size":    len(X),
                "fitted_at":     datetime.utcnow().isoformat(),
            }, mpath)
            log.info(f"  Model + leakage-free scaler saved → {mpath}")
        except Exception as e:
            log.warning(f"  Model save failed: {e}")

    metrics = {
        "r2": r2, "rmse": rmse, "mae": mae,
        "cv_r2": cv_r2, "cv_std": cv_std,
        "n_samples": len(X), "features": avail,
    }
    df.drop(columns=["_target"], errors="ignore", inplace=True)
    log.info(f"  ML (patched): R²={r2:.4f} RMSE={rmse:.4f}")
    return df, ensemble, metrics


# ─────────────────────────────────────────────────────────────────────────────
# PATCH 2: INFERENCE ENGINE (predict on NEW molecules without re-training)
# ─────────────────────────────────────────────────────────────────────────────

class InferenceEngine:
    """
    Loads a saved CEREBRO-X model and runs inference on NEW drug data.
    Uses .transform() ONLY — never re-fits the scaler.

    Usage:
        engine = InferenceEngine.load("CEREBRO_RESULTS/models/ensemble_20240101_120000.pkl")
        new_profile = {"MW_Da": 379.49, "LogP": 4.34, "Half_Life_Days": 3.0,
                       "Docking_Affinity_kcal": -8.5}
        score = engine.predict_single(new_profile)
        print(f"ML Success Probability: {score:.1f}%")
    """

    def __init__(self, model, scaler: TrainAwareScaler,
                 features: list[str], run_id: str = ""):
        self.model    = model
        self.scaler   = scaler
        self.features = features
        self.run_id   = run_id

    @classmethod
    def load(cls, pkl_path: str) -> "InferenceEngine":
        try:
            import joblib
            bundle = joblib.load(pkl_path)
        except Exception as e:
            raise RuntimeError(f"Cannot load model from {pkl_path}: {e}")

        # Restore TrainAwareScaler
        if "scaler_state" in bundle:
            scaler = TrainAwareScaler.from_state(bundle["scaler_state"])
        elif isinstance(bundle.get("scaler"), TrainAwareScaler):
            scaler = bundle["scaler"]
        else:
            # Legacy: raw MinMaxScaler — wrap it
            log.warning("  Legacy scaler detected — wrapping in TrainAwareScaler")
            scaler = TrainAwareScaler()
            scaler._scaler  = bundle["scaler"]
            scaler._fitted  = True
            scaler._train_min = 0.0
            scaler._train_max = 1.0

        return cls(model=bundle["model"], scaler=scaler,
                   features=bundle["features"], run_id=bundle.get("run_id",""))

    def predict_single(self, profile: dict[str, float]) -> float:
        """
        Predict ML_Success_Probability for ONE new molecule profile.
        Uses .transform() on the TRAINING scaler — NO re-fitting.
        """
        X = np.array([[profile.get(f, 0.0) for f in self.features]])
        raw_pred = self.model.predict(X)
        score    = self.scaler.transform(raw_pred)[0]
        return round(float(np.clip(score, 45, 98)), 2)

    def predict_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Predict ML_Success_Probability for a DataFrame of new molecules.
        Safe: uses .transform() only, scaler NOT re-fitted.
        """
        df = df.copy()
        avail = [f for f in self.features if f in df.columns]
        X     = df[avail].fillna(0).values
        raw   = self.model.predict(X)
        df["ML_Success_Probability"] = np.clip(
            self.scaler.transform(raw), 45, 98)
        return df


# ─────────────────────────────────────────────────────────────────────────────
# PATCH 3: EXCEL READER  (Template → dict → pipeline)
# ─────────────────────────────────────────────────────────────────────────────

class ExcelReader:
    """
    Reads CEREBRO_Input_Template.xlsx and returns:
      (drug_profile_dict, formulations_list)

    drug_profile_dict  → passed to cerebro_molecule_engine.analyze_molecule()
    formulations_list  → equivalent to YAML formulations section
    """

    DRUG_FIELD_MAP = {
        "Drug Name":                    "name",
        "Molecule Class":               "molecule_class",
        "Molecule Input (SMILES / FASTA / PDB / HELM / InChIKey)": "molecule_input",
        "Indication (Disease Target)":  "indication",
        "Target Protein":               "target_protein",
        "Target PDB ID":                "target_pdb_id",
        "Native BBB Penetration %":     "bbb_native_penetration_pct",
        "Clinical Phase":               "clinical_phase",
        "FDA Approval Date":            "fda_approval_date",
    }

    @classmethod
    def read(cls, xlsx_path: str) -> tuple[dict, list[dict]]:
        """
        Returns:
            drug_profile  : dict with drug identity (for MoleculeEngine)
            formulations  : list of dicts (one per DDS system)
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("pip install openpyxl")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        drug_profile = cls._read_drug_sheet(wb)
        formulations = cls._read_dds_sheet(wb)
        log.info(f"  Excel loaded: drug={drug_profile.get('name')}, "
                 f"formulations={len(formulations)}")
        return drug_profile, formulations

    # v18 FIX-1 — section-marker prefixes that must be skipped BEFORE empty-val check
    _SECTION_MARKER_PREFIXES: tuple = (
        "▶", "►", "●", "•", "─", "═",
        "Fill", "NOTE", "CEREBRO", "(fetched", "Format", "Brand",
    )
    # Regex: row[0] starts with a digit or arrow followed by capital letter = section header
    _MARKER_RE = re.compile(r"^[\d▶►●•\-–—]+[\s\.]*[A-Z]", re.UNICODE)
    # Regex: extract a clean INN/drug name from the value cell
    _DRUG_NAME_CLEAN_RE = re.compile(r"[^a-zA-Z0-9\-\s]")

    @classmethod
    def _is_section_marker(cls, label: str) -> bool:
        """Return True when a cell contains a layout/section marker, not a field name."""
        for prefix in cls._SECTION_MARKER_PREFIXES:
            if label.startswith(prefix):
                return True
        return bool(cls._MARKER_RE.match(label))

    @classmethod
    def _read_drug_sheet(cls, wb) -> dict:
        ws = wb["1_Drug_Input"]
        profile = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            cell0 = row[0]

            # ── [FIX-1-A] MARKER GUARD — evaluated BEFORE empty-val skip ─────
            # Rows like "▶  A.  Drug Identity" are section headers, not fields.
            label_raw = str(cell0).strip() if cell0 is not None else ""
            if cls._is_section_marker(label_raw):
                log.debug("  [parser] section marker skipped: %r", label_raw[:50])
                continue

            # ── [FIX-1-B] EMPTY-VAL SKIP — after marker guard ────────────────
            if not cell0 or not label_raw:
                continue

            # ── [FIX-1-C] REGEX FIELD EXTRACTION ─────────────────────────────
            label = label_raw
            value = row[1] if len(row) > 1 else None
            val_str = str(value).strip() if value is not None else ""

            _SKIP_VALS = {"", "nan", "none", "(fetched automatically)", "—", "-"}
            if val_str.lower() in _SKIP_VALS:
                continue

            key = cls.DRUG_FIELD_MAP.get(label, label.lower().replace(" ", "_"))

            # For the drug name field specifically: strip stray unicode via regex
            if key == "name":
                val_str = cls._DRUG_NAME_CLEAN_RE.sub("", val_str).strip()
                if not val_str:
                    continue

            profile[key] = val_str
        return profile

    @classmethod
    def _read_dds_sheet(cls, wb) -> list[dict]:
        ws = wb["2_DDS_Formulations"]
        # Row 3 = headers
        headers = [str(c.value or "").strip()
                   for c in ws[3]]
        formulations = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0] or str(row[0]).strip() in ("", "(auto)"):
                continue
            rec = {}
            for i, h in enumerate(headers):
                v = row[i] if i < len(row) else None
                if v is not None and str(v).strip() not in ("", "(auto)"):
                    try:
                        rec[h] = float(v)
                    except (ValueError, TypeError):
                        rec[h] = str(v).strip()
                else:
                    rec[h] = None
            formulations.append(rec)
        return formulations

    @classmethod
    def to_yaml_dict(cls, drug_profile: dict,
                     formulations: list[dict]) -> dict:
        """Convert Excel data to the same dict structure as dds_config.yaml."""
        return {
            "drug": {
                "name":           drug_profile.get("name", ""),
                "molecule_class": drug_profile.get("molecule_class", ""),
                "smiles":         (drug_profile.get("molecule_input")
                                   if "smiles" in str(drug_profile.get("molecule_input","")).lower()
                                   or (drug_profile.get("molecule_input","") or "").startswith("C")
                                   else None),
                "fasta":          (drug_profile.get("molecule_input")
                                   if str(drug_profile.get("molecule_input","")).startswith(">")
                                   else None),
                "molecule_input": drug_profile.get("molecule_input"),
                "indication":     drug_profile.get("indication"),
                "target_protein": drug_profile.get("target_protein"),
                "target_pdb_id":  drug_profile.get("target_pdb_id"),
                "mw_da":          None,   # always fetch live
                "logp":           None,
                "half_life_days": None,
            },
            "formulations": formulations,
        }


# ─────────────────────────────────────────────────────────────────────────────
# PATCH 4: GIF / VIDEO ENGINE
# ─────────────────────────────────────────────────────────────────────────────

class AnimationEngine:
    """
    Generates animated GIFs and MP4 videos from matplotlib figures.
    All animations are also exportable as Base64 for REST API responses.

    Outputs (each with companion _DOCUMENTATION.txt):
      pkpd_animation.gif      — PK/PD kinetics animated curve build-up
      radar_animation.gif     — radar fingerprint rotating reveal
      bbb_score_animation.gif — BBB score bar chart animated ranking
      formulation_mp4.mp4     — optional video (requires imageio-ffmpeg)
    """

    def __init__(self, output_dir: Path):
        self.out = Path(output_dir) / "figures"
        self.out.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _write_doc(path: Path, doc: dict):
        sep = "="*70
        lines = [sep, "  CEREBRO-X |  FILE DOCUMENTATION",
                 f"  File      : {path.name}",
                 f"  Generated : {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}",
                 sep, ""]
        for title, key in [
            ("OVERVIEW","overview"),("SIGNIFICANCE","significance"),
            ("THEORETICAL & PRACTICAL SCIENCE","theoretical_science"),
            ("METHODOLOGY","methodology"),
            ("COMPUTATIONAL ARCHITECTURE","computational_architecture"),
        ]:
            body = doc.get(key,"")
            if body:
                lines += ["─"*70, f"  {title}", "─"*70, body.strip(), ""]
        lines.append(sep)
        (str(path)+"_DOCUMENTATION.txt").replace("//","/")
        with open(str(path)+"_DOCUMENTATION.txt","w",encoding="utf-8") as f:
            f.write("\n".join(lines))

    def pkpd_animation(self, df_pk: pd.DataFrame,
                       fps: int = 15) -> Path | None:
        """
        Animated GIF: PK/PD kinetics curves drawn incrementally.
        Shows each drug's concentration curve being drawn over time.
        """
        try:
            import imageio.v2 as imageio
        except ImportError:
            log.warning("  GIF skipped — install imageio: pip install imageio")
            return None

        drugs  = df_pk["Drug"].unique().tolist()
        t_vals = sorted(df_pk["Day"].unique())
        n_steps = min(60, len(t_vals))
        step_idx = [int(i*len(t_vals)/n_steps) for i in range(1, n_steps+1)]
        frames = []

        for si in step_idx:
            t_sub = t_vals[:si]
            fig, ax = plt.subplots(figsize=(10, 6))
            for drug in drugs:
                sub = df_pk[(df_pk["Drug"]==drug) & (df_pk["Day"].isin(t_sub))]
                if sub.empty: continue
                ax.plot(sub["Day"], sub["Concentration_Pct"],
                        label=drug, lw=2.5)
            ax.axhline(50, color="red", linestyle="--", lw=1.5,
                       label="Threshold 50%")
            ax.set_xlim(0, df_pk["Day"].max())
            ax.set_ylim(0, df_pk["Concentration_Pct"].max()*1.05)
            ax.set_title("Brain PK/PD Kinetics — Vexosome Release",
                         fontweight="bold")
            ax.set_xlabel("Days Post-Administration")
            ax.set_ylabel("Effective Brain Concentration (%)")
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)
            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=80)
            plt.close(fig)
            buf.seek(0)
            frames.append(imageio.imread(buf))

        out_path = self.out / "pkpd_animation.gif"
        imageio.mimsave(str(out_path), frames, fps=fps, loop=0)
        log.info(f"  PK/PD GIF → {out_path} ({len(frames)} frames)")
        self._write_doc(out_path, {
            "overview":
                "Animated GIF showing brain PK/PD concentration kinetics "
                "being drawn incrementally for each drug candidate.",
            "significance":
                "Communicates the temporal dynamics of drug exposure — "
                "how quickly each candidate rises and falls relative to "
                "the 50% therapeutic threshold.",
            "theoretical_science":
                "C(t) = C₀·e^(−kt),  k = ln2/t½,  C₀ = 100·(150kDa/MW_Da)\n"
                "One-compartment first-order elimination model. "
                "Animation built by incrementally revealing time points "
                "from t=0 to t=60 days in 60 equal steps.",
            "methodology":
                "1. Compute C(t) for each drug at 500 time points.\n"
                "2. For each animation frame, plot only t[0..i].\n"
                "3. Encode each matplotlib figure as PNG bytes.\n"
                "4. imageio.mimsave() assembles frames into GIF.",
            "computational_architecture":
                "matplotlib Agg backend · imageio v2 · io.BytesIO frame buffer.",
        })
        return out_path

    def bbb_score_animation(self, df_dds: pd.DataFrame,
                            fps: int = 8) -> Path | None:
        """
        Animated GIF: BBB Engineering Score bar chart revealed one bar at a time.
        Ranked from lowest to highest, building suspense toward the top candidate.
        """
        try:
            import imageio.v2 as imageio
        except ImportError:
            return None

        if "BBB_Engineering_Score" not in df_dds.columns:
            return None

        top = df_dds.nlargest(20, "BBB_Engineering_Score")
        sorted_df = top.sort_values("BBB_Engineering_Score", ascending=True)
        names  = sorted_df["Formulation_Name"].str[:25].tolist()
        scores = sorted_df["BBB_Engineering_Score"].tolist()
        frames = []

        for n in range(1, len(names)+1):
            fig, ax = plt.subplots(figsize=(11, 7))
            colours = ["#1B7A4A" if i == n-1 else "#0f2040"
                       for i in range(n)]
            bars = ax.barh(names[:n], scores[:n], color=colours, edgecolor="white")
            ax.set_xlim(0, 100)
            ax.set_xlabel("BBB Engineering Score (0–100)", fontsize=11)
            ax.set_title(f"CEREBRO-X  |  Top DDS Rankings  "
                         f"(showing {n}/{len(names)})",
                         fontweight="bold")
            for bar, val in zip(bars, scores[:n]):
                ax.text(val+0.5, bar.get_y()+bar.get_height()/2,
                        f"{val:.1f}", va="center", fontsize=8)
            ax.axvline(75, color="gold", linestyle="--", lw=1.5,
                       label="Target score ≥ 75")
            ax.legend(fontsize=8)
            ax.grid(True, axis="x", alpha=0.3)
            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=80)
            plt.close(fig)
            buf.seek(0)
            frames.append(imageio.imread(buf))

        # hold last frame
        frames.extend([frames[-1]] * (fps * 2))

        out_path = self.out / "bbb_score_animation.gif"
        imageio.mimsave(str(out_path), frames, fps=fps, loop=0)
        log.info(f"  BBB Score GIF → {out_path}")
        self._write_doc(out_path, {
            "overview":
                "Animated GIF revealing DDS formulations ranked by BBB Engineering "
                "Score, one bar at a time from lowest to highest.",
            "significance":
                "Visually communicates competitive advantage of top formulations "
                "for investor presentations and PI review.",
            "theoretical_science":
                "BBB Engineering Score (0–100) computed from Pardridge 2012 framework: "
                "size, zeta, PEGylation, ligand specificity, ApoE affinity, "
                "CARPA risk, liver off-target penalty, phase transition temperature.",
            "methodology":
                "1. Select top-20 by BBB score.\n"
                "2. Sort ascending (most suspenseful reveal).\n"
                "3. One frame per bar addition.\n"
                "4. Final frame held for 2 seconds.",
            "computational_architecture":
                "matplotlib barh · imageio.mimsave · io.BytesIO.",
        })
        return out_path

    def radar_animation(self, df_ml: pd.DataFrame,
                        fps: int = 10) -> Path | None:
        """
        Animated GIF: radar fingerprint chart rotating through each drug candidate.
        Each frame highlights one drug while dimming the others.
        """
        try:
            import imageio.v2 as imageio
        except ImportError:
            return None

        from sklearn.preprocessing import MinMaxScaler

        features = [c for c in ["Half_Life_Days","ML_Success_Probability",
                                 "Docking_Affinity_kcal","Binding_Affinity_kcal",
                                 "Estimated_Affinity_kcal"]
                    if c in df_ml.columns]
        if len(features) < 2 or "Drug" not in df_ml.columns:
            return None

        df_r = df_ml.drop_duplicates(subset=["Drug"]).copy()
        scaler = MinMaxScaler()
        df_r[features] = scaler.fit_transform(abs(df_r[features]))
        N = len(features)
        angles = [n/N*2*np.pi for n in range(N)] + [0]
        colours = ["#0f2040","#C9A84C","#0D6E6E","#F57C00","#1A7A4A",
                   "#C62828","#5C2D91","#1565C0"]
        frames = []

        for focus_idx, (_, focus_row) in enumerate(df_r.iterrows()):
            fig = plt.figure(figsize=(8, 8))
            ax  = plt.subplot(111, polar=True)
            for idx, (_, row) in enumerate(df_r.iterrows()):
                vals = row[features].tolist() + [row[features].iloc[0]]
                alpha = 1.0 if idx == focus_idx else 0.2
                lw    = 3.0 if idx == focus_idx else 1.0
                col   = colours[idx % len(colours)]
                ax.plot(angles, vals, color=col, lw=lw, alpha=alpha,
                        label=row["Drug"] if idx==focus_idx else "")
                ax.fill(angles, vals, color=col, alpha=0.15 if idx==focus_idx else 0.03)
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(features, fontsize=8)
            ax.set_title(f"Molecule Fingerprint  —  {focus_row['Drug']}",
                         fontweight="bold", pad=20)
            plt.legend(loc="upper right", bbox_to_anchor=(1.35, 1.15))
            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=80, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            frames.append(imageio.imread(buf))

        # Hold 1s per frame
        slow_frames = []
        for f in frames:
            slow_frames.extend([f] * fps)

        out_path = self.out / "radar_animation.gif"
        imageio.mimsave(str(out_path), slow_frames, fps=fps, loop=0)
        log.info(f"  Radar GIF → {out_path}")
        self._write_doc(out_path, {
            "overview":
                "Animated GIF cycling through each drug candidate's "
                "multi-attribute radar fingerprint.",
            "significance":
                "Shows how each drug's molecular 'shape' compares — "
                "ideal for at-a-glance comparison in presentations.",
            "theoretical_science":
                "Values Min-Max normalised across all candidates to [0,1]. "
                "Absolute values used so all axes are positive. "
                "Larger enclosed area = superior profile.",
            "methodology":
                "1. Normalise features (MinMaxScaler on training data).\n"
                "2. One frame per drug — highlighted drug at full opacity.\n"
                "3. All other drugs dimmed to alpha=0.2.\n"
                "4. Each frame held for 1 second.",
            "computational_architecture":
                "matplotlib polar projection · MinMaxScaler · imageio.",
        })
        return out_path


# ─────────────────────────────────────────────────────────────────────────────
# PATCH 5: BASE64 EXPORT (for REST API responses)
# ─────────────────────────────────────────────────────────────────────────────

def encode_file_base64(path: Path) -> str | None:
    """Read file and return Base64-encoded string for JSON transport."""
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return None

def collect_results_as_json(output_root: Path) -> dict:
    """
    Collect all pipeline outputs into a single JSON-serialisable dict.
    PNG/GIF files are Base64-encoded for direct embedding in API response.
    CSV files are included as list-of-dicts.

    This dict is what the /upload-formulation-excel endpoint returns to the web client.
    """
    result = {
        "generated_at": datetime.utcnow().isoformat(),
        "version":      "CEREBRO-X",
        "figures":      {},
        "data":         {},
        "reports":      {},
        "dds_analysis": {},
    }

    # ── Figures (PNG + GIF → Base64) ────────────────────────────────────────
    figs_dir = output_root / "figures"
    if figs_dir.exists():
        for ext in ("*.png", "*.gif", "*.mp4"):
            for fp in figs_dir.glob(ext):
                if "_DOCUMENTATION" in fp.name:
                    continue
                key = fp.stem
                result["figures"][key] = {
                    "filename": fp.name,
                    "format":   fp.suffix.lstrip("."),
                    "base64":   encode_file_base64(fp),
                    "size_kb":  round(fp.stat().st_size / 1024, 1),
                }

    # ── CSVs → list-of-dicts ────────────────────────────────────────────────
    for subdir, target in [("data", "data"), ("results", "data"),
                            ("dds_analysis", "dds_analysis")]:
        d = output_root / subdir
        if d.exists():
            for csv_file in d.glob("*.csv"):
                if "_DOCUMENTATION" in csv_file.name:
                    continue
                try:
                    df_csv = pd.read_csv(csv_file, nrows=200)
                    result["data"][csv_file.stem] = df_csv.to_dict(orient="records")
                except Exception as _exc_bare:
                    pass

    # ── Master report (text) ─────────────────────────────────────────────────
    rp = output_root / "reports" / "Master_Report.txt"
    if rp.exists():
        result["reports"]["master_report"] = rp.read_text(encoding="utf-8")

    # ── Project config (JSON) ────────────────────────────────────────────────
    cfg = output_root / "deliverable" / "project_config.json"
    if cfg.exists():
        try:
            result["project_config"] = json.loads(cfg.read_text())
        except Exception as _exc_bare:
            pass

    return result


# ─────────────────────────────────────────────────────────────────────────────
# PATCH 6: IN-MEMORY PDF REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_pdf_report(results_json: dict,
                        title: str = "CEREBRO-X Analysis Report") -> bytes:
    """
    Generate a PDF report IN MEMORY from the results JSON.
    Returns raw bytes — never written to disk.
    Sent directly to client via FastAPI StreamingResponse.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Image as RLImage
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        raise ImportError("pip install reportlab")

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Styles
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  fontSize=18, textColor=colors.HexColor("#0f2040"),
                                  spaceAfter=12)
    h1_style    = ParagraphStyle("H1", parent=styles["Heading1"],
                                  fontSize=13, textColor=colors.HexColor("#0D6E6E"),
                                  spaceAfter=6)
    body_style  = ParagraphStyle("Body", parent=styles["Normal"],
                                  fontSize=9, leading=13, spaceAfter=4)
    note_style  = ParagraphStyle("Note", parent=styles["Normal"],
                                  fontSize=8, textColor=colors.grey,
                                  leftIndent=20, spaceAfter=4)

    # Title page
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(
        f"Generated: {results_json.get('generated_at', '')}", note_style))
    story.append(Paragraph(
        f"Version: {results_json.get('version', '')}", note_style))
    story.append(Spacer(1, 0.5*cm))

    # Project config table
    cfg = results_json.get("project_config", {})
    if cfg:
        story.append(Paragraph("Executive Summary", h1_style))
        tdata = [[k, str(v)] for k, v in cfg.items()]
        t = Table(tdata, colWidths=[7*cm, 10*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0f2040")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.HexColor("#F5F5F5"), colors.white]),
            ("GRID",       (0,0), (-1,-1), 0.25, colors.lightgrey),
            ("LEFTPADDING",(0,0),(-1,-1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

    # Master report text
    master = results_json.get("reports", {}).get("master_report", "")
    if master:
        story.append(Paragraph("Full Pipeline Report", h1_style))
        # Only include first 3000 chars in PDF (rest available in text file)
        for para in master[:3000].split("\n"):
            if para.strip():
                story.append(Paragraph(para.strip(), body_style))
        story.append(Paragraph(
            "(Full report available in reports/Master_Report.txt)",
            note_style))
        story.append(Spacer(1, 0.4*cm))

    # Figures page (Base64 → PIL → ReportLab Image)
    figs = results_json.get("figures", {})
    if figs:
        story.append(PageBreak())
        story.append(Paragraph("Visualisations", h1_style))
        for fig_name, fig_data in list(figs.items())[:8]:  # max 8 figs in PDF
            b64 = fig_data.get("base64")
            fmt = fig_data.get("format","png")
            if not b64 or fmt not in ("png","jpg","jpeg"):
                continue
            try:
                img_bytes = base64.b64decode(b64)
                img_buf   = io.BytesIO(img_bytes)
                rl_img    = RLImage(img_buf, width=15*cm, height=9*cm)
                story.append(Paragraph(fig_name.replace("_"," ").title(),
                                        h1_style))
                story.append(rl_img)
                story.append(Spacer(1, 0.3*cm))
            except Exception as _exc_bare:
                pass

    # DDS top 10
    top10 = results_json.get("dds_analysis", {}).get("top10_formulations", [])
    if top10:
        story.append(PageBreak())
        story.append(Paragraph("Top 10 DDS Formulations", h1_style))
        if top10:
            keys = list(top10[0].keys())[:6]  # first 6 columns
            tdata = [keys] + [[str(r.get(k,"")) for k in keys] for r in top10[:10]]
            t = Table(tdata, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0f2040")),
                ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS", (0,1), (-1,-1),
                 [colors.HexColor("#F5F5F5"), colors.white]),
                ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
                ("LEFTPADDING",(0,0),(-1,-1), 4),
            ]))
            story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# APPLY PATCHES  (monkey-patch the pipeline module)
# ─────────────────────────────────────────────────────────────────────────────

def apply_patches(cp_module) -> None:
    """
    Apply all patches to the CEREBRO_Pipeline module.
    Call after: import CEREBRO_Pipeline as cp
    """
    # Patch AdvancedMLEngine.train with the leakage-free version
    cp_module.AdvancedMLEngine.train = classmethod(
        lambda cls, *a, **kw: patched_train(cls, *a, **kw))
    log.info("[PATCH] AdvancedMLEngine.train → leakage-free TrainAwareScaler")

    # Attach new engines to module namespace
    cp_module.TrainAwareScaler   = TrainAwareScaler
    cp_module.InferenceEngine    = InferenceEngine
    cp_module.ExcelReader        = ExcelReader
    cp_module.AnimationEngine    = AnimationEngine
    cp_module.collect_results_as_json = collect_results_as_json
    cp_module.generate_pdf_report     = generate_pdf_report
    log.info("[PATCH] InferenceEngine, ExcelReader, AnimationEngine, "
             "PDF + JSON exporters attached to CEREBRO_Pipeline module")


__all__ = [
    "AnimationEngine",
    "ExcelReader",
    "InferenceEngine",
    "TrainAwareScaler",
    "apply_patches",
    "collect_results_as_json",
    "generate_pdf_report",
    "patched_train",
]