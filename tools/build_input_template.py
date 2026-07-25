# -*- coding: utf-8 -*-
"""
================================================================================
CEREBRO-X |  build_input_template.py
================================================================================
Created by: Muhammad Talaat (BPharm, R&D Computational Lead) — CEREBRO-X

Generates the polished, dashboard-style researcher input Excel template:
    CEREBRO_Input_Template.xlsx

The template is fully backward-compatible with the existing parser
(`excel_to_yaml` in run.py): same field labels in column A, same DDS column
order in 2_DDS_Formulations row 3, dynamic Drug-N section detection.

Visual upgrades over the v18-v20 template:
  • Branded title band with project tagline
  • Colour-banded sections (Identity / Auto-fetched / Multi-drug)
  • Two-tier instruction strip directly under the header
  • Data-validation dropdowns for Molecule Class, Clinical Phase, Carrier
    Type, Release Kinetics, Surface Ligand
  • Cell-level comments on every field explaining what to enter
  • Frozen header rows and auto-filters on DDS sheet
  • Conditional formatting on numeric DDS columns to spot outliers
  • Banded row striping for readability
  • Print-ready page setup (A3 landscape with fitted scaling)

Run:
    python build_input_template.py

Output:
    ./CEREBRO_Input_Template.xlsx

Run the pipeline normally — the parser auto-detects the new template by
sheet name (it matches the existing schema exactly).
================================================================================
"""
from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              NamedStyle)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import (ColorScaleRule, CellIsRule, FormulaRule)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

# ── Brand palette  (CEREBRO-X — see cerebro_brand.py for the full spec) ──
# Excel-friendly variant: deep-space navy/gold for HEADERS, light tints for
# data cells (Excel is a working surface, not a presentation surface).
NAVY    = "0F2040"   # void panel — header band, section bars
TEAL    = "0D6E6E"   # neuro-positive — subtitle, success accents
GOLD    = "C9A84C"   # signature gold — primary accent
GOLD_L  = "F5EAC0"   # light gold tint for input-cell fills
TEAL_L  = "DCEFEF"   # light teal tint for auto-fetched cell fills
ROSE_L  = "F8DADA"   # light rose tint for biologic warning cells
GREY    = "F4F4F6"   # banded-row fill (off-white)
WHITE   = "FFFFFF"
DARK    = "060610"   # void base — used as a deep accent only
ACCENT  = GOLD       # alias for backward-compat
SKY     = TEAL_L     # alias for backward-compat
CREAM   = GOLD_L     # alias for backward-compat
MINT    = "DAF7DC"   # legacy banded fill
ROSE    = ROSE_L     # alias for backward-compat

# Borders
THIN  = Side(border_style="thin",  color="888888")
THICK = Side(border_style="medium", color=NAVY)

# ──────────────────────────────────────────────────────────────────────────
def _brand_title(ws, text: str, subtitle: str = "", colspan: int = 8) -> None:
    """
    Drop a polished, brand-aligned title band into rows 1–2 of a sheet.

    Row 1: navy panel background (#0F2040) with GOLD title text — matches
           the "deep-space + signature gold" brand spec.
    Row 2: teal background (#0D6E6E) with white italic subtitle.
    """
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=colspan)
    c = ws.cell(1, 1, text)
    c.font = Font(bold=True, size=20, color=GOLD, name="Calibri")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 42

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=colspan)
        s = ws.cell(2, 1, subtitle)
        s.font = Font(italic=True, size=10, color=WHITE, name="Calibri")
        s.fill = PatternFill("solid", fgColor=TEAL)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 22


def _section_header(ws, row: int, text: str, colour: str = SKY,
                     colspan: int = 4) -> None:
    """Draw a coloured section divider row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, size=12, color=DARK)
    c.fill = PatternFill("solid", fgColor=colour)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26


def _input_row(ws, row: int, label: str, example: str = "",
                required: str = "", note: str = "",
                comment_text: str = "") -> None:
    """
    Write a labeled-input row.
    Column A: label (parser key)
    Column B: empty input cell (yellow, with light border)
    Column C: example
    Column D: required flag
    Column E: notes (free text, parser ignores)
    """
    a = ws.cell(row, 1, label)
    a.font = Font(bold=True, size=10, color=DARK)
    a.fill = PatternFill("solid", fgColor=GREY)
    a.alignment = Alignment(vertical="center", indent=1)
    a.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    b = ws.cell(row, 2)
    b.fill = PatternFill("solid", fgColor=CREAM)
    b.alignment = Alignment(vertical="center", indent=1)
    b.border = Border(left=THICK, right=THICK, top=THIN, bottom=THIN)
    b.font = Font(bold=False, size=11, color=DARK)
    if comment_text:
        b.comment = Comment(comment_text, "CEREBRO-X")

    c = ws.cell(row, 3, example)
    c.font = Font(italic=True, size=9, color="9CA3AF")
    c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)

    d = ws.cell(row, 4, required)
    d.font = Font(size=9, color=("C62828" if "REQUIRED" in required.upper() else "9CA3AF"))
    d.alignment = Alignment(horizontal="center", vertical="center")

    if note:
        e = ws.cell(row, 5, note)
        e.font = Font(size=9, italic=True, color="9CA3AF")
        e.alignment = Alignment(vertical="center", wrap_text=True)


def _autofetch_row(ws, row: int, label: str, example: str = "",
                     comment_text: str = "") -> None:
    """Same layout but value column shows '(fetched automatically)'."""
    a = ws.cell(row, 1, label)
    a.font = Font(size=10, color="404040")
    a.fill = PatternFill("solid", fgColor=GREY)
    a.alignment = Alignment(vertical="center", indent=1)
    a.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    b = ws.cell(row, 2, "(fetched automatically)")
    b.fill = PatternFill("solid", fgColor=GREY)
    b.alignment = Alignment(vertical="center", indent=1)
    b.font = Font(italic=True, size=10, color="888888")
    b.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    if comment_text:
        b.comment = Comment(comment_text, "CEREBRO-X")

    c = ws.cell(row, 3, example)
    c.font = Font(italic=True, size=9, color="9CA3AF")
    c.alignment = Alignment(vertical="center", indent=1)


# ──────────────────────────────────────────────────────────────────────────
# SHEET 1 — Drug Input (dashboard polish)
# ──────────────────────────────────────────────────────────────────────────
def build_drug_input(ws, n_drug_slots: int = 3) -> None:
    ws.title = "1_Drug_Input"

    _brand_title(ws,
        "CEREBRO-X   ⟶   Drug Input",
        "Fill the yellow cells. Grey cells are auto-fetched by the pipeline. "
        "Section headers (▶) are read by the parser — leave structure intact.",
        colspan=5)

    # Header row
    hdrs = ["Field", "Your Input", "Format / Example", "Required",
            "Notes (optional, ignored by parser)"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(3, j, h)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=THICK, bottom=THICK)
    ws.row_dimensions[3].height = 24

    row = 4
    # ── Drug 1 (always primary) ──────────────────────────────────────────
    _section_header(ws, row, "  ▶  A.  Drug 1 — Primary Drug Identity (required)",
                     colour=SKY, colspan=5)
    row += 1
    _input_row(ws, row, "Drug Name", "e.g. Naloxegol",
                "YES — REQUIRED",
                "Generic or trade name",
                "Free-text generic or trade name. The pipeline uses this to "
                "fetch data from PubChem/ChEMBL/UniProt/DrugBank.")
    row += 1
    _input_row(ws, row, "Molecule Class",
                "small_molecule | biologic | peptide",
                "YES",
                "Determines which PK fallback class is used",
                "One of: small_molecule, biologic, peptide. "
                "If left blank, the pipeline assumes small_molecule.")
    # Data validation dropdown for Molecule Class
    dv_class = DataValidation(type="list",
        formula1='"small_molecule,biologic,peptide,monoclonal_antibody"',
        allow_blank=True, showDropDown=False)
    dv_class.add(f"B{row}")
    ws.add_data_validation(dv_class)
    row += 1
    _input_row(ws, row,
                "Molecule Input (SMILES / FASTA / PDB / HELM / InChIKey)",
                "SMILES: COc1cc2c… | FASTA: >seq\\nMVLS… | PDB: 2NAO",
                "STRONGLY RECOMMENDED",
                "Pipeline auto-detects the type",
                "Paste a SMILES string, a FASTA sequence (starting with '>'), "
                "a 4-char PDB ID, or HELM peptide notation. "
                "If left blank, the pipeline tries to resolve from Drug Name.")
    row += 1
    _input_row(ws, row, "Indication (Disease Target)", "Alzheimer's Disease",
                "Optional", "CNS focus recommended")
    row += 1
    _input_row(ws, row, "Target Protein", "Amyloid-β protofibrils", "Optional", "")
    row += 1
    _input_row(ws, row, "Target PDB ID", "2NAO", "Optional",
                "4-character RCSB PDB code")
    row += 1
    _input_row(ws, row, "Native BBB Penetration %", "0.1", "Optional",
                "If unknown, pipeline predicts via QSAR")
    row += 1
    _input_row(ws, row, "Clinical Phase", "4 | 3 | 2 | 1 | preclinical",
                "Optional", "")
    dv_phase = DataValidation(type="list",
        formula1='"4,3,2,1,preclinical"', allow_blank=True)
    dv_phase.add(f"B{row}")
    ws.add_data_validation(dv_phase)
    row += 1
    _input_row(ws, row, "FDA Approval Date", "2023-07-06 (YYYY-MM-DD)",
                "Optional", "")
    row += 1

    # ── Auto-fetched physchem (do not edit unless you have in-vitro values) ─
    _section_header(ws, row,
        "  ▶  B.  Drug 1 — Molecular Properties (auto-fetched; "
        "override with in-vitro values if available)",
        colour=GREY, colspan=5)
    row += 1
    AUTO_FIELDS = [
        ("MW (Da)",            "143379",  "Molecular weight in Daltons"),
        ("LogP",               "−0.7",    "Octanol/water partition coefficient"),
        ("Half-Life (days)",   "7.0",     "Elimination half-life in days"),
        ("H-Bond Donors",      "—",       "Lipinski H-bond donor count"),
        ("H-Bond Acceptors",   "—",       "Lipinski H-bond acceptor count"),
        ("TPSA (Å²)",          "—",       "Topological polar surface area"),
        ("pI",                 "—",       "Isoelectric point (biologics)"),
        ("Instability Index",  "—",       "Stability metric (biologics)"),
        ("UniProt ID",         "—",       "UniProt accession (biologics)"),
        ("LogBB",              "—",       "Brain/blood partition log ratio"),
        ("BBB Penetration %",  "—",       "Computed BBB permeability"),
    ]
    for label, ex, comment in AUTO_FIELDS:
        _autofetch_row(ws, row, label, f"Example: {ex}",
                        comment_text=(f"{comment}. "
                                      "If you have an in-vitro measurement, "
                                      "REPLACE the '(fetched automatically)' text "
                                      "with the value — the pipeline will treat it "
                                      "as a researcher override (Tier 0, 100% confidence)."))
        row += 1

    # ── Multi-drug optional slots ────────────────────────────────────────
    row += 1
    _section_header(ws, row,
        "  ▶  C.  Optional — Additional Drugs for Multi-Drug Comparison",
        colour=MINT, colspan=5)
    row += 1
    note_cell = ws.cell(row, 1,
        "If you enter Drug 2 / Drug 3 / etc. below, the pipeline runs the full "
        "62-principle pipeline for EACH drug separately, then produces a "
        "cross-drug comparison Excel. Leave blank to skip.")
    note_cell.font = Font(italic=True, size=9, color="9CA3AF")
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 30
    row += 1

    for n in range(2, n_drug_slots + 1):
        _section_header(ws, row, f"  Drug {n}  (OPTIONAL — leave blank to skip)",
                         colour=MINT, colspan=5)
        row += 1
        _input_row(ws, row, "Drug Name", f"e.g. Donepezil", "Optional", "")
        row += 1
        _input_row(ws, row, "Molecule Class",
                    "small_molecule | biologic | peptide", "Optional", "")
        dv = DataValidation(type="list",
            formula1='"small_molecule,biologic,peptide,monoclonal_antibody"',
            allow_blank=True)
        dv.add(f"B{row}"); ws.add_data_validation(dv)
        row += 1
        _input_row(ws, row,
                    "Molecule Input (SMILES / FASTA / PDB / InChIKey)",
                    "SMILES or sequence", "Optional", "")
        row += 1
        _input_row(ws, row, "Indication", "Alzheimer's Disease", "Optional", "")
        row += 1
        _input_row(ws, row, "Native BBB %", "8.0", "Optional", "")
        row += 1
        _input_row(ws, row, "Clinical Phase", "4 | 3 | 2 | 1 | preclinical",
                    "Optional", "")
        dv2 = DataValidation(type="list", formula1='"4,3,2,1,preclinical"',
                              allow_blank=True)
        dv2.add(f"B{row}"); ws.add_data_validation(dv2)
        row += 1

    # Footer note
    row += 1
    foot = ws.cell(row, 1,
        "⚡ For Drug 2+: MW, LogP, half-life, TPSA, H-bond donors/acceptors, "
        "pI, UniProt, LogBB, BBB% are ALL auto-fetched. Add them as override "
        "values only if you have measured them in-vitro for the same compound.")
    foot.font = Font(italic=True, size=9, color=ACCENT)
    foot.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 28

    # Column widths
    widths = [56, 38, 36, 18, 40]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    # Freeze panes
    ws.freeze_panes = "A4"
    # Hide gridlines
    ws.sheet_view.showGridLines = False
    # Print setup
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"


# ──────────────────────────────────────────────────────────────────────────
# SHEET 2 — DDS Formulations  (the big editable grid)
# ──────────────────────────────────────────────────────────────────────────
def build_dds_formulations(ws, n_rows: int = 100) -> None:
    ws.title = "2_DDS_Formulations"

    _brand_title(ws,
        "CEREBRO-X   ⟶   DDS Formulations",
        "Each row = one drug-delivery system. Fill the YELLOW columns; the "
        "pipeline computes the GREY columns from your specs.",
        colspan=21)

    # Header row (row 3 — parser hardcoded to expect headers here)
    HEADERS = [
        # Researcher fills these — first 21 columns
        "Formulation_ID", "Formulation_Name", "Carrier_Type", "Surface_Ligand",
        "Size_nm", "Zeta_Potential_mV", "PDI", "Elasticity_kPa",
        "Encapsulation_Efficiency_pct", "PEGylation_Degree_mol_pct",
        "PEG_Chain_Length_Da", "Release_Kinetics", "pH_Trigger",
        "Phase_Transition_Temp_C", "Surface_Ligand_Density_per_nm2",
        "Endosomal_Escape_Eff", "Drug_Loading_Pct",
        "CNS_Bioavailability_Pct", "Manufacturing_Cost_USD_per_mg",
        "Scale_Up_Readiness", "Notes",
        # Auto-computed by pipeline — columns 22+
        "BBB_Engineering_Score", "Toxicity_Index", "Tanimoto_Similarity",
        "Composite_Score", "Principle_Composite_Score", "Principle_Rank",
        "G1_CNS_Delivery_Score", "G2_Release_Score", "G3_Stability_Score",
        "G4_Safety_Score", "G5_Glymphatic_Score", "G6_Manufacturability_Score",
        "G7_DrugDDS_Fit_Score",
    ]
    RESEARCHER_COL_END = 21    # parser keys off this — keep as-is

    for j, h in enumerate(HEADERS, 1):
        c = ws.cell(3, j, h)
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid",
            fgColor=(NAVY if j <= RESEARCHER_COL_END else "5E5E5E"))
        c.alignment = Alignment(horizontal="center", wrap_text=True,
                                 vertical="center")
        c.border = Border(left=THIN, right=THIN, top=THICK, bottom=THICK)
    ws.row_dimensions[3].height = 42

    # Data validation for the carrier and release-kinetics columns
    dv_carrier = DataValidation(type="list",
        formula1='"liposome,plga,polymer,micelle,dendrimer,nanogel,solid_lipid,metallic,vexosome"',
        allow_blank=True)
    dv_carrier.add(f"C4:C{3 + n_rows}")
    ws.add_data_validation(dv_carrier)

    dv_release = DataValidation(type="list",
        formula1='"sustained,zero-order,first-order,burst,ph-responsive,thermo"',
        allow_blank=True)
    dv_release.add(f"L4:L{3 + n_rows}")
    ws.add_data_validation(dv_release)

    dv_scale = DataValidation(type="list",
        formula1='"lab,pilot,clinical,commercial"', allow_blank=True)
    dv_scale.add(f"T4:T{3 + n_rows}")
    ws.add_data_validation(dv_scale)

    # Researcher input columns: cream (yellow-ish) fill, light borders.
    # Auto-computed columns: light grey fill, italic.
    for r in range(4, 4 + n_rows):
        for j in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, j)
            cell.alignment = Alignment(vertical="center")
            if j <= RESEARCHER_COL_END:
                cell.fill = PatternFill("solid", fgColor=CREAM)
                cell.border = Border(left=THIN, right=THIN,
                                      top=THIN, bottom=THIN)
            else:
                cell.fill = PatternFill("solid", fgColor="ECECEC")
                cell.font = Font(italic=True, size=9, color="888888")
                cell.border = Border(left=THIN, right=THIN,
                                      top=THIN, bottom=THIN)

    # Add an example row (row 4) so the researcher sees the expected format.
    # Marked as "DELETE BEFORE RUN" via the Notes column — the row stays so
    # the researcher can copy-paste real data, but the warning is loud.
    EXAMPLE = ["EXAMPLE-DELETE", "Tf-Liposome-V1 (sample)", "liposome",
                "Transferrin", 100, -10.5, 0.18, 0.5,
                82, 5.0, 2000, "sustained", 6.5, 42, 0.8,
                0.6, 12, 25, 4.50, "pilot",
                "⚠ EXAMPLE ROW — DELETE THIS ROW BEFORE RUNNING THE PIPELINE"]
    for j, v in enumerate(EXAMPLE, 1):
        cell = ws.cell(4, j, v)
        # Make the example row visually distinctive — italic, bordered red
        cell.font = Font(italic=True, size=9, color="C62828")
        cell.border = Border(left=Side(border_style="dashed", color="C62828"),
                              right=Side(border_style="dashed", color="C62828"),
                              top=Side(border_style="dashed", color="C62828"),
                              bottom=Side(border_style="dashed", color="C62828"))

    # Conditional formatting on key numeric columns:
    # Size_nm column (E): green if 50–150, yellow if outside but <250, red if >250
    size_col = "E"
    rng_size = f"{size_col}5:{size_col}{3 + n_rows}"
    rule_size = ColorScaleRule(start_type="num", start_value=20, start_color="F8696B",
                                 mid_type="num", mid_value=100, mid_color="63BE7B",
                                 end_type="num", end_value=400, end_color="F8696B")
    ws.conditional_formatting.add(rng_size, rule_size)

    # Zeta column (F): green if -25..-5 or +5..+25, yellow otherwise
    rng_zeta = f"F5:F{3 + n_rows}"
    rule_zeta = ColorScaleRule(start_type="num", start_value=-50, start_color="F8696B",
                                 mid_type="num", mid_value=-10, mid_color="63BE7B",
                                 end_type="num", end_value=50, end_color="F8696B")
    ws.conditional_formatting.add(rng_zeta, rule_zeta)

    # EE column (I): green high, red low
    rng_ee = f"I5:I{3 + n_rows}"
    rule_ee = ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                               mid_type="num", mid_value=60, mid_color="FFEB84",
                               end_type="num", end_value=95, end_color="63BE7B")
    ws.conditional_formatting.add(rng_ee, rule_ee)

    # Auto-filter on the header row
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}{3 + n_rows}"

    # Column widths
    custom_widths = {
        1:14, 2:24, 3:14, 4:14, 5:10, 6:14, 7:8, 8:14,
        9:18, 10:16, 11:14, 12:16, 13:10, 14:14, 15:18,
        16:16, 17:14, 18:18, 19:18, 20:14, 21:36
    }
    for j in range(1, len(HEADERS) + 1):
        w = custom_widths.get(j, 14)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3


# ──────────────────────────────────────────────────────────────────────────
# SHEET 3 — Instructions (researcher onboarding)
# ──────────────────────────────────────────────────────────────────────────
def build_instructions(ws) -> None:
    ws.title = "3_Instructions"
    _brand_title(ws, "CEREBRO-X   ⟶   How to Use This Workbook",
                  "Three steps. Five minutes. The pipeline does the rest.",
                  colspan=4)

    sections = [
        ("STEP 1 — Fill Sheet 1 (Drug Identity)",
            "Open the '1_Drug_Input' tab. Fill the yellow cells for at least "
            "Drug 1. The most important cells are Drug Name and Molecule Input "
            "(SMILES/FASTA/PDB). Leave grey '(fetched automatically)' cells "
            "alone unless you have an in-vitro measurement to override."),
        ("STEP 2 — Fill Sheet 2 (DDS Formulations)",
            "Open the '2_DDS_Formulations' tab. Each row is one formulation. "
            "Fill the YELLOW columns (specs). Grey columns are computed by the "
            "pipeline. Add as many rows as you want — 1 to 1000+."),
        ("STEP 3 — Save and run the pipeline",
            "Save the file as 'CEREBRO_Input_<your-tag>.xlsx' and place it in "
            "the CEREBRO-X working directory. The pipeline picks it up "
            "automatically and produces a Completed-Data Excel workbook with "
            "every property resolved, a 62-principle C+ Flow analysis, and "
            "(if multi-drug) a cross-drug comparison report."),
        ("OUTPUT — What you'll get back",
            "Per drug: PDF report, HTML5 dashboard, Completed-Data Excel "
            "(every property with provenance), DDS×Principle matrix, "
            "Top-10 reasoning narrative, Class B Deep Validation sheet, "
            "Class C Translational sheet, and Top-N Fallback Audit. "
            "For multi-drug runs: a combined Completed Excel and a "
            "Cross-Drug Comparison with C+ Flow comparison sheets. "
            "All outputs include color-coded tier provenance — every value "
            "is traceable to its source."),
        ("THE 62-PRINCIPLE C+ FLOW — How the system makes a recommendation",
            "Class A (Surrogate, fast): all 62 principles are evaluated for "
            "every DDS in your input, producing a CNS-weighted composite score "
            "and a ranking. Class B (Deep Physics): the Top-1 DDS is then "
            "re-validated through full PBPK ODE solvers, FEP+, MM/GBSA, and "
            "Stokes-Einstein glymphatic kinetics. If the Top-1 fails the 70% "
            "validation threshold, the orchestrator falls back to Top-2, then "
            "Top-3 — with explicit failure & transition reasons recorded. "
            "Class C (Translational): only after a DDS passes deep validation "
            "do we generate the administrative deliverables — Pre-IND outline, "
            "FTO analysis, 21 CFR Part 11 audit, NIH grant outline, "
            "patentability score. All three classes appear in every output: "
            "Excel, PDF, HTML5 dashboard, and cross-drug comparison."),
        ("OVERRIDES — When you have in-vitro data",
            "For any '(fetched automatically)' cell, just type your in-vitro "
            "value over the placeholder text. The pipeline detects the "
            "override and tags it as Tier 0 (100% confidence, researcher "
            "in-vitro). Tier-6 (orange, class-mean) values in the output "
            "report are exactly what you should consider overriding next."),
        ("SUPPORT",
            "Every metric in the output is documented in the "
            "'Principle_Explanations' sheet of the Completed Excel — open "
            "it any time you want to know what a metric means or how it "
            "was computed."),
    ]

    row = 4
    for title, body in sections:
        # Title line
        c = ws.cell(row, 1, title)
        c.font = Font(bold=True, size=12, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=4)
        ws.row_dimensions[row].height = 24
        row += 1
        # Body
        b = ws.cell(row, 1, body)
        b.font = Font(size=10, color=DARK)
        b.fill = PatternFill("solid", fgColor=GREY)
        b.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=4)
        ws.row_dimensions[row].height = 70
        row += 2

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────
# SHEET 4 — Material Library (reference, not parsed)
# ──────────────────────────────────────────────────────────────────────────
def build_material_library(ws) -> None:
    ws.title = "4_Material_Library"
    _brand_title(ws, "CEREBRO-X   ⟶   Material Science Library",
                  "Reference table — materials commonly used in CNS DDS. "
                  "Use these names in the 'Carrier_Type' column of Sheet 2.",
                  colspan=6)

    hdrs = ["Material Name", "CAS Number", "MW (Da)", "LogP",
            "Tm (°C)", "Source Database / Notes"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(3, j, h)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    MATERIALS = [
        ("DSPC (lipid)",      "816-94-4",   "790.1",  "9.5",  "55",
            "Avanti Polar Lipids — liposomal carrier, neutral"),
        ("Cholesterol",       "57-88-5",    "386.65", "9.1",  "148",
            "Sigma — membrane stabilizer for liposomes"),
        ("DOPE",              "4004-05-1",  "744.0",  "10.2", "−16",
            "Avanti — fusogenic phospholipid for endosomal escape"),
        ("PLGA 50:50",        "26780-50-7", "30000",  "—",    "45-55",
            "Boehringer — biodegradable copolymer"),
        ("PEG-2000",          "25322-68-3", "2000",   "—",    "45",
            "Sigma — surface stealth coating"),
        ("Chitosan",          "9012-76-4",  "150000", "—",    "—",
            "Sigma — natural cationic polymer"),
        ("Albumin (HSA)",     "70024-90-7", "66500",  "—",    "—",
            "Sigma — biologic carrier"),
        ("Iron oxide (SPION)","1309-37-1",  "159.69", "—",    "1565",
            "Magnetic targeting nanoparticle"),
        ("PAMAM-G4",          "163442-67-9","14215",  "—",    "—",
            "Sigma — dendrimer scaffold"),
    ]
    for i, m in enumerate(MATERIALS, 4):
        for j, v in enumerate(m, 1):
            cell = ws.cell(i, j, v)
            cell.alignment = Alignment(vertical="center", indent=1)
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)

    widths = [22, 18, 12, 10, 10, 50]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────────────────────────────────
def build_template(output_path: Path = Path("CEREBRO_Input_Template.xlsx"),
                    n_drug_slots: int = 3,
                    n_dds_rows: int = 100) -> Path:
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet
    build_drug_input(wb.create_sheet(), n_drug_slots=n_drug_slots)
    build_dds_formulations(wb.create_sheet(), n_rows=n_dds_rows)
    build_instructions(wb.create_sheet())
    build_material_library(wb.create_sheet())
    wb.save(str(output_path))
    return output_path


if __name__ == "__main__":
    import sys
    _default = Path(__file__).resolve().parent.parent / "inputs" / "CEREBRO_Input_Template.xlsx"
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else _default
    output.parent.mkdir(parents=True, exist_ok=True)
    p = build_template(output)
    print(f"✅ Built polished template → {p}  ({p.stat().st_size:,} bytes)")
